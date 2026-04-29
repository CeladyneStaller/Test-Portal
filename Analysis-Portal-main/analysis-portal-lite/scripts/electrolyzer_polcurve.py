#!/usr/bin/env python3
"""
PEM Electrolyzer Polarization Curve Analyzer
=============================================
Loads raw potentiostat CSV data (e.g. Biologic, Gamry, Squidstat),
auto-detects voltage and current columns, extracts representative
data from each dwell period, identifies polcurve cycles, and plots
each cycle as a separate V vs j curve.

Usage:
  python electrolyzer_polcurve.py                  # interactive
  python electrolyzer_polcurve.py --file data.csv  # direct
"""

import argparse, csv, gc
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from pathlib import Path


def run(input_dir: str, output_dir: str, params: dict = None) -> dict:
    """Analyze PEM electrolyzer polarization curve data."""
    import traceback as tb
    import os
    p = params or {}
    inp = Path(input_dir)
    out = Path(output_dir)

    geo_area = float(p.get('geo_area', 5.0))
    T_C = float(p.get('T_C', 80.0))
    p_cath = float(p.get('p_cathode_barg', 0.0))
    p_an = float(p.get('p_anode_barg', 0.0))
    cell_id = p.get('cell_id', 'a1')
    eis_ref_v = p.get('eis_ref_voltage', '')
    eis_ref_v = float(eis_ref_v) if eis_ref_v else None

    # Find all data files recursively
    all_files = sorted([f for f in inp.rglob("*")
                        if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv', '.fcd')])

    if not all_files:
        return {"status": "error", "message": "No data files found"}

    # Try scan_folder approach: find the actual data directory
    # (could be input_dir itself or a subfolder if a folder was uploaded)
    scan_dirs = set(f.parent for f in all_files)

    polcurve_file = None
    eis_files = []

    for d in scan_dirs:
        try:
            pf, ef = scan_folder(str(d), cell_id=cell_id)
            if pf:
                polcurve_file = pf
                eis_files = ef
                break
        except Exception:
            pass

    # Fallback: if scan_folder didn't find anything, use heuristics
    if polcurve_file is None:
        if len(all_files) == 1:
            polcurve_file = str(all_files[0])
        else:
            # Largest file is likely the polcurve
            polcurve_file = str(max(all_files, key=lambda f: f.stat().st_size))
            eis_files = [str(f) for f in all_files if str(f) != polcurve_file]

    from scripts.helpers.conditions import img_ext_from_params
    image_ext = img_ext_from_params(p)

    try:
        analyze(polcurve_file, geo_area=geo_area, save_dir=str(out),
                title=Path(polcurve_file).stem, T_C=T_C,
                p_cathode_barg=p_cath, p_anode_barg=p_an,
                eis_files=eis_files if eis_files else None,
                eis_ref_voltage=eis_ref_v, image_ext=image_ext)
        plt.close('all')
    except Exception as e:
        raise RuntimeError(
            f"Analysis failed for {Path(polcurve_file).name}: {e}\n{tb.format_exc()}"
        )

    # Verify expected outputs
    ext = image_ext or 'png'
    EXPECTED = {'analysis_data.xlsx'}
    if image_ext:
        EXPECTED.update({f'polcurve.{ext}', f'j_vs_cycle.{ext}',
                        f'model_fit.{ext}', f'nyquist.{ext}',
                        f'losses_vs_cycle.{ext}', f'ir_correction.{ext}'})
    output_files = [f.name for f in out.iterdir()
                    if f.is_file() and f.name in EXPECTED]

    if not output_files:
        raise RuntimeError(
            f"Analysis produced no output files. "
            f"Polcurve: {Path(polcurve_file).name}, "
            f"EIS files: {len(eis_files)}"
        )

    return {
        "status": "success",
        "polcurve_file": Path(polcurve_file).name,
        "eis_files": len(eis_files),
        "files_produced": output_files,
    }


# ═══════════════════════════════════════════════════════════════════
#  Column detection
# ═══════════════════════════════════════════════════════════════════

def _clean_path(p):
    p = p.strip()
    if p.startswith('& '): p = p[2:]
    return p.strip().strip('"').strip("'").strip('\u2018\u2019\u201c\u201d\u202a\u200b')


def _match_col(header, candidates):
    """Return the first column name in header that matches any candidate (case-insensitive substring)."""
    hl = header.lower()
    for c in candidates:
        if c in hl:
            return True
    return False


def detect_columns(fieldnames):
    """
    Auto-detect voltage, current, step, and repeat columns.

    Returns dict with keys: 'v_col', 'i_col', 'step_col', 'repeat_col', 'step_name_col'
    (values are column name strings or None).
    """
    result = {k: None for k in ('v_col', 'i_col', 'step_col', 'repeat_col', 'step_name_col', 'time_col')}

    v_candidates = ['working electrode (v)', 'voltage (v)', 'v_cell', 'vcell',
                     'e_stack', 'ewe (v)', 'potential (v)', 'working electrode vs']
    # Match "Current (A)" but not "Current Density"
    i_candidates = ['current (a)', 'i (a)', 'current(a)']
    step_candidates = ['step number', 'step_number', 'step no', 'ns']
    repeat_candidates = ['repeats', 'repeat', 'cycle number', 'cycle']
    step_name_candidates = ['step name', 'step_name', 'technique']
    time_candidates = ['elapsed time', 'time (s)', 'elapsed_time', 'time(s)']

    for fn in fieldnames:
        fl = fn.lower().strip()

        # Voltage — but skip "current density" columns that contain "a/m"
        if result['v_col'] is None:
            for c in v_candidates:
                if c in fl:
                    result['v_col'] = fn
                    break

        # Current — must NOT be "current density"
        if result['i_col'] is None and 'density' not in fl:
            for c in i_candidates:
                if c in fl:
                    result['i_col'] = fn
                    break

        if result['step_col'] is None:
            for c in step_candidates:
                if c in fl:
                    result['step_col'] = fn
                    break

        if result['repeat_col'] is None:
            for c in repeat_candidates:
                if c in fl:
                    result['repeat_col'] = fn
                    break

        if result['step_name_col'] is None:
            for c in step_name_candidates:
                if c in fl:
                    result['step_name_col'] = fn
                    break

        if result['time_col'] is None:
            for c in time_candidates:
                if c in fl:
                    result['time_col'] = fn
                    break

    return result


# ═══════════════════════════════════════════════════════════════════
#  Data loading
# ═══════════════════════════════════════════════════════════════════

def load_data(filepath):
    """
    Load CSV/TSV, keeping only the columns needed for analysis.

    Streams line-by-line to avoid loading the entire file into memory.

    Returns
    -------
    data : dict of column_name → numpy array
    fieldnames : list of all column name strings (for detection)
    """
    path = Path(_clean_path(filepath))
    if not path.exists():
        raise FileNotFoundError(
            f"File not found: {path}\n"
            f"  On Windows: right-click → Copy as path, then paste.")

    # Read header to detect delimiter and columns
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        first_line = f.readline()

    delim = '\t' if '\t' in first_line else ','
    fn = [c.strip() for c in first_line.strip().split(delim)]

    # Detect which columns we actually need
    cols = detect_columns(fn)
    keep = set(v for v in cols.values() if v is not None)

    # Map needed column names to their index in each row
    col_idx = {name: i for i, name in enumerate(fn) if name in keep}

    # Stream file, only storing needed columns
    raw = {name: [] for name in col_idx}
    n_rows = 0

    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        next(f)  # skip header
        for line in f:
            parts = line.strip().split(delim)
            if len(parts) < max(col_idx.values()) + 1:
                continue
            for name, idx in col_idx.items():
                raw[name].append(parts[idx].strip())
            n_rows += 1

    # Convert to numpy
    data = {}
    for name, vals in raw.items():
        try:
            data[name] = np.array([float(x) if x.lower() != 'nan' else np.nan
                                   for x in vals])
        except ValueError:
            data[name] = np.array(vals)
    del raw

    print(f"  Loaded {n_rows:,} rows from '{path.name}'  ({len(col_idx)} of {len(fn)} columns kept)")

    return data, fn


# ═══════════════════════════════════════════════════════════════════
#  Control mode detection
# ═══════════════════════════════════════════════════════════════════

def detect_control_mode(data, cols):
    """
    Auto-detect whether the polcurve is potentiostatic or galvanostatic.

    Strategy:
      1. If step_name column exists and contains 'constant potential' or
         'constant current', use that.
      2. Otherwise, compare the staircase-ness of V vs I: the controlled
         variable will have discrete steps with low intra-step variance,
         while the response variable will drift/ramp within each step.

    Returns 'potentiostatic' or 'galvanostatic'.
    """
    # Method 1: step name column
    if cols.get('step_name_col') is not None:
        names = data[cols['step_name_col']]
        unique_lower = set(n.lower().strip() for n in names)
        has_potential = any('constant potential' in n for n in unique_lower)
        has_current = any('constant current' in n for n in unique_lower)
        if has_potential and not has_current:
            return 'potentiostatic'
        if has_current and not has_potential:
            return 'galvanostatic'
        if has_potential and has_current:
            # Both present — the dominant type (more segments) is the polcurve
            # The minority type is typically recovery/baseline holds
            n_potential = sum(1 for n in names if 'constant potential' in n.lower())
            n_current = sum(1 for n in names if 'constant current' in n.lower())
            if n_current > n_potential:
                return 'galvanostatic'
            else:
                return 'potentiostatic'

    # Method 2: compare staircase quality of V vs I
    V = data[cols['v_col']]
    I = data[cols['i_col']]

    v_step = _detect_step(V)
    i_step = _detect_step(I)

    if v_step is not None and i_step is None:
        return 'potentiostatic'
    if i_step is not None and v_step is None:
        return 'galvanostatic'

    # Both have steps — compare coefficient of variation within segments
    # The controlled variable will have lower CV within each step
    def _intra_step_cv(signal, n_seg=50):
        """Compute mean coefficient of variation within equal-length segments."""
        clean = signal[~np.isnan(signal)]
        if len(clean) < n_seg * 10:
            return 1.0
        seg_len = len(clean) // n_seg
        cvs = []
        for i in range(n_seg):
            seg = clean[i * seg_len:(i + 1) * seg_len]
            mu = np.mean(seg)
            if abs(mu) > 1e-9:
                cvs.append(np.std(seg) / abs(mu))
        return np.median(cvs) if cvs else 1.0

    cv_v = _intra_step_cv(V)
    cv_i = _intra_step_cv(I)

    # The controlled variable has lower intra-segment variation
    if cv_v < cv_i:
        return 'potentiostatic'
    else:
        return 'galvanostatic'


# ═══════════════════════════════════════════════════════════════════
#  Dwell extraction
# ═══════════════════════════════════════════════════════════════════

def extract_dwells_from_steps(data, cols, geo_area, mode='potentiostatic'):
    """
    Use step/repeat columns to define dwells. Extract representative
    (V, j) from the stable tail of each dwell.

    mode: 'potentiostatic' (V controlled, I response),
          'galvanostatic' (I controlled, V response), or
          'dual' (extract both, tagged with dwell['mode'])

    Returns list of dicts: [{'V': float, 'j': float, 'step': int, 'repeat': int,
                             'mode': str, 't_mid': float}, ...]
    """
    V_raw = data[cols['v_col']]
    I_raw = data[cols['i_col']]
    step = data[cols['step_col']].astype(int)
    repeat = data[cols['repeat_col']].astype(int)
    has_time = cols['time_col'] is not None
    if has_time:
        T_raw = data[cols['time_col']]

    has_names = cols['step_name_col'] is not None
    if has_names:
        names = data[cols['step_name_col']]

    # Identify unique (step, repeat) segments in time order
    segments = []
    cur_key = (step[0], repeat[0])
    seg_start = 0

    for i in range(1, len(step)):
        key = (step[i], repeat[i])
        if key != cur_key:
            segments.append((cur_key, seg_start, i))
            cur_key = key
            seg_start = i
    segments.append((cur_key, seg_start, len(step)))

    # For dual mode, detect the mode of each segment from step name
    # For single mode, filter to matching step names
    dwells = []
    for (s, r), start, end in segments:
        n_pts = end - start

        # Skip very short segments (< 10 pts)
        if n_pts < 10:
            continue

        # Determine the mode for this segment
        if mode == 'dual' and has_names:
            sname = names[start].lower().strip()
            if 'constant current' in sname:
                seg_mode = 'galvanostatic'
            elif 'constant potential' in sname:
                seg_mode = 'potentiostatic'
            else:
                continue  # skip unknown step types
        elif mode == 'dual':
            seg_mode = 'potentiostatic'  # fallback
        else:
            seg_mode = mode
            # Filter by step name for single mode
            if has_names:
                sname = names[start].lower().strip()
                if mode == 'galvanostatic' and 'constant current' not in sname:
                    continue
                if mode == 'potentiostatic' and 'constant potential' not in sname:
                    continue

        V_seg = V_raw[start:end]
        I_seg = I_raw[start:end]

        if seg_mode == 'galvanostatic':
            # Current is controlled → mean of segment
            I_clean = I_seg[~np.isnan(I_seg)]
            if len(I_clean) == 0:
                continue
            j_rep = np.mean(I_clean) / geo_area

            # Voltage is response → find stable tail
            n = len(V_seg)
            stable_start = n - 1
            for k in range(n - 2, max(0, n // 2) - 1, -1):
                tail = V_seg[k:]
                tail_clean = tail[~np.isnan(tail)]
                if len(tail_clean) < 2:
                    continue
                t_mean = np.mean(tail_clean)
                t_std = np.std(tail_clean)
                if t_std > max(0.03 * abs(t_mean), 0.001):
                    break
                stable_start = k

            n_tail = min(20, n - stable_start)
            V_tail = V_seg[n - n_tail:n]
            V_tail = V_tail[~np.isnan(V_tail)]
            if len(V_tail) == 0:
                continue
            V_sp = np.mean(V_tail)
        else:
            # Potentiostatic: Voltage is controlled → mean of segment
            V_sp = np.nanmean(V_seg)

            # Current is response → find stable tail
            n = len(I_seg)
            stable_start = n - 1
            for k in range(n - 2, max(0, n // 2) - 1, -1):
                tail = I_seg[k:]
                tail_clean = tail[~np.isnan(tail)]
                if len(tail_clean) < 2:
                    continue
                t_mean = np.mean(tail_clean)
                t_std = np.std(tail_clean)
                if t_std > max(0.03 * abs(t_mean), 0.001):
                    break
                stable_start = k

            n_tail = min(20, n - stable_start)
            tail_sl = slice(n - n_tail, n)
            I_tail = I_seg[tail_sl]
            I_tail = I_tail[~np.isnan(I_tail)]
            if len(I_tail) == 0:
                continue

            I_rep = np.mean(I_tail)
            j_rep = I_rep / geo_area

        # For dual/potentiostatic mode, skip recovery holds:
        # segments with only 1 distinct setpoint and short duration
        if seg_mode == 'potentiostatic' and mode == 'dual':
            # Recovery holds have very low |j| at a fixed V
            if abs(j_rep) < 0.001 and n_pts < 200:
                continue

        dwells.append({
            'V': V_sp,
            'j': j_rep,
            'step': s,
            'repeat': r,
            'n_pts': n_pts,
            't_mid': float(np.nanmean(T_raw[start:end])) if has_time else None,
            'mode': seg_mode,
        })

    return dwells


def extract_dwells_generic(data, cols, geo_area, mode='potentiostatic'):
    """
    Fallback when no step/repeat columns exist.
    Uses stability grouping on the controlled variable.

    mode: 'potentiostatic' (group by V stability) or
          'galvanostatic' (group by I stability)
    """
    V_raw = data[cols['v_col']]
    I_raw = data[cols['i_col']]

    # Choose grouping signal based on mode
    if mode == 'galvanostatic':
        group_signal = I_raw
    else:
        group_signal = V_raw

    # Detect step size from the controlled variable
    step_size = _detect_step(group_signal)

    if step_size and step_size > 0:
        af, rf = step_size * 0.25, 0.01
    else:
        ds = np.abs(np.diff(group_signal[~np.isnan(group_signal)]))
        ne = np.percentile(ds, 50) if len(ds) > 0 else 0.01
        af, rf = max(ne * 10, 0.003), 0.03

    signal = group_signal
    grps, gs, gm, gc = [], 0, signal[0], 1
    for i in range(1, len(signal)):
        if np.isnan(signal[i]):
            continue
        if gc == 0:
            gs, gm, gc = i, signal[i], 1
            continue
        mu = gm / gc
        if abs(signal[i] - mu) <= max(af, rf * abs(mu)):
            gm += signal[i]; gc += 1
        else:
            grps.append((gs, i))
            gs, gm, gc = i, signal[i], 1
    if gc > 0:
        grps.append((gs, len(signal)))

    dwells = []
    for start, end in grps:
        n = end - start
        if n < 10:
            continue

        if mode == 'galvanostatic':
            # Current is controlled → mean of segment
            I_seg = I_raw[start:end]
            I_clean = I_seg[~np.isnan(I_seg)]
            if len(I_clean) == 0:
                continue
            j_rep = np.mean(I_clean) / geo_area

            # Voltage is response → stable tail
            V_seg = V_raw[start:end]
            n_tail = min(20, n)
            V_tail = V_seg[n - n_tail:n]
            V_tail = V_tail[~np.isnan(V_tail)]
            if len(V_tail) == 0:
                continue
            V_rep = np.mean(V_tail)
        else:
            # Voltage is controlled → mean of segment
            V_rep = np.nanmean(V_raw[start:end])

            # Current is response → stable tail
            I_seg = I_raw[start:end]
            n_tail = min(20, n)
            I_tail = I_seg[n - n_tail:n]
            I_tail = I_tail[~np.isnan(I_tail)]
            if len(I_tail) == 0:
                continue
            j_rep = np.mean(I_tail) / geo_area

        dwells.append({
            'V': V_rep,
            'j': j_rep,
            'step': 0,
            'repeat': len(dwells),
            'n_pts': n,
        })

    return dwells


def _detect_step(signal):
    """Auto-detect step size via histogram peak spacing."""
    s = signal[~np.isnan(signal)]
    if len(s) < 100:
        return None
    span = s.max() - s.min()
    if span < 1e-6:
        return None
    bw = max(span / 200, 1e-4)
    bins = np.arange(s.min() - bw, s.max() + 2 * bw, bw)
    if len(bins) < 10:
        return None
    hist, edges = np.histogram(s, bins=bins)
    ctrs = (edges[:-1] + edges[1:]) / 2
    thr = np.percentile(hist[hist > 0], 50)
    pks = [ctrs[i] for i in range(1, len(hist) - 1)
           if hist[i] > thr and hist[i] >= hist[i - 1] and hist[i] >= hist[i + 1]]
    if len(pks) < 4:
        return None
    pks = np.sort(pks)
    gaps = np.diff(pks)
    gr = np.round(gaps / bw) * bw
    ug, uc = np.unique(gr, return_counts=True)
    m = ug > bw * 0.5
    return float(ug[m][np.argmax(uc[m])]) if m.any() else None


# ═══════════════════════════════════════════════════════════════════
#  Cycle detection
# ═══════════════════════════════════════════════════════════════════

def detect_cycles(dwells, mode='potentiostatic'):
    """
    Group dwells into polcurve cycles.

    For potentiostatic mode, cycle boundaries are detected from voltage
    resets. For galvanostatic mode, from current density resets.

    A cycle is a sequence of dwells with monotonically changing setpoints.
    Dwells below the sweep range (e.g. recovery holds) are treated as
    cycle boundaries.

    Returns list of cycles, each a list of dwell dicts sorted by V.
    """
    if not dwells:
        return []

    # Choose the controlled variable for cycle detection
    if mode == 'galvanostatic':
        setpoints = np.array(sorted(set(round(d['j'], 4) for d in dwells)))
        get_sp = lambda d: d['j']
    else:
        setpoints = np.array(sorted(set(round(d['V'], 3) for d in dwells)))
        get_sp = lambda d: d['V']

    if len(setpoints) < 3:
        return [sorted(dwells, key=lambda d: d['V'])]

    # For potentiostatic mode, find boundary between recovery holds and
    # polcurve setpoints (e.g., 1.25V recovery vs 1.40-1.80V sweep).
    # For galvanostatic mode, skip boundary detection — current setpoints
    # span a log scale where gap-based boundary finding doesn't work,
    # and recovery holds are already filtered by step name.
    if mode == 'galvanostatic':
        sp_boundary = setpoints.min() - 0.001  # keep everything
    else:
        gaps = np.diff(setpoints)
        median_gap = np.median(gaps)
        largest_gap_idx = np.argmax(gaps)
        largest_gap = gaps[largest_gap_idx]

        if largest_gap > max(3.0 * median_gap, 0.05):
            sp_boundary = setpoints[largest_gap_idx] + largest_gap * 0.5
        else:
            sp_boundary = setpoints.min() - 0.01

    # Determine sweep range (above boundary)
    pc_sp = setpoints[setpoints > sp_boundary]
    if len(pc_sp) < 2:
        pc_sp = setpoints
    sp_lo = pc_sp.min()
    sp_hi = pc_sp.max()
    sp_span = sp_hi - sp_lo

    # Walk through dwells, grouping into cycles.
    # For potentiostatic: step changes and voltage resets mark boundaries.
    # For galvanostatic: only current density resets mark boundaries
    #   (step numbers often change every dwell since each setpoint is a
    #    separate step in the potentiostat program).
    cycles = []
    current_cycle = []

    for i, d in enumerate(dwells):
        # Skip baseline/recovery dwells
        if get_sp(d) < sp_boundary:
            if len(current_cycle) >= 3:
                cycles.append(current_cycle)
            current_cycle = []
            continue

        if current_cycle:
            prev = current_cycle[-1]

            if mode == 'galvanostatic':
                # Galvanostatic: only use j reset as cycle boundary
                if get_sp(d) < get_sp(prev) - 0.30 * sp_span:
                    if len(current_cycle) >= 3:
                        cycles.append(current_cycle)
                    current_cycle = []
            else:
                # Potentiostatic: step changes and V resets
                if d['step'] != prev['step']:
                    if len(current_cycle) >= 3:
                        cycles.append(current_cycle)
                    current_cycle = []
                elif get_sp(d) < get_sp(prev) - 0.30 * sp_span:
                    if len(current_cycle) >= 3:
                        cycles.append(current_cycle)
                    current_cycle = []

        current_cycle.append(d)

    # Flush last cycle
    if len(current_cycle) >= 3:
        cycles.append(current_cycle)

    # Sort each cycle by voltage ascending
    for cyc in cycles:
        cyc.sort(key=lambda d: d['V'])

    return cycles


def detect_cycles_dual(dwells):
    """
    For dual-mode data: split dwells by mode, detect cycles independently,
    then merge and number chronologically by elapsed time.

    Returns list of cycles (each a list of dwell dicts sorted by V).
    Each cycle's dwells carry a 'cycle_label' tag ('CC' or 'CP').
    The returned list is ordered by median elapsed time of each cycle.
    """
    # Split by mode
    galv_dwells = [d for d in dwells if d.get('mode') == 'galvanostatic']
    pots_dwells = [d for d in dwells if d.get('mode') == 'potentiostatic']

    # Detect cycles in each set independently
    galv_cycles = detect_cycles(galv_dwells, mode='galvanostatic') if galv_dwells else []
    pots_cycles = detect_cycles(pots_dwells, mode='potentiostatic') if pots_dwells else []

    # Tag cycles and compute timing
    tagged = []
    for cyc in galv_cycles:
        t_mid = _cycle_time(cyc)
        for d in cyc:
            d['cycle_label'] = 'CC'
        tagged.append((t_mid, cyc))
    for cyc in pots_cycles:
        t_mid = _cycle_time(cyc)
        for d in cyc:
            d['cycle_label'] = 'CP'
        tagged.append((t_mid, cyc))

    # Sort by median elapsed time for chronological numbering
    tagged.sort(key=lambda x: x[0] if x[0] is not None else 0)
    cycles = [cyc for _, cyc in tagged]

    print(f"    Galvanostatic (CC): {len(galv_cycles)} cycles from {len(galv_dwells)} dwells")
    print(f"    Potentiostatic (CP): {len(pots_cycles)} cycles from {len(pots_dwells)} dwells")

    return cycles


def _cycle_time(cyc):
    """Median elapsed time for a cycle's dwells."""
    times = [d['t_mid'] for d in cyc if d.get('t_mid') is not None]
    return float(np.median(times)) if times else None

def plot_cycles(cycles, geo_area, title=None, save_path=None):
    """
    Plot all polcurve cycles on one figure.
    V on y-axis, j on x-axis, one curve per cycle.
    """
    n_cyc = len(cycles)
    if n_cyc == 0:
        print("  No cycles to plot.")
        return

    fig, ax = plt.subplots(figsize=(10, 6), dpi=120)
    cmap = plt.cm.viridis(np.linspace(0, 0.9, n_cyc))

    for i, cyc in enumerate(cycles):
        j_arr = np.array([d['j'] for d in cyc])
        V_arr = np.array([d['V'] for d in cyc])
        mode_tag = cyc[0].get('cycle_label', '') if cyc else ''
        label = f'Cycle {i + 1}'
        if mode_tag:
            label += f' ({mode_tag})'
        marker = 's' if mode_tag == 'CP' else 'o'
        ax.plot(j_arr, V_arr, marker=marker, linestyle='-', ms=4, lw=1.2,
                color=cmap[i], label=label)

    ax.set_xlabel('Current density  j  [A/cm²]', fontsize=12)
    ax.set_ylabel('Cell voltage  V  [V]', fontsize=12)
    ax.set_xlim(left=0)
    ax.grid(True, alpha=0.3)

    # Smart legend: if many cycles, put colorbar instead
    if n_cyc <= 15:
        ax.legend(fontsize=8, loc='upper left', ncol=max(1, n_cyc // 8))
    else:
        sm = plt.cm.ScalarMappable(cmap=plt.cm.viridis,
                                    norm=plt.Normalize(1, n_cyc))
        sm.set_array([])
        cbar = fig.colorbar(sm, ax=ax, pad=0.02)
        cbar.set_label('Cycle number', fontsize=11)

    ttl = title or f'Electrolyzer Polarization  ({n_cyc} cycles, {geo_area:.1f} cm²)'
    ax.set_title(ttl, fontsize=12, fontweight='bold')

    plt.tight_layout()
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()

    return fig


def select_analysis_voltage(cycles, candidates=None, min_coverage=0.6):
    """
    Choose the primary analysis voltage based on cycle coverage.

    If more than (1 - min_coverage) of cycles don't reach the first
    candidate, fall back to the next. Returns the first candidate
    with sufficient coverage.

    Parameters
    ----------
    cycles : list of cycles
    candidates : list of voltages in preference order (default [1.8, 1.7])
    min_coverage : fraction of cycles that must reach the voltage

    Returns
    -------
    v_primary : float — the chosen analysis voltage
    """
    if candidates is None:
        candidates = [1.8, 1.7, 1.6]

    n_total = len(cycles)
    if n_total == 0:
        return candidates[0]

    for v in candidates:
        n_reach = sum(1 for cyc in cycles if max(d['V'] for d in cyc) >= v - 0.02)
        frac = n_reach / n_total
        if frac >= min_coverage:
            return v

    # Nothing met the threshold — use the lowest candidate
    return candidates[-1]


def extract_j_at_voltage(cycles, v_target, v_tol=0.015, interpolate=True):
    """
    Extract current density at a target voltage from each cycle.

    For measured data: uses the dwell closest to v_target (within v_tol).
    For interpolated data (when interpolate=True): if v_target falls within
    the cycle's voltage range but no dwell is within v_tol, linearly
    interpolates j from the two surrounding dwells.

    Returns
    -------
    cycle_nums : array of cycle numbers (1-indexed)
    j_values   : array of j at v_target for each cycle
    is_interp  : array of bool — True if interpolated, False if measured
    """
    cycle_nums = []
    j_values = []
    is_interp = []

    for i, cyc in enumerate(cycles):
        V_arr = np.array([d['V'] for d in cyc])
        j_arr = np.array([d['j'] for d in cyc])
        is_galv = cyc[0].get('mode', '') == 'galvanostatic'

        # Sort by voltage for interpolation
        order = np.argsort(V_arr)
        V_sorted = V_arr[order]
        j_sorted = j_arr[order]

        if is_galv and interpolate and len(V_sorted) >= 2:
            # Galvanostatic: j values are discrete setpoints, so a "direct
            # match" always returns the same setpoint and hides degradation.
            # Always interpolate to get a continuous j(V) value.
            if V_sorted.min() <= v_target <= V_sorted.max():
                j_interp = float(np.interp(v_target, V_sorted, j_sorted))
                cycle_nums.append(i + 1)
                j_values.append(j_interp)
                is_interp.append(True)
        else:
            # Potentiostatic or non-interpolating: check for direct match first
            dv = np.abs(V_sorted - v_target)
            best_idx = np.argmin(dv)

            if dv[best_idx] <= v_tol:
                cycle_nums.append(i + 1)
                j_values.append(j_sorted[best_idx])
                is_interp.append(False)
            elif interpolate and V_sorted.min() < v_target < V_sorted.max():
                j_interp = float(np.interp(v_target, V_sorted, j_sorted))
                cycle_nums.append(i + 1)
                j_values.append(j_interp)
                is_interp.append(True)
        # else: target outside cycle range — skip

    return np.array(cycle_nums), np.array(j_values), np.array(is_interp)


# ── Galvanostatic analysis helpers ────────────────────────────────

def select_analysis_current(cycles, candidates=None, min_coverage=0.6):
    """
    Choose the primary analysis current density for galvanostatic data.
    Picks the highest j that ≥ min_coverage fraction of cycles reach.
    """
    if candidates is None:
        candidates = [5.0, 4.5, 4.0, 3.5, 3.0, 2.5, 2.0, 1.5, 1.0]
    n_total = len(cycles)
    if n_total == 0:
        return candidates[-1]
    for j in candidates:
        n_reach = sum(1 for cyc in cycles
                      if max(d['j'] for d in cyc) >= j * 0.95)
        if n_reach / n_total >= min_coverage:
            return j
    return candidates[-1]


def extract_v_at_current(cycles, j_target, j_tol=0.005, interpolate=True):
    """
    Extract voltage at a target current density from each cycle.

    For galvanostatic data, j values are discrete setpoints so direct
    matching works well. Falls back to interpolation if needed.

    Returns
    -------
    cycle_nums : array of cycle numbers (1-indexed)
    v_values   : array of V at j_target
    is_interp  : array of bool
    """
    cycle_nums = []
    v_values = []
    is_interp = []

    for i, cyc in enumerate(cycles):
        j_arr = np.array([d['j'] for d in cyc])
        V_arr = np.array([d['V'] for d in cyc])
        order = np.argsort(j_arr)
        j_sorted = j_arr[order]
        V_sorted = V_arr[order]

        dj = np.abs(j_sorted - j_target)
        best_idx = np.argmin(dj)

        if dj[best_idx] <= j_tol * j_target + 0.001:
            cycle_nums.append(i + 1)
            v_values.append(V_sorted[best_idx])
            is_interp.append(False)
        elif interpolate and j_sorted.min() < j_target < j_sorted.max():
            v_interp = float(np.interp(j_target, j_sorted, V_sorted))
            cycle_nums.append(i + 1)
            v_values.append(v_interp)
            is_interp.append(True)

    return np.array(cycle_nums), np.array(v_values), np.array(is_interp)


def detect_stabilization(cycle_nums, j_values, n_stable=5,
                         n_lookback=10, threshold_pct=1.0):
    """
    Detect the cycle at which performance stabilizes.

    At each candidate cycle i, the reference is the mean of the
    preceding `n_lookback` cycles. Stabilization is the first cycle
    where j for that cycle and the next `n_stable`−1 cycles are all
    within `threshold_pct`% of that rolling reference.

    Example with defaults: stable at cycle 20 means cycles 20–24
    are each within 1% of mean(cycles 10–19).

    Parameters
    ----------
    cycle_nums : array of cycle numbers
    j_values   : array of j at each cycle
    n_stable   : consecutive cycles that must be within band (default 3)
    n_lookback : number of preceding cycles for rolling reference (default 10)
    threshold_pct : max deviation from reference as % (default 0.5)

    Returns
    -------
    stable_cycle : int or None
    """
    n = len(j_values)
    if n < n_lookback + n_stable:
        return None

    frac = threshold_pct / 100.0

    for i in range(n_lookback, n - n_stable + 1):
        j_ref = np.mean(j_values[i - n_lookback:i])
        if abs(j_ref) < 1e-9:
            continue
        band = frac * abs(j_ref)

        # Check if cycles i, i+1, ..., i+n_stable-1 are all within band
        window = j_values[i:i + n_stable]
        if np.all(np.abs(window - j_ref) <= band):
            return int(cycle_nums[i])

    return None


def plot_j_vs_cycle(cycles, v_targets, save_path=None):
    """
    Plot current density at specified voltages vs cycle number,
    with a vertical indicator where performance stabilizes.

    Parameters
    ----------
    cycles : list of cycle dicts
    v_targets : list of voltages to track (e.g. [1.8, 1.7])
    save_path : str or None
    """
    fig, ax = plt.subplots(figsize=(9, 5), dpi=120)
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd']
    markers = ['o', 's', '^', 'D', 'v']

    stable_cycles = []

    for k, vt in enumerate(v_targets):
        cn, jv, interp = extract_j_at_voltage(cycles, vt)
        if len(cn) == 0:
            print(f"  No data at V = {vt:.3f} V")
            continue
        c = colors[k % len(colors)]
        m = markers[k % len(markers)]

        # Plot measured points (solid)
        meas = ~interp
        if meas.any():
            ax.plot(cn[meas], jv[meas], marker=m, linestyle='none', color=c,
                    ms=5, label=f'j @ {vt:.2f} V')
        # Plot interpolated points (open markers)
        if interp.any():
            ax.plot(cn[interp], jv[interp], marker=m, linestyle='none', color=c,
                    ms=5, fillstyle='none', markeredgewidth=1.2,
                    label=f'j @ {vt:.2f} V (interp)')
        # Connecting line through all points
        ax.plot(cn, jv, '-', color=c, lw=1.0, alpha=0.5)

        # Detect stabilization
        sc = detect_stabilization(cn, jv)
        if sc is not None:
            stable_cycles.append((vt, sc))
            sc_idx = np.where(cn == sc)[0]
            if len(sc_idx) > 0:
                ax.plot(sc, jv[sc_idx[0]], '*', color=c, ms=14,
                        markeredgecolor='k', markeredgewidth=0.8, zorder=6)
            print(f"  Stabilization at {vt:.2f} V: cycle {sc}")

    ax.set_xlabel('Cycle number', fontsize=12)
    ax.set_ylabel('Current density  j  [A/cm²]', fontsize=12)
    ax.ticklabel_format(axis='y', useOffset=False)
    ax.set_xlim(left=0)
    ax.grid(True, alpha=0.3)
    ax.set_title('Current Density vs. Cycle', fontsize=12, fontweight='bold')

    handles, labels = ax.get_legend_handles_labels()
    fig.legend(handles, labels,
               loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=len(handles), fontsize=9,
               frameon=True, fancybox=True)

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()

    return fig


# ── Galvanostatic-specific plots ──────────────────────────────────

def plot_v_vs_cycle(cycles, j_targets, save_path=None):
    """
    Plot voltage at specified current densities vs cycle number.
    Galvanostatic counterpart to plot_j_vs_cycle.
    """
    fig, ax = plt.subplots(figsize=(9, 5), dpi=120)
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd']
    markers = ['o', 's', '^', 'D', 'v']

    stable_cycles = []

    for k, jt in enumerate(j_targets):
        cn, vv, interp = extract_v_at_current(cycles, jt)
        if len(cn) == 0:
            print(f"  No data at j = {jt:.3f} A/cm²")
            continue
        c = colors[k % len(colors)]
        m = markers[k % len(markers)]

        meas = ~interp
        if meas.any():
            ax.plot(cn[meas], vv[meas], marker=m, linestyle='none', color=c,
                    ms=5, label=f'V @ {jt:.2f} A/cm²')
        if interp.any():
            ax.plot(cn[interp], vv[interp], marker=m, linestyle='none', color=c,
                    ms=5, fillstyle='none', markeredgewidth=1.2,
                    label=f'V @ {jt:.2f} A/cm² (interp)')
        ax.plot(cn, vv, '-', color=c, lw=1.0, alpha=0.5)

        sc = detect_stabilization(cn, vv)
        if sc is not None:
            stable_cycles.append((jt, sc))
            sc_idx = np.where(cn == sc)[0]
            if len(sc_idx) > 0:
                ax.plot(sc, vv[sc_idx[0]], '*', color=c, ms=14,
                        markeredgecolor='k', markeredgewidth=0.8, zorder=6)
            print(f"  Stabilization at {jt:.2f} A/cm²: cycle {sc}")

    ax.set_xlabel('Cycle number', fontsize=12)
    ax.set_ylabel('Cell voltage  V  [V]', fontsize=12)
    ax.ticklabel_format(axis='y', useOffset=False)
    ax.set_xlim(left=0)
    ax.grid(True, alpha=0.3)
    ax.set_title('Voltage vs. Cycle', fontsize=12, fontweight='bold')

    handles, labels = ax.get_legend_handles_labels()
    fig.legend(handles, labels, loc='lower center',
               bbox_to_anchor=(0.5, -0.02), ncol=min(len(handles), 4),
               fontsize=9, frameon=True, fancybox=True)

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


def plot_v_and_hfr_vs_cycle(cycles, j_targets, eis_mapped, save_path=None):
    """
    Dual-axis: V at j targets (left) and ASR (R₀+R₁) (right) vs cycle.
    Galvanostatic counterpart to plot_j_and_hfr_vs_cycle.
    """
    fig, ax1 = plt.subplots(figsize=(10, 5.5), dpi=120)
    ax2 = ax1.twinx()

    colors_v = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
    markers_v = ['o', 's', '^', 'D']

    for k, jt in enumerate(j_targets):
        cn, vv, interp = extract_v_at_current(cycles, jt)
        if len(cn) == 0:
            continue
        c = colors_v[k % len(colors_v)]
        m = markers_v[k % len(markers_v)]

        meas = ~interp
        if meas.any():
            ax1.plot(cn[meas], vv[meas], marker=m, linestyle='none', color=c,
                     ms=5, label=f'V @ {jt:.2f} A/cm²')
        if interp.any():
            ax1.plot(cn[interp], vv[interp], marker=m, linestyle='none', color=c,
                     ms=5, fillstyle='none', markeredgewidth=1.2,
                     label=f'V @ {jt:.2f} A/cm² (interp)')
        ax1.plot(cn, vv, '-', color=c, lw=1.0, alpha=0.5)

    ax1.set_xlabel('Cycle number', fontsize=12)
    ax1.set_ylabel('Cell voltage  V  [V]', fontsize=12, color='#1f77b4')
    ax1.tick_params(axis='y', labelcolor='#1f77b4')
    ax1.ticklabel_format(axis='y', useOffset=False)
    ax1.set_xlim(left=0)
    ax1.grid(True, alpha=0.3)

    # Right axis: ASR
    eis_cn = [e['cycle'] for e in eis_mapped]
    eis_asr = [e['asr_mohm_cm2'] for e in eis_mapped]
    ax2.plot(eis_cn, eis_asr, 'v--', color='#9467bd', ms=5, lw=1.2,
             label='ASR (R₀+R₁)', alpha=0.8)
    ax2.set_ylabel('ASR (R₀+R₁)  [mΩ·cm²]', fontsize=12, color='#9467bd')
    ax2.tick_params(axis='y', labelcolor='#9467bd')

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    fig.legend(lines1 + lines2, labels1 + labels2,
               loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=min(len(lines1) + len(lines2), 5), fontsize=9,
               frameon=True, fancybox=True)

    ax1.set_title('Voltage & ASR vs. Cycle', fontsize=12, fontweight='bold')
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


def extract_losses_at_current(cycles, j_target, T_C=80.0,
                               p_cathode_barg=0.0, p_anode_barg=0.0,
                               fix_ASR=None):
    """
    Fit each complete cycle and extract V + loss breakdown at j_target.
    Galvanostatic counterpart to extract_losses_vs_cycle.
    """
    from scipy.optimize import least_squares

    T_K = T_C + 273.15
    E = E_rev(T_C, p_cathode_barg, p_anode_barg)

    max_pts = max(len(c) for c in cycles) if cycles else 0
    min_pts = max(5, int(max_pts * 0.8))

    cycle_nums = []
    v_values = []
    losses = {'eta_anode_mV': [], 'eta_cathode_mV': [],
              'V_ohmic_mV': [], 'V_mt_mV': []}

    n_fittable = sum(1 for c in cycles if len(c) >= min_pts)
    print(f"\n  Fitting {n_fittable} cycles "
          f"for loss tracking at {j_target:.2f} A/cm²...")

    for ci, cyc in enumerate(cycles):
        if len(cyc) < min_pts:
            continue

        j_arr = np.array([d['j'] for d in cyc])
        V_arr = np.array([d['V'] for d in cyc])

        # Check cycle reaches the target j
        if j_arr.max() < j_target * 0.95:
            continue

        # Filter j > 0
        mask = j_arr > 0
        j_fit = j_arr[mask]
        V_fit = V_arr[mask]
        if len(j_fit) < 3:
            continue

        def model(j, x):
            return _electrolyzer_model(j, x, E, T_K)

        x0 = [70.0, -7.0, 0.5, -3.0, 0.0]
        lo = [10.0, -12.0, 0.2, -6.0, 0.0]
        hi = [500.0, -3.0, 2.0, -0.5, 0.05]
        if fix_ASR is not None:
            x0[0], lo[0], hi[0] = fix_ASR, fix_ASR - 0.01, fix_ASR + 0.01

        try:
            res = least_squares(lambda x: model(j_fit, x) - V_fit, x0,
                                bounds=(lo, hi), method='trf',
                                loss='soft_l1', f_scale=0.01)
        except Exception:
            continue

        if not res.success:
            continue

        xf = res.x
        j0a, j0c = 10**xf[1], 10**xf[3]
        ba = (_R * T_K) / (xf[2] * _n_e * _F)
        bc = (_R * T_K) / (0.5 * _n_e * _F)

        # Compute losses at j_target
        eta_a = ba * np.log10(j_target / j0a) if j_target > j0a else 0.0
        eta_c = bc * np.log10(j_target / j0c) if j_target > j0c else 0.0
        v_ohm = j_target * xf[0] / 1000.0
        v_mt = xf[4] * j_target**2

        # Get V at j_target (from model)
        V_at_j = float(model(np.array([j_target]), xf)[0])

        cycle_nums.append(ci + 1)
        v_values.append(V_at_j)
        losses['eta_anode_mV'].append(eta_a * 1000)
        losses['eta_cathode_mV'].append(eta_c * 1000)
        losses['V_ohmic_mV'].append(v_ohm * 1000)
        losses['V_mt_mV'].append(v_mt * 1000)

    for k in losses:
        losses[k] = np.array(losses[k])
    print(f"    {len(cycle_nums)} cycles fitted successfully")

    return cycle_nums, v_values, losses


def compute_eis_loss_decomposition(cycles, eis_circuit_by_cycle, j_target,
                                    T_C=80.0, p_cathode_barg=0.0,
                                    p_anode_barg=0.0, geo_area=5.0):
    """
    Decompose polcurve losses using EIS circuit fit R0, R1, R2 values.

    For each cycle that has current-dependent EIS circuit fits,
    interpolates R0(j), R1(j), R2(j) at j_target and computes:
      V_R0  = j × R0      (membrane ohmic)
      V_R1  = j × R1      (HF arc / charge transfer)
      V_CL  = j × [√(R1·R2)·coth(√(R2/R1)) - R1]  (CL ionic, transmission line)
      V_kin = V - E_rev - V_R0 - V_R1 - V_CL  (kinetic / activation)

    Parameters
    ----------
    cycles : list of polcurve cycles
    eis_circuit_by_cycle : dict {cycle_idx: list of fit result dicts}
    j_target : float — target current density for loss evaluation
    T_C, p_cathode_barg, p_anode_barg : thermodynamic conditions
    geo_area : float — cell area cm²

    Returns
    -------
    cycle_nums, dep_values, losses  (same format as extract_losses_at_current)
    losses keys: 'V_R0_mV', 'V_R1_mV', 'V_CL_mV', 'V_kinetic_mV'
    """
    E = E_rev(T_C, p_cathode_barg, p_anode_barg)

    cycle_nums = []
    dep_values = []
    losses = {'V_R0_mV': [], 'V_R1_mV': [], 'V_CL_mV': [], 'V_kinetic_mV': []}

    for ci, cyc in enumerate(cycles):
        if ci not in eis_circuit_by_cycle:
            continue

        fits = eis_circuit_by_cycle[ci]
        if len(fits) < 1:
            continue

        # Get j and R values from circuit fits
        j_eis = np.array([f['j'] for f in fits if f.get('j') is not None])
        R0_eis = np.array([f['R0_ohm'] for f in fits if f.get('j') is not None])
        R1_eis = np.array([f['R1_ohm'] for f in fits if f.get('j') is not None])
        R2_eis = np.array([f['R2_ohm'] for f in fits if f.get('j') is not None])

        if len(j_eis) < 1:
            continue

        # Use lowest-j EIS fit (constant R values)
        idx_min = np.argmin(j_eis)
        R0_j = R0_eis[idx_min]
        R1_j = R1_eis[idx_min]
        R2_j = R2_eis[idx_min]

        # Compute voltage losses: V = j (A/cm²) × R (Ω) × A (cm²)
        V_R0 = j_target * R0_j * geo_area
        V_R1 = j_target * R1_j * geo_area
        V_CL = j_target * coth_cl_ionic_only(R1_j, R2_j) * geo_area  # coth model

        # Get V at j_target from polcurve data
        j_arr = np.array([d['j'] for d in cyc])
        V_arr = np.array([d['V'] for d in cyc])
        order_v = np.argsort(j_arr)
        j_sorted = j_arr[order_v]
        V_sorted = V_arr[order_v]

        if j_sorted.max() < j_target * 0.95:
            continue

        if j_sorted.min() <= j_target <= j_sorted.max():
            V_at_j = float(np.interp(j_target, j_sorted, V_sorted))
        else:
            continue

        V_kinetic = V_at_j - E - V_R0 - V_R1 - V_CL

        cycle_nums.append(ci + 1)
        dep_values.append(V_at_j)
        losses['V_R0_mV'].append(V_R0 * 1000)
        losses['V_R1_mV'].append(V_R1 * 1000)
        losses['V_CL_mV'].append(V_CL * 1000)
        losses['V_kinetic_mV'].append(V_kinetic * 1000)

    losses = {k: np.array(v) for k, v in losses.items()}

    if cycle_nums:
        print(f"    EIS-based loss decomposition: {len(cycle_nums)} cycle(s)")

    return cycle_nums, dep_values, losses


def plot_v_and_losses_vs_cycle(cycle_nums, v_values, losses,
                                j_target, save_path=None):
    """
    Dual-axis: V at target j (left) and losses (right) vs cycle.
    Galvanostatic counterpart to plot_j_and_losses_vs_cycle.
    """
    fig, ax1 = plt.subplots(figsize=(10, 6), dpi=120)
    ax2 = ax1.twinx()

    ax1.plot(cycle_nums, v_values, 'o-', color='#1f77b4', ms=5, lw=1.5,
             label=f'V @ {j_target:.2f} A/cm²')
    ax1.set_xlabel('Cycle number', fontsize=12)
    ax1.set_ylabel(f'Cell voltage at {j_target:.2f} A/cm²  [V]',
                   fontsize=12, color='#1f77b4')
    ax1.tick_params(axis='y', labelcolor='#1f77b4')
    ax1.ticklabel_format(axis='y', useOffset=False)
    ax1.set_xlim(left=0)

    loss_styles = [
        ('V_ohmic_mV',      'Ohmic',            '#4CAF50', 's-'),
        ('eta_anode_mV',    'η anode (OER)',     '#FF5722', '^-'),
        ('eta_cathode_mV',  'η cathode (HER)',   '#FF9800', 'v-'),
        ('V_mt_mV',         'Mass transport',    '#9C27B0', 'D-'),
    ]

    for key, label, color, fmt in loss_styles:
        vals = losses[key]
        if len(vals) > 0 and np.max(vals) < 0.1:
            continue
        ax2.plot(cycle_nums, vals, fmt, color=color, ms=4, lw=1.2,
                 label=label, alpha=0.85)

    eta_total = losses['eta_anode_mV'] + losses['eta_cathode_mV']
    ax2.plot(cycle_nums, eta_total, 'p-', color='#d62728', ms=5, lw=1.5,
             label='η kinetic (total)', alpha=0.85)

    ax2.set_ylabel('Voltage loss  [mV]', fontsize=12)

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    fig.legend(lines1 + lines2, labels1 + labels2,
               loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=len(lines1) + len(lines2), fontsize=9,
               frameon=True, fancybox=True)

    ax1.grid(True, alpha=0.3)
    ax1.set_title(f'Voltage & Losses at {j_target:.2f} A/cm² vs. Cycle',
                  fontsize=12, fontweight='bold')

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


def plot_eis_losses_vs_cycle(cycle_nums, dep_values, losses,
                              j_target=None, v_target=None, save_path=None):
    """
    Dual-axis: V (or j) at target (left) and EIS-decomposed losses (right).
    Uses R0, R1, R2 from circuit fits instead of polcurve model.
    """
    fig, ax1 = plt.subplots(figsize=(10, 6), dpi=120)
    ax2 = ax1.twinx()

    if j_target is not None:
        ax1.plot(cycle_nums, dep_values, 'o-', color='#1f77b4', ms=5, lw=1.5,
                 label=f'V @ {j_target:.2f} A/cm²')
        ax1.set_ylabel(f'Cell voltage at {j_target:.2f} A/cm²  [V]',
                       fontsize=12, color='#1f77b4')
        target_str = f'{j_target:.2f} A/cm²'
    else:
        ax1.plot(cycle_nums, dep_values, 'o-', color='#1f77b4', ms=5, lw=1.5,
                 label=f'j @ {v_target:.2f} V')
        ax1.set_ylabel(f'Current density at {v_target:.2f} V  [A/cm²]',
                       fontsize=12, color='#1f77b4')
        target_str = f'{v_target:.2f} V'

    ax1.tick_params(axis='y', labelcolor='#1f77b4')
    ax1.ticklabel_format(axis='y', useOffset=False)
    ax1.set_xlim(left=0)
    ax1.set_xlabel('Cycle number', fontsize=12)

    loss_styles = [
        ('V_R0_mV',       'R₀ (membrane)',         '#4CAF50', 's-'),
        ('V_R1_mV',       'R₁ (charge transfer)',   '#FF5722', '^-'),
        ('V_CL_mV',       'CL ionic (coth)',        '#FF9800', 'v-'),
        ('V_kinetic_mV',  'Kinetic residual',       '#9C27B0', 'D-'),
    ]

    for key, label, color, fmt in loss_styles:
        vals = losses.get(key)
        if vals is not None and len(vals) > 0:
            ax2.plot(cycle_nums, vals, fmt, color=color, ms=4, lw=1.2,
                     label=label, alpha=0.85)

    ax2.set_ylabel('Voltage loss  [mV]', fontsize=12)

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    fig.legend(lines1 + lines2, labels1 + labels2,
               loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=len(lines1) + len(lines2), fontsize=9,
               frameon=True, fancybox=True)

    ax1.grid(True, alpha=0.3)
    ax1.set_title(f'EIS Loss Decomposition at {target_str} vs. Cycle',
                  fontsize=12, fontweight='bold')

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


# ═══════════════════════════════════════════════════════════════════
#  Data export
# ═══════════════════════════════════════════════════════════════════

def export_excel(filepath, cycles, v_targets=[1.8, 1.7], j_targets=None,
                 eis_mapped=None, loss_data=None, fit_result=None,
                 eis_results=None, ir_data=None, geo_area=5.0,
                 eis_fit_results=None, eis_loss_data=None,
                 coth_results=None):
    """
    Export all analysis data to a multi-sheet Excel workbook.

    Sheets:
      - Polcurve Data: cycle, V, j for every setpoint in every cycle
      - j vs Cycle / V vs Cycle: tracking at target voltages or currents
      - ASR vs Cycle: cycle number, ASR (R₀+R₁) from EIS (if available)
      - Losses vs Cycle: per-cycle fitted loss breakdown
      - Model Fit: last cycle fit — data, model, residuals, components
      - EIS: frequency, Z', -Z'' for each EIS measurement
    """
    from openpyxl import Workbook
    wb = Workbook()

    # ── Sheet 1: Polcurve Data ──
    ws = wb.active
    ws.title = "Polcurve Data"
    ws.append(['Cycle', 'Mode', 'V_setpoint [V]', 'j [A/cm²]', 'Step', 'Repeat'])
    for i, cyc in enumerate(cycles):
        mode_tag = cyc[0].get('cycle_label', '') if cyc else ''
        for d in cyc:
            ws.append([i + 1, mode_tag, round(d['V'], 4), round(d['j'], 6),
                       d['step'], d['repeat']])

    # ── Sheet 2: j vs Cycle ──
    ws2 = wb.create_sheet("j vs Cycle")
    header = ['Cycle', 'Mode']
    for vt in v_targets:
        header.append(f'j @ {vt:.2f} V [A/cm²]')
        header.append(f'Source ({vt:.2f} V)')
    ws2.append(header)

    # Build mode lookup
    cycle_modes = {}
    for i, cyc in enumerate(cycles):
        cycle_modes[i + 1] = cyc[0].get('cycle_label', '') if cyc else ''

    if j_targets:
        # ── Galvanostatic: V vs Cycle sheet ──
        ws2 = wb.create_sheet("V vs Cycle")
        header = ['Cycle', 'Mode']
        for jt in j_targets:
            header.append(f'V @ {jt:.2f} A/cm² [V]')
            header.append(f'Source ({jt:.2f} A/cm²)')
        ws2.append(header)

        v_at_j = {}
        interp_at_j = {}
        for jt in j_targets:
            cn, vv, interp = extract_v_at_current(cycles, jt)
            v_at_j[jt] = dict(zip(cn.astype(int), vv))
            interp_at_j[jt] = dict(zip(cn.astype(int), interp))

        all_cycle_nums = sorted(set().union(*(v_at_j[jt].keys() for jt in j_targets)))
        for cn in all_cycle_nums:
            row = [cn, cycle_modes.get(cn, '')]
            for jt in j_targets:
                v_val = v_at_j[jt].get(cn, float('nan'))
                is_int = interp_at_j[jt].get(cn, False)
                row.append(round(v_val, 4) if not np.isnan(v_val) else '')
                row.append('interpolated' if is_int else 'measured')
            ws2.append(row)

        ws2.append([])
        ws2.append(['Stability Detection'])
        for jt in j_targets:
            cn_arr, vv_arr, _ = extract_v_at_current(cycles, jt)
            sc = detect_stabilization(cn_arr, vv_arr)
            ws2.append([f'{jt:.2f} A/cm²',
                        f'Stable @ cycle {sc}' if sc else 'Not stabilized'])
    else:
        # ── Potentiostatic: j vs Cycle sheet ──
        ws2 = wb.create_sheet("j vs Cycle")
        header = ['Cycle', 'Mode']
        for vt in v_targets:
            header.append(f'j @ {vt:.2f} V [A/cm²]')
            header.append(f'Source ({vt:.2f} V)')
        ws2.append(header)

        j_at_v = {}
        interp_at_v = {}
        for vt in v_targets:
            cn, jv, interp = extract_j_at_voltage(cycles, vt)
            j_at_v[vt] = dict(zip(cn.astype(int), jv))
            interp_at_v[vt] = dict(zip(cn.astype(int), interp))

        all_cycle_nums = sorted(set().union(*(j_at_v[vt].keys() for vt in v_targets)))
        for cn in all_cycle_nums:
            row = [cn, cycle_modes.get(cn, '')]
            for vt in v_targets:
                j_val = j_at_v[vt].get(cn, float('nan'))
                is_int = interp_at_v[vt].get(cn, False)
                row.append(round(j_val, 6) if not np.isnan(j_val) else '')
                row.append('interpolated' if is_int else 'measured')
            ws2.append(row)

        ws2.append([])
        ws2.append(['Stability Detection'])
        for vt in v_targets:
            cn_arr, jv_arr, _ = extract_j_at_voltage(cycles, vt)
            sc = detect_stabilization(cn_arr, jv_arr)
            ws2.append([f'{vt:.2f} V',
                        f'Stable @ cycle {sc}' if sc else 'Not stabilized'])

    # ── Sheet 3: ASR vs Cycle ──
    if eis_mapped:
        ws3 = wb.create_sheet("ASR vs Cycle")
        ws3.append(['Cycle', 'ASR [mΩ·cm²]', 'Elapsed Time [s]'])
        for em in eis_mapped:
            ws3.append([em['cycle'], round(em['asr_mohm_cm2'], 2),
                        round(em['t_eis'], 1) if em['t_eis'] else ''])

    # ── Sheet 4: Losses vs Cycle ──
    if loss_data is not None:
        cn_loss, dep_values, losses = loss_data
        if len(cn_loss) > 0:
            ws4 = wb.create_sheet("Losses vs Cycle")
            if j_targets:
                dep_label = 'V [V]'
                dep_round = 4
            else:
                dep_label = 'j [A/cm²]'
                dep_round = 6
            ws4.append(['Cycle', dep_label,
                        'η_anode [mV]', 'η_cathode [mV]', 'η_kinetic_total [mV]',
                        'V_ohmic [mV]', 'V_mt [mV]'])
            for i in range(len(cn_loss)):
                eta_total = losses['eta_anode_mV'][i] + losses['eta_cathode_mV'][i]
                ws4.append([
                    int(cn_loss[i]),
                    round(dep_values[i], dep_round),
                    round(losses['eta_anode_mV'][i], 2),
                    round(losses['eta_cathode_mV'][i], 2),
                    round(eta_total, 2),
                    round(losses['V_ohmic_mV'][i], 2),
                    round(losses['V_mt_mV'][i], 2),
                ])

    # ── Sheet 5: Model Fit ──
    if fit_result is not None:
        ws5 = wb.create_sheet("Model Fit")

        # Fitted parameters
        ws5.append(['Fitted Parameters'])
        xf = fit_result['x']
        T_K = fit_result['T_K']
        params = [
            ('E_rev [V]', round(fit_result['E_rev'], 4)),
            ('ASR_total [mΩ·cm²]', round(xf[0], 1)),
            ('j0_anode [A/cm²]', f"{10**xf[1]:.3e}"),
            ('α_anode', round(xf[2], 3)),
            ('Anode Tafel slope [mV/dec]', round((_R*T_K)/(xf[2]*_n_e*_F)*1000, 1)),
            ('j0_cathode [A/cm²]', f"{10**xf[3]:.3e}"),
            ('Cathode Tafel slope [mV/dec]', round((_R*T_K)/(0.5*_n_e*_F)*1000, 1)),
            ('c_mt [V·cm⁴/A²]', round(xf[4], 5)),
            ('RMSE [mV]', round(fit_result['rmse_mV'], 2)),
            ('MAE [mV]', round(fit_result['mae_mV'], 2)),
            ('Max |error| [mV]', round(fit_result['max_err_mV'], 2)),
        ]
        for name, val in params:
            ws5.append([name, val])

        # Data vs model
        ws5.append([])
        ws5.append(['j [A/cm²]', 'V_data [V]', 'V_model [V]', 'Residual [mV]'])
        for j_val, v_dat, v_mod, res in zip(
                fit_result['j_data'], fit_result['V_data'],
                fit_result['V_model'], fit_result['residual']):
            ws5.append([round(j_val, 6), round(v_dat, 5), round(v_mod, 5),
                        round(res * 1000, 2)])

        # Model curve components
        ws5.append([])
        comp = fit_result['components']
        ws5.append(['j_model [A/cm²]', 'V_total [V]', 'E_rev [V]',
                    'η_anode [V]', 'η_cathode [V]', 'V_ohmic [V]', 'V_mt [V]'])
        # Export every 10th point of the smooth curve to keep file manageable
        step = max(1, len(comp['j']) // 50)
        for idx in range(0, len(comp['j']), step):
            ws5.append([
                round(comp['j'][idx], 6),
                round(comp['V_total'][idx], 5),
                round(comp['E_rev'][idx], 5),
                round(comp['eta_anode'][idx], 5),
                round(comp['eta_cathode'][idx], 5),
                round(comp['V_ohmic'][idx], 5),
                round(comp['V_mt'][idx], 5),
            ])

    # ── Sheet 6: EIS ──
    if eis_results:
        ws6 = wb.create_sheet("EIS")
        ws6.append(['File', 'Frequency [Hz]', "Z' [Ω]", "-Z'' [Ω]",
                     "Z' [mΩ·cm²]", "-Z'' [mΩ·cm²]",
                     'DC Voltage [V]', 'Elapsed Time [s]'])
        for er in eis_results:
            eis = er['eis_data']
            fname = Path(er['file']).name
            for i in range(len(eis['freq'])):
                row = [
                    fname,
                    round(eis['freq'][i], 3),
                    round(eis['zre'][i], 8),
                    round(eis['zim'][i], 8),
                    round(eis['zre'][i] * geo_area * 1000, 4),
                    round(eis['zim'][i] * geo_area * 1000, 4),
                ]
                if eis.get('dc_v') is not None:
                    row.append(round(eis['dc_v'][i], 5))
                else:
                    row.append('')
                if eis.get('time') is not None:
                    row.append(round(eis['time'][i], 2))
                else:
                    row.append('')
                ws6.append(row)

    # ── Sheet 7: iR Correction ──
    if ir_data is not None:
        # ir_data can be a list of per-cycle dicts or a single dict (legacy)
        ir_list = ir_data if isinstance(ir_data, list) else [ir_data]
        ws7 = wb.create_sheet("iR Correction")

        for gi, ir_entry in enumerate(ir_list):
            if gi > 0:
                ws7.append([])  # blank row between groups

            cyc_label = ir_entry.get('cycle_num', gi + 1)
            ws7.append([f'Cycle {cyc_label}'])
            ws7.append(['j [A/cm²]', 'V_raw [V]', 'V_iR-free [V]',
                        'ASR_interp [mΩ·cm²]', 'iR_drop [mV]'])
            for i in range(len(ir_entry['j_pol'])):
                j_val = ir_entry['j_pol'][i]
                ir_drop = j_val * ir_entry['asr_interp'][i]
                ws7.append([
                    round(j_val, 6),
                    round(ir_entry['V_pol'][i], 5),
                    round(ir_entry['V_irfree'][i], 5),
                    round(ir_entry['asr_interp'][i], 2),
                    round(ir_drop, 2),
                ])
            ws7.append([])
            ws7.append(['R₀+R₁ Measurements'])
            ws7.append(['j [A/cm²]', 'ASR [mΩ·cm²]'])
            for i in range(len(ir_entry['j_hfr'])):
                ws7.append([round(ir_entry['j_hfr'][i], 6),
                            round(ir_entry['asr_hfr'][i], 2)])

    # ── Sheet 8: EIS Circuit Fit ──
    if eis_fit_results:
        ws8 = wb.create_sheet("EIS Circuit Fit")
        ws8.append(['Parameter', 'Units'] +
                   [efr.get('label', f'EIS {i+1}')
                    for i, efr in enumerate(eis_fit_results)])

        # Add j and V_dc context rows if available
        j_row = ['j', 'A/cm²'] + [round(efr['j'], 4) if efr.get('j') is not None else ''
                                    for efr in eis_fit_results]
        v_row = ['V_dc', 'V'] + [round(efr['dc_v'], 4) if efr.get('dc_v') is not None else ''
                                  for efr in eis_fit_results]
        ws8.append(j_row)
        ws8.append(v_row)
        ws8.append([])  # blank separator

        rows = [
            ('R₀ (ohmic)', 'mΩ·cm²', [efr['R0_asr'] for efr in eis_fit_results]),
            ('R₁ (HF arc)', 'mΩ·cm²', [efr['R1_asr'] for efr in eis_fit_results]),
            ('Q₁', 'S·s^n', [efr['Q1'] for efr in eis_fit_results]),
            ('n₁', '—', [efr['n1'] for efr in eis_fit_results]),
            ('f_c1', 'Hz', [efr['fc1_Hz'] for efr in eis_fit_results]),
            ('R₂ (CL ionic)', 'mΩ·cm²', [efr['R2_asr'] for efr in eis_fit_results]),
            ('Q₂', 'S·s^n', [efr['Q2'] for efr in eis_fit_results]),
            ('n₂', '—', [efr['n2'] for efr in eis_fit_results]),
            ('f_c2', 'Hz', [efr['fc2_Hz'] for efr in eis_fit_results]),
            ('R_total', 'mΩ·cm²', [efr['R_total_asr'] for efr in eis_fit_results]),
            ('RMSE', 'mΩ·cm²', [efr['rmse_mohm_cm2'] for efr in eis_fit_results]),
        ]
        for name, unit, vals in rows:
            ws8.append([name, unit] + [round(v, 4) for v in vals])

    # ── Sheet 9: EIS Loss Decomposition ──
    if eis_loss_data is not None:
        cn_eis, dep_eis, eis_losses = eis_loss_data
        if len(cn_eis) > 0:
            ws9 = wb.create_sheet("EIS Losses")
            dep_label = 'V [V]' if j_targets else 'j [A/cm²]'
            ws9.append(['Cycle', dep_label,
                        'R₀ membrane [mV]', 'R₁ charge transfer [mV]',
                        'CL ionic coth [mV]', 'Kinetic residual [mV]',
                        'Total non-E_rev [mV]'])
            for i in range(len(cn_eis)):
                total = (eis_losses['V_R0_mV'][i] + eis_losses['V_R1_mV'][i] +
                         eis_losses['V_CL_mV'][i] + eis_losses['V_kinetic_mV'][i])
                ws9.append([
                    int(cn_eis[i]),
                    round(dep_eis[i], 4),
                    round(eis_losses['V_R0_mV'][i], 2),
                    round(eis_losses['V_R1_mV'][i], 2),
                    round(eis_losses['V_CL_mV'][i], 2),
                    round(eis_losses['V_kinetic_mV'][i], 2),
                    round(total, 2),
                ])

    # ── Sheet 10: Coth iR Correction ──
    if coth_results:
        for gi, cr_entry in enumerate(coth_results):
            cr = cr_entry['coth_result']
            tfr = cr_entry['tafel_result']
            ci = cr_entry['cycle_idx']

            sheet_name = f'Coth Cycle {ci+1}'
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            ws10 = wb.create_sheet(sheet_name)

            # Tafel summary
            ws10.append(['Transmission-Line iR Correction'])
            ws10.append([])
            if tfr is not None:
                ws10.append(['Tafel Fit Results'])
                ws10.append(['j₀,anode (A/cm²)', f"{tfr['j0_a']:.4e}"])
                ws10.append(['α_anode', round(tfr['alpha_a'], 4)])
                ws10.append(['b_anode (mV/dec)', round(tfr['ba_mVdec'], 1)])
                ws10.append(['j₀,cathode (A/cm²)', f"{tfr['j0_c']:.4e}"])
                ws10.append(['b_cathode (mV/dec)', round(tfr['bc_mVdec'], 1)])
                ws10.append(['c_mt (V·cm⁴/A²)', round(tfr['c_mt'], 6)])
                ws10.append(['E_rev (V)', round(tfr['E_rev'], 4)])
                ws10.append(['RMSE (mV)', round(tfr['rmse_mV'], 2)])
                ws10.append([])

            # Polcurve data with corrections
            V_mt = tfr['V_mt'] if tfr is not None else np.zeros(len(cr['j']))
            ws10.append(['j (A/cm²)', 'V_raw (V)', 'V_R0 (mV)', 'V_R1 (mV)',
                         'V_CL_coth (mV)', 'V_mt (mV)', 'V_irfree (V)',
                         'R₀ (mΩ·cm²)', 'R₁ (mΩ·cm²)', 'R₂ (mΩ·cm²)'])
            for i in range(len(cr['j'])):
                ws10.append([
                    round(float(cr['j'][i]), 6),
                    round(float(cr['V_raw'][i]), 5),
                    round(float(cr['V_R0'][i]) * 1000, 2),
                    round(float(cr['V_R1'][i]) * 1000, 2),
                    round(float(cr['V_CL_ionic'][i]) * 1000, 2),
                    round(float(V_mt[i]) * 1000, 2),
                    round(float(cr['V_irfree'][i]), 5),
                    round(float(cr['R0_interp'][i]), 2),
                    round(float(cr['R1_interp'][i]), 2),
                    round(float(cr['R2_interp'][i]), 2),
                ])

    wb.save(filepath)
    print(f"  Data exported: {filepath}")


# ═══════════════════════════════════════════════════════════════════
#  EIS loading & EIS analysis
# ═══════════════════════════════════════════════════════════════════

def detect_eis_columns(fieldnames):
    """Auto-detect EIS columns: frequency, Z', -Z'', elapsed time, DC voltage."""
    result = {k: None for k in ('freq_col', 'zre_col', 'zim_col', 'time_col', 'dc_v_col')}

    freq_cands = ['frequency (hz)', 'freq (hz)', 'frequency(hz)', 'freq(hz)']
    zre_cands  = ["z' (ohms)", "z'(ohms)", "zre (ohms)", "z' (ohm)", "zreal"]
    zim_cands  = ['-z" (ohms)', '-z"(ohms)', "-z'' (ohms)", '-zim', '-z" (ohm)',
                  "z\" (ohms)", "z''"]
    time_cands = ['elapsed time', 'time (s)', 'elapsed_time', 'time(s)']
    dcv_cands  = ['dc working electrode (v)', 'working electrode (v)',
                  'dc voltage', 'ewe (v)', 'dc potential']

    for fn in fieldnames:
        fl = fn.lower().strip()

        if result['freq_col'] is None:
            for c in freq_cands:
                if c in fl:
                    result['freq_col'] = fn; break

        if result['zre_col'] is None:
            for c in zre_cands:
                if c in fl:
                    result['zre_col'] = fn; break

        if result['zim_col'] is None:
            for c in zim_cands:
                if c in fl:
                    result['zim_col'] = fn; break

        if result['time_col'] is None:
            for c in time_cands:
                if c in fl:
                    result['time_col'] = fn; break

        if result['dc_v_col'] is None:
            for c in dcv_cands:
                if c in fl:
                    result['dc_v_col'] = fn; break

    return result


def load_eis_data(filepath):
    """
    Load EIS data from CSV. Returns dict with arrays:
    freq, zre, zim, time (all sorted by descending frequency).
    """
    path = Path(_clean_path(filepath))
    if not path.exists():
        raise FileNotFoundError(f"EIS file not found: {path}")

    try:    text = path.read_text(encoding='utf-8')
    except: text = path.read_text(encoding='latin-1')

    delim = '\t' if '\t' in text.split('\n')[0] else ','
    lines = text.strip().split('\n')
    reader = csv.DictReader(lines, delimiter=delim)
    fn = [f.strip() for f in (reader.fieldnames or [])]
    reader.fieldnames = fn

    cols = detect_eis_columns(fn)
    if cols['freq_col'] is None or cols['zre_col'] is None or cols['zim_col'] is None:
        raise ValueError(
            f"Could not detect EIS columns.\n"
            f"  Available: {fn}\n"
            f"  Need: Frequency (Hz), Z' (Ohms), -Z\" (Ohms)")

    freq, zre, zim, time, dc_v = [], [], [], [], []
    for row in reader:
        try:
            freq.append(float(row[cols['freq_col']].strip()))
            zre.append(float(row[cols['zre_col']].strip()))
            zim.append(float(row[cols['zim_col']].strip()))
            if cols['time_col']:
                time.append(float(row[cols['time_col']].strip()))
            if cols['dc_v_col']:
                dc_v.append(float(row[cols['dc_v_col']].strip()))
        except (ValueError, KeyError):
            continue

    freq = np.array(freq)
    zre = np.array(zre)
    zim = np.array(zim)
    time = np.array(time) if time else None
    dc_v = np.array(dc_v) if dc_v else None

    # Sort by descending frequency
    order = np.argsort(freq)[::-1]
    freq, zre, zim = freq[order], zre[order], zim[order]
    if time is not None:
        time = time[order]
    if dc_v is not None:
        dc_v = dc_v[order]

    # Mean DC voltage during EIS measurement
    dc_v_mean = float(np.nanmean(dc_v)) if dc_v is not None else None

    print(f"  EIS: {len(freq)} points from '{path.name}'")
    print(f"    Columns: freq='{cols['freq_col']}', "
          f"Z'='{cols['zre_col']}', -Z''='{cols['zim_col']}'")
    if dc_v_mean is not None:
        print(f"    DC voltage: {dc_v_mean:.4f} V")
    print(f"    Freq range: {freq.min():.1f} – {freq.max():.0f} Hz")
    if time is not None:
        print(f"    Time range: {time.min():.1f} – {time.max():.1f} s")

    return {'freq': freq, 'zre': zre, 'zim': zim, 'time': time,
            'dc_v': dc_v, 'dc_v_mean': dc_v_mean, 'cols': cols}


def extract_hfr(eis, geo_area=5.0):
    """
    Extract HFR (high-frequency resistance) from EIS data.

    Finds the Z' intercept where -Z'' crosses zero from negative
    (inductive) to positive (capacitive). Uses linear interpolation
    between the two points bracketing the crossing.

    Returns
    -------
    hfr_ohm : float — HFR in Ohms
    asr_ohm_cm2 : float — ASR in Ω·cm²
    asr_mohm_cm2 : float — ASR in mΩ·cm²
    f_hfr : float — frequency at the crossing [Hz]
    """
    zre = eis['zre']
    zim = eis['zim']
    freq = eis['freq']

    # -Z'' convention: in the file, -Z'' is stored directly.
    # Inductive region: -Z'' < 0 (high frequency)
    # Capacitive region: -Z'' > 0 (lower frequency)
    # HFR = Z' where -Z'' crosses zero from negative to positive

    # Scan from high to low frequency for sign change in zim
    crossing_idx = None
    for i in range(len(zim) - 1):
        if zim[i] <= 0 and zim[i + 1] > 0:
            crossing_idx = i
            break

    if crossing_idx is not None:
        # Linear interpolation
        z1, z2 = zim[crossing_idx], zim[crossing_idx + 1]
        zr1, zr2 = zre[crossing_idx], zre[crossing_idx + 1]
        f1, f2 = freq[crossing_idx], freq[crossing_idx + 1]

        frac = -z1 / (z2 - z1)  # fraction from point 1 to point 2
        hfr = zr1 + frac * (zr2 - zr1)
        f_hfr = f1 + frac * (f2 - f1)
    else:
        # No sign change found — use minimum |Z''| point
        min_idx = np.argmin(np.abs(zim))
        hfr = zre[min_idx]
        f_hfr = freq[min_idx]

    # Sanity check: if the intercept frequency is below 1500 Hz,
    # the crossing is likely between arcs (cathode/anode) rather than
    # the true high-frequency ohmic intercept. Fall back to Z' at ~1000 Hz.
    if f_hfr < 1500:
        idx_1k = np.argmin(np.abs(freq - 1000.0))
        hfr_fallback = zre[idx_1k]
        f_fallback = freq[idx_1k]
        print(f"    HFR intercept at {f_hfr:.0f} Hz is below 1500 Hz — "
              f"likely between arcs. Using Z' at {f_fallback:.0f} Hz instead.")
        hfr = hfr_fallback
        f_hfr = f_fallback

    asr = hfr * geo_area
    asr_m = asr * 1000

    print(f"\n  EIS HFR Analysis (seed for circuit fit):")
    print(f"    R₀+R₁ (approx): {hfr*1000:.3f} mΩ  ({hfr:.6f} Ω)")
    print(f"    ASR           : {asr_m:.1f} mΩ·cm²  ({asr:.4f} Ω·cm²)")
    print(f"    Intercept freq: {f_hfr:.0f} Hz")

    return {'hfr_ohm': hfr, 'asr_ohm_cm2': asr, 'asr_mohm_cm2': asr_m,
            'f_hfr': f_hfr}


# ═══════════════════════════════════════════════════════════════════
#  EIS Equivalent Circuit Fitting — R₀-(R₁Q₁)-(R₂Q₂)
# ═══════════════════════════════════════════════════════════════════

def _zarc(omega, R, Q, n):
    """Impedance of a parallel R-CPE element (ZARC)."""
    Z_cpe = 1.0 / (Q * (1j * omega) ** n)
    return R * Z_cpe / (R + Z_cpe)


def _circuit_r_rq_rq(omega, params):
    """
    R₀-(R₁Q₁)-(R₂Q₂) equivalent circuit.

    params: [R0, R1, Q1, n1, R2, Q2, n2]
      R0  — ohmic resistance (Ω)
      R1  — high-frequency arc resistance (Ω)
      Q1  — high-frequency CPE magnitude (S·s^n)
      n1  — high-frequency CPE exponent (0-1)
      R2  — low-frequency arc resistance (Ω)
      Q2  — low-frequency CPE magnitude (S·s^n)
      n2  — low-frequency CPE exponent (0-1)
    """
    R0, R1, Q1, n1, R2, Q2, n2 = params
    Z = R0 + _zarc(omega, R1, Q1, n1) + _zarc(omega, R2, Q2, n2)
    return Z


def fit_eis_circuit(eis, geo_area=5.0, hfr_seed=None):
    """
    Fit R₀-(R₁Q₁)-(R₂Q₂) to EIS data.

    If -Z'' crosses zero between arcs, the Z' value at that crossing
    constrains R₀+R₁ (±5%).

    Parameters
    ----------
    eis : dict with 'freq', 'zre', 'zim' arrays
    geo_area : float — cell area in cm²
    hfr_seed : float or None — seed R0 from HFR (Ω)

    Returns
    -------
    dict with fit parameters, uncertainties, and diagnostic info,
    or None if fitting fails.
    """
    from scipy.optimize import least_squares

    freq = eis['freq']
    zre = eis['zre']
    zim = eis['zim']
    omega = 2 * np.pi * freq

    # Target: complex impedance (Z = Z' - jZ'')
    Z_data = zre - 1j * zim

    # ── Find Z'' zero crossing between arcs ──
    # zim is stored as -Z'' (positive = capacitive arc).
    # Sort by frequency high→low to walk through the spectrum.
    order = np.argsort(freq)[::-1]
    freq_s = freq[order]
    zre_s = zre[order]
    zim_s = zim[order]  # -Z'': positive during arcs, ~0 at intercepts

    # ── Find R₀+R₁ constraint from EIS intercept ──
    # Uses the same rules as extract_hfr:
    #   1. Z'' zero crossing (inductive→capacitive or between arcs)
    #   2. If crossing freq < 1500 Hz, fall back to Z' at ~1000 Hz
    #   3. If no crossing, use Z' at minimum |Z''|, with 1000 Hz fallback
    R0R1_crossing = None
    zim_max = max(abs(zim_s)) if len(zim_s) > 0 else 1.0

    # Check for zero crossing
    crossing_idx = None
    for i in range(len(zim_s) - 1):
        # HF crossing: -Z'' goes from negative (inductive) to positive (arc)
        if zim_s[i] <= 0 and zim_s[i + 1] > 0:
            crossing_idx = i
            break
        # Mid-spectrum crossing: -Z'' goes from positive back to ≤0
        if zim_s[i] > 0.05 * zim_max and zim_s[i + 1] <= 0:
            if i < len(zim_s) - 3 and zre_s[i] < 0.95 * zre_s.max():
                crossing_idx = i
                break

    if crossing_idx is not None:
        frac = abs(zim_s[crossing_idx]) / abs(zim_s[crossing_idx + 1] - zim_s[crossing_idx])
        zre_cross = zre_s[crossing_idx] + frac * (zre_s[crossing_idx + 1] - zre_s[crossing_idx])
        f_crossing = freq_s[crossing_idx] + frac * (freq_s[crossing_idx + 1] - freq_s[crossing_idx])

        if f_crossing >= 1500:
            R0R1_crossing = zre_cross
            print(f"    R₀+R₁ from Z'' crossing: {R0R1_crossing*geo_area*1000:.1f} mΩ·cm² "
                  f"at {f_crossing:.0f} Hz")
        else:
            # Crossing below 1500 Hz — likely between arcs, use Z' at ~1000 Hz
            idx_1k = np.argmin(np.abs(freq_s - 1000.0))
            R0R1_crossing = zre_s[idx_1k]
            print(f"    R₀+R₁ from Z' at {freq_s[idx_1k]:.0f} Hz: "
                  f"{R0R1_crossing*geo_area*1000:.1f} mΩ·cm² "
                  f"(crossing at {f_crossing:.0f} Hz < 1500 Hz)")
    else:
        # No crossing — use Z' at frequency nearest to 1000 Hz
        idx_1k = np.argmin(np.abs(freq_s - 1000.0))
        R0R1_crossing = zre_s[idx_1k]
        print(f"    R₀+R₁ from Z' at {freq_s[idx_1k]:.0f} Hz: "
              f"{R0R1_crossing*geo_area*1000:.1f} mΩ·cm² (no Z'' crossing)")

    # Initial guesses
    R0_guess = hfr_seed if hfr_seed else zre.min()
    z_span = zre.max() - zre.min()

    if R0R1_crossing is not None:
        # Use crossing to set better initial guesses
        R1_guess = max(R0R1_crossing - R0_guess, z_span * 0.1)
        R2_guess = max(z_span - R1_guess, z_span * 0.1)
    else:
        R1_guess = z_span * 0.4
        R2_guess = z_span * 0.6

    # CPE guesses
    Q1_guess = 0.01
    Q2_guess = 0.1
    n1_guess = 0.85
    n2_guess = 0.85

    x0 = [R0_guess, R1_guess, Q1_guess, n1_guess, R2_guess, Q2_guess, n2_guess]

    # Bounds
    lo = [R0_guess * 0.5, 1e-6,   1e-6, 0.3, 1e-6,   1e-5, 0.3]
    hi = [R0_guess * 2.0, z_span * 3, 10.0, 1.0, z_span * 3, 100.0, 1.0]

    # Constraint weight for R0+R1 crossing
    margin = 0.05  # ±5%
    constraint_weight = 10.0  # strong penalty

    def residuals(x):
        Z_model = _circuit_r_rq_rq(omega, x)
        diff = Z_model - Z_data
        weights = 1.0 / np.maximum(np.abs(Z_data), 1e-10)
        res = np.concatenate([diff.real * weights, diff.imag * weights])

        # Add R0+R1 constraint if crossing was found
        if R0R1_crossing is not None:
            R0_fit, R1_fit = x[0], x[1]
            R0R1_fit = R0_fit + R1_fit
            # Penalty: deviation from crossing value, normalized
            deviation = (R0R1_fit - R0R1_crossing) / R0R1_crossing
            # Zero penalty within ±margin, linear outside
            if abs(deviation) > margin:
                penalty = constraint_weight * (abs(deviation) - margin)
            else:
                penalty = 0.0
            res = np.append(res, penalty)

        return res

    try:
        result = least_squares(residuals, x0, bounds=(lo, hi),
                                method='trf', loss='soft_l1',
                                f_scale=0.1, max_nfev=5000)
    except Exception as e:
        print(f"    EIS fit failed: {e}")
        return None

    if not result.success:
        print(f"    EIS fit did not converge: {result.message}")
        return None

    xf = result.x
    R0, R1, Q1, n1, R2, Q2, n2 = xf

    # Ensure R1 is the high-frequency arc (higher characteristic freq)
    fc1 = (1.0 / (R1 * Q1)) ** (1.0 / n1) / (2 * np.pi) if R1 > 0 and Q1 > 0 else 0
    fc2 = (1.0 / (R2 * Q2)) ** (1.0 / n2) / (2 * np.pi) if R2 > 0 and Q2 > 0 else 0

    # Swap if arc 2 has higher characteristic frequency
    if fc2 > fc1:
        R1, Q1, n1, R2, Q2, n2 = R2, Q2, n2, R1, Q1, n1
        fc1, fc2 = fc2, fc1

    # Convert to ASR (mΩ·cm²)
    R0_asr = R0 * geo_area * 1000
    R1_asr = R1 * geo_area * 1000
    R2_asr = R2 * geo_area * 1000
    R_total_asr = (R0 + R1 + R2) * geo_area * 1000

    # Compute model curve for plotting
    Z_model = _circuit_r_rq_rq(omega, [R0, R1, Q1, n1, R2, Q2, n2])

    # Goodness of fit
    err = Z_model - Z_data
    rmse = np.sqrt(np.mean(np.abs(err) ** 2)) * geo_area * 1000

    fit_result = {
        'R0_ohm': R0, 'R1_ohm': R1, 'R2_ohm': R2,
        'Q1': Q1, 'n1': n1, 'Q2': Q2, 'n2': n2,
        'R0_asr': R0_asr, 'R1_asr': R1_asr, 'R2_asr': R2_asr,
        'R_total_asr': R_total_asr,
        'fc1_Hz': fc1, 'fc2_Hz': fc2,
        'rmse_mohm_cm2': rmse,
        'Z_model': Z_model,
        'freq': freq, 'Z_data': Z_data,
        'converged': result.success,
        'R0R1_crossing': R0R1_crossing,
    }

    return fit_result


def print_eis_fit_summary(fr, geo_area=5.0):
    """Print formatted EIS fit results."""
    print(f"\n    EIS Circuit Fit: R₀-(R₁Q₁)-(R₂Q₂)")
    print(f"    {'─' * 50}")
    print(f"    R₀ (ohmic)     : {fr['R0_asr']:.1f} mΩ·cm²")
    print(f"    R₁ (HF arc)    : {fr['R1_asr']:.1f} mΩ·cm²  "
          f"(f_c = {fr['fc1_Hz']:.0f} Hz, n = {fr['n1']:.3f})")
    print(f"    R₂ (CL ionic)  : {fr['R2_asr']:.1f} mΩ·cm²  "
          f"(f_c = {fr['fc2_Hz']:.0f} Hz, n = {fr['n2']:.3f})")
    print(f"    R_total        : {fr['R_total_asr']:.1f} mΩ·cm²")
    print(f"    RMSE           : {fr['rmse_mohm_cm2']:.2f} mΩ·cm²")


def plot_eis_fit(eis, fit_result, geo_area=5.0, title=None, save_path=None):
    """
    Three-panel EIS fit plot:
      Left   — Nyquist: data + model + individual arcs
      Top R  — Bode magnitude
      Bot R  — Bode phase
    """
    freq = eis['freq']
    zre_data = eis['zre'] * geo_area * 1000   # mΩ·cm²
    zim_data = eis['zim'] * geo_area * 1000   # -Z'' positive = capacitive

    Z_model = fit_result['Z_model']
    zre_model = Z_model.real * geo_area * 1000
    zim_model = -Z_model.imag * geo_area * 1000  # convert to -Z'' convention

    R0 = fit_result['R0_ohm']
    R1, Q1, n1 = fit_result['R1_ohm'], fit_result['Q1'], fit_result['n1']
    R2, Q2, n2 = fit_result['R2_ohm'], fit_result['Q2'], fit_result['n2']

    omega = 2 * np.pi * freq
    Z_arc1 = _zarc(omega, R1, Q1, n1)
    Z_arc2 = _zarc(omega, R2, Q2, n2)

    zre_arc1 = (R0 + Z_arc1.real) * geo_area * 1000
    zim_arc1 = -Z_arc1.imag * geo_area * 1000
    zre_arc2 = (R0 + R1 + Z_arc2.real) * geo_area * 1000
    zim_arc2 = -Z_arc2.imag * geo_area * 1000

    fig = plt.figure(figsize=(14, 5.5), dpi=120)
    gs = fig.add_gridspec(2, 2, width_ratios=[1.3, 1], hspace=0.35)
    ax1 = fig.add_subplot(gs[:, 0])    # Nyquist (spans both rows)
    ax2 = fig.add_subplot(gs[0, 1])    # Bode |Z|
    ax3 = fig.add_subplot(gs[1, 1])    # Bode phase

    ttl = title or 'EIS Equivalent Circuit Fit'
    fig.suptitle(ttl, fontsize=12, fontweight='bold')

    # ── Nyquist ──
    ax1.plot(zre_data, zim_data, 'o', ms=4, color='#1f77b4',
             label='Data', zorder=3)
    ax1.plot(zre_model, zim_model, '-', lw=2, color='#d62728',
             label='Fit (R₀-RQ-RQ)', zorder=4)

    # Individual arcs
    ax1.plot(zre_arc1, zim_arc1, '--', lw=1.2, color='#ff7f0e', alpha=0.7,
             label=f'HF arc ({fit_result["R1_asr"]:.1f} mΩ·cm²)')
    ax1.plot(zre_arc2, zim_arc2, '--', lw=1.2, color='#2ca02c', alpha=0.7,
             label=f'LF arc / CL ionic ({fit_result["R2_asr"]:.1f} mΩ·cm²)')

    # Mark R0
    R0_asr = fit_result['R0_asr']
    ax1.plot(R0_asr, 0, '*', ms=12, color='#d62728', markeredgecolor='k',
             markeredgewidth=0.6, zorder=5,
             label=f'R₀ = {R0_asr:.1f} mΩ·cm²')

    ax1.axhline(0, color='k', lw=0.5, alpha=0.5)
    ax1.set_xlabel("Z'  [mΩ·cm²]", fontsize=11)
    ax1.set_ylabel("-Z''  [mΩ·cm²]", fontsize=11)
    ax1.set_title('Nyquist', fontsize=10)
    ax1.legend(fontsize=8, loc='upper left')
    ax1.grid(True, alpha=0.3)
    ax1.set_aspect('equal', adjustable='datalim')

    # ── Bode magnitude ──
    Z_data_complex = eis['zre'] - 1j * eis['zim']
    Z_mag_data = np.abs(Z_data_complex) * geo_area * 1000
    Z_mag_model = np.abs(Z_model) * geo_area * 1000

    ax2.loglog(freq, Z_mag_data, 'o', ms=3, color='#1f77b4', label='Data')
    ax2.loglog(freq, Z_mag_model, '-', lw=1.5, color='#d62728', label='Fit')
    ax2.set_ylabel('|Z|  [mΩ·cm²]', fontsize=10)
    ax2.set_title('Bode Magnitude', fontsize=10)
    ax2.legend(fontsize=8)
    ax2.grid(True, alpha=0.3, which='both')

    # ── Bode phase ──
    phase_data = np.degrees(np.arctan2(eis['zim'], eis['zre']))
    phase_model = np.degrees(np.arctan2(-Z_model.imag, Z_model.real))

    ax3.semilogx(freq, phase_data, 'o', ms=3, color='#1f77b4', label='Data')
    ax3.semilogx(freq, phase_model, '-', lw=1.5, color='#d62728', label='Fit')
    ax3.set_xlabel('Frequency  [Hz]', fontsize=10)
    ax3.set_ylabel('Phase  [°]', fontsize=10)
    ax3.set_title('Bode Phase', fontsize=10)
    ax3.legend(fontsize=8)
    ax3.grid(True, alpha=0.3, which='both')

    plt.subplots_adjust(left=0.08, right=0.95, top=0.90, bottom=0.12,
                        wspace=0.35, hspace=0.35)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"    Plot saved: {save_path}")
    else:
        plt.show()
    return fig


def plot_nyquist(eis, hfr_result, geo_area=5.0, save_path=None):
    """Nyquist plot with R₀+R₁ intercept marked."""
    zre = eis['zre'] * geo_area * 1000   # → mΩ·cm²
    zim = eis['zim'] * geo_area * 1000

    fig, ax = plt.subplots(figsize=(7, 6), dpi=120)

    ax.plot(zre, zim, 'o-', ms=4, lw=1.2, color='#1f77b4')

    # Mark R₀+R₁ intercept
    hfr_asr = hfr_result['asr_mohm_cm2']
    ax.plot(hfr_asr, 0, '*', ms=14, color='#d62728', markeredgecolor='k',
            markeredgewidth=0.8, zorder=5,
            label=f'R₀+R₁ = {hfr_asr:.1f} mΩ·cm²')

    ax.axhline(0, color='k', lw=0.5, alpha=0.5)
    ax.set_xlabel("Z'  [mΩ·cm²]", fontsize=12)
    ax.set_ylabel("-Z''  [mΩ·cm²]", fontsize=12)
    ax.set_title('Nyquist Plot (EIS)', fontsize=12, fontweight='bold')
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.set_aspect('equal', adjustable='datalim')

    plt.tight_layout()
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


def plot_eis_for_ir_correction(eis_at_j, geo_area=5.0, cycle_num=None,
                                save_path=None):
    """
    Two-panel figure showing EIS data used for iR correction:
      Left  — Overlaid Nyquist curves colored by current density, R₀+R₁ marked
      Right — ASR (R₀+R₁) vs current density
    """
    if not eis_at_j:
        return None

    n = len(eis_at_j)
    cmap = plt.cm.viridis(np.linspace(0.1, 0.9, n))

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5.5), dpi=120,
                                    gridspec_kw={'width_ratios': [1.3, 1]})

    ttl = 'EIS for iR Correction'
    if cycle_num is not None:
        ttl += f' — Cycle {cycle_num}'
    fig.suptitle(ttl, fontsize=12, fontweight='bold')

    # ── Left: Nyquist curves ──
    for i, e in enumerate(eis_at_j):
        eis = e.get('eis_data')
        if eis is None:
            continue
        zre = eis['zre'] * geo_area * 1000   # mΩ·cm²
        zim = eis['zim'] * geo_area * 1000

        ax1.plot(zre, zim, 'o-', ms=3, lw=1.0, color=cmap[i],
                 label=f'j = {e["j"]:.3f} A/cm²')

        # Mark R₀+R₁ intercept
        hfr_asr = e['asr_mohm_cm2']
        ax1.plot(hfr_asr, 0, 's', ms=7, color=cmap[i],
                 markeredgecolor='k', markeredgewidth=0.6, zorder=5)

    ax1.axhline(0, color='k', lw=0.5, alpha=0.5)
    ax1.set_xlabel("Z'  [mΩ·cm²]", fontsize=11)
    ax1.set_ylabel("-Z''  [mΩ·cm²]", fontsize=11)
    ax1.set_title('Nyquist Plots', fontsize=10)
    ax1.grid(True, alpha=0.3)
    ax1.set_aspect('equal', adjustable='datalim')

    if n <= 12:
        ax1.legend(fontsize=7, loc='upper left', ncol=1)
    else:
        sm = plt.cm.ScalarMappable(cmap=plt.cm.viridis,
                                    norm=plt.Normalize(
                                        eis_at_j[0]['j'], eis_at_j[-1]['j']))
        sm.set_array([])
        cbar = fig.colorbar(sm, ax=ax1, pad=0.02, fraction=0.04)
        cbar.set_label('j  [A/cm²]', fontsize=9)

    # ── Right: ASR vs j ──
    j_vals = [e['j'] for e in eis_at_j]
    asr_vals = [e['asr_mohm_cm2'] for e in eis_at_j]

    ax2.plot(j_vals, asr_vals, 'o-', ms=7, lw=1.5, color='#d62728',
             markeredgecolor='k', markeredgewidth=0.5)
    ax2.set_xlabel('Current density  j  [A/cm²]', fontsize=11)
    ax2.set_ylabel('ASR (R₀+R₁)  [mΩ·cm²]', fontsize=11)
    ax2.set_title('ASR (R₀+R₁) vs. Current Density', fontsize=10)
    ax2.set_xlim(left=0)
    ax2.grid(True, alpha=0.3)
    ax2.ticklabel_format(axis='y', useOffset=False)

    for j, asr in zip(j_vals, asr_vals):
        ax2.annotate(f'{asr:.1f}', (j, asr), textcoords='offset points',
                     xytext=(5, 5), fontsize=7, color='#666')

    plt.tight_layout()
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


# ═══════════════════════════════════════════════════════════════════
#  Folder scanning & multi-EIS processing
# ═══════════════════════════════════════════════════════════════════

def scan_folder(folder_path, cell_id='a1'):
    """
    Scan a folder to find polcurve and EIS files for a given cell ID.

    Convention:
      - Polcurve file: starts with "1_" + cell_id (e.g. "1_a1_...")
      - EIS files: contain cell_id but don't start with "1_" + cell_id

    Returns
    -------
    polcurve_file : str or None
    eis_files : list of str (sorted by filename)
    """
    import glob, os
    folder = Path(_clean_path(folder_path))
    if not folder.is_dir():
        raise FileNotFoundError(f"Folder not found: {folder}")

    # Find all CSV files containing the cell ID
    all_csvs = sorted(glob.glob(str(folder / '*.csv')) +
                      glob.glob(str(folder / '*.CSV')))

    polcurve_file = None
    eis_files = []
    prefix = f'1_{cell_id}'

    for fp in all_csvs:
        base = os.path.basename(fp)
        if cell_id.lower() not in base.lower():
            continue

        if base.lower().startswith(prefix.lower()):
            polcurve_file = fp
        else:
            eis_files.append(fp)

    print(f"  Folder: {folder}")
    print(f"  Cell ID: '{cell_id}'")
    if polcurve_file:
        print(f"  Polcurve: {os.path.basename(polcurve_file)}")
    else:
        print(f"  WARNING: No polcurve file found (expected '1_{cell_id}_...')")
    print(f"  EIS files: {len(eis_files)}")
    for fp in eis_files:
        print(f"    {os.path.basename(fp)}")

    return polcurve_file, eis_files


def load_all_eis(eis_files, geo_area=5.0):
    """
    Load multiple EIS files, extract R₀+R₁ from each.

    Returns list of dicts with EIS results + timestamp + DC voltage,
    sorted by time.
    """
    results = []

    for fp in eis_files:
        try:
            eis = load_eis_data(fp)
            hfr = extract_hfr(eis, geo_area=geo_area)

            t_eis = float(np.mean(eis['time'])) if eis['time'] is not None else None

            results.append({
                'file': fp,
                'hfr_ohm': hfr['hfr_ohm'],
                'asr_mohm_cm2': hfr['asr_mohm_cm2'],
                'f_hfr': hfr['f_hfr'],
                't_eis': t_eis,
                'dc_v_mean': eis.get('dc_v_mean'),
                'eis_data': eis,
            })
        except Exception as e:
            print(f"    WARNING: Failed to process {Path(fp).name}: {e}")

    # Sort by timestamp
    results.sort(key=lambda r: r['t_eis'] if r['t_eis'] is not None else 0)

    return results


def map_eis_to_cycles(eis_results, cycles):
    """
    Map each EIS measurement to the polcurve cycle that follows it.

    EIS characterizes the cell state before the next polcurve sweep,
    so each EIS is assigned to the first cycle whose start time is
    after the EIS measurement time.

    Returns list of dicts with cycle number and ASR.
    """
    # Compute start time for each cycle (earliest dwell midpoint)
    cycle_times = []
    for ci, cyc in enumerate(cycles):
        times = [d.get('t_mid') for d in cyc if d.get('t_mid') is not None]
        if times:
            cycle_times.append((ci + 1, min(times)))
        else:
            cycle_times.append((ci + 1, None))

    mapped = []
    for er in eis_results:
        t_eis = er['t_eis']
        if t_eis is None:
            continue

        # Find the first cycle that starts after the EIS measurement
        assigned_cyc = None
        for cn, ct in cycle_times:
            if ct is not None and ct > t_eis:
                assigned_cyc = cn
                break

        # If no cycle follows the EIS, assign to the last cycle
        if assigned_cyc is None:
            assigned_cyc = cycle_times[-1][0] if cycle_times else 0

        mapped.append({
            'cycle': assigned_cyc,
            'asr_mohm_cm2': er['asr_mohm_cm2'],
            't_eis': t_eis,
        })

    return mapped


def plot_j_and_hfr_vs_cycle(cycles, v_targets, eis_mapped, save_path=None):
    """
    Dual-axis plot: j at target voltages (left) and ASR (R₀+R₁) (right)
    vs cycle number.
    """
    fig, ax1 = plt.subplots(figsize=(10, 5.5), dpi=120)
    ax2 = ax1.twinx()

    colors_j = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
    markers_j = ['o', 's', '^', 'D']

    # ── Left axis: j at target voltages ──
    stable_cycles = []
    for k, vt in enumerate(v_targets):
        cn, jv, interp = extract_j_at_voltage(cycles, vt)
        if len(cn) == 0:
            continue
        c = colors_j[k % len(colors_j)]
        m = markers_j[k % len(markers_j)]

        # Measured (solid) and interpolated (open)
        meas = ~interp
        if meas.any():
            ax1.plot(cn[meas], jv[meas], marker=m, linestyle='none', color=c,
                     ms=5, label=f'j @ {vt:.2f} V')
        if interp.any():
            ax1.plot(cn[interp], jv[interp], marker=m, linestyle='none', color=c,
                     ms=5, fillstyle='none', markeredgewidth=1.2,
                     label=f'j @ {vt:.2f} V (interp)')
        ax1.plot(cn, jv, '-', color=c, lw=1.0, alpha=0.5)

        sc = detect_stabilization(cn, jv)
        if sc is not None:
            stable_cycles.append(sc)
            sc_idx = np.where(cn == sc)[0]
            if len(sc_idx) > 0:
                ax1.plot(sc, jv[sc_idx[0]], '*', color=c, ms=14,
                         markeredgecolor='k', markeredgewidth=0.8, zorder=6)

    ax1.set_xlabel('Cycle number', fontsize=12)
    ax1.set_ylabel('Current density  j  [A/cm²]', fontsize=12, color='#1f77b4')
    ax1.tick_params(axis='y', labelcolor='#1f77b4')
    ax1.ticklabel_format(axis='y', useOffset=False)
    ax1.set_xlim(left=0)
    ax1.grid(True, alpha=0.3)

    # ── Right axis: ASR (R₀+R₁) ──
    if eis_mapped:
        cyc_nums = [em['cycle'] for em in eis_mapped]
        asr_vals = [em['asr_mohm_cm2'] for em in eis_mapped]
        ax2.plot(cyc_nums, asr_vals, 'v--', color='#9467bd', ms=8, lw=1.5,
                 markeredgecolor='k', markeredgewidth=0.5,
                 label='ASR (R₀+R₁)')
        ax2.set_ylabel('ASR  [mΩ·cm²]', fontsize=12, color='#9467bd')
        ax2.tick_params(axis='y', labelcolor='#9467bd')

    # Combined legend below x-axis
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    fig.legend(lines1 + lines2, labels1 + labels2,
               loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=len(lines1) + len(lines2), fontsize=9,
               frameon=True, fancybox=True)

    ax1.set_title('Current Density & ASR vs. Cycle', fontsize=12, fontweight='bold')

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig



# ═══════════════════════════════════════════════════════════════════
#  Current-dependent EIS & iR correction
# ═══════════════════════════════════════════════════════════════════

def detect_current_dependent_eis(eis_results, cycles, v_tol=0.03):
    """
    Match each EIS measurement to its nearest polcurve cycle by elapsed time,
    then identify cycles with current-dependent EIS (≥3 distinct DC voltages).

    EIS at low DC voltage (~1.25V, recovery holds) are excluded.

    Returns
    -------
    groups : list of dicts, each with:
        'cycle_idx' : int — index into cycles list
        'eis_at_j'  : list of {'j', 'V', 'asr_mohm_cm2', 'hfr_ohm', 'dc_v_eis'}
    """
    if not eis_results or not cycles:
        return []

    # Filter: only EIS above ~1.30V (exclude recovery holds at ~1.25V)
    valid = [er for er in eis_results
             if er.get('dc_v_mean') is not None
             and er.get('t_eis') is not None
             and er['dc_v_mean'] > 1.30]
    if len(valid) < 3:
        return []

    # Build cycle time ranges
    cycle_times = []
    for ci, cyc in enumerate(cycles):
        times = [d.get('t_mid') for d in cyc if d.get('t_mid') is not None]
        if times:
            cycle_times.append((ci, min(times), max(times)))

    if not cycle_times:
        return []

    # Assign each EIS to the nearest polcurve cycle by elapsed time
    from collections import defaultdict
    eis_by_cycle = defaultdict(list)

    for er in valid:
        t_eis = er['t_eis']
        best_ci = None
        best_dist = float('inf')
        for ci, t_start, t_end in cycle_times:
            if t_start <= t_eis <= t_end:
                dist = 0  # EIS falls within cycle
            else:
                dist = min(abs(t_eis - t_start), abs(t_eis - t_end))
            if dist < best_dist:
                best_dist = dist
                best_ci = ci
        if best_ci is not None:
            eis_by_cycle[best_ci].append(er)

    # For each cycle, check if it has current-dependent EIS (≥3 unique voltages)
    results = []
    for ci, eis_list in sorted(eis_by_cycle.items()):
        unique_v = set(round(er['dc_v_mean'], 2) for er in eis_list)
        if len(unique_v) < 3:
            continue

        ref_cyc = cycles[ci]

        # Match each EIS to a polcurve dwell by DC voltage
        eis_at_j = []
        for er in eis_list:
            v_eis = er['dc_v_mean']
            best_d, best_dv = None, v_tol + 1
            for d in ref_cyc:
                dv = abs(d['V'] - v_eis)
                if dv < best_dv:
                    best_dv = dv
                    best_d = d
            if best_d is not None and best_dv <= v_tol:
                eis_at_j.append({
                    'j': best_d['j'],
                    'V': best_d['V'],
                    'asr_mohm_cm2': er['asr_mohm_cm2'],
                    'hfr_ohm': er['hfr_ohm'],
                    'dc_v_eis': v_eis,
                    'eis_data': er['eis_data'],
                })

        eis_at_j.sort(key=lambda x: x['j'])

        if len(eis_at_j) >= 3:
            results.append({
                'cycle_idx': ci,
                'eis_at_j': eis_at_j,
            })

    if results:
        print(f"\n  Current-dependent EIS: {len(results)} set(s) detected")
        for gi, g in enumerate(results):
            cyc_num = g['cycle_idx'] + 1
            pts = len(g['eis_at_j'])
            asr_range = (min(e['asr_mohm_cm2'] for e in g['eis_at_j']),
                         max(e['asr_mohm_cm2'] for e in g['eis_at_j']))
            print(f"    Set {gi+1}: {pts} pts matched to cycle {cyc_num}, "
                  f"ASR = {asr_range[0]:.1f}–{asr_range[1]:.1f} mΩ·cm²")

    return results


def compute_ir_corrected_polcurve(j_pol, V_pol, eis_at_j, geo_area=5.0):
    """
    Compute iR-corrected polarization curve using current-dependent ASR (R₀+R₁).

    Interpolates ASR(j) from the EIS measurements, then computes:
      V_iR-free(j) = V(j) - j × ASR(j)

    Returns
    -------
    j_pol : array — current density [A/cm²]
    V_pol : array — raw cell voltage [V]
    V_irfree : array — iR-corrected voltage [V]
    j_hfr : array — j values where HFR was measured
    asr_hfr : array — ASR at each j [mΩ·cm²]
    asr_interp : array — interpolated ASR at each polcurve j [mΩ·cm²]
    """
    j_hfr = np.array([e['j'] for e in eis_at_j])
    asr_hfr = np.array([e['asr_mohm_cm2'] for e in eis_at_j])

    # Interpolate ASR to all polcurve j values
    # Clip to the range of measured ASR (no extrapolation)
    asr_interp = np.interp(j_pol, j_hfr, asr_hfr)

    # iR correction: V_irfree = V - j × ASR (convert mΩ·cm² to Ω·cm²)
    V_irfree = V_pol - j_pol * asr_interp / 1000.0

    return j_pol, V_pol, V_irfree, j_hfr, asr_hfr, asr_interp


def plot_ir_correction(j_pol, V_pol, V_irfree, j_hfr, asr_hfr, asr_interp,
                       cycle_label=None, save_path=None):
    """
    Three-panel plot:
      Left  — Raw and iR-corrected polcurves
      Center — ASR (R₀+R₁) vs current density
      Right — iR drop vs current density
    """
    if len(j_pol) == 0 or len(j_hfr) == 0:
        print("  Skipping iR correction plot: empty data arrays")
        return None

    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(16, 5.5), dpi=120)
    ttl = cycle_label or 'Current-Dependent EIS Analysis'
    fig.suptitle(ttl, fontsize=12, fontweight='bold')

    # ── Left: polcurves ──
    ax1.plot(j_pol, V_pol, 'ko-', ms=5, lw=1.5, label='Raw V(j)')
    ax1.plot(j_pol, V_irfree, 's-', color='#2196F3', ms=5, lw=1.5,
             label='iR-corrected')
    ax1.set_xlabel('j  [A/cm²]', fontsize=11)
    ax1.set_ylabel('Cell voltage  [V]', fontsize=11)
    ax1.set_xlim(left=0)
    ax1.legend(fontsize=9)
    ax1.grid(True, alpha=0.3)
    ax1.set_title('Polarization Curve', fontsize=10)

    # ── Center: ASR vs j ──
    ax2.plot(j_hfr, asr_hfr, 'o', ms=7, color='#d62728',
             markeredgecolor='k', markeredgewidth=0.5, label='ASR measured (R₀+R₁)')
    j_smooth = np.linspace(j_pol.min(), j_pol.max(), 200)
    asr_smooth = np.interp(j_smooth, j_hfr, asr_hfr)
    ax2.plot(j_smooth, asr_smooth, '--', color='#d62728', lw=1, alpha=0.6,
             label='Interpolated')
    ax2.set_xlabel('j  [A/cm²]', fontsize=11)
    ax2.set_ylabel('ASR  [mΩ·cm²]', fontsize=11)
    ax2.set_xlim(left=0)
    ax2.legend(fontsize=9)
    ax2.grid(True, alpha=0.3)
    ax2.set_title('ASR (R₀+R₁) vs. Current Density', fontsize=10)

    # ── Right: iR drop ──
    ir_drop = j_pol * asr_interp  # mV
    ax3.plot(j_pol, ir_drop, 'o-', ms=5, color='#4CAF50', lw=1.5)
    ax3.set_xlabel('j  [A/cm²]', fontsize=11)
    ax3.set_ylabel('iR drop  [mV]', fontsize=11)
    ax3.set_xlim(left=0)
    ax3.grid(True, alpha=0.3)
    ax3.set_title('Ohmic Loss', fontsize=10)

    plt.tight_layout()
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


# ═══════════════════════════════════════════════════════════════════
#  Transmission-line iR correction (coth model)
# ═══════════════════════════════════════════════════════════════════

def _coth(x):
    """Hyperbolic cotangent, safe for large x."""
    x = np.asarray(x, dtype=float)
    result = np.ones_like(x)
    small = np.abs(x) < 50
    result[small] = np.cosh(x[small]) / np.sinh(x[small])
    return result


def coth_cl_impedance(R1, R2):
    """
    Transmission line CL impedance: √(R1·R2) × coth(√(R2/R1))

    R1 = charge transfer resistance (Ω or Ω·cm²)
    R2 = CL ionic resistance (Ω or Ω·cm²)

    Returns the full CL impedance (includes R1 contribution).
    """
    if R1 <= 0 or R2 <= 0:
        return R1 + R2 / 3.0  # fallback
    xi = np.sqrt(R2 / R1)
    return np.sqrt(R1 * R2) * float(_coth(np.array([xi]))[0])


def coth_cl_ionic_only(R1, R2):
    """
    CL ionic voltage loss only (excluding charge transfer R1).

    V_CL_ionic = j × [√(R1·R2)·coth(√(R2/R1)) - R1]

    At low j (full utilization): ≈ R2/3
    At high j (under-utilized): ≈ √(R1·R2) - R1
    """
    Z_total = coth_cl_impedance(R1, R2)
    return max(0.0, Z_total - R1)


def compute_coth_corrections(j_pol, V_pol, eis_fits, geo_area=5.0):
    """
    Compute transmission-line iR correction for a polcurve using
    EIS circuit fit R0, R1, R2 from the lowest current density.

    Uses only the lowest-j EIS fit and assumes R0, R1, R2 are constant
    across the polcurve. This avoids pre-subtracting the j-dependent
    charge transfer resistance that Tafel should capture.

    Parameters
    ----------
    j_pol : array — polcurve current densities (A/cm²)
    V_pol : array — polcurve voltages (V)
    eis_fits : list of circuit fit dicts with 'j', 'R0_ohm', 'R1_ohm', 'R2_ohm'
    geo_area : float — cell area (cm²)

    Returns
    -------
    dict with arrays:
        j, V_raw, V_R0, V_R1, V_CL_ionic, V_ohmic_total, V_irfree,
        R0_interp, R1_interp, R2_interp (all in native units)
    """
    # Use the lowest-j EIS fit (constant R values)
    j_eis = np.array([f['j'] for f in eis_fits])
    idx_min = np.argmin(j_eis)

    R0 = eis_fits[idx_min]['R0_ohm']
    R1 = eis_fits[idx_min]['R1_ohm']
    R2 = eis_fits[idx_min]['R2_ohm']

    j_ref = j_eis[idx_min]
    print(f"    Using EIS at j = {j_ref:.4f} A/cm² for iR correction "
          f"(R₀={R0*geo_area*1000:.1f}, R₁={R1*geo_area*1000:.1f}, "
          f"R₂={R2*geo_area*1000:.1f} mΩ·cm²)")

    # Compute voltage losses at each j (constant R values)
    # V = j (A/cm²) × R (Ω) × A (cm²)  →  V
    V_R0 = j_pol * R0 * geo_area                  # membrane ohmic
    V_R1 = j_pol * R1 * geo_area                  # charge transfer
    V_CL_ionic = j_pol * coth_cl_ionic_only(R1, R2) * geo_area  # CL ionic (coth)
    V_ohmic_total = V_R0 + V_CL_ionic             # total non-kinetic ohmic (coth)

    # iR-free for Tafel fitting: subtract R0 + R1 + CL ionic (coth)
    V_irfree = V_pol - V_R0 - V_R1 - V_CL_ionic

    R0_arr = np.full_like(j_pol, R0 * geo_area * 1000)
    R1_arr = np.full_like(j_pol, R1 * geo_area * 1000)
    R2_arr = np.full_like(j_pol, R2 * geo_area * 1000)

    return {
        'j': j_pol,
        'V_raw': V_pol,
        'V_R0': V_R0,
        'V_R1': V_R1,
        'V_CL_ionic': V_CL_ionic,
        'V_ohmic_total': V_ohmic_total,
        'V_irfree': V_irfree,
        'R0_interp': R0_arr,  # mΩ·cm² (constant)
        'R1_interp': R1_arr,
        'R2_interp': R2_arr,
    }


def fit_tafel_irfree(j, V_irfree, T_C=80.0, p_cathode_barg=0.0,
                      p_anode_barg=0.0, j_min=0.01):
    """
    Fit Tafel kinetics + mass transport to an iR-corrected electrolyzer polcurve.

    V_irfree = E_rev + b_a × log10(j/j0_a) + b_c × log10(j/j0_c) + c_mt × j²

    where b_c is fixed at RT/(0.5·nF) and b_a, j0_a, j0_c, c_mt are fitted.

    Parameters
    ----------
    j : array — current density (A/cm²)
    V_irfree : array — iR-free voltage (V)
    T_C : float — temperature (°C)
    j_min : float — minimum j for fitting (A/cm²)

    Returns
    -------
    dict with fit results or None if fitting fails.
    """
    from scipy.optimize import least_squares

    E = E_rev(T_C, p_cathode_barg, p_anode_barg)
    T_K = T_C + 273.15

    mask = j >= j_min
    j_fit = j[mask]
    V_fit = V_irfree[mask]

    if len(j_fit) < 3:
        print(f"    Tafel fit: only {len(j_fit)} points with j >= {j_min}, skipping")
        return None

    # Model: V = E_rev + b_a·log10(j/j0_a) + b_c·log10(j/j0_c) + c_mt·j²
    # b_c fixed at RT/(0.5·nF)
    bc = (_R * T_K) / (0.5 * _n_e * _F) / np.log(10)  # V/decade

    def model(x):
        log_j0a, alpha_a, log_j0c, c_mt = x
        j0a = 10**log_j0a
        j0c = 10**log_j0c
        ba = (_R * T_K) / (alpha_a * _n_e * _F) / np.log(10)
        eta_a = np.where(j_fit > j0a, ba * np.log10(j_fit / j0a), 0.0)
        eta_c = np.where(j_fit > j0c, bc * np.log10(j_fit / j0c), 0.0)
        return E + eta_a + eta_c + c_mt * j_fit**2 - V_fit

    x0 = [-7.0, 0.5, -3.0, 0.0]
    lo = [-12.0, 0.2, -8.0, 0.0]
    hi = [-3.0, 2.0, -0.5, 0.1]

    try:
        res = least_squares(model, x0, bounds=(lo, hi), method='trf',
                            loss='soft_l1', f_scale=0.005)
    except Exception as e:
        print(f"    Tafel fit failed: {e}")
        return None

    if not res.success:
        print(f"    Tafel fit did not converge")
        return None

    log_j0a, alpha_a, log_j0c, c_mt = res.x
    j0a = 10**log_j0a
    j0c = 10**log_j0c
    ba_Vdec = (_R * T_K) / (alpha_a * _n_e * _F) / np.log(10)

    # Compute model over full j range
    j_smooth = np.linspace(max(j.min(), 1e-4), j.max() * 1.05, 200)
    V_model_smooth = E + np.where(j_smooth > j0a,
                                   ba_Vdec * np.log10(j_smooth / j0a), 0.0) \
                       + np.where(j_smooth > j0c,
                                   bc * np.log10(j_smooth / j0c), 0.0) \
                       + c_mt * j_smooth**2

    # Residuals on fit data
    V_model_fit = E + np.where(j_fit > j0a,
                                ba_Vdec * np.log10(j_fit / j0a), 0.0) \
                    + np.where(j_fit > j0c,
                                bc * np.log10(j_fit / j0c), 0.0) \
                    + c_mt * j_fit**2
    residuals = V_fit - V_model_fit
    rmse = np.sqrt(np.mean(residuals**2)) * 1000  # mV

    # Compute mass transport contribution at each polcurve j
    V_mt = c_mt * j**2

    return {
        'j0_a': j0a,
        'alpha_a': alpha_a,
        'ba_mVdec': ba_Vdec * 1000,
        'j0_c': j0c,
        'bc_mVdec': bc * 1000,
        'c_mt': c_mt,
        'E_rev': E,
        'rmse_mV': rmse,
        'j_fit': j_fit, 'V_fit': V_fit,
        'j_smooth': j_smooth, 'V_model_smooth': V_model_smooth,
        'V_mt': V_mt,
        'residuals': residuals,
    }


def plot_coth_analysis(coth_result, tafel_result, eis_fits, cycle_num=None,
                        geo_area=5.0, T_C=80.0, p_cathode_barg=0.0,
                        p_anode_barg=0.0, save_path=None):
    """
    Four-panel plot for transmission-line iR correction analysis:
      Top-left:  Polcurve — raw, R₀-corrected, fully corrected, Tafel fit
      Top-right: Loss breakdown vs j — stacked losses
      Bot-left:  R₀, R₁, R₂ vs j from EIS fits
      Bot-right: Tafel fit residuals
    """
    cr = coth_result
    j = cr['j']
    E = E_rev(T_C, p_cathode_barg, p_anode_barg)

    fig, axes = plt.subplots(2, 2, figsize=(14, 10), dpi=120)
    ax1, ax2, ax3, ax4 = axes.flat

    ttl = 'Transmission-Line iR Correction'
    if cycle_num is not None:
        ttl += f' — Cycle {cycle_num}'
    fig.suptitle(ttl, fontsize=13, fontweight='bold')

    # ── Top-left: Polcurves ──
    ax1.plot(j, cr['V_raw'], 'o-', ms=4, lw=1.5, color='#1f77b4',
             label='Raw polcurve')
    ax1.plot(j, cr['V_irfree'], '^-', ms=4, lw=1.5, color='#2ca02c',
             label='iR-free (R₀+R₁+CL coth)')

    if tafel_result is not None:
        ax1.plot(tafel_result['j_smooth'], tafel_result['V_model_smooth'],
                 '-', lw=2, color='#d62728', alpha=0.8,
                 label=f'Tafel fit (b_a={tafel_result["ba_mVdec"]:.0f}, '
                       f'b_c={tafel_result["bc_mVdec"]:.0f} mV/dec)')

    ax1.axhline(E, color='gray', ls=':', lw=1, alpha=0.5, label=f'E_rev = {E:.3f} V')
    ax1.set_xlabel('Current density  j  [A/cm²]', fontsize=10)
    ax1.set_ylabel('Cell voltage  [V]', fontsize=10)
    ax1.set_title('Polcurve & iR Correction', fontsize=10)
    ax1.legend(fontsize=7.5, loc='upper left')
    ax1.grid(True, alpha=0.3)
    ax1.set_xlim(left=0)

    # ── Top-right: Loss breakdown ──
    V_total_eta = cr['V_raw'] - E

    # Mass transport from Tafel fit (if available)
    V_mt = tafel_result['V_mt'] if tafel_result is not None else np.zeros_like(j)
    V_kinetic_residual = V_total_eta - cr['V_R0'] - cr['V_R1'] - cr['V_CL_ionic'] - V_mt
    V_kinetic_residual = np.maximum(V_kinetic_residual, 0)

    ax2.fill_between(j, 0, cr['V_R0'] * 1000, alpha=0.4, color='#4CAF50',
                     label='R₀ (membrane)')
    base = cr['V_R0'] * 1000
    ax2.fill_between(j, base, base + cr['V_R1'] * 1000, alpha=0.4,
                     color='#FF5722', label='R₁ (charge transfer)')
    base2 = base + cr['V_R1'] * 1000
    ax2.fill_between(j, base2, base2 + cr['V_CL_ionic'] * 1000, alpha=0.4,
                     color='#FF9800', label='CL ionic (coth)')
    base3 = base2 + cr['V_CL_ionic'] * 1000
    ax2.fill_between(j, base3, base3 + V_mt * 1000, alpha=0.4,
                     color='#9C27B0', label='Mass transport')
    base4 = base3 + V_mt * 1000
    ax2.fill_between(j, base4, base4 + V_kinetic_residual * 1000, alpha=0.4,
                     color='#E91E63', label='Kinetic (activation)')

    ax2.plot(j, cr['V_R0'] * 1000, '-', color='#4CAF50', lw=1.2)
    ax2.plot(j, (cr['V_R0'] + cr['V_R1']) * 1000, '-', color='#FF5722', lw=1.2)
    ax2.plot(j, (cr['V_R0'] + cr['V_R1'] + cr['V_CL_ionic']) * 1000,
             '-', color='#FF9800', lw=1.2)
    ax2.plot(j, (cr['V_R0'] + cr['V_R1'] + cr['V_CL_ionic'] + V_mt) * 1000,
             '-', color='#9C27B0', lw=1.2)
    ax2.plot(j, V_total_eta * 1000, 'o-', ms=3, color='k', lw=1.2,
             label='Total overpotential')

    ax2.set_xlabel('Current density  j  [A/cm²]', fontsize=10)
    ax2.set_ylabel('Voltage loss  [mV]', fontsize=10)
    ax2.set_title('Loss Breakdown', fontsize=10)
    ax2.legend(fontsize=7.5, loc='upper left')
    ax2.grid(True, alpha=0.3)
    ax2.set_xlim(left=0)

    # ── Bottom-left: R values vs j ──
    j_eis = [f['j'] for f in eis_fits]
    R0_eis = [f['R0_asr'] for f in eis_fits]
    R1_eis = [f['R1_asr'] for f in eis_fits]
    R2_eis = [f['R2_asr'] for f in eis_fits]

    ax3.plot(j_eis, R0_eis, 'o-', ms=6, lw=1.5, color='#4CAF50',
             label='R₀ (membrane)')
    ax3.plot(j_eis, R1_eis, 's-', ms=6, lw=1.5, color='#FF5722',
             label='R₁ (charge transfer)')
    ax3.plot(j_eis, R2_eis, '^-', ms=6, lw=1.5, color='#FF9800',
             label='R₂ (CL ionic)')

    # Mark where R1 = R2 (utilization transition)
    for i in range(len(j_eis) - 1):
        if (R1_eis[i] - R2_eis[i]) * (R1_eis[i+1] - R2_eis[i+1]) < 0:
            j_cross = j_eis[i] + (j_eis[i+1] - j_eis[i]) * \
                      (R1_eis[i] - R2_eis[i]) / ((R1_eis[i] - R2_eis[i]) -
                       (R1_eis[i+1] - R2_eis[i+1]))
            ax3.axvline(j_cross, color='gray', ls='--', lw=1, alpha=0.7)
            ax3.annotate(f'R₁=R₂ @ {j_cross:.2f}', (j_cross, max(R1_eis)),
                         fontsize=7, ha='center', va='bottom')

    ax3.set_xlabel('Current density  j  [A/cm²]', fontsize=10)
    ax3.set_ylabel('Resistance  [mΩ·cm²]', fontsize=10)
    ax3.set_title('EIS Fit Parameters vs. j', fontsize=10)
    ax3.legend(fontsize=8)
    ax3.grid(True, alpha=0.3)
    ax3.set_xlim(left=0)

    # ── Bottom-right: Tafel residuals ──
    if tafel_result is not None:
        ax4.stem(tafel_result['j_fit'], tafel_result['residuals'] * 1000,
                 linefmt='C0-', markerfmt='C0o', basefmt='k-')
        ax4.axhline(0, color='k', lw=0.5)
        ax4.set_xlabel('Current density  j  [A/cm²]', fontsize=10)
        ax4.set_ylabel('Residual  [mV]', fontsize=10)
        ax4.set_title(f'Tafel Fit Residuals (RMSE = {tafel_result["rmse_mV"]:.1f} mV)',
                      fontsize=10)
        ax4.grid(True, alpha=0.3)
    else:
        ax4.text(0.5, 0.5, 'Tafel fit unavailable', ha='center', va='center',
                 fontsize=12, color='gray', transform=ax4.transAxes)
        ax4.set_title('Tafel Fit Residuals', fontsize=10)

    plt.tight_layout()
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


# ═══════════════════════════════════════════════════════════════════
#  Electrolyzer model & fitting
# ═══════════════════════════════════════════════════════════════════

# Physical constants
_F   = 96485.3329   # C/mol
_R   = 8.31446      # J/(mol·K)
_n_e = 2            # electrons per H2O split


def E_rev(T_C=80.0, p_cathode_barg=0.0, p_anode_barg=0.0):
    """Nernst reversible potential (LeRoy & Bowen 1980 + pressure correction)."""
    T = T_C + 273.15
    E0 = 1.5184 - 1.5421e-3*T + 9.523e-5*T*np.log(T) + 9.84e-8*T**2
    p_H2 = p_cathode_barg + 1.01325
    p_O2 = p_anode_barg + 1.01325
    nernst = (_R * T) / (_n_e * _F) * np.log(p_H2 * p_O2**0.5)
    return E0 + nernst


def _electrolyzer_model(j, x, E, T_K):
    """V(j) = E_rev + η_a + η_c + j·ASR + c_mt·j²"""
    ASR, log_j0a, alpha_a, log_j0c, c_mt = x
    j0a, j0c = 10**log_j0a, 10**log_j0c
    ba = (_R * T_K) / (alpha_a * _n_e * _F)
    bc = (_R * T_K) / (0.5 * _n_e * _F)
    eta_a = np.where(j > j0a, ba * np.log10(j / j0a), 0.0)
    eta_c = np.where(j > j0c, bc * np.log10(j / j0c), 0.0)
    return E + eta_a + eta_c + j * ASR / 1000.0 + c_mt * j**2


def fit_polcurve(j_data, V_data, T_C=80.0, p_cathode_barg=0.0,
                 p_anode_barg=0.0, fix_ASR=None):
    """
    Fit electrolyzer model to a single polcurve.

    Model: V = E_rev(T,p) + η_a(Tafel) + η_c(Tafel) + j·ASR + c_mt·j²

    Parameters
    ----------
    j_data, V_data : arrays of current density [A/cm²] and voltage [V]
    T_C            : cell temperature [°C]
    p_cathode_barg : cathode H2 pressure [barg]
    p_anode_barg   : anode O2 pressure [barg]
    fix_ASR        : if set, pin ASR to this value [mΩ·cm²]

    Returns
    -------
    result : dict with fitted parameters, model curve, residuals, stats
    """
    from scipy.optimize import least_squares

    E = E_rev(T_C, p_cathode_barg, p_anode_barg)
    T_K = T_C + 273.15

    # Filter j > 0
    mask = j_data > 0
    j_fit = j_data[mask]
    V_fit = V_data[mask]

    if len(j_fit) < 3:
        print(f"  Skipping model fit: only {len(j_fit)} points with j > 0 "
              f"(need >= 3). j range: {j_data.min():.4f} to {j_data.max():.4f}")
        return None

    def model(j, x):
        return _electrolyzer_model(j, x, E, T_K)

    def residuals(x):
        return model(j_fit, x) - V_fit

    # x = [ASR_mOhm, log10(j0_a), alpha_a, log10(j0_c), c_mt]
    x0 = [70.0, -7.0, 0.5, -3.0, 0.0]
    lo = [10.0, -12.0, 0.2, -6.0, 0.0]
    hi = [500.0, -3.0, 2.0, -0.5, 0.05]

    if fix_ASR is not None:
        x0[0], lo[0], hi[0] = fix_ASR, fix_ASR - 0.01, fix_ASR + 0.01

    res = least_squares(residuals, x0, bounds=(lo, hi), method='trf',
                        loss='soft_l1', f_scale=0.01)

    xf = res.x
    V_model = model(j_fit, xf)
    rd = V_fit - V_model
    rmse = np.sqrt(np.mean(rd**2))

    # Compute individual loss components over a smooth j range
    j_smooth = np.linspace(0, j_fit.max() * 1.1, 500)
    j0a, j0c = 10**xf[1], 10**xf[3]
    ba = (_R * T_K) / (xf[2] * _n_e * _F)
    bc = (_R * T_K) / (0.5 * _n_e * _F)

    components = {
        'j': j_smooth,
        'E_rev': np.full_like(j_smooth, E),
        'eta_anode': np.where(j_smooth > j0a, ba * np.log10(j_smooth / j0a), 0.0),
        'eta_cathode': np.where(j_smooth > j0c, bc * np.log10(j_smooth / j0c), 0.0),
        'V_ohmic': j_smooth * xf[0] / 1000.0,
        'V_mt': xf[4] * j_smooth**2,
    }
    components['V_total'] = (components['E_rev'] + components['eta_anode'] +
                              components['eta_cathode'] + components['V_ohmic'] +
                              components['V_mt'])

    return {
        'x': xf, 'E_rev': E, 'T_C': T_C, 'T_K': T_K,
        'p_cathode_barg': p_cathode_barg, 'p_anode_barg': p_anode_barg,
        'j_data': j_fit, 'V_data': V_fit,
        'V_model': V_model, 'residual': rd,
        'rmse_mV': rmse * 1000,
        'mae_mV': np.mean(np.abs(rd)) * 1000,
        'max_err_mV': np.max(np.abs(rd)) * 1000,
        'components': components,
        'converged': res.success,
        'message': res.message,
    }


def print_fit_summary(fr):
    """Print fitted model parameters and loss breakdown at key current densities."""
    xf = fr['x']
    T_K = fr['T_K']

    print("=" * 60)
    print("  Electrolyzer Model Fit — Last Cycle")
    print("=" * 60)
    print(f"  Temperature          : {fr['T_C']:.1f} °C")
    print(f"  Cathode pressure     : {fr['p_cathode_barg']:.1f} barg")
    print(f"  Anode pressure       : {fr['p_anode_barg']:.1f} barg")
    print(f"  E_rev (fixed)        : {fr['E_rev']:.4f} V")
    print()
    print(f"  ASR_total            : {xf[0]:.1f} mΩ·cm²")
    print(f"  j0_anode             : {10**xf[1]:.3e} A/cm²")
    print(f"  α_anode              : {xf[2]:.3f}")
    print(f"  Anode Tafel slope    : {(_R*T_K)/(xf[2]*_n_e*_F)*1000:.1f} mV/dec")
    print(f"  j0_cathode           : {10**xf[3]:.3e} A/cm²")
    print(f"  Cathode Tafel slope  : {(_R*T_K)/(0.5*_n_e*_F)*1000:.1f} mV/dec (α_c = 0.5)")
    print(f"  c_mt                 : {xf[4]:.5f} V·cm⁴/A²")
    print()
    print(f"  RMSE                 : {fr['rmse_mV']:.2f} mV")
    print(f"  MAE                  : {fr['mae_mV']:.2f} mV")
    print(f"  Max |error|          : {fr['max_err_mV']:.2f} mV")
    print(f"  Converged            : {fr['converged']}")
    print()

    # Loss breakdown at key current densities
    E = fr['E_rev']
    V_tn = 1.481
    j0a, j0c = 10**xf[1], 10**xf[3]
    ba = (_R * T_K) / (xf[2] * _n_e * _F) * 1000  # mV/dec
    bc = (_R * T_K) / (0.5 * _n_e * _F) * 1000

    print(f"  {'j [A/cm²]':>12} {'V [V]':>10} {'η_a [mV]':>10} {'η_c [mV]':>10} "
          f"{'Ohm [mV]':>10} {'MT [mV]':>10} {'η_HHV [%]':>10}")
    print("  " + "-" * 75)

    for jj in [0.1, 0.5, 1.0, 2.0, 3.0, 4.0, 5.0]:
        if jj > fr['j_data'].max() * 1.2:
            continue
        j_arr = np.array([jj])
        V = _electrolyzer_model(j_arr, xf, E, T_K)[0]
        ea = ba/1000 * np.log10(jj / j0a) if jj > j0a else 0
        ec = bc/1000 * np.log10(jj / j0c) if jj > j0c else 0
        vo = jj * xf[0] / 1000
        vm = xf[4] * jj**2
        eff = V_tn / V * 100 if V > 0 else 0
        print(f"  {jj:12.2f} {V:10.4f} {ea*1000:10.1f} {ec*1000:10.1f} "
              f"{vo*1000:10.1f} {vm*1000:10.1f} {eff:10.1f}")
    print("=" * 60)
    print()


def plot_fit(fr, save_path=None):
    """
    Three-panel plot: stacked-area loss breakdown, data vs model, residuals.
    """
    comp = fr['components']
    j_m = comp['j']
    jd, Vd = fr['j_data'], fr['V_data']

    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(18, 5.5), dpi=120)
    fig.suptitle(
        f"Electrolyzer Model Fit — Last Cycle\n"
        f"RMSE = {fr['rmse_mV']:.1f} mV,  ASR = {fr['x'][0]:.1f} mΩ·cm²,  "
        f"T = {fr['T_C']:.0f} °C",
        fontsize=11, fontweight='bold')

    # Left: stacked area
    labels = ['E_rev', 'η anode (OER)', 'η cathode (HER)', 'Ohmic', 'Mass transport']
    colors = ['#2196F3', '#FF5722', '#FF9800', '#4CAF50', '#9C27B0']
    arrays = [comp['E_rev'], comp['eta_anode'], comp['eta_cathode'],
              comp['V_ohmic'], comp['V_mt']]
    ax1.stackplot(j_m, *arrays, labels=labels, colors=colors, alpha=0.7)
    ax1.plot(jd, Vd, 'ko', ms=5, label='Data', zorder=5)
    ax1.set_xlabel('j  [A/cm²]'); ax1.set_ylabel('V  [V]')
    ax1.set_xlim(0, j_m.max()); ax1.legend(loc='upper left', fontsize=7)
    ax1.grid(True, alpha=0.3); ax1.set_title('Loss breakdown', fontsize=10)

    # Center: data vs model
    ax2.plot(jd, Vd, 'ko', ms=5, label='Data', zorder=5)
    ax2.plot(j_m, comp['V_total'], 'r-', lw=2, label='Fitted model')
    ax2.set_xlabel('j  [A/cm²]'); ax2.set_ylabel('V  [V]')
    ax2.set_xlim(0, j_m.max()); ax2.legend(loc='upper left', fontsize=9)
    ax2.grid(True, alpha=0.3); ax2.set_title('Data vs. model', fontsize=10)

    # Right: residuals
    rm = fr['residual'] * 1000
    ax3.stem(jd, rm, linefmt='C0-', markerfmt='C0o', basefmt='k-')
    ax3.axhline(0, color='k', lw=0.5)
    ax3.axhspan(-fr['rmse_mV'], fr['rmse_mV'], alpha=0.15, color='green',
                label=f'±RMSE ({fr["rmse_mV"]:.1f} mV)')
    ax3.set_xlabel('j  [A/cm²]'); ax3.set_ylabel('Residual [mV]')
    ax3.legend(fontsize=9); ax3.grid(True, alpha=0.3)
    ax3.set_title('Residuals', fontsize=10)

    plt.tight_layout()
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig



# ═══════════════════════════════════════════════════════════════════
#  Per-cycle loss tracking
# ═══════════════════════════════════════════════════════════════════

def compute_losses_at_j(j_val, xf, T_K):
    """
    Compute individual voltage losses at a single current density
    from fitted parameters.

    Returns dict with eta_anode, eta_cathode, V_ohmic, V_mt in mV.
    """
    ASR, log_j0a, alpha_a, log_j0c, c_mt = xf
    j0a, j0c = 10**log_j0a, 10**log_j0c
    ba = (_R * T_K) / (alpha_a * _n_e * _F)
    bc = (_R * T_K) / (0.5 * _n_e * _F)

    eta_a = ba * np.log10(j_val / j0a) * 1000 if j_val > j0a else 0.0
    eta_c = bc * np.log10(j_val / j0c) * 1000 if j_val > j0c else 0.0
    v_ohm = j_val * ASR  # ASR in mΩ·cm², j in A/cm² → mV
    v_mt  = c_mt * j_val**2 * 1000  # V → mV

    return {'eta_anode_mV': eta_a, 'eta_cathode_mV': eta_c,
            'V_ohmic_mV': v_ohm, 'V_mt_mV': v_mt}


def extract_losses_vs_cycle(cycles, v_target, T_C=80.0,
                            p_cathode_barg=0.0, p_anode_barg=0.0,
                            fix_ASR=None, v_tol=0.015):
    """
    Fit each complete cycle and extract j + loss breakdown at v_target.

    Returns
    -------
    cycle_nums : list of cycle numbers
    j_values   : list of j at v_target [A/cm²]
    losses     : dict of loss name → list of values [mV]
    """
    from scipy.optimize import least_squares

    T_K = T_C + 273.15
    E = E_rev(T_C, p_cathode_barg, p_anode_barg)

    # Determine which cycles are "complete" (max setpoint count)
    max_pts = max(len(c) for c in cycles) if cycles else 0
    min_pts = max(5, int(max_pts * 0.8))

    cycle_nums = []
    j_values = []
    losses = {'eta_anode_mV': [], 'eta_cathode_mV': [],
              'V_ohmic_mV': [], 'V_mt_mV': []}

    print(f"\n  Fitting {sum(1 for c in cycles if len(c) >= min_pts)} cycles "
          f"for loss tracking at {v_target:.2f} V...")

    for ci, cyc in enumerate(cycles):
        if len(cyc) < min_pts:
            continue

        j_arr = np.array([d['j'] for d in cyc])
        V_arr = np.array([d['V'] for d in cyc])

        # Find j at target voltage — try direct match, then interpolation
        order = np.argsort(V_arr)
        V_sorted = V_arr[order]
        j_sorted = j_arr[order]

        dv = np.abs(V_sorted - v_target)
        best_idx = np.argmin(dv)

        if dv[best_idx] <= v_tol:
            best_j = j_sorted[best_idx]
        elif V_sorted.min() < v_target < V_sorted.max():
            best_j = float(np.interp(v_target, V_sorted, j_sorted))
        else:
            continue

        if best_j is None or best_j <= 0:
            continue

        # Fit this cycle (silent)
        mask = j_arr > 0
        j_fit, V_fit = j_arr[mask], V_arr[mask]
        if len(j_fit) < 4:
            continue

        def model(j, x):
            return _electrolyzer_model(j, x, E, T_K)

        x0 = [70.0, -7.0, 0.5, -3.0, 0.0]
        lo = [10.0, -12.0, 0.2, -6.0, 0.0]
        hi = [500.0, -3.0, 2.0, -0.5, 0.05]
        if fix_ASR is not None:
            x0[0], lo[0], hi[0] = fix_ASR, fix_ASR - 0.01, fix_ASR + 0.01

        try:
            res = least_squares(lambda x: model(j_fit, x) - V_fit,
                                x0, bounds=(lo, hi), method='trf',
                                loss='soft_l1', f_scale=0.01)
            if not res.success:
                continue
        except Exception:
            continue

        # Compute losses at the target j
        loss = compute_losses_at_j(best_j, res.x, T_K)

        cycle_nums.append(ci + 1)
        j_values.append(best_j)
        for k in losses:
            losses[k].append(loss[k])

    # Convert to arrays
    cycle_nums = np.array(cycle_nums)
    j_values = np.array(j_values)
    losses = {k: np.array(v) for k, v in losses.items()}

    print(f"    {len(cycle_nums)} cycles fitted successfully")

    return cycle_nums, j_values, losses


def plot_j_and_losses_vs_cycle(cycle_nums, j_values, losses,
                                v_target, save_path=None):
    """
    Dual-axis plot: j at target voltage (left) and voltage losses (right)
    vs cycle number.
    """
    fig, ax1 = plt.subplots(figsize=(10, 6), dpi=120)
    ax2 = ax1.twinx()

    # ── Left axis: j ──
    ax1.plot(cycle_nums, j_values, 'o-', color='#1f77b4', ms=5, lw=1.5,
             label=f'j @ {v_target:.2f} V')
    ax1.set_xlabel('Cycle number', fontsize=12)
    ax1.set_ylabel(f'Current density at {v_target:.2f} V  [A/cm²]',
                   fontsize=12, color='#1f77b4')
    ax1.tick_params(axis='y', labelcolor='#1f77b4')
    ax1.ticklabel_format(axis='y', useOffset=False)
    ax1.set_xlim(left=0)

    # ── Right axis: losses ──
    loss_styles = [
        ('V_ohmic_mV',      'Ohmic',            '#4CAF50', 's-'),
        ('eta_anode_mV',    'η anode (OER)',     '#FF5722', '^-'),
        ('eta_cathode_mV',  'η cathode (HER)',   '#FF9800', 'v-'),
        ('V_mt_mV',         'Mass transport',    '#9C27B0', 'D-'),
    ]

    for key, label, color, fmt in loss_styles:
        vals = losses[key]
        if np.max(vals) < 0.1:  # skip if negligible
            continue
        ax2.plot(cycle_nums, vals, fmt, color=color, ms=4, lw=1.2,
                 label=label, alpha=0.85)

    # Total kinetic losses (anode + cathode)
    eta_total_kinetic = losses['eta_anode_mV'] + losses['eta_cathode_mV']
    ax2.plot(cycle_nums, eta_total_kinetic, 'p-', color='#d62728', ms=5, lw=1.5,
             label='η kinetic (total)', alpha=0.85)

    ax2.set_ylabel('Voltage loss  [mV]', fontsize=12)
    ax2.tick_params(axis='y')

    # Combined legend below x-axis
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    fig.legend(lines1 + lines2, labels1 + labels2,
               loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=len(lines1) + len(lines2), fontsize=9,
               frameon=True, fancybox=True)

    ax1.grid(True, alpha=0.3)
    ax1.set_title(f'Current Density & Losses at {v_target:.2f} V vs. Cycle',
                  fontsize=12, fontweight='bold')

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


# ═══════════════════════════════════════════════════════════════════
#  Main pipeline
# ═══════════════════════════════════════════════════════════════════

def analyze(filepath, geo_area=5.0, save_dir=None, title=None,
            T_C=80.0, p_cathode_barg=0.0, p_anode_barg=0.0,
            eis_files=None, eis_ref_voltage=None, image_ext='png'):
    """Full pipeline: load → extract → cycles → EIS → plot → fit."""

    # Load polcurve
    data, fieldnames = load_data(filepath)
    cols = detect_columns(fieldnames)

    if cols['v_col'] is None:
        raise ValueError(f"Could not auto-detect voltage column.\n  Available: {fieldnames}")
    if cols['i_col'] is None:
        raise ValueError(f"Could not auto-detect current column.\n  Available: {fieldnames}")

    print(f"\n  Detected columns:")
    print(f"    Voltage : '{cols['v_col']}'")
    print(f"    Current : '{cols['i_col']}'")
    if cols['step_col']:   print(f"    Step    : '{cols['step_col']}'")
    if cols['repeat_col']: print(f"    Repeat  : '{cols['repeat_col']}'")
    if cols['time_col']:   print(f"    Time    : '{cols['time_col']}'")
    print(f"\n  Cell area : {geo_area:.2f} cm²")

    # Detect control mode
    mode = detect_control_mode(data, cols)
    print(f"  Control mode: {mode}")

    # Check for dual mode (both CC and CP in same file)
    is_dual = False
    if cols.get('step_name_col') is not None:
        names = data[cols['step_name_col']]
        unique_lower = set(n.lower().strip() for n in names)
        has_potential = any('constant potential' in n for n in unique_lower)
        has_current = any('constant current' in n for n in unique_lower)
        if has_potential and has_current:
            # Check if CP segments are real polcurves (multiple distinct V setpoints)
            # vs just recovery holds (single V)
            cp_voltages = set()
            for n_val, v_val in zip(names, data[cols['v_col']]):
                if 'constant potential' in n_val.lower():
                    cp_voltages.add(round(v_val, 2))
            if len(cp_voltages) > 2:
                is_dual = True
                print(f"  Dual mode detected: CC + CP polcurves in same file")

    # Extract dwells
    if cols['step_col'] and cols['repeat_col']:
        if is_dual:
            dwells = extract_dwells_from_steps(data, cols, geo_area, mode='dual')
            n_cc = sum(1 for d in dwells if d.get('mode') == 'galvanostatic')
            n_cp = sum(1 for d in dwells if d.get('mode') == 'potentiostatic')
            print(f"  Dwells extracted: {len(dwells)} ({n_cc} CC + {n_cp} CP)")
        else:
            dwells = extract_dwells_from_steps(data, cols, geo_area, mode=mode)
            print(f"  Dwells extracted: {len(dwells)} (from step/repeat structure)")
    else:
        dwells = extract_dwells_generic(data, cols, geo_area, mode=mode)
        grouping = 'current' if mode == 'galvanostatic' else 'voltage'
        print(f"  Dwells extracted: {len(dwells)} (from {grouping} grouping)")

    if not dwells:
        print("  ERROR: No valid dwells found.")
        return None, None, None

    # Free raw data — no longer needed after dwell extraction
    del data
    gc.collect()

    # Detect cycles
    if is_dual:
        cycles = detect_cycles_dual(dwells)
    else:
        cycles = detect_cycles(dwells, mode=mode)
    print(f"  Cycles detected: {len(cycles)}")
    for i, cyc in enumerate(cycles):
        V_lo, V_hi = min(d['V'] for d in cyc), max(d['V'] for d in cyc)
        j_lo, j_hi = min(d['j'] for d in cyc), max(d['j'] for d in cyc)
        mode_tag = cyc[0].get('cycle_label', '')
        tag_str = f' [{mode_tag}]' if mode_tag else ''
        print(f"    Cycle {i+1:3d}{tag_str}: {len(cyc):2d} pts, "
              f"V = {V_lo:.3f}–{V_hi:.3f} V, j = {j_lo:.3f}–{j_hi:.3f} A/cm²")

    # ── Select analysis targets based on mode ──
    is_galv = (mode == 'galvanostatic')

    if is_galv:
        j_primary = select_analysis_current(cycles)
        # Secondary: next lower candidate or half of primary
        j_candidates = [5.0, 4.5, 4.0, 3.5, 3.0, 2.5, 2.0, 1.5, 1.0]
        j_secondary = None
        for jc in j_candidates:
            if jc < j_primary:
                j_secondary = jc
                break
        # If primary ≤ 2.0, only plot one curve
        if j_primary <= 2.0:
            j_targets_galv = [j_primary]
        else:
            j_targets_galv = [j_primary] + ([j_secondary] if j_secondary else [])
        v_targets = [1.8, 1.7]  # kept for Excel fallback
        print(f"  Analysis current: {j_primary:.2f} A/cm²"
              + (f" (secondary: {j_secondary:.2f} A/cm²)" if j_secondary and len(j_targets_galv) > 1 else ""))
    else:
        v_primary = select_analysis_voltage(cycles, candidates=[1.8, 1.7, 1.6])
        v_secondary = 1.7 if v_primary == 1.8 else (1.6 if v_primary == 1.7 else 1.5)
        v_targets = [v_primary, v_secondary]
        print(f"  Analysis voltage: {v_primary:.2f} V "
              f"(secondary: {v_secondary:.2f} V)")

    # Resolve output paths
    import os
    if save_dir:
        os.makedirs(save_dir, exist_ok=True)
        _ext = image_ext or 'png'
        polcurve_path = os.path.join(save_dir, f'polcurve.{_ext}') if image_ext else None
        jvc_path = os.path.join(save_dir, f'j_vs_cycle.{_ext}') if image_ext else None
        xlsx_path = os.path.join(save_dir, 'analysis_data.xlsx')
        fit_path = os.path.join(save_dir, f'model_fit.{_ext}') if image_ext else None
        nyquist_path = os.path.join(save_dir, f'nyquist.{_ext}') if image_ext else None
        losses_path = os.path.join(save_dir, f'losses_vs_cycle.{_ext}') if image_ext else None
        ir_path = os.path.join(save_dir, f'ir_correction.{_ext}') if image_ext else None
    else:
        polcurve_path = jvc_path = xlsx_path = fit_path = nyquist_path = losses_path = ir_path = None

    # Plot polcurves
    plot_cycles(cycles, geo_area, title=title, save_path=polcurve_path)
    plt.close('all')
    gc.collect()

    # ── EIS analysis ──
    # Normalize eis_files to a list
    if eis_files is None:
        eis_files_list = []
    elif isinstance(eis_files, str):
        eis_files_list = [eis_files]
    else:
        eis_files_list = list(eis_files)

    eis_mapped = []
    fix_ASR = None
    eis_results_for_export = []
    eis_fit_results = []
    eis_circuit_by_cycle = {}

    if eis_files_list:
        print(f"\n  Loading {len(eis_files_list)} EIS file(s)...")
        eis_results = load_all_eis(eis_files_list, geo_area=geo_area)

        if eis_results:
            # ── Filter EIS by reference voltage ──
            if eis_ref_voltage is not None:
                v_tol = 0.02 * abs(eis_ref_voltage)  # 2% tolerance
                matched = [er for er in eis_results
                           if er['dc_v_mean'] is not None and
                           abs(er['dc_v_mean'] - eis_ref_voltage) <= v_tol]
                excluded = len(eis_results) - len(matched)
                print(f"\n  EIS voltage filter: ref = {eis_ref_voltage:.4f} V "
                      f"(±{v_tol*1000:.1f} mV, 2%)")
                print(f"    Matched : {len(matched)} file(s)")
                if excluded > 0:
                    matched_ids = set(id(m) for m in matched)
                    for er in eis_results:
                        if id(er) not in matched_ids:
                            v_str = f"{er['dc_v_mean']:.4f}" if er['dc_v_mean'] else "unknown"
                            print(f"    Excluded: {Path(er['file']).name} "
                                  f"(V = {v_str} V)")
                eis_results_for_tracking = matched
            else:
                eis_results_for_tracking = eis_results

            # Map filtered EIS to cycles by elapsed time
            eis_mapped = map_eis_to_cycles(eis_results_for_tracking, cycles)

            if eis_mapped:
                print(f"\n  EIS-to-cycle mapping:")
                for em in eis_mapped:
                    lbl = f"cycle {em['cycle']}" if em['cycle'] > 0 else "pre-conditioning"
                    print(f"    t = {em['t_eis']:.0f} s → {lbl}, "
                          f"ASR (extract_hfr) = {em['asr_mohm_cm2']:.1f} mΩ·cm²")

            eis_results_for_export = eis_results_for_tracking
            plt.close('all')

            # ── EIS equivalent circuit fitting ──
            if eis_results_for_tracking:
                print(f"\n  Fitting EIS equivalent circuit R₀-(R₁Q₁)-(R₂Q₂)...")
                eis_fit_results = []
                for ei, er in enumerate(eis_results_for_tracking):
                    eis_data = er['eis_data']
                    hfr_seed = er['hfr_ohm']
                    dc_v = er.get('dc_v_mean')
                    label = f"V_dc={dc_v:.3f}V" if dc_v else f"EIS {ei+1}"

                    efr = fit_eis_circuit(eis_data, geo_area=geo_area,
                                          hfr_seed=hfr_seed)
                    if efr is not None:
                        efr['dc_v'] = dc_v
                        efr['t_eis'] = er.get('t_eis')
                        efr['label'] = label
                        eis_fit_results.append(efr)
                        print_eis_fit_summary(efr, geo_area=geo_area)

                        if image_ext and save_dir:
                            if len(eis_results_for_tracking) == 1:
                                eis_fit_path = os.path.join(
                                    save_dir, f'eis_fit.{image_ext}')
                            else:
                                safe = label.replace('=', '').replace(' ', '_')
                                eis_fit_path = os.path.join(
                                    save_dir, f'eis_fit_{safe}.{image_ext}')
                            plot_eis_fit(eis_data, efr, geo_area=geo_area,
                                         title=f'EIS Fit — {label}',
                                         save_path=eis_fit_path)
                            plt.close('all')

                # Update eis_mapped and fix_ASR from circuit fits (R₀+R₁)
                if eis_fit_results:
                    # Update eis_mapped ASR with R₀+R₁ from circuit fits
                    # Match each mapped EIS to nearest circuit fit by elapsed time
                    for em in eis_mapped:
                        t_em = em.get('t_eis')
                        best_efr = None
                        best_dt = float('inf')
                        for efr in eis_fit_results:
                            t_efr = efr.get('t_eis')
                            if t_em is not None and t_efr is not None:
                                dt = abs(t_em - t_efr)
                                if dt < best_dt:
                                    best_dt = dt
                                    best_efr = efr
                        if best_efr is not None:
                            em['asr_mohm_cm2'] = best_efr['R0_asr'] + best_efr['R1_asr']
                        else:
                            em['asr_mohm_cm2'] = (eis_fit_results[-1]['R0_asr'] +
                                                   eis_fit_results[-1]['R1_asr'])

                    # Set fix_ASR from mean R₀+R₁ of tracking fits
                    r0r1_values = [efr['R0_asr'] + efr['R1_asr']
                                   for efr in eis_fit_results]
                    fix_ASR = float(np.mean(r0r1_values))
                    print(f"\n  ASR from circuit fit (R₀+R₁): {fix_ASR:.1f} mΩ·cm²")

                    # Plot Nyquist for last EIS with R₀+R₁ from fit
                    last_eis = eis_results_for_tracking[-1]
                    last_fit = eis_fit_results[-1]
                    plot_nyquist(last_eis['eis_data'],
                                 {'asr_mohm_cm2': last_fit['R0_asr'] + last_fit['R1_asr'],
                                  'hfr_ohm': (last_fit['R0_ohm'] + last_fit['R1_ohm'])},
                                 geo_area=geo_area, save_path=nyquist_path)
                    plt.close('all')
                else:
                    # No circuit fits succeeded — fall back to extract_hfr
                    if eis_results_for_tracking:
                        fix_ASR = eis_results_for_tracking[-1]['asr_mohm_cm2']
                        print(f"\n  ASR from extract_hfr (fallback): {fix_ASR:.1f} mΩ·cm²")
                        last_eis = eis_results_for_tracking[-1]
                        plot_nyquist(last_eis['eis_data'],
                                     {'asr_mohm_cm2': fix_ASR,
                                      'hfr_ohm': last_eis['hfr_ohm']},
                                     geo_area=geo_area, save_path=nyquist_path)
                        plt.close('all')

                if not eis_fit_results:
                    print("    No EIS spectra could be fitted")

    # ── Current-dependent EIS → circuit fit + iR correction per cycle ──
    ir_data_list = []
    if eis_files_list and eis_results:
        eis_groups = detect_current_dependent_eis(eis_results, cycles)
        for gi, group in enumerate(eis_groups):
            cd_cyc_idx = group['cycle_idx']
            eis_at_j = group['eis_at_j']

            # Fit equivalent circuit for each EIS FIRST to get R₀+R₁
            for ei, e in enumerate(eis_at_j):
                eis_data = e.get('eis_data')
                if eis_data is None:
                    continue
                label = f"Cycle{cd_cyc_idx+1}_j={e['j']:.3f}"
                efr = fit_eis_circuit(eis_data, geo_area=geo_area,
                                      hfr_seed=e['hfr_ohm'])
                if efr is not None:
                    efr['dc_v'] = e.get('dc_v_eis')
                    efr['j'] = e['j']
                    efr['label'] = label
                    eis_fit_results.append(efr)
                    print_eis_fit_summary(efr, geo_area=geo_area)

                    # Update eis_at_j with R₀+R₁ from circuit fit
                    e['asr_mohm_cm2'] = efr['R0_asr'] + efr['R1_asr']

                    if image_ext and save_dir:
                        safe = label.replace('=', '').replace(' ', '_')
                        eis_fit_path = os.path.join(
                            save_dir, f'eis_fit_{safe}.{image_ext}')
                        plot_eis_fit(eis_data, efr, geo_area=geo_area,
                                     title=f'EIS Fit — Cycle {cd_cyc_idx+1}, '
                                           f'j = {e["j"]:.3f} A/cm²',
                                     save_path=eis_fit_path)
                        plt.close('all')

            # Now compute iR correction using updated R₀+R₁ values
            ref_cyc = cycles[cd_cyc_idx]
            j_pol = np.array([d['j'] for d in ref_cyc])
            V_pol = np.array([d['V'] for d in ref_cyc])
            mask = j_pol > 0
            j_pol, V_pol = j_pol[mask], V_pol[mask]

            j_p, V_p, V_irf, j_h, asr_h, asr_i = compute_ir_corrected_polcurve(
                j_pol, V_pol, eis_at_j, geo_area=geo_area)

            if len(j_p) == 0 or len(j_h) == 0:
                print(f"  Skipping iR correction for cycle {cd_cyc_idx + 1}: "
                      f"no valid polcurve points in EIS range")
                continue

            ir_entry = {
                'cycle_idx': cd_cyc_idx,
                'cycle_num': cd_cyc_idx + 1,
                'j_pol': j_p, 'V_pol': V_p, 'V_irfree': V_irf,
                'j_hfr': j_h, 'asr_hfr': asr_h, 'asr_interp': asr_i,
            }
            ir_data_list.append(ir_entry)

            # Save per-group plots
            if image_ext and save_dir:
                import os
                _ext = image_ext
                if len(eis_groups) == 1:
                    ir_save = os.path.join(save_dir, f'ir_correction.{_ext}')
                    eis_save = os.path.join(save_dir, f'eis_for_ir.{_ext}')
                else:
                    ir_save = os.path.join(save_dir,
                                           f'ir_correction_cycle{cd_cyc_idx + 1}.{_ext}')
                    eis_save = os.path.join(save_dir,
                                            f'eis_for_ir_cycle{cd_cyc_idx + 1}.{_ext}')
                plot_ir_correction(
                    j_p, V_p, V_irf, j_h, asr_h, asr_i,
                    cycle_label=f'iR Correction — Cycle {cd_cyc_idx + 1}',
                    save_path=ir_save)
                plt.close('all')

                plot_eis_for_ir_correction(
                    eis_at_j, geo_area=geo_area,
                    cycle_num=cd_cyc_idx + 1,
                    save_path=eis_save)
                plt.close('all')

    # Legacy single ir_data for export (use first group if available)
    ir_data = ir_data_list if ir_data_list else None

    # Build per-cycle circuit fit lookup for loss decomposition
    from collections import defaultdict
    eis_circuit_by_cycle = defaultdict(list)
    for efr in eis_fit_results:
        if efr.get('j') is not None:
            # Find which cycle this fit belongs to by matching label
            label = efr.get('label', '')
            for ci in range(len(cycles)):
                if f'Cycle{ci+1}_' in label:
                    eis_circuit_by_cycle[ci].append(efr)
                    break

    # Update fix_ASR from current-dependent EIS circuit fits if available
    if eis_circuit_by_cycle:
        r0r1_values = []
        for ci, fits in eis_circuit_by_cycle.items():
            for f in fits:
                r0r1_values.append(f['R0_asr'] + f['R1_asr'])
        if r0r1_values:
            fix_ASR_cd = float(np.mean(r0r1_values))
            print(f"\n  ASR from current-dep EIS (R₀+R₁): {fix_ASR_cd:.1f} mΩ·cm²"
                  + (f" (was {fix_ASR:.1f} from tracking EIS)" if fix_ASR else ""))
            fix_ASR = fix_ASR_cd

    # ── Transmission-line (coth) iR correction per cycle ──
    coth_results = []
    if eis_circuit_by_cycle:
        print(f"\n  Transmission-line iR correction (coth model)...")
        for ci, fits in sorted(eis_circuit_by_cycle.items()):
            if len(fits) < 3:
                continue

            ref_cyc = cycles[ci]
            j_pol = np.array([d['j'] for d in ref_cyc])
            V_pol = np.array([d['V'] for d in ref_cyc])
            mask = j_pol > 0
            j_pol, V_pol = j_pol[mask], V_pol[mask]

            if len(j_pol) < 3:
                continue

            # Sort by j
            order = np.argsort(j_pol)
            j_pol, V_pol = j_pol[order], V_pol[order]

            cr = compute_coth_corrections(j_pol, V_pol, fits, geo_area=geo_area)

            # Fit Tafel to fully corrected curve
            tfr = fit_tafel_irfree(cr['j'], cr['V_irfree'], T_C=T_C,
                                    p_cathode_barg=p_cathode_barg,
                                    p_anode_barg=p_anode_barg)

            if tfr is not None:
                print(f"\n    Cycle {ci+1} Tafel fit:")
                print(f"      j₀,a = {tfr['j0_a']:.2e} A/cm²")
                print(f"      α_a  = {tfr['alpha_a']:.3f}")
                print(f"      b_a  = {tfr['ba_mVdec']:.1f} mV/dec")
                print(f"      j₀,c = {tfr['j0_c']:.2e} A/cm²")
                print(f"      b_c  = {tfr['bc_mVdec']:.1f} mV/dec (fixed α_c=0.5)")
                print(f"      c_mt = {tfr['c_mt']:.5f} V·cm⁴/A²")
                print(f"      RMSE = {tfr['rmse_mV']:.1f} mV")

            coth_results.append({
                'cycle_idx': ci,
                'coth_result': cr,
                'tafel_result': tfr,
                'eis_fits': fits,
            })

            if image_ext and save_dir:
                if len(eis_circuit_by_cycle) == 1:
                    coth_path = os.path.join(save_dir, f'coth_analysis.{image_ext}')
                else:
                    coth_path = os.path.join(save_dir,
                                              f'coth_analysis_cycle{ci+1}.{image_ext}')
                plot_coth_analysis(cr, tfr, fits, cycle_num=ci+1,
                                    geo_area=geo_area, T_C=T_C,
                                    p_cathode_barg=p_cathode_barg,
                                    p_anode_barg=p_anode_barg,
                                    save_path=coth_path)
                plt.close('all')

    # ── Plot j/V and ASR vs cycle ──
    if len(cycles) >= 2:
        if is_galv:
            if eis_mapped:
                plot_v_and_hfr_vs_cycle(cycles, j_targets_galv, eis_mapped,
                                         save_path=jvc_path)
            else:
                plot_v_vs_cycle(cycles, j_targets_galv, save_path=jvc_path)
        else:
            if eis_mapped:
                plot_j_and_hfr_vs_cycle(cycles, v_targets, eis_mapped,
                                         save_path=jvc_path)
            else:
                plot_j_vs_cycle(cycles, v_targets, save_path=jvc_path)
        plt.close('all')

    # ── Loss breakdown vs cycle ──
    loss_data = None
    eis_loss_data = None
    if len(cycles) >= 3:
        if is_galv:
            cn_loss, v_loss, losses = extract_losses_at_current(
                cycles, j_target=j_primary, T_C=T_C,
                p_cathode_barg=p_cathode_barg, p_anode_barg=p_anode_barg,
                fix_ASR=fix_ASR)
            if len(cn_loss) >= 2:
                loss_data = (cn_loss, v_loss, losses)
                plot_v_and_losses_vs_cycle(cn_loss, v_loss, losses,
                                            j_target=j_primary, save_path=losses_path)
                plt.close('all')
        else:
            cn_loss, j_loss, losses = extract_losses_vs_cycle(
                cycles, v_target=v_primary, T_C=T_C,
                p_cathode_barg=p_cathode_barg, p_anode_barg=p_anode_barg,
                fix_ASR=fix_ASR)
            if len(cn_loss) >= 2:
                loss_data = (cn_loss, j_loss, losses)
                plot_j_and_losses_vs_cycle(cn_loss, j_loss, losses,
                                           v_target=v_primary, save_path=losses_path)
                plt.close('all')

    # ── EIS-based loss decomposition (when circuit fits available) ──
    if eis_circuit_by_cycle and len(cycles) >= 2:
        if is_galv:
            cn_eis, v_eis, eis_losses = compute_eis_loss_decomposition(
                cycles, eis_circuit_by_cycle, j_target=j_primary,
                T_C=T_C, p_cathode_barg=p_cathode_barg,
                p_anode_barg=p_anode_barg, geo_area=geo_area)
            if len(cn_eis) >= 1:
                eis_loss_data = (cn_eis, v_eis, eis_losses)
                if image_ext and save_dir:
                    eis_loss_path = os.path.join(save_dir,
                                                  f'eis_losses_vs_cycle.{image_ext}')
                    plot_eis_losses_vs_cycle(cn_eis, v_eis, eis_losses,
                                             j_target=j_primary,
                                             save_path=eis_loss_path)
                    plt.close('all')
        else:
            # Potentiostatic EIS decomposition not yet implemented
            pass

    # ── Fit last complete cycle ──
    fr = None
    if cycles:
        max_pts = max(len(c) for c in cycles)
        full_cycles = [c for c in cycles if len(c) == max_pts]
        if not full_cycles:
            full_cycles = [c for c in cycles if len(c) >= max_pts * 0.8]

        if full_cycles:
            last_cyc = full_cycles[-1]
            cyc_num = cycles.index(last_cyc) + 1
            j_last = np.array([d['j'] for d in last_cyc])
            V_last = np.array([d['V'] for d in last_cyc])

            if fix_ASR:
                print(f"\n  Fitting cycle {cyc_num} ({len(last_cyc)} pts)"
                      f" with ASR (R₀+R₁) = {fix_ASR:.1f} mΩ·cm²...")
            else:
                print(f"\n  Fitting cycle {cyc_num} ({len(last_cyc)} pts)...")

            fr = fit_polcurve(j_last, V_last, T_C=T_C,
                              p_cathode_barg=p_cathode_barg,
                              p_anode_barg=p_anode_barg,
                              fix_ASR=fix_ASR)
            if fr is not None:
                print_fit_summary(fr)
                if fit_path:
                    plot_fit(fr, save_path=fit_path)
                    plt.close('all')
            else:
                print("  Model fit skipped — insufficient data points")

    # ── Export Excel ──
    if xlsx_path:
        export_excel(xlsx_path, cycles, v_targets=v_targets,
                     j_targets=j_targets_galv if is_galv else None,
                     eis_mapped=eis_mapped if eis_mapped else None,
                     loss_data=loss_data, fit_result=fr,
                     eis_results=eis_results_for_export if eis_results_for_export else None,
                     ir_data=ir_data,
                     geo_area=geo_area,
                     eis_fit_results=eis_fit_results if eis_fit_results else None,
                     eis_loss_data=eis_loss_data,
                     coth_results=coth_results if coth_results else None)

    return cycles, fr, eis_mapped


# ═══════════════════════════════════════════════════════════════════
#  CLI & Interactive
# ═══════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description='PEM Electrolyzer Polcurve Analyzer')
    ap.add_argument('--file', type=str, default=None, help='Path to polcurve CSV')
    ap.add_argument('--eis', type=str, nargs='*', default=None,
                    help='Path(s) to EIS CSV file(s)')
    ap.add_argument('--folder', type=str, default=None,
                    help='Folder containing polcurve (1_a1_...) and EIS (*a1*) files')
    ap.add_argument('--cell-id', type=str, default='a1',
                    help='Cell identifier for folder scanning (default: a1)')
    ap.add_argument('--eis-ref-voltage', type=float, default=None,
                    help='Reference DC voltage for EIS filtering [V] (only include '
                         'EIS measured at this voltage ±2%%)')
    ap.add_argument('--area', type=float, default=None,
                    help='Geometric electrode area [cm²]')
    ap.add_argument('--T', type=float, default=80.0, help='Cell temperature [°C]')
    ap.add_argument('--p-cathode', type=float, default=0.0,
                    help='Cathode H₂ pressure [barg]')
    ap.add_argument('--p-anode', type=float, default=0.0,
                    help='Anode O₂ pressure [barg]')
    ap.add_argument('--save-dir', type=str, default=None,
                    help='Directory to save plots and data')
    ap.add_argument('--title', type=str, default=None, help='Plot title')
    args = ap.parse_args()

    eis_ref_v = None

    # ── Resolve inputs ──
    if args.folder:
        # Folder mode
        geo_area = args.area
        if geo_area is None:
            area_str = input("  Geometric electrode area [cm²] (5.0): ").strip()
            geo_area = float(area_str) if area_str else 5.0
        fp, eis_fps = scan_folder(args.folder, cell_id=args.cell_id)
        if fp is None:
            print("  ERROR: No polcurve file found.")
            return
        eis_ref_v = args.eis_ref_voltage
        if eis_ref_v is None and eis_fps:
            v_str = input("  EIS reference voltage [V] for filtering (1.25): ").strip()
            eis_ref_v = float(v_str) if v_str else 1.25
        save = args.save_dir
        T_C, p_cath, p_an = args.T, args.p_cathode, args.p_anode
        title = args.title

    elif args.file:
        # Direct file mode
        fp = _clean_path(args.file)
        eis_fps = [_clean_path(e) for e in args.eis] if args.eis else []
        geo_area = args.area
        if geo_area is None:
            area_str = input("  Geometric electrode area [cm²] (5.0): ").strip()
            geo_area = float(area_str) if area_str else 5.0
        eis_ref_v = args.eis_ref_voltage
        if eis_ref_v is None and eis_fps:
            v_str = input("  EIS reference voltage [V] for filtering (1.25): ").strip()
            eis_ref_v = float(v_str) if v_str else 1.25
        save = args.save_dir
        T_C, p_cath, p_an = args.T, args.p_cathode, args.p_anode
        title = args.title

    else:
        # Interactive mode
        print("=" * 60)
        print("  PEM Electrolyzer Polarization Curve Analyzer")
        print("=" * 60)
        print()

        mode = input("  [1] Single polcurve file  [2] Folder with polcurve + EIS  (2): ").strip() or '2'

        if mode == '2':
            folder = input("  Folder path: ").strip()
            folder = _clean_path(folder)
            cell_id = input("  Cell ID (a1): ").strip() or 'a1'
            area_str = input("  Geometric electrode area [cm²] (5.0): ").strip()
            geo_area = float(area_str) if area_str else 5.0

            fp, eis_fps = scan_folder(folder, cell_id=cell_id)
            if fp is None:
                print("  ERROR: No polcurve file found.")
                return

            if eis_fps:
                v_str = input("  EIS reference voltage [V] for filtering (1.25): ").strip()
                eis_ref_v = float(v_str) if v_str else 1.25
        else:
            fp = input("  Polcurve data file path: ").strip()
            fp = _clean_path(fp)
            if not fp:
                print("  No file provided.")
                return

            eis_input = input("  EIS data file path (Enter = none): ").strip()
            eis_fps = [_clean_path(eis_input)] if eis_input else []

            area_str = input("  Geometric electrode area [cm²] (5.0): ").strip()
            geo_area = float(area_str) if area_str else 5.0

            if eis_fps:
                v_str = input("  EIS reference voltage [V] for filtering (1.25): ").strip()
                eis_ref_v = float(v_str) if v_str else 1.25

        print("\n  Operating conditions for model fit (Enter = default):")
        t_str = input("    Temperature [°C] (80.0): ").strip()
        T_C = float(t_str) if t_str else 80.0
        pc_str = input("    Cathode pressure [barg] (0.0): ").strip()
        p_cath = float(pc_str) if pc_str else 0.0
        pa_str = input("    Anode pressure [barg] (0.0): ").strip()
        p_an = float(pa_str) if pa_str else 0.0

        save = input("\n  Save directory (Enter = display only): ").strip()
        save = _clean_path(save) if save else None
        title = input("  Plot title (Enter = auto): ").strip() or None
        print()

    analyze(fp, geo_area=geo_area, save_dir=save, title=title,
            T_C=T_C, p_cathode_barg=p_cath, p_anode_barg=p_an,
            eis_files=eis_fps if eis_fps else None,
            eis_ref_voltage=eis_ref_v)


if __name__ == '__main__':
    main()
