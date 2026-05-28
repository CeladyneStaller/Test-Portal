"""
Generic Plot Comparison Tool
============================
Reads sidecar JSON files written by `save_with_sidecar()` and overlays
them in a same-shape figure with distinct colors per source.

Auto-groups selected plots by `plot_type` (the filename stem) and runs
one comparison per group via the generic overlay renderer.

Custom per-type generators may be registered in COMPARISON_GENERATORS;
unknown types fall back to the generic overlay renderer.
"""

import os
import json
from pathlib import Path

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from scripts.helpers.plot_compare import (
    find_sidecar, load_sidecar, render_overlay_comparison
)


def _parse_condition_signature(filename):
    """Extract a condition signature from a filename for matching.

    Captures:
      - Cell letter prefix (e.g. 'a' from 'a10', 'b' from 'b14')
      - Temperature (e.g. '80C')
      - Relative humidity (e.g. '100RH')
      - Pressure (e.g. '0kPa', '150kPa')

    Returns a string like 'b_80C_100RH_0kPa', or '' if no conditions found.
    Two filenames with the same signature are considered "same conditions".

    Example:
      'polcurve_b14_IV_80C_100RH_OCV_0o3V_0o2H2_0o2Air_0kPa.png'
      → 'b_80C_100RH_0kPa'
      'polcurve_b8_IV_80C_100RH_OCV_0o3V_0o2H2_0o2Air_0kPa.png'
      → 'b_80C_100RH_0kPa'  (same: only number differs)
      'polcurve_a10_IV_80C_100RH_OCV_0o3V_0o2H2_0o2Air_0kPa.png'
      → 'a_80C_100RH_0kPa'  (different: letter differs)
    """
    import re
    if not filename:
        return ''

    # Strip path and extension
    name = os.path.basename(filename)
    name = os.path.splitext(name)[0]

    parts = []

    # Cell letter prefix: pattern like 'b14', 'a10', 'c2' — letter followed by digits
    # We want the letter only. Use a word-boundary match.
    cell_match = re.search(r'(?<![A-Za-z])([A-Za-z])(\d+)(?![A-Za-z])', name)
    if cell_match:
        parts.append(cell_match.group(1).lower())

    # Temperature: digits followed by C (with optional o/° decimal)
    t_match = re.search(r'(\d+(?:o\d+)?)C(?![A-Za-z])', name)
    if t_match:
        parts.append(f'{t_match.group(1)}C')

    # Relative humidity: digits followed by RH
    rh_match = re.search(r'(\d+(?:o\d+)?)RH', name, re.IGNORECASE)
    if rh_match:
        parts.append(f'{rh_match.group(1)}RH')

    # Pressure: digits followed by kPa, barg, or psi
    p_match = re.search(r'(\d+(?:o\d+)?)\s*(kPa|barg|psi|bar)', name, re.IGNORECASE)
    if p_match:
        parts.append(f'{p_match.group(1)}{p_match.group(2)}')

    return '_'.join(parts) if parts else ''


def _parse_full_conditions(filename):
    """Extract all parseable conditions from a filename.

    Returns dict with possible keys:
      step (e.g. 'b14'), T (numeric), RH (numeric),
      P_value (numeric), P_unit (str like 'kPa'),
      H2 (numeric), Air, N2, O2, V (numeric), V_unit
    """
    import re
    if not filename:
        return {}

    name = os.path.splitext(os.path.basename(filename))[0]

    def parse_num(s):
        try:
            return float(s.replace('o', '.'))
        except (ValueError, TypeError):
            return None

    cond = {}

    # Step (cell letter + digits, e.g. b14, c6)
    m = re.search(r'(?<![A-Za-z])([A-Za-z])(\d+)(?![A-Za-z])', name)
    if m:
        cond['step'] = m.group(0).lower()

    # Temperature
    m = re.search(r'(\d+(?:o\d+)?)C(?![A-Za-z])', name)
    if m:
        v = parse_num(m.group(1))
        if v is not None:
            cond['T'] = v

    # Relative humidity
    m = re.search(r'(\d+(?:o\d+)?)RH', name, re.IGNORECASE)
    if m:
        v = parse_num(m.group(1))
        if v is not None:
            cond['RH'] = v

    # Pressure
    m = re.search(r'(\d+(?:o\d+)?)\s*(kPa|barg|psi|bar)', name, re.IGNORECASE)
    if m:
        v = parse_num(m.group(1))
        if v is not None:
            cond['P_value'] = v
            cond['P_unit'] = m.group(2)

    # Gas flows
    for key, pat in (('H2', r'(\d+(?:o\d+)?)H2(?![A-Za-z])'),
                      ('Air', r'(\d+(?:o\d+)?)Air'),
                      ('N2', r'(\d+(?:o\d+)?)N2(?![A-Za-z])'),
                      ('O2', r'(\d+(?:o\d+)?)O2(?![A-Za-z])')):
        m = re.search(pat, name, re.IGNORECASE)
        if m:
            v = parse_num(m.group(1))
            if v is not None:
                cond[key] = v

    # Voltage setpoint (digits followed by V, not preceded by alpha,
    # not part of a unit like 'kV')
    m = re.search(r'(?<![A-Za-z])(\d+(?:o\d+)?)V(?![A-Za-z])', name)
    if m:
        v = parse_num(m.group(1))
        if v is not None:
            cond['V'] = v

    return cond


def _format_conditions_subtitle(cond):
    """Format conditions dict as a human-readable subtitle.

    Example: '80°C | 100% RH | H₂: 0.2 / Air: 0.2 slpm | 0kPa'
    """
    parts = []
    if 'T' in cond:
        parts.append(f"{cond['T']:g}°C")
    if 'RH' in cond:
        parts.append(f"{cond['RH']:g}% RH")

    gas_parts = []
    if 'H2' in cond:
        gas_parts.append(f"H₂: {cond['H2']:g}")
    if 'Air' in cond:
        gas_parts.append(f"Air: {cond['Air']:g}")
    if 'N2' in cond:
        gas_parts.append(f"N₂: {cond['N2']:g}")
    if 'O2' in cond:
        gas_parts.append(f"O₂: {cond['O2']:g}")
    if gas_parts:
        parts.append(' / '.join(gas_parts) + ' slpm')

    if 'P_value' in cond:
        parts.append(f"{cond['P_value']:g} {cond['P_unit']}")

    return ' | '.join(parts)


def _friendly_plot_type(plot_type):
    """Map sidecar plot_type to a human-readable name for plot titles."""
    mapping = {
        'polcurve': 'Polarization Curve',
        'polcurve_overlay': 'Polarization Curve Overlay',
        'polcurve_down': 'Polarization Curve (Downswing)',
        'polcurve_down_overlay': 'Polarization Curve Overlay (Downswing)',
        'eis': 'EIS',
        'eis_fit': 'EIS Fit',
        'eis_for_ir': 'EIS (for iR Correction)',
        'nyquist': 'Nyquist',
        'nyquist_overlay': 'Nyquist Overlay',
        'crossover': 'H₂ Crossover',
        'activation': 'Activation',
        'activation_sequential': 'Activation (Sequential)',
        'ecsa_hupd': 'ECSA (HUPD)',
        'ecsa_co': 'ECSA (CO Stripping)',
        'ecsa_degradation': 'ECSA Degradation',
        'ecsa_overlay': 'ECSA Overlay',
        'ocv': 'OCV',
        'ocv_overlay': 'OCV Overlay',
        'losses_vs_cycle': 'Losses vs Cycle',
        'j_vs_cycle': 'Current Density vs Cycle',
        'model_fit': 'Model Fit',
        'ir_correction': 'iR Correction',
        'CLR_analysis': 'Catalyst Layer Resistance',
        'CLR_model_fit': 'CLR Model Fit',
        'durability_voltage': 'Durability — Voltage vs Time',
        'durability_hfr': 'Durability — HFR vs Time',
        'durability_degradation': 'Durability — Degradation Rate',
        'durability_polcurve_evolution': 'Durability — Polcurve Evolution',
        'durability_asr_evolution': 'Durability — ASR Evolution',
        'durability_loss_evolution': 'Durability — Loss Evolution',
        'cleaning_cycles': 'Electrode Cleaning Cycles',
        'cleaning_diagnostics': 'Cleaning Diagnostics',
    }
    return mapping.get(plot_type, plot_type.replace('_', ' ').title())


def _build_clean_label(sample_name, filename, grouping_mode):
    """Build a minimal-info legend label.

    grouping_mode 'plot_type' (All Together):
        {sample_name}_{step}_{T}C_{RH}RH_{P}{P_unit}
    grouping_mode 'plot_type_conditions' (Same Conditions):
        {sample_name}_{step}
    """
    if not sample_name:
        sample_name = 'sample'
    cond = _parse_full_conditions(filename)
    step = cond.get('step', '')

    parts = [sample_name]
    if step:
        parts.append(step)

    if grouping_mode != 'plot_type_conditions':
        cond_parts = []
        if 'T' in cond:
            cond_parts.append(f"{cond['T']:g}C")
        if 'RH' in cond:
            cond_parts.append(f"{cond['RH']:g}RH")
        if 'P_value' in cond:
            cond_parts.append(f"{cond['P_value']:g}{cond['P_unit']}")
        if cond_parts:
            parts.append('_'.join(cond_parts))

    return '_'.join(parts)


def _parse_metrics_from_text(text_str):
    """Parse 'KEY = VALUE' style readouts into a dict.

    Handles patterns like:
      'OCV = 0.95 V'
      'Peak P = 560 mW/cm²'
      'V @ 1 A/cm² = 0.559 V'
      'b = 5851 mV/dec\nj₀ = 1.40e-02 A/cm²'
    """
    import re
    result = {}
    if not text_str:
        return result
    # Split on newlines, then look for "key = value" or "key: value"
    for line in text_str.split('\n'):
        line = line.strip()
        if not line:
            continue
        # Try " = " first, then ":"
        m = re.match(r'^(.+?)\s*=\s*(.+)$', line) or re.match(r'^(.+?):\s*(.+)$', line)
        if m:
            key = m.group(1).strip()
            val = m.group(2).strip()
            if key and val:
                result[key] = val
    return result


def export_comparison_excel(items, plot_type, filepath):
    """Side-by-side Excel: Metrics sheet + one sheet per source axis."""
    wb = Workbook()
    wb.remove(wb.active)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    label_fill = PatternFill("solid", fgColor="2E75B6")
    label_font = Font(bold=True, color="FFFFFF")

    ref_axes = items[0]['sidecar'].get('data', {}).get('axes', [])
    primary_axes = [a for a in ref_axes if not a.get('is_twin')]

    # ─────────────────────────────────────────────────────────────────
    # Sheet 1: Metrics summary parsed from plot text annotations + ref lines
    # ─────────────────────────────────────────────────────────────────
    sample_metrics = {}  # {sample_label: {metric_key: value}}
    for item in items:
        s_label = item['label']
        sample_metrics[s_label] = {}
        s_axes = [a for a in item['sidecar'].get('data', {}).get('axes', [])
                  if not a.get('is_twin')]
        for ax_idx, s_ax in enumerate(s_axes):
            ax_title = primary_axes[ax_idx].get('title', '') if ax_idx < len(primary_axes) else ''
            ax_prefix = f'[{ax_title}] ' if ax_title else f'[Axis {ax_idx+1}] '
            # 1. Parse from text box annotations
            for txt in s_ax.get('texts', []):
                parsed = _parse_metrics_from_text(txt.get('text', ''))
                for k, v in parsed.items():
                    sample_metrics[s_label][f'{ax_prefix}{k}'] = v
            # 2. Parse from axhline/axvline labels (e.g. EIS HFR markers)
            for ref_line in s_ax.get('axhlines', []) + s_ax.get('axvlines', []):
                lbl = (ref_line.get('label') or '').strip()
                if lbl and '=' in lbl:
                    parsed = _parse_metrics_from_text(lbl)
                    for k, v in parsed.items():
                        sample_metrics[s_label][f'{ax_prefix}{k}'] = v

    # Collect union of all metric keys across samples
    all_keys = []
    seen = set()
    for s_label, mets in sample_metrics.items():
        for k in mets:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    if all_keys:  # only create Metrics sheet if any metrics found
        ws = wb.create_sheet('Metrics')
        # Header row: Sample | metric1 | metric2 | ...
        ws.cell(row=1, column=1, value='Sample').font = hdr_font
        ws.cell(row=1, column=1).fill = hdr_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        for ki, k in enumerate(all_keys):
            c = ws.cell(row=1, column=ki + 2, value=k)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(ki + 2)].width = 22
        ws.column_dimensions['A'].width = 28
        # Data rows
        for ri, item in enumerate(items):
            s_label = item['label']
            ws.cell(row=ri + 2, column=1, value=s_label).font = Font(bold=True)
            for ki, k in enumerate(all_keys):
                v = sample_metrics.get(s_label, {}).get(k, '')
                # Try numeric conversion
                try:
                    # Strip units (e.g. "0.95 V" → 0.95)
                    import re as _re
                    m = _re.match(r'^([-+]?\d+\.?\d*(?:[eE][-+]?\d+)?)', str(v).strip())
                    if m:
                        ws.cell(row=ri + 2, column=ki + 2, value=float(m.group(1)))
                    else:
                        ws.cell(row=ri + 2, column=ki + 2, value=str(v))
                except Exception:
                    ws.cell(row=ri + 2, column=ki + 2, value=str(v))
        ws.row_dimensions[1].height = 30

    # ─────────────────────────────────────────────────────────────────
    # Per-axis trace data sheets
    # ─────────────────────────────────────────────────────────────────
    for ax_idx, ref_ax in enumerate(primary_axes):
        sheet_name = f'Axis {ax_idx + 1}'
        if ref_ax.get('title'):
            t = str(ref_ax['title'])[:25]
            sheet_name = f'A{ax_idx + 1} - {t}'[:31]
        ws = wb.create_sheet(sheet_name)

        col = 1
        for item in items:
            label = item['label']
            s_axes = [a for a in item['sidecar'].get('data', {}).get('axes', [])
                      if not a.get('is_twin')]
            if ax_idx >= len(s_axes):
                continue
            s_ad = s_axes[ax_idx]
            traces = s_ad.get('lines', []) + s_ad.get('scatters', [])
            if not traces:
                continue

            n_cols_per_sample = 2 * len(traces)
            col_end = col + n_cols_per_sample - 1

            cell = ws.cell(row=1, column=col, value=label)
            cell.font = label_font
            cell.fill = label_fill
            cell.alignment = Alignment(horizontal='center')
            if col_end > col:
                ws.merge_cells(start_row=1, start_column=col,
                               end_row=1, end_column=col_end)

            xlabel = s_ad.get('xlabel', 'x')
            ylabel = s_ad.get('ylabel', 'y')

            for ti, trace in enumerate(traces):
                trace_label = trace.get('label') or f'Series {ti + 1}'
                tcol = col + 2 * ti
                tc = ws.cell(row=2, column=tcol, value=trace_label)
                tc.font = Font(bold=True, italic=True, color="333333")
                tc.alignment = Alignment(horizontal='center')
                ws.merge_cells(start_row=2, start_column=tcol,
                               end_row=2, end_column=tcol + 1)
                for cc, h in enumerate([xlabel, ylabel]):
                    hc = ws.cell(row=3, column=tcol + cc, value=h)
                    hc.font = hdr_font
                    hc.fill = hdr_fill
                    hc.alignment = Alignment(horizontal='center')
                    ws.column_dimensions[get_column_letter(tcol + cc)].width = 14

                xs = trace.get('x', [])
                ys = trace.get('y', [])
                for ri, (x, y) in enumerate(zip(xs, ys)):
                    if x is not None and isinstance(x, (int, float)):
                        ws.cell(row=ri + 4, column=tcol, value=round(float(x), 6))
                    elif x is not None:
                        ws.cell(row=ri + 4, column=tcol, value=x)
                    if y is not None and isinstance(y, (int, float)):
                        ws.cell(row=ri + 4, column=tcol + 1, value=round(float(y), 6))
                    elif y is not None:
                        ws.cell(row=ri + 4, column=tcol + 1, value=y)

            col = col_end + 2

    if not wb.sheetnames:
        wb.create_sheet('Empty')
    wb.save(filepath)
    print(f"  Excel exported: {filepath}")


def _compare_generic(items, output_dir, params):
    """Default generator: overlay + Excel. Writes outputs into a
    subfolder named after the plot_type so the portal UI auto-groups them."""
    plot_type = items[0]['sidecar'].get('plot_type', 'plot')
    fname_suffix = params.get('_filename_suffix', '')
    grouping_mode = params.get('grouping_mode', 'plot_type')
    image_format = params.get('image_format', 'png')

    # Build friendly title (e.g. 'Polarization Curve Comparison').
    # If user provided a custom title in the modal, prefer that.
    user_title = params.get('title', '').strip()
    friendly = _friendly_plot_type(plot_type)
    if user_title and not user_title.startswith('Comparison ('):
        # User typed something custom — use it as-is (no friendly mapping)
        title = user_title
    else:
        title = f'{friendly} Comparison'

    # Subtitle: only for "Same Conditions" mode — describes the shared
    # conditions of the group. Parsed from the first item's filename
    # (all items in the group share the same condition signature).
    subtitle = None
    if grouping_mode == 'plot_type_conditions':
        first_filename = items[0].get('filename', '')
        cond = _parse_full_conditions(first_filename)
        subtitle = _format_conditions_subtitle(cond)

    # Build per-row labels:
    #   - If the source dict carries a non-empty 'label' (user typed an
    #     override in the modal), use it verbatim.
    #   - Otherwise auto-generate the clean default from
    #     sample_name + filename + grouping_mode.
    # This allows mixed rows: some manually labeled, others auto-defaulted.
    clean_items = []
    for it in items:
        sample_name = it.get('sample_name', '') or ''
        filename = it.get('filename', '')
        user_label = (it.get('label') or '').strip()
        if user_label:
            label_to_use = user_label
        else:
            label_to_use = _build_clean_label(sample_name, filename, grouping_mode)
        clean_items.append({
            'label': label_to_use,
            'filename': filename,
            'sidecar': it['sidecar'],
            'sample_name': sample_name,
        })

    # Group outputs by plot_type into subfolders so the portal UI can
    # display "Activation", "Polarization Curve", "EIS", etc. as separate
    # buckets when a single comparison job spans multiple plot types.
    subfolder = _plot_type_to_subfolder(plot_type)
    sub_dir = os.path.join(output_dir, subfolder)
    os.makedirs(sub_dir, exist_ok=True)

    print(f"    {len(clean_items)} samples (output → {subfolder}/):")
    for it in clean_items:
        print(f"      - {it['label']}")
    if subtitle:
        print(f"    Subtitle: {subtitle}")

    out_files = []

    if image_format and image_format != 'none':
        plot_path = os.path.join(sub_dir,
                                 f'{plot_type}{fname_suffix}.{image_format}')
        try:
            fig = render_overlay_comparison(clean_items, save_path=plot_path,
                                             title=title, subtitle=subtitle)
            if fig:
                plt.close(fig)
                out_files.append(os.path.relpath(plot_path, output_dir))
        except Exception as e:
            print(f"    ✗ Failed to render comparison: {e}")
            import traceback; traceback.print_exc()

    try:
        xlsx_path = os.path.join(sub_dir,
                                 f'{plot_type}{fname_suffix}.xlsx')
        export_comparison_excel(clean_items, plot_type, xlsx_path)
        out_files.append(os.path.relpath(xlsx_path, output_dir))
    except Exception as e:
        print(f"    ✗ Failed to export Excel: {e}")
        import traceback; traceback.print_exc()

    return out_files


def _plot_type_to_subfolder(plot_type):
    """Map a sidecar plot_type to an output subfolder name.

    Different plot_types from the same characterization share a folder so
    the portal UI buckets them together. E.g., 'eis', 'eis_fit',
    'nyquist', 'nyquist_overlay' all live in 'eis/'.
    """
    pt = (plot_type or 'plot').lower()
    # Explicit mapping for known characterization buckets
    if pt.startswith('activation'):
        return 'activation'
    if pt.startswith('cleaning'):
        return 'cleaning'
    if pt.startswith('crossover') or 'h2x' in pt:
        return 'crossover'
    if pt.startswith('eis') or pt.startswith('nyquist'):
        return 'eis'
    if pt.startswith('ecsa'):
        return 'ecsa'
    if pt.startswith('ocv'):
        return 'ocv'
    if pt.startswith('polcurve'):
        return 'polcurve'
    if pt.startswith('durability'):
        return 'durability'
    if pt.startswith('clr') or pt.startswith('coth'):
        return 'CLR'
    if pt in ('losses_vs_cycle', 'j_vs_cycle', 'model_fit', 'ir_correction'):
        return 'polcurve'
    # Fallback: use the plot_type itself as the folder name (sanitized)
    safe = ''.join(c if c.isalnum() or c in '-_' else '_' for c in pt)
    return safe or 'other'


COMPARISON_GENERATORS = {
    # Override generic for specific plot types here if needed.
}


def run(input_dir, output_dir, params=None):
    p = params or {}
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    sources_raw = p.get('sources', '[]')
    if isinstance(sources_raw, str):
        try:
            sources = json.loads(sources_raw)
        except Exception as e:
            raise RuntimeError(f"Invalid sources JSON: {e}")
    else:
        sources = sources_raw

    if not sources:
        raise RuntimeError("No sources provided")

    # Grouping mode:
    #   'plot_type'             — group by plot_type only (default — all data combined)
    #   'plot_type_conditions'  — also split by parsed condition signature
    grouping_mode = p.get('grouping_mode', 'plot_type')

    print(f"\n{'='*60}")
    print(f"  Plot Comparison")
    print(f"{'='*60}")
    print(f"  Total selected: {len(sources)}")
    print(f"  Grouping mode: {grouping_mode}")

    # Map: group_key → list of items
    # group_key is either plot_type or (plot_type, condition_sig)
    by_group = {}
    skipped = []
    for src in sources:
        label = src.get('label', src.get('job_id', '?'))
        filename = src.get('filename', '')
        output_dir_s = src.get('output_dir', '')

        sidecar_path = find_sidecar(output_dir_s, filename)
        if sidecar_path is None:
            print(f"  ✗ {label} ({filename}): no sidecar")
            skipped.append(f"{label}/{filename}")
            continue
        sidecar = load_sidecar(sidecar_path)
        if sidecar is None:
            skipped.append(f"{label}/{filename}")
            continue

        plot_type = sidecar.get('plot_type', 'unknown')

        if grouping_mode == 'plot_type_conditions':
            cond_sig = _parse_condition_signature(filename)
            group_key = (plot_type, cond_sig)
        else:
            group_key = (plot_type, '')

        by_group.setdefault(group_key, []).append({
            'label': label,
            'filename': filename,
            'sidecar': sidecar,
            'sample_name': src.get('sample_name', ''),
        })

    if not by_group:
        raise RuntimeError(
            "No comparable plot data found. Re-run source analyses with the "
            "latest scripts that write sidecar JSON files."
            + (f"\nSkipped: {', '.join(skipped)}" if skipped else "")
        )

    print(f"\n  Grouped:")
    for (ptype, csig), items in by_group.items():
        sig_part = f" [{csig}]" if csig else ""
        print(f"    {ptype}{sig_part}: {len(items)} items")

    all_output_files = []
    for (plot_type, cond_sig), items in by_group.items():
        if len(items) < 2:
            sig_part = f" [{cond_sig}]" if cond_sig else ""
            print(f"\n  Skipping '{plot_type}{sig_part}': only {len(items)} item")
            continue
        gen = COMPARISON_GENERATORS.get(plot_type, _compare_generic)
        # Pass condition sig in params so output filenames include it
        gen_params = dict(p)
        if cond_sig:
            gen_params['_filename_suffix'] = f'_{cond_sig}'
            gen_params['_title_suffix'] = f' [{cond_sig}]'
        sig_part = f" [{cond_sig}]" if cond_sig else ""
        print(f"\n  Generating '{plot_type}{sig_part}' comparison...")
        try:
            out_files = gen(items, str(out), gen_params)
            all_output_files.extend(out_files)
        except Exception as e:
            print(f"    ✗ Generator failed: {e}")
            import traceback; traceback.print_exc()

    if not all_output_files:
        raise RuntimeError(
            "No comparison outputs generated. Check that ≥2 plots of the "
            "same type were selected and their sidecars are valid."
        )

    # Collect all output files (recursively, since each plot type now goes
    # into its own subfolder for the portal UI grouping).
    output_files = [str(f.relative_to(out)) for f in out.rglob('*')
                    if f.is_file() and '_plot_data' not in f.parts]
    return {
        'status': 'success',
        'files_produced': output_files,
        'plot_types': sorted(set(k[0] for k in by_group.keys())),
    }