#!/usr/bin/env python3
"""
PEM Electrolyzer Current Hold (Durability) Analyzer
====================================================

Analyzes long-duration constant-current hold tests with periodic
EIS and polarization curve interruptions.

File classification by "Step name" column:
  - Current hold:  "Constant Current"
  - EIS:           contains "EIS"
  - Polarization:  "Constant Potential"

Usage:
  python electrolyzer_durability.py --folder data/ --area 25.0
  python electrolyzer_durability.py   # interactive
"""

# ─── Imports ───────────────────────────────────────────────────
import argparse, csv, gc
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scripts.helpers.plot_compare import save_with_sidecar
from pathlib import Path


def run(input_dir: str, output_dir: str, params: dict = None) -> dict:
    """Analyze PEM electrolyzer durability (current hold) data from sequential folders."""
    import traceback as tb
    p = params or {}

    inp = Path(input_dir)
    geo_area = float(p.get('geo_area', 25.0))
    eis_ref_v = p.get('eis_ref_voltage', '1.25')
    eis_ref_v = float(eis_ref_v) if eis_ref_v not in ('', None) else 1.25
    data_interval = p.get('data_interval_min', '')
    data_interval = float(data_interval) if data_interval not in ('', None) else None

    # Discover folder structure:
    # Could be: input/folder1/*.csv, input/folder2/*.csv (sequential folders)
    # Or:       input/parent/folder1/*.csv, input/parent/folder2/*.csv
    # Or:       input/*.csv (single folder)

    # Find all data files recursively
    all_files = sorted([f for f in inp.rglob('*')
                        if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv')])

    if not all_files:
        return {"status": "error", "message": "No data files found"}

    # Group files by their parent directory
    folders_with_data = {}
    for f in all_files:
        parent = f.parent
        folders_with_data.setdefault(str(parent), []).append(f)

    # If user specified folder_order, use that order
    user_order = p.get('folder_order', [])
    if user_order and isinstance(user_order, list) and len(user_order) > 0:
        # Map user-supplied relative folder paths to absolute paths
        ordered_paths = []
        for rel_folder in user_order:
            # Find the matching absolute path
            matched = None
            for abs_path in folders_with_data:
                if abs_path.endswith(rel_folder) or rel_folder in abs_path:
                    matched = abs_path
                    break
            if matched:
                ordered_paths.append(matched)
        folder_paths = ordered_paths if ordered_paths else sorted(folders_with_data.keys())
    elif len(folders_with_data) == 1 and str(inp) in folders_with_data:
        # All files in input_dir itself
        folder_paths = [str(inp)]
    else:
        # Sort folders alphabetically (fallback)
        folder_paths = sorted(folders_with_data.keys())

    print(f"  Discovered {len(folder_paths)} data folder(s):")
    for i, fp in enumerate(folder_paths):
        n = len(folders_with_data.get(fp, []))
        print(f"    {i+1}. {Path(fp).name}/ ({n} files)")

    from scripts.helpers.conditions import img_ext_from_params
    image_ext = img_ext_from_params(p)

    try:
        analyze(folder_paths, geo_area=geo_area,
                eis_ref_voltage=eis_ref_v,
                T_C=80.0, v_target=1.8,
                save_dir=str(output_dir),
                data_interval_min=data_interval,
                image_ext=image_ext)
        plt.close('all')
    except Exception as e:
        raise RuntimeError(
            f"Durability analysis failed: {e}\n{tb.format_exc()}"
        )

    out = Path(output_dir)
    output_files = [str(f.relative_to(out)) for f in out.rglob('*') if f.is_file()]
    if not output_files:
        raise RuntimeError(
            f"Analysis produced no output. Found {len(all_files)} data file(s) "
            f"in {len(folder_paths)} folder(s)."
        )
    return {"status": "success", "folders_processed": len(folder_paths),
            "files_produced": output_files}


# ═══════════════════════════════════════════════════════════════════
#  Utilities
# ═══════════════════════════════════════════════════════════════════

def _clean_path(p):
    if p is None: return None
    p = p.strip()
    if p.startswith('& '): p = p[2:]
    elif p.startswith('&'): p = p[1:]
    while len(p) >= 2 and ((p[0] == '"' and p[-1] == '"') or
                            (p[0] == "'" and p[-1] == "'")):
        p = p[1:-1]
    return p.strip().strip('\u2018\u2019\u201c\u201d\u202a\u200b')


def _get_header(filepath):
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        line = f.readline()
    delim = '\t' if '\t' in line else ','
    return [h.strip() for h in line.strip().split(delim)], delim


# ═══════════════════════════════════════════════════════════════════
#  Column detection
# ═══════════════════════════════════════════════════════════════════

def detect_columns(fieldnames):
    result = {k: None for k in (
        'v_col', 'i_col', 'step_col', 'repeat_col',
        'step_name_col', 'time_col',
        'freq_col', 'zre_col', 'zim_col', 'dc_v_col')}
    mapping = {
        'v_col':         ['working electrode (v)', 'voltage (v)', 'ewe (v)', 'potential (v)'],
        'i_col':         ['current (a)', 'i (a)', 'current(a)'],
        'step_col':      ['step number', 'step_number', 'step no'],
        'repeat_col':    ['repeats', 'repeat', 'cycle number'],
        'step_name_col': ['step name', 'step_name', 'technique'],
        'time_col':      ['elapsed time', 'time (s)', 'elapsed_time'],
        'freq_col':      ['frequency (hz)', 'freq (hz)', 'frequency(hz)'],
        'zre_col':       ["z' (ohms)", "z'(ohms)", "zre (ohms)", "z' (ohm)"],
        'zim_col':       ['-z" (ohms)', '-z"(ohms)', "-z'' (ohms)", '-z" (ohm)'],
        'dc_v_col':      ['dc working electrode (v)', 'dc voltage', 'dc potential'],
    }
    for fn in fieldnames:
        fl = fn.lower().strip()
        for key, cands in mapping.items():
            if result[key] is not None: continue
            if key == 'i_col' and 'density' in fl: continue
            for c in cands:
                if c in fl:
                    result[key] = fn; break
    return result


# ═══════════════════════════════════════════════════════════════════
#  Unified file loader — reads once, splits by step type
# ═══════════════════════════════════════════════════════════════════

def load_and_split_file(filepath, geo_area):
    """
    Load a CSV once, split into current_hold / EIS / polcurve segments
    by "Step name". Current converted: j = |I_total(A)| / geo_area(cm2).
    Uses raw elapsed time (s) — no offset applied.
    """
    headers, delim = _get_header(filepath)
    cols = detect_columns(headers)

    has_echem = cols['v_col'] is not None and cols['i_col'] is not None
    has_eis = cols['freq_col'] is not None and cols['zre_col'] is not None
    has_step_name = cols['step_name_col'] is not None
    has_time = cols['time_col'] is not None

    if not has_echem and not has_eis:
        return {'current_hold': [], 'eis': [], 'polcurve': []}, 'unknown'

    keep = set(v for v in cols.values() if v is not None)
    col_idx = {}
    raw = {}

    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        hdr = [h.strip() for h in f.readline().strip().split(delim)]
        for name in keep:
            if name in hdr:
                col_idx[name] = hdr.index(name)
                raw[name] = []
        n_cols = len(hdr)
        for line in f:
            parts = line.strip().split(delim)
            if len(parts) < n_cols: continue
            for name, idx in col_idx.items():
                raw[name].append(parts[idx].strip())

    data = {}
    for name, vals in raw.items():
        try:
            data[name] = np.array([float(x) if x.lower() not in ('nan', '')
                                   else np.nan for x in vals])
        except ValueError:
            data[name] = np.array(vals)
    del raw
    n_rows = len(data[list(data.keys())[0]])

    step_names = (data[cols['step_name_col']]
                  if has_step_name and cols['step_name_col'] in data
                  else np.array(['unknown'] * n_rows))

    segments = {'current_hold': [], 'eis': [], 'polcurve': []}
    types_found = set()

    # ── Current hold (Constant Current) ──
    if has_echem and has_time:
        cc_mask = np.array([str(s).lower().strip() == 'constant current'
                            for s in step_names])
        if np.sum(cc_mask) > 0:
            types_found.add('current_hold')
            segments['current_hold'].append({
                't_s': data[cols['time_col']][cc_mask],
                'voltage': data[cols['v_col']][cc_mask],
                'j_Acm2': np.abs(data[cols['i_col']][cc_mask]) / geo_area,
            })

    # ── EIS ──
    if has_eis:
        eis_mask = np.array(['eis' in str(s).lower() for s in step_names])
        if np.sum(eis_mask) > 0:
            types_found.add('eis')
            freq_all = data[cols['freq_col']][eis_mask]
            zre_all = data[cols['zre_col']][eis_mask]
            zim_all = data[cols['zim_col']][eis_mask]
            t_eis = data[cols['time_col']][eis_mask] if has_time else None
            dcv_eis = (data[cols['dc_v_col']][eis_mask]
                       if cols['dc_v_col'] and cols['dc_v_col'] in data else None)

            # Split sweeps by frequency resets (jump from low to high)
            split_idx = [0]
            for k in range(1, len(freq_all)):
                if not np.isnan(freq_all[k]) and not np.isnan(freq_all[k-1]):
                    if freq_all[k] > freq_all[k-1] * 5:
                        split_idx.append(k)
            split_idx.append(len(freq_all))

            for si in range(len(split_idx) - 1):
                s, e = split_idx[si], split_idx[si + 1]
                if e - s < 5: continue
                freq = freq_all[s:e]; zre = zre_all[s:e]; zim = zim_all[s:e]
                t_sw = t_eis[s:e] if t_eis is not None else None
                dcv = dcv_eis[s:e] if dcv_eis is not None else None

                valid = ~np.isnan(freq) & (freq > 0)
                if np.sum(valid) < 5: continue
                freq, zre, zim = freq[valid], zre[valid], zim[valid]
                if t_sw is not None: t_sw = t_sw[valid]
                if dcv is not None: dcv = dcv[valid]

                order = np.argsort(freq)[::-1]
                freq, zre, zim = freq[order], zre[order], zim[order]
                if t_sw is not None: t_sw = t_sw[order]
                if dcv is not None: dcv = dcv[order]

                segments['eis'].append({
                    'freq': freq, 'zre': zre, 'zim': zim, 'time': t_sw,
                    'dc_v_mean': float(np.nanmean(dcv)) if dcv is not None else None,
                    't_mid_s': float(np.nanmean(t_sw)) if t_sw is not None else None,
                })

    # ── Polcurve (Constant Potential) ──
    if has_echem and has_time and cols['step_col'] in data and cols['repeat_col'] in data:
        cp_mask = np.array([str(s).lower().strip() == 'constant potential'
                            for s in step_names])
        if np.sum(cp_mask) > 0:
            types_found.add('polcurve')
            V = data[cols['v_col']][cp_mask]
            I_A = data[cols['i_col']][cp_mask]
            t_s = data[cols['time_col']][cp_mask]
            step = data[cols['step_col']][cp_mask].astype(int)
            rep = data[cols['repeat_col']][cp_mask].astype(int)

            setpoints = []
            prev_key, start = None, 0
            for idx in range(len(V)):
                key = (step[idx], rep[idx])
                if key != prev_key:
                    if prev_key is not None and idx - start >= 3:
                        tail = start + (idx - start) // 2
                        V_sp = float(np.nanmedian(V[tail:idx]))
                        j_sp = float(np.nanmedian(np.abs(I_A[tail:idx]))) / geo_area
                        t_sp = float(np.nanmean(t_s[tail:idx]))
                        setpoints.append({'V': V_sp, 'j': j_sp, 't_s': t_sp})
                    start = idx; prev_key = key
            if prev_key is not None and len(V) - start >= 3:
                tail = start + (len(V) - start) // 2
                V_sp = float(np.nanmedian(V[tail:]))
                j_sp = float(np.nanmedian(np.abs(I_A[tail:]))) / geo_area
                t_sp = float(np.nanmean(t_s[tail:]))
                setpoints.append({'V': V_sp, 'j': j_sp, 't_s': t_sp})

            if setpoints:
                v_arr = np.array([s['V'] for s in setpoints])
                v_span = v_arr.max() - v_arr.min() if len(v_arr) > 1 else 0.1
                sweeps = [[]]
                for k, sp in enumerate(setpoints):
                    if k > 0 and v_arr[k] < v_arr[k-1] - 0.3 * v_span:
                        sweeps.append([])
                    sweeps[-1].append(sp)
                for sw in sweeps:
                    if len(sw) >= 3:
                        sw.sort(key=lambda d: d['V'])
                        segments['polcurve'].append(sw)

    del data; gc.collect()
    if len(types_found) == 0: ftype = 'unknown'
    elif len(types_found) == 1: ftype = types_found.pop()
    else: ftype = 'mixed'
    return segments, ftype


# ═══════════════════════════════════════════════════════════════════
#  Folder scanner
# ═══════════════════════════════════════════════════════════════════

def scan_folder(folder_path):
    folder = Path(_clean_path(str(folder_path)))
    if not folder.is_dir():
        raise FileNotFoundError(f"Folder not found: {folder}")
    csvs = sorted([str(f) for f in folder.iterdir()
                   if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv')])
    print(f"\n  Scanning: {folder}")
    print(f"  Data files: {len(csvs)}")
    return csvs


# ═══════════════════════════════════════════════════════════════════
#  HFR extraction
# ═══════════════════════════════════════════════════════════════════

def extract_hfr(eis_sweep, geo_area):
    freq, zre, zim = eis_sweep['freq'], eis_sweep['zre'], eis_sweep['zim']
    hfr, f_hfr = None, None
    for i in range(len(zim) - 1):
        if zim[i] <= 0 and zim[i + 1] > 0:
            frac = -zim[i] / (zim[i + 1] - zim[i])
            hfr = zre[i] + frac * (zre[i + 1] - zre[i])
            f_hfr = freq[i] + frac * (freq[i + 1] - freq[i])
            break
    if hfr is None:
        min_idx = np.argmin(np.abs(zim))
        hfr, f_hfr = zre[min_idx], freq[min_idx]
    return hfr, hfr * geo_area * 1000, f_hfr


# ═══════════════════════════════════════════════════════════════════
#  Degradation rate
# ═══════════════════════════════════════════════════════════════════

def compute_degradation_rate(t_hours, voltage, window_hours=50):
    from scipy.stats import linregress
    res = linregress(t_hours, voltage)
    rate_uV_hr = res.slope * 1e6  # V/hr → μV/hr
    # 1 μV/hr = 1e-6 V/hr; 1 mV/khr = 1e-3 V / 1000 hr = 1e-6 V/hr → same number
    rate_mV_khr = rate_uV_hr
    V0 = res.intercept
    # V%/khr = (slope / V0) × 100% × 1000 hr/khr
    rate_Vpct_khr = (res.slope / V0 * 100 * 1000) if abs(V0) > 0.01 else 0.0
    print(f"\n  Degradation rate (linear):")
    print(f"    {rate_uV_hr:.1f} μV/hr  ({rate_mV_khr:.1f} mV/khr)")
    print(f"    {rate_Vpct_khr:.2f} V%/khr  (V₀ = {V0:.4f} V)")
    print(f"    R² = {res.rvalue**2:.4f}")
    t_r, r_r = [], []
    t0, tN = t_hours[0], t_hours[-1]
    ts = t0
    while ts + window_hours <= tN:
        m = (t_hours >= ts) & (t_hours < ts + window_hours)
        if np.sum(m) > 10:
            r = linregress(t_hours[m], voltage[m])
            t_r.append(ts + window_hours / 2); r_r.append(r.slope * 1e6)
        ts += window_hours / 2
    return {'rate_uV_hr': rate_uV_hr, 'rate_mV_khr': rate_mV_khr,
            'rate_Vpct_khr': rate_Vpct_khr, 'V0': V0,
            'slope': res.slope, 'intercept': res.intercept,
            'r_squared': res.rvalue**2,
            't_rolling': np.array(t_r), 'rate_rolling': np.array(r_r)}


# ═══════════════════════════════════════════════════════════════════
#  Electrolyzer model
# ═══════════════════════════════════════════════════════════════════

_F, _R, _n_e = 96485.3329, 8.31446, 2

def E_rev(T_C=80.0, p_cath=0.0, p_an=0.0):
    T = T_C + 273.15
    E0 = 1.5184 - 1.5421e-3*T + 9.523e-5*T*np.log(T) + 9.84e-8*T**2
    return E0 + (_R*T)/(_n_e*_F)*np.log((p_cath+1.01325)*(p_an+1.01325)**0.5)

def _model(j, x, E, T_K):
    ASR, lj0a, aa, lj0c, cmt = x
    j0a, j0c = 10**lj0a, 10**lj0c
    ba = (_R*T_K)/(aa*_n_e*_F); bc = (_R*T_K)/(0.5*_n_e*_F)
    ea = np.where(j > j0a, ba*np.log10(j/j0a), 0.0)
    ec = np.where(j > j0c, bc*np.log10(j/j0c), 0.0)
    return E + ea + ec + j*ASR/1000 + cmt*j**2

def fit_polcurve(j, V, T_C=80.0, fix_ASR=None):
    from scipy.optimize import least_squares
    E = E_rev(T_C); T_K = T_C + 273.15
    m = j > 0; jf, Vf = j[m], V[m]
    if len(jf) < 4: return None
    x0 = [70, -7, 0.5, -3, 0]
    lo = [10, -12, 0.2, -6, 0]; hi = [500, -3, 2, -0.5, 0.05]
    if fix_ASR: x0[0], lo[0], hi[0] = fix_ASR, fix_ASR-0.01, fix_ASR+0.01
    try:
        r = least_squares(lambda x: _model(jf, x, E, T_K) - Vf,
                          x0, bounds=(lo, hi), method='trf', loss='soft_l1', f_scale=0.01)
        if not r.success: return None
    except: return None
    Vm = _model(jf, r.x, E, T_K)
    return {'x': r.x, 'E_rev': E, 'T_K': T_K,
            'rmse_mV': np.sqrt(np.mean((Vf-Vm)**2))*1000,
            'j_data': jf, 'V_data': Vf, 'V_model': Vm}

def compute_losses(j_val, xf, T_K):
    ASR, lj0a, aa, lj0c, cmt = xf
    j0a, j0c = 10**lj0a, 10**lj0c
    ba = (_R*T_K)/(aa*_n_e*_F)*1000; bc = (_R*T_K)/(0.5*_n_e*_F)*1000
    ea = ba*np.log10(j_val/j0a) if j_val > j0a else 0
    ec = bc*np.log10(j_val/j0c) if j_val > j0c else 0
    return {'eta_anode_mV': ea, 'eta_cathode_mV': ec, 'eta_kinetic_mV': ea+ec,
            'V_ohmic_mV': j_val*ASR, 'V_mt_mV': cmt*j_val**2*1000}


# ═══════════════════════════════════════════════════════════════════
#  Plotting
# ═══════════════════════════════════════════════════════════════════

def plot_voltage_vs_time(t_h, V, deg, j_hold, pc_t=None, save_path=None):
    fig, ax = plt.subplots(figsize=(12, 5.5), dpi=120)
    step = max(1, len(t_h)//5000)
    ax.plot(t_h[::step], V[::step], '-', color='#1f77b4', lw=0.5, alpha=0.6, label='Cell voltage')
    tf = np.array([t_h[0], t_h[-1]])
    V0 = deg['intercept']
    Vpct_khr = (deg['slope'] / V0) * 100 * 1000  # V%/khr
    ax.plot(tf, V0+deg['slope']*tf, '--', color='#d62728', lw=2,
            label=f"Fit: {deg['rate_uV_hr']:.1f} μV/hr ({Vpct_khr:.2f} V%/khr)")
    if pc_t:
        for t in pc_t: ax.axvline(t, color='#2ca02c', alpha=0.3, lw=0.8)
        ax.axvline(pc_t[0], color='#2ca02c', alpha=0.3, lw=0.8, label='Characterization')
    ax.set_xlabel('Time [hours]', fontsize=12); ax.set_ylabel('Cell voltage [V]', fontsize=12)
    ax.set_title(f'Durability Test — j = {j_hold:.2f} A/cm²', fontsize=12, fontweight='bold')
    ax.set_xlim(left=0); ax.grid(True, alpha=0.3)
    h, l = ax.get_legend_handles_labels()
    fig.legend(h, l, loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=len(h), fontsize=9, frameon=True, fancybox=True)
    plt.tight_layout(); plt.subplots_adjust(bottom=0.18)
    if save_path: save_with_sidecar(fig, save_path, bbox_inches='tight'); print(f"  Plot saved: {save_path}")
    else: plt.show()

def plot_deg_rolling(deg, save_path=None):
    if len(deg['t_rolling']) < 2: return
    fig, ax = plt.subplots(figsize=(10, 5), dpi=120)
    ax.plot(deg['t_rolling'], deg['rate_rolling'], 'o-', color='#d62728', ms=4, lw=1.5)
    ax.axhline(deg['rate_uV_hr'], color='k', ls='--', lw=1,
               label=f"Overall: {deg['rate_uV_hr']:.1f} μV/hr")
    ax.set_xlabel('Time [hours]'); ax.set_ylabel('Rate [μV/hr]')
    ax.set_title('Rolling Degradation Rate', fontweight='bold')
    ax.set_xlim(left=0); ax.legend(); ax.grid(True, alpha=0.3); plt.tight_layout()
    if save_path: save_with_sidecar(fig, save_path, bbox_inches='tight'); print(f"  Plot saved: {save_path}")
    else: plt.show()

def plot_hfr_vs_time(eis_p, save_path=None):
    if not eis_p: return
    fig, ax = plt.subplots(figsize=(10, 5), dpi=120)
    ax.plot([e['t_hours'] for e in eis_p], [e['asr_mohm_cm2'] for e in eis_p],
            'o-', color='#9467bd', ms=6, lw=1.5, markeredgecolor='k', markeredgewidth=0.5)
    ax.set_xlabel('Time [hours]'); ax.set_ylabel('ASR [mΩ·cm²]')
    ax.set_title('HFR (ASR) vs. Time', fontweight='bold')
    ax.set_xlim(left=0); ax.grid(True, alpha=0.3); plt.tight_layout()
    if save_path: save_with_sidecar(fig, save_path, bbox_inches='tight'); print(f"  Plot saved: {save_path}")
    else: plt.show()

def plot_polcurve_evolution(pcs, save_path=None):
    if not pcs: return
    fig, ax = plt.subplots(figsize=(10, 6), dpi=120)
    ts = [p['t_hours'] for p in pcs]; t0, t1 = min(ts), max(ts)
    cmap = plt.cm.viridis
    for p in pcs:
        j = np.array([s['j'] for s in p['setpoints']])
        V = np.array([s['V'] for s in p['setpoints']])
        c = cmap((p['t_hours']-t0)/(t1-t0+1e-9))
        ax.plot(j, V, 'o-', color=c, ms=4, lw=1.2)
        if 'V_irfree' in p:
            ax.plot(j, p['V_irfree'], 's--', color=c, ms=3, lw=0.8, alpha=0.7)
    sm = plt.cm.ScalarMappable(cmap=cmap, norm=plt.Normalize(t0, t1)); sm.set_array([])
    plt.colorbar(sm, ax=ax, pad=0.02).set_label('Time [hours]')
    if any('V_irfree' in p for p in pcs):
        ax.plot([], [], 'ko-', ms=4, label='Raw'); ax.plot([], [], 'ks--', ms=3, alpha=0.7, label='iR-free')
        ax.legend(fontsize=9)
    ax.set_xlabel('j [A/cm²]'); ax.set_ylabel('V [V]')
    ax.set_title('Polarization Curve Evolution', fontweight='bold'); ax.grid(True, alpha=0.3)
    plt.tight_layout()
    if save_path: save_with_sidecar(fig, save_path, bbox_inches='tight'); print(f"  Plot saved: {save_path}")
    else: plt.show()

def plot_asr_vs_j_evolution(pcs, save_path=None):
    has = [p for p in pcs if 'asr_interp' in p]
    if not has: return
    fig, ax = plt.subplots(figsize=(10, 6), dpi=120)
    ts = [p['t_hours'] for p in has]; t0, t1 = min(ts), max(ts)
    cmap = plt.cm.viridis
    for p in has:
        j = np.array([s['j'] for s in p['setpoints']])
        c = cmap((p['t_hours']-t0)/(t1-t0+1e-9))
        ax.plot(j, p['asr_interp'], 'o-', color=c, ms=4, lw=1.2)
    sm = plt.cm.ScalarMappable(cmap=cmap, norm=plt.Normalize(t0, t1)); sm.set_array([])
    plt.colorbar(sm, ax=ax, pad=0.02).set_label('Time [hours]')
    ax.set_xlabel('j [A/cm²]'); ax.set_ylabel('ASR [mΩ·cm²]')
    ax.set_title('ASR vs. Current Density Evolution', fontweight='bold'); ax.grid(True, alpha=0.3)
    plt.tight_layout()
    if save_path: save_with_sidecar(fig, save_path, bbox_inches='tight'); print(f"  Plot saved: {save_path}")
    else: plt.show()

def plot_loss_evolution(pcs, v_target=1.8, save_path=None):
    has = [p for p in pcs if 'losses' in p and p['losses']]
    if len(has) < 2: return
    t_a = [p['t_hours'] for p in has]; j_a = [p['j_at_target'] for p in has]
    loss_k = ['V_ohmic_mV', 'eta_anode_mV', 'eta_cathode_mV', 'eta_kinetic_mV', 'V_mt_mV']
    loss_d = {k: [p['losses'][k] for p in has] for k in loss_k}
    fig, ax1 = plt.subplots(figsize=(10, 6), dpi=120)
    ax2 = ax1.twinx()
    ax1.plot(t_a, j_a, 'o-', color='#1f77b4', ms=5, lw=1.5, label=f'j @ {v_target:.2f} V')
    ax1.set_xlabel('Time [hours]'); ax1.set_ylabel(f'j at {v_target:.2f} V [A/cm²]', color='#1f77b4')
    ax1.tick_params(axis='y', labelcolor='#1f77b4'); ax1.set_xlim(left=0)
    sty = [('V_ohmic_mV','Ohmic','#4CAF50','s-'), ('eta_anode_mV','η anode','#FF5722','^-'),
           ('eta_cathode_mV','η cathode','#FF9800','v-'), ('eta_kinetic_mV','η kinetic (total)','#d62728','p-'),
           ('V_mt_mV','Mass transport','#9C27B0','D-')]
    for k, lb, co, fm in sty:
        v = np.array(loss_d[k])
        if np.max(np.abs(v)) < 0.1: continue
        ax2.plot(t_a, v, fm, color=co, ms=4, lw=1.2, label=lb, alpha=0.85)
    ax2.set_ylabel('Voltage loss [mV]'); ax1.grid(True, alpha=0.3)
    h1, l1 = ax1.get_legend_handles_labels(); h2, l2 = ax2.get_legend_handles_labels()
    fig.legend(h1+h2, l1+l2, loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=min(len(h1)+len(h2), 6), fontsize=9, frameon=True, fancybox=True)
    ax1.set_title(f'Performance & Loss Evolution at {v_target:.2f} V', fontweight='bold')
    plt.tight_layout(); plt.subplots_adjust(bottom=0.18)
    if save_path: save_with_sidecar(fig, save_path, bbox_inches='tight'); print(f"  Plot saved: {save_path}")
    else: plt.show()


# ═══════════════════════════════════════════════════════════════════
#  Excel export
# ═══════════════════════════════════════════════════════════════════

def export_excel(fp, t_hours=None, voltage=None, j_hold=None,
                 deg=None, eis_p=None, pcs=None, geo_area=5.0):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active; ws.title = "Current Hold"
    if t_hours is not None:
        ws.append(['Time [hr]', 'Voltage [V]'])
        step = max(1, len(t_hours)//5000)
        for i in range(0, len(t_hours), step):
            ws.append([round(t_hours[i], 4), round(voltage[i], 5)])
    if deg:
        ws2 = wb.create_sheet("Degradation"); ws2.append(['Parameter', 'Value'])
        ws2.append(['j_hold [A/cm²]', round(j_hold, 3) if j_hold else ''])
        ws2.append(['Rate [μV/hr]', round(deg['rate_uV_hr'], 2)])
        ws2.append(['Rate [mV/khr]', round(deg['rate_mV_khr'], 2)])
        ws2.append(['Rate [V%/khr]', round(deg['rate_Vpct_khr'], 3)])
        ws2.append(['V₀ (intercept) [V]', round(deg['V0'], 5)])
        ws2.append(['R²', round(deg['r_squared'], 5)])
        if len(deg['t_rolling']) > 0:
            ws2.append([]); ws2.append(['Time [hr]', 'Rate [μV/hr]'])
            for t, r in zip(deg['t_rolling'], deg['rate_rolling']):
                ws2.append([round(t, 1), round(r, 2)])
    if eis_p:
        ws3 = wb.create_sheet("HFR vs Time")
        ws3.append(['Time [hr]', 'ASR [mΩ·cm²]', 'HFR [mΩ]', 'Freq [Hz]', 'DC V [V]'])
        for e in eis_p:
            ws3.append([round(e['t_hours'], 2), round(e['asr_mohm_cm2'], 2),
                        round(e['hfr_ohm']*1000, 3), round(e['f_hfr'], 0),
                        round(e['dc_v_mean'], 4) if e['dc_v_mean'] else ''])
    if pcs:
        ws4 = wb.create_sheet("Polcurve Evolution")
        hdr = ['Time [hr]', 'V [V]', 'j [A/cm²]']
        if any('V_irfree' in p for p in pcs): hdr += ['V_iR-free [V]', 'ASR [mΩ·cm²]']
        ws4.append(hdr)
        for p in pcs:
            j_a = np.array([s['j'] for s in p['setpoints']])
            V_a = np.array([s['V'] for s in p['setpoints']])
            Vi = p.get('V_irfree'); ai = p.get('asr_interp')
            for k in range(len(j_a)):
                row = [round(p['t_hours'], 2), round(V_a[k], 4), round(j_a[k], 6)]
                if Vi is not None: row += [round(Vi[k], 4), round(ai[k], 2)]
                ws4.append(row)
        has_loss = [p for p in pcs if 'losses' in p and p['losses']]
        if has_loss:
            ws5 = wb.create_sheet("Loss Evolution")
            ws5.append(['Time [hr]', 'j [A/cm²]', 'η_anode [mV]', 'η_cathode [mV]',
                        'η_kinetic [mV]', 'V_ohmic [mV]', 'V_mt [mV]'])
            for p in has_loss:
                lo = p['losses']
                ws5.append([round(p['t_hours'], 2), round(p['j_at_target'], 4),
                            round(lo['eta_anode_mV'], 2), round(lo['eta_cathode_mV'], 2),
                            round(lo['eta_kinetic_mV'], 2), round(lo['V_ohmic_mV'], 2),
                            round(lo['V_mt_mV'], 2)])
    if eis_p:
        ws6 = wb.create_sheet("EIS Spectra")
        ws6.append(['Time [hr]', 'Freq [Hz]', "Z' [Ω]", "-Z'' [Ω]",
                    "Z' [mΩ·cm²]", "-Z'' [mΩ·cm²]"])
        for e in eis_p:
            for i in range(len(e['freq'])):
                ws6.append([round(e['t_hours'], 2), round(e['freq'][i], 3),
                            round(e['zre'][i], 8), round(e['zim'][i], 8),
                            round(e['zre'][i]*geo_area*1000, 4),
                            round(e['zim'][i]*geo_area*1000, 4)])
    wb.save(fp); print(f"  Data exported: {fp}")


# ═══════════════════════════════════════════════════════════════════
#  Main pipeline
# ═══════════════════════════════════════════════════════════════════

def analyze(folder_paths, geo_area=5.0, eis_ref_voltage=1.25,
            T_C=80.0, v_target=1.8, save_dir=None, data_interval_min=None,
            image_ext='png'):
    """
    Full durability analysis pipeline.

    Parameters
    ----------
    folder_paths : str or list of str
        Single folder or list of folders in chronological order.
        For multiple folders, elapsed times restart at zero in each;
        the script adds offsets so time is continuous across folders.
    """
    # Normalize to list
    if isinstance(folder_paths, str):
        folder_paths = [folder_paths]

    hold_segs, eis_raw, pc_raw = [], [], []
    t_offset_s = 0.0  # cumulative time offset in seconds

    for fi, folder in enumerate(folder_paths):
        all_csvs = scan_folder(folder)
        if not all_csvs:
            print(f"  No CSV files in folder {fi+1}.")
            continue

        # Find max elapsed time in this folder (for offset to next folder)
        folder_max_t_s = 0.0

        print(f"\n  ── Folder {fi+1}/{len(folder_paths)} "
              f"(offset = {t_offset_s/3600:.1f} hr) ──")

        for fp in all_csvs:
            fn = Path(fp).name
            try:
                segs, ft = load_and_split_file(fp, geo_area)
                ncc = sum(len(s['t_s']) for s in segs['current_hold'])
                print(f"    {fn:50s} → {ft:12s}  "
                      f"[CC:{ncc:,} EIS:{len(segs['eis'])} PC:{len(segs['polcurve'])}]")

                # Apply time offset and track max time
                for s in segs['current_hold']:
                    s['t_s'] = s['t_s'] + t_offset_s
                    folder_max_t_s = max(folder_max_t_s, s['t_s'].max())
                for s in segs['eis']:
                    if s['time'] is not None:
                        s['time'] = s['time'] + t_offset_s
                        folder_max_t_s = max(folder_max_t_s, s['time'].max())
                    if s['t_mid_s'] is not None:
                        s['t_mid_s'] = s['t_mid_s'] + t_offset_s
                for s in segs['polcurve']:
                    for sp in s:
                        sp['t_s'] = sp['t_s'] + t_offset_s
                        folder_max_t_s = max(folder_max_t_s, sp['t_s'])

                hold_segs.extend(segs['current_hold'])
                eis_raw.extend(segs['eis'])
                pc_raw.extend(segs['polcurve'])
            except Exception as e:
                print(f"    {fn:50s} → ERROR: {e}")

        # Set offset for next folder = max time seen in this folder
        t_offset_s = folder_max_t_s

    if not hold_segs and not eis_raw and not pc_raw:
        print("  No data found."); return

    # ── Current hold ──
    t_hours, voltage, j_hold, deg = None, None, None, None
    if hold_segs:
        at = np.concatenate([s['t_s']/3600 for s in hold_segs])
        av = np.concatenate([s['voltage'] for s in hold_segs])
        aj = np.concatenate([s['j_Acm2'] for s in hold_segs])
        o = np.argsort(at); t_hours, voltage = at[o], av[o]
        j_hold = float(np.median(aj))

        # Thin data by time interval if requested
        if data_interval_min is not None and data_interval_min > 0:
            n_before = len(t_hours)
            interval_hr = data_interval_min / 60.0
            keep = [0]
            for i in range(1, len(t_hours)):
                if t_hours[i] - t_hours[keep[-1]] >= interval_hr:
                    keep.append(i)
            keep = np.array(keep)
            t_hours, voltage = t_hours[keep], voltage[keep]
            print(f"\n  Data thinned: {n_before:,} → {len(t_hours):,} pts "
                  f"(interval = {data_interval_min:.1f} min)")

        print(f"\n  Current hold: {t_hours[-1]-t_hours[0]:.1f} hr, {len(t_hours):,} pts, "
              f"j = {j_hold:.3f} A/cm² (I = {j_hold*geo_area:.3f} A)")
        if len(t_hours) > 10: deg = compute_degradation_rate(t_hours, voltage)
    del hold_segs; gc.collect()

    # ── EIS ──
    eis_p = []
    for sw in eis_raw:
        if eis_ref_voltage and sw['dc_v_mean'] is not None:
            if abs(sw['dc_v_mean'] - eis_ref_voltage) > 0.02*abs(eis_ref_voltage):
                continue
        hfr, asr, fhfr = extract_hfr(sw, geo_area)
        t_hr = sw['t_mid_s']/3600 if sw['t_mid_s'] else None
        if t_hr is None: continue
        eis_p.append({'freq': sw['freq'], 'zre': sw['zre'], 'zim': sw['zim'],
                      'dc_v_mean': sw['dc_v_mean'], 't_hours': t_hr,
                      'hfr_ohm': hfr, 'asr_mohm_cm2': asr, 'f_hfr': fhfr})
    eis_p.sort(key=lambda e: e['t_hours'])
    if eis_p:
        print(f"\n  EIS: {len(eis_p)} sweep(s) at ref V = {eis_ref_voltage} V")
        for e in eis_p:
            print(f"    t = {e['t_hours']:.1f} hr, ASR = {e['asr_mohm_cm2']:.1f} mΩ·cm²")
    del eis_raw; gc.collect()

    # ── Polcurves + iR correction + model fit ──
    pcs = []
    for sw in pc_raw:
        t_hr = np.mean([s['t_s'] for s in sw]) / 3600
        j_a = np.array([s['j'] for s in sw])
        V_a = np.array([s['V'] for s in sw])
        entry = {'setpoints': sw, 't_hours': t_hr, 'n_pts': len(sw),
                 'j_arr': j_a, 'V_arr': V_a}

        # iR correct from nearest EIS
        fix_ASR = None
        if eis_p:
            best = min(eis_p, key=lambda e: abs(e['t_hours'] - t_hr))
            asr_c = best['asr_mohm_cm2']
            entry['V_irfree'] = V_a - j_a * asr_c / 1000
            entry['asr_interp'] = np.full_like(j_a, asr_c)
            fix_ASR = asr_c

        # Model fit
        fr = fit_polcurve(j_a, V_a, T_C=T_C, fix_ASR=fix_ASR)
        if fr is not None:
            best_j, best_dv = None, 0.02
            for s in sw:
                dv = abs(s['V'] - v_target)
                if dv < best_dv: best_dv = dv; best_j = s['j']
            if best_j:
                entry['losses'] = compute_losses(best_j, fr['x'], fr['T_K'])
                entry['j_at_target'] = best_j
                entry['fit_result'] = fr
        pcs.append(entry)

    pcs.sort(key=lambda p: p['t_hours'])
    if pcs:
        print(f"\n  Polcurves: {len(pcs)} sweep(s)")
        for p in pcs:
            ir = ", iR-corr" if 'V_irfree' in p else ""
            lo = ", fitted" if 'losses' in p else ""
            print(f"    t = {p['t_hours']:.1f} hr, {p['n_pts']} pts{ir}{lo}")
    del pc_raw; gc.collect()

    # ── Paths ──
    import os
    if save_dir:
        os.makedirs(save_dir, exist_ok=True)
        P = lambda n: os.path.join(save_dir, n)
    else:
        P = lambda n: None

    _ext = image_ext or 'png'

    # ── Plots ──
    pt = [p['t_hours'] for p in pcs] if pcs else None

    if image_ext:
        if t_hours is not None and deg:
            plot_voltage_vs_time(t_hours, voltage, deg, j_hold, pc_t=pt,
                                 save_path=P(f'voltage_vs_time.{_ext}'))
            plt.close('all')
            plot_deg_rolling(deg, P(f'degradation_rate.{_ext}')); plt.close('all')
        if eis_p:
            plot_hfr_vs_time(eis_p, P(f'hfr_vs_time.{_ext}')); plt.close('all')
        if pcs:
            plot_polcurve_evolution(pcs, P(f'polcurve_evolution.{_ext}')); plt.close('all')
            plot_asr_vs_j_evolution(pcs, P(f'asr_vs_j_evolution.{_ext}')); plt.close('all')
            plot_loss_evolution(pcs, v_target, P(f'loss_evolution.{_ext}')); plt.close('all')

    if P('x'):
        export_excel(P('durability_data.xlsx'), t_hours, voltage, j_hold,
                     deg, eis_p, pcs, geo_area)
    gc.collect()


# ═══════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description='PEM Electrolyzer Durability Analyzer')
    ap.add_argument('--folder', type=str, nargs='+', default=None,
                    help='Folder(s) containing test data, in chronological order')
    ap.add_argument('--area', type=float, default=None)
    ap.add_argument('--eis-ref-voltage', type=float, default=None)
    ap.add_argument('--T', type=float, default=80.0)
    ap.add_argument('--data-interval', type=float, default=None,
                    help='Data point interval for current hold [minutes]. '
                         'Thins data to one point per interval.')
    ap.add_argument('--save-dir', type=str, default=None)
    args = ap.parse_args()

    if args.folder:
        folders = [_clean_path(f) for f in args.folder]
        geo = args.area or float(input("  Area [cm²] (25.0): ").strip() or '25')
        erv = args.eis_ref_voltage
        if erv is None: erv = float(input("  EIS ref V (1.25): ").strip() or '1.25')
        di = args.data_interval
        if di is None:
            di_str = input("  Data point interval [min] (Enter = keep all): ").strip()
            di = float(di_str) if di_str else None
        save = args.save_dir; T = args.T
    else:
        print("=" * 60)
        print("  PEM Electrolyzer Durability Analyzer")
        print("=" * 60)

        print("\n  Enter data folder path(s) in chronological order.")
        print("  Press Enter on an empty line when done.\n")

        folders = []
        while True:
            n = len(folders) + 1
            fp = input(f"  Folder {n} (Enter = done): ").strip()
            fp = _clean_path(fp)
            if not fp:
                break
            folders.append(fp)

        if not folders:
            print("  No folders provided."); return

        if len(folders) > 1:
            print(f"\n  {len(folders)} folders entered (will be stitched in order):")
            for i, f in enumerate(folders):
                print(f"    {i+1}. {f}")

        geo = float(input("\n  Area [cm²] (25.0): ").strip() or '25')
        erv = float(input("  EIS ref V (1.25): ").strip() or '1.25')
        T = float(input("  Temperature [°C] (80): ").strip() or '80')
        di_str = input("  Data point interval [min] (Enter = keep all): ").strip()
        di = float(di_str) if di_str else None
        save = _clean_path(input("  Save dir (Enter=display): ").strip()) or None

    analyze(folders, geo_area=geo, eis_ref_voltage=erv, T_C=T,
            save_dir=save, data_interval_min=di)

if __name__ == '__main__':
    main()
