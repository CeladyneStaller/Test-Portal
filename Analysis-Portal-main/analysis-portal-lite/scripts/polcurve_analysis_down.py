#!/usr/bin/env python3
"""
Fuel Cell Polarization Curve & HFR Analysis
============================================
Analyzes polarization curve data for PEM fuel cells.

Features:
  - Polarization curve and power density plots
  - HFR overlay and iR-free voltage correction
  - Tafel slope extraction (kinetic region)
  - Loss breakdown: kinetic, ohmic, mass transport
  - Batch processing (folder or file list)
  - Multi-file overlay for degradation tracking

Usage:
  python polcurve_analysis.py                    # interactive mode
  python polcurve_analysis.py --file data.csv    # analyze pol curve
  python polcurve_analysis.py --demo             # run built-in demo
"""

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import csv
import os
import glob

from scripts.helpers.plot_compare import save_with_sidecar


# ═══════════════════════════════════════════════════════════════════════
#  Data I/O
# ═══════════════════════════════════════════════════════════════════════

def parse_fcd_header(filepath):
    """
    Parse Scribner .fcd file header to extract skip count and column indices.

    Returns dict with 'skip' and column indices, or None for non-.fcd files.
    Column keys: j_col, v_col, hfr_col, freq_col, zreal_col, zimag_col,
                 mode_col, and condition_cols dict.
    """
    if not filepath.lower().endswith('.fcd'):
        return None
    try:
        with open(filepath, 'r', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                lines.append(line)
                if 'End Comments' in line:
                    skip = i + 1
                    break
            else:
                return None
    except Exception:
        return None

    # Column header line is immediately before "End Comments"
    if len(lines) < 2:
        return None
    header_line = lines[-2].strip()
    cols = header_line.split('\t')

    result = {'skip': skip}
    cond_cols = {}

    for ci, name in enumerate(cols):
        n = name.strip()
        # Current (total, in A) — match "I (A)" but not "I (mA/cm²)"
        if n == 'I (A)':
            result['j_col'] = ci
        # Voltage — stack voltage
        elif n == 'E_Stack (V)':
            result['v_col'] = ci
        # HFR — non-area-normalized (mOhm)
        elif n == 'HFR (mOhm)':
            result['hfr_col'] = ci
        # EIS columns
        elif n.startswith('Z_Freq'):
            result['freq_col'] = ci
        elif n.startswith('Z_Real'):
            result['zreal_col'] = ci
        elif n.startswith('Z_Imag'):
            result['zimag_col'] = ci
        # Control mode
        elif n == 'Ctrl_Mode':
            result['mode_col'] = ci
        # Condition columns
        elif n == 'Cell (C)':
            cond_cols['T_cell (C)'] = ci
        elif n.startswith('Temp_Anode'):
            cond_cols['T_anode_dp (C)'] = ci
        elif n.startswith('Flow_Anode'):
            cond_cols['H2_flow (slpm)'] = ci
        elif n.startswith('Temp_Cathode'):
            cond_cols['T_cathode_dp (C)'] = ci
        elif n.startswith('Flow_Cathode'):
            cond_cols['Air_flow (slpm)'] = ci

    if cond_cols:
        result['condition_cols'] = cond_cols

    return result


def _clean_path(p):
    """Clean a file path from drag-drop artifacts and smart quotes."""
    p = p.strip()
    if p.startswith('& '):
        p = p[2:]
    p = p.strip().strip('"').strip("'")
    p = p.strip('\u2018\u2019\u201c\u201d')
    p = p.strip('\u202a\u200b')
    return p


def load_polcurve_data(filepath, j_col=0, v_col=1, hfr_col=None,
                       delimiter=',', skip_header=1,
                       current_is_total=False, geo_area=5.0,
                       condition_cols=None, hfr_scale=1.0,
                       mode_col=None, mode_exclude=None,
                       j_scale=1.0, v_scale=1.0):
    """
    Load polarization curve data from a CSV/TSV file.

    Parameters
    ----------
    filepath : str
        Path to data file.
    j_col : int
        Column index for current density (A/cm²) or total current (A).
    v_col : int
        Column index for cell voltage (V).
    hfr_col : int or None
        Column index for HFR (Ohm). None if no HFR data.
    delimiter : str
        Column delimiter.
    skip_header : int
        Number of header rows to skip.
    current_is_total : bool
        If True, column is total current (A) and will be divided by geo_area.
    geo_area : float
        Geometric area in cm² (used only if current_is_total=True).
    condition_cols : dict or None
        Mapping of condition name to column index.
    hfr_scale : float
        Multiplier for HFR values (e.g. 0.001 to convert mΩ → Ω).
    mode_col : int or None
        Column index for control mode (e.g. Scribner Ctrl_Mode).
        Used with mode_exclude to filter out non-polcurve data.
    mode_exclude : set/list or None
        Mode values to exclude (e.g. {5} removes HFR measurement points).

    Returns
    -------
    j_raw, V_raw, HFR_raw, conditions_raw
    """
    filepath = filepath.strip()
    if filepath.startswith('& '):
        filepath = filepath[2:]
    filepath = filepath.strip().strip('"').strip("'")
    filepath = filepath.strip('\u2018\u2019\u201c\u201d')
    filepath = filepath.strip('\u202a\u200b')

    if not os.path.isfile(filepath):
        raise FileNotFoundError(
            f'File not found: "{filepath}"\n'
            f'  Check the path and try again. On Windows, right-click the file\n'
            f'  in Explorer → Copy as path, then paste into the terminal.')

    j_data, v_data, hfr_data = [], [], []
    cond_data = {name: [] for name in (condition_cols or {})}
    mode_excl = set(mode_exclude) if mode_exclude else set()

    with open(filepath, 'r', errors='replace') as f:
        reader = csv.reader(f, delimiter=delimiter)
        for _ in range(skip_header):
            next(reader)
        for row in reader:
            try:
                # Filter by control mode if specified
                if mode_col is not None and mode_excl:
                    mode_val = int(float(row[mode_col]))
                    if mode_val in mode_excl:
                        continue
                j_val = float(row[j_col])
                v_val = float(row[v_col])
                j_data.append(j_val)
                v_data.append(v_val)
                if hfr_col is not None:
                    hfr_data.append(float(row[hfr_col]))
                for name, col in (condition_cols or {}).items():
                    cond_data[name].append(float(row[col]))
            except (ValueError, IndexError):
                continue

    j = np.array(j_data) * j_scale
    V = np.array(v_data) * v_scale

    if current_is_total:
        j = j / geo_area

    HFR = np.array(hfr_data) * hfr_scale if hfr_col is not None and len(hfr_data) > 0 else None
    conditions = {name: np.array(vals) for name, vals in cond_data.items()} if condition_cols else None

    return j, V, HFR, conditions


def extract_dwell_endpoints(j_raw, V_raw, HFR_raw=None, conditions_raw=None,
                            min_dwell_pts=5, v_step=None):
    """
    Extract the last data point from each voltage or current dwell.

    When *v_step* is provided (voltage-step protocol detected), groups
    consecutive points by voltage stability — the controlled variable.
    Otherwise, groups by current density with a relative tolerance.

    Parameters
    ----------
    j_raw, V_raw : array
        Time-ordered raw data.
    HFR_raw : array or None
    conditions_raw : dict of arrays or None
    min_dwell_pts : int
        Minimum points for a group to be considered a valid dwell.
    v_step : float or None
        Voltage step size.  When provided, uses V-based grouping (tighter
        tolerance, better separation of adjacent voltage setpoints).

    Returns
    -------
    j, V, HFR, conditions : arrays (one point per dwell)
    """
    if len(j_raw) < 2:
        return j_raw, V_raw, HFR_raw, conditions_raw

    # ── Choose grouping signal: V (controlled) or j (response) ──
    if v_step is not None:
        # Voltage-step protocol: group by voltage stability
        signal = V_raw
        abs_floor = v_step * 0.25
        rel_frac = 0.01

        # Pre-filter: detect HFR impedance measurement zones by high local
        # V variance.  These zones have rapidly alternating V (±50 mV between
        # consecutive points) and contaminate the high-V setpoint region.
        hw = 5  # half-window for rolling std
        n = len(V_raw)
        stable = np.ones(n, dtype=bool)
        var_threshold = v_step * 0.4  # ~10 mV for 25 mV step
        for i in range(n):
            lo, hi = max(0, i - hw), min(n, i + hw + 1)
            if np.std(V_raw[lo:hi]) > var_threshold:
                stable[i] = False
    else:
        # Current-step or unknown: group by current stability
        signal = j_raw
        dj_consecutive = np.abs(np.diff(j_raw))
        noise_est = np.percentile(dj_consecutive, 50)
        abs_floor = max(noise_est * 10, 0.003)
        rel_frac = 0.03
        stable = np.ones(len(j_raw), dtype=bool)

    # ── Group consecutive STABLE points by running-mean comparison ──
    groups = []
    group_start = 0
    group_sum = signal[0]
    group_count = 1

    for i in range(1, len(signal)):
        if not stable[i]:
            # Unstable point (HFR zone): end current group, skip this point
            if group_count > 0:
                groups.append((group_start, i))
            group_start = i + 1
            group_sum = 0
            group_count = 0
            continue
        if group_count == 0:
            # Starting new group after unstable region
            group_start = i
            group_sum = signal[i]
            group_count = 1
            continue
        group_mean = group_sum / group_count
        tol = max(abs_floor, rel_frac * abs(group_mean))
        if abs(signal[i] - group_mean) <= tol:
            group_sum += signal[i]
            group_count += 1
        else:
            groups.append((group_start, i))
            group_start = i
            group_sum = signal[i]
            group_count = 1

    if group_count > 0:
        groups.append((group_start, len(signal)))

    # ── Extract averaged endpoint from each valid dwell ──
    # Average the last N points (up to 5) for stability
    avg_pts = 5
    j_out, v_out, hfr_out = [], [], []
    cond_out = {name: [] for name in (conditions_raw or {})}

    for start, end in groups:
        n_pts = end - start
        if n_pts < min_dwell_pts:
            continue
        tail = min(avg_pts, n_pts)
        sl = slice(end - tail, end)
        j_out.append(np.mean(j_raw[sl]))
        v_out.append(np.mean(V_raw[sl]))
        if HFR_raw is not None:
            hfr_out.append(np.mean(HFR_raw[sl]))
        for name in cond_out:
            cond_out[name].append(np.mean(conditions_raw[name][sl]))

    j = np.array(j_out)
    V = np.array(v_out)
    HFR = np.array(hfr_out) if HFR_raw is not None else None
    conditions = {name: np.array(vals) for name, vals in cond_out.items()} if conditions_raw else None

    return j, V, HFR, conditions


def extract_polcurve_cycles(j, V, HFR=None, conditions=None,
                           min_sweep_pts=4, v_step=None):
    """
    Detect up-sweep and down-sweep cycles from dwell-extracted pol curve data.

    When *v_step* is provided, uses voltage for direction detection (V
    decreasing = up-sweep, V increasing = down-sweep).  This is more
    robust than current-based detection at high current density where
    j noise is large.

    Parameters
    ----------
    j, V : array
        Dwell-extracted data (time-ordered, one point per setpoint).
    HFR : array or None
    conditions : dict of arrays or None
    min_sweep_pts : int
        Minimum number of monotonic steps to recognize a direction change.
    v_step : float or None
        Voltage step size. When provided, enables V-based cycle detection.

    Returns
    -------
    cycles : list of dicts
    """
    def _slice_conditions(cond, idx):
        if cond is None:
            return None
        return {name: arr[idx] for name, arr in cond.items()}

    if len(j) < 3:
        direction = 'down' if len(j) < 2 or j[-1] >= j[0] else 'up'
        return [{'j': j, 'V': V, 'HFR': HFR, 'conditions': conditions,
                 'direction': direction, 'cycle_number': 1}]

    # Choose signal for direction detection
    if v_step is not None:
        # V-based: V decreasing = up-sweep (j increasing)
        dsig = -np.diff(V)  # negate so positive = "up" (V dropping)
    else:
        dsig = np.diff(j)

    signs = np.sign(dsig)
    signs[signs == 0] = 1

    # Find sustained direction runs using hysteresis
    turn_indices = [0]
    current_dir = signs[0]
    run_start = 0

    for i in range(1, len(signs)):
        if signs[i] == current_dir:
            continue
        run_len = i - run_start
        if run_len >= min_sweep_pts:
            if i + 1 != turn_indices[-1]:
                turn_indices.append(i + 1)
        current_dir = signs[i]
        run_start = i

    turn_indices.append(len(j))

    # Build cycle list — share boundary points between adjacent cycles
    # The last point of a sweep matches the first point of the next sweep
    cycles = []
    down_count = 0
    for seg in range(len(turn_indices) - 1):
        i_start = turn_indices[seg]
        i_end = turn_indices[seg + 1]

        # Extend to include boundary point from previous segment
        if seg > 0:
            i_start = max(i_start - 1, turn_indices[seg - 1])
        # Extend to include boundary point from next segment
        if seg < len(turn_indices) - 2:
            i_end = min(i_end + 1, turn_indices[seg + 2])

        if i_end - i_start < 2:
            continue

        j_seg = j[i_start:i_end]
        V_seg = V[i_start:i_end]
        HFR_seg = HFR[i_start:i_end] if HFR is not None else None
        cond_seg = _slice_conditions(conditions, slice(i_start, i_end))

        if j_seg[-1] > j_seg[0]:
            # Current increased → voltage decreased → "down" sweep
            direction = 'down'
            down_count += 1
            cycle_num = down_count
        else:
            # Current decreased → voltage increased → "up" sweep
            direction = 'up'
            cycle_num = down_count

        idx = np.argsort(j_seg)
        cycles.append({
            'j': j_seg[idx],
            'V': V_seg[idx],
            'HFR': HFR_seg[idx] if HFR_seg is not None else None,
            'conditions': _slice_conditions(cond_seg, idx) if cond_seg else None,
            'direction': direction,
            'cycle_number': max(cycle_num, 1),
        })

    if not cycles:
        idx = np.argsort(j)
        cycles = [{'j': j[idx], 'V': V[idx],
                   'HFR': HFR[idx] if HFR is not None else None,
                   'conditions': _slice_conditions(conditions, idx),
                   'direction': 'up', 'cycle_number': 1}]

    # ── Merge adjacent cycles with the same direction ──
    merged = [cycles[0]]
    for c in cycles[1:]:
        prev = merged[-1]
        if c['direction'] == prev['direction']:
            j_m = np.concatenate([prev['j'], c['j']])
            V_m = np.concatenate([prev['V'], c['V']])
            HFR_m = None
            if prev['HFR'] is not None and c['HFR'] is not None:
                HFR_m = np.concatenate([prev['HFR'], c['HFR']])
            cond_m = None
            if prev['conditions'] is not None and c['conditions'] is not None:
                cond_m = {name: np.concatenate([prev['conditions'][name], c['conditions'][name]])
                          for name in prev['conditions']}
            idx = np.argsort(j_m)
            merged[-1] = {
                'j': j_m[idx], 'V': V_m[idx],
                'HFR': HFR_m[idx] if HFR_m is not None else None,
                'conditions': _slice_conditions(cond_m, idx) if cond_m else None,
                'direction': prev['direction'],
                'cycle_number': prev['cycle_number'],
            }
        else:
            merged.append(c)

    # Re-number cycles
    down_count = 0
    for c in merged:
        if c['direction'] == 'down':
            down_count += 1
            c['cycle_number'] = down_count
        else:
            c['cycle_number'] = down_count

    return merged


def enforce_consensus_setpoints(cycles, v_step):
    """
    Ensure all cycles share the same voltage setpoints.

    1.  Dedup each cycle independently to find its setpoints.
    2.  Pool and cluster setpoints across all cycles.
    3.  Keep only setpoints present in a majority of cycles (consensus).
    4.  For each cycle, snap raw points to consensus setpoints (one point
        per setpoint, nearest V wins) and discard non-matching points.

    This guarantees every cycle has the same number of points at the
    same voltage levels, eliminating HFR artifacts and boundary noise.

    Parameters
    ----------
    cycles : list of dicts
        Output of extract_polcurve_cycles().
    v_step : float
        Voltage step size for clustering tolerance.

    Returns
    -------
    list of dicts — cleaned cycles with uniform setpoint count.
    """
    if v_step is None or len(cycles) < 2:
        return cycles

    tol_V = v_step * 0.45

    # Step 1: dedup each cycle, collect V setpoints
    deduped_V = []
    for c in cycles:
        _, V_d, _, _, _ = deduplicate_polcurve(c['j'], c['V'], v_step=v_step)
        deduped_V.append(V_d)

    # Step 2: pool all V values and cluster
    v_to_cycles = {}
    for ci, vs in enumerate(deduped_V):
        for v in vs:
            v_to_cycles.setdefault(round(v, 6), set()).add(ci)

    all_V = np.sort(np.concatenate(deduped_V))[::-1]
    clusters = []
    grp = [all_V[0]]
    for i in range(1, len(all_V)):
        if abs(all_V[i] - np.mean(grp)) <= tol_V:
            grp.append(all_V[i])
        else:
            clusters.append(grp)
            grp = [all_V[i]]
    clusters.append(grp)

    # Count unique cycles per cluster
    n_cycles = len(cycles)
    min_cycles = max(2, n_cycles // 2)
    consensus_V = []
    for grp in clusters:
        cyc_set = set()
        for v in grp:
            cyc_set |= v_to_cycles.get(round(v, 6), set())
        if len(cyc_set) >= min_cycles:
            consensus_V.append(np.mean(grp))
    consensus_V = np.array(sorted(consensus_V, reverse=True))

    if len(consensus_V) == 0:
        return cycles

    # Step 3: for each cycle, snap raw points to consensus setpoints
    def _slice(arr, idx):
        return arr[idx] if arr is not None else None

    cleaned = []
    for c in cycles:
        j_c, V_c = c['j'], c['V']
        HFR_c = c['HFR']
        cond_c = c.get('conditions')

        keep_idx = []
        used = set()
        for cv in consensus_V:
            dists = np.abs(V_c - cv)
            candidates = np.argsort(dists)
            for best in candidates:
                if dists[best] > tol_V:
                    break
                if best not in used:
                    keep_idx.append(best)
                    used.add(best)
                    break

        if not keep_idx:
            cleaned.append(c)
            continue

        idx = np.array(sorted(keep_idx))
        order = np.argsort(j_c[idx])
        idx = idx[order]

        cleaned.append({
            'j': j_c[idx], 'V': V_c[idx],
            'HFR': _slice(HFR_c, idx),
            'conditions': ({name: vals[idx] for name, vals in cond_c.items()}
                           if cond_c else None),
            'direction': c['direction'],
            'cycle_number': c['cycle_number'],
        })

    return cleaned


def select_representative_cycle(cycles, choice='last_down'):
    """
    Select a representative cycle from extracted pol curve cycles.

    Parameters
    ----------
    cycles : list of dicts
        Output of extract_polcurve_cycles().
    choice : str
        'last_down' — last down-sweep (default, OCV → high current)
        'last_up'  — last up-sweep
        'first_up' — first up-sweep
        'all'       — return list of all cycles

    Returns
    -------
    dict (single cycle) or list of dicts (if choice='all').
    """
    if choice == 'all':
        return cycles

    up_sweeps = [c for c in cycles if c['direction'] == 'up']
    down_sweeps = [c for c in cycles if c['direction'] == 'down']

    if choice == 'last_up':
        return up_sweeps[-1] if up_sweeps else cycles[-1]
    elif choice == 'first_up':
        return up_sweeps[0] if up_sweeps else cycles[0]
    elif choice == 'last_down':
        return down_sweeps[-1] if down_sweeps else cycles[-1]
    else:
        raise ValueError(f'Unknown choice: {choice}')


def detect_voltage_step(V_raw):
    """
    Auto-detect the voltage setpoint step size from raw time-series voltage.

    Uses histogram peak analysis: each voltage setpoint produces a cluster
    of raw V values. The most common spacing between adjacent peaks gives
    the step size.

    Returns step_size (V) or None if detection fails.
    """
    V_min, V_max = np.min(V_raw), np.max(V_raw)
    bins = np.arange(V_min - 0.005, V_max + 0.01, 0.005)
    if len(bins) < 4:
        return None
    hist, edges = np.histogram(V_raw, bins=bins)
    centers = (edges[:-1] + edges[1:]) / 2

    # Find peaks: bins with more counts than both neighbours
    threshold = np.percentile(hist[hist > 0], 50)
    peaks = []
    for i in range(1, len(hist) - 1):
        if hist[i] > threshold and hist[i] >= hist[i - 1] and hist[i] >= hist[i + 1]:
            peaks.append(centers[i])

    if len(peaks) < 4:
        return None

    peaks = np.sort(peaks)
    gaps = np.diff(peaks)
    # Mode of gaps rounded to nearest 0.005 V
    gaps_rounded = np.round(gaps / 0.005) * 0.005
    unique_gaps, counts = np.unique(gaps_rounded, return_counts=True)
    return float(unique_gaps[np.argmax(counts)])


def deduplicate_polcurve(j, V, HFR=None, conditions=None, v_step=None):
    """
    Average duplicate points that correspond to the same voltage setpoint.

    When *v_step* is provided (auto-detected or manual), deduplication works
    on voltage — the controlled variable — by snapping each point to a
    regular voltage grid and averaging points that map to the same grid
    value.  This is robust regardless of current noise.

    When *v_step* is None, falls back to current-based grouping with a
    fixed tolerance of 0.005 A/cm².

    Parameters
    ----------
    j, V : array
        Sorted by ascending j.
    HFR : array or None
    conditions : dict of arrays or None
    v_step : float or None
        Voltage setpoint spacing (V).  When provided, enables V-based
        snap-to-grid deduplication.  Use detect_voltage_step() on the
        raw voltage time-series to obtain this automatically.

    Returns
    -------
    j, V, HFR, conditions : deduplicated arrays (sorted by ascending j)
    n_merged : int
        Number of points removed by merging.
    """
    if len(j) < 2:
        return j, V, HFR, conditions, 0

    if v_step is not None and v_step > 0:
        return _dedup_voltage_grid(j, V, HFR, conditions, v_step)
    else:
        return _dedup_current(j, V, HFR, conditions, tol=0.005)


def _dedup_current(j, V, HFR, conditions, tol=0.005):
    """Fallback: group by current density proximity."""
    j_out, v_out, hfr_out = [], [], []
    cond_out = {name: [] for name in (conditions or {})}

    group_j = [j[0]]
    group_v = [V[0]]
    group_hfr = [HFR[0]] if HFR is not None else []
    group_cond = {name: [conditions[name][0]] for name in (conditions or {})}

    for i in range(1, len(j)):
        if abs(j[i] - np.mean(group_j)) <= tol:
            group_j.append(j[i])
            group_v.append(V[i])
            if HFR is not None:
                group_hfr.append(HFR[i])
            for name in group_cond:
                group_cond[name].append(conditions[name][i])
        else:
            j_out.append(np.mean(group_j))
            v_out.append(np.mean(group_v))
            if HFR is not None:
                hfr_out.append(np.mean(group_hfr))
            for name in cond_out:
                cond_out[name].append(np.mean(group_cond[name]))
            group_j = [j[i]]
            group_v = [V[i]]
            group_hfr = [HFR[i]] if HFR is not None else []
            group_cond = {name: [conditions[name][i]] for name in (conditions or {})}

    j_out.append(np.mean(group_j))
    v_out.append(np.mean(group_v))
    if HFR is not None:
        hfr_out.append(np.mean(group_hfr))
    for name in cond_out:
        cond_out[name].append(np.mean(group_cond[name]))

    n_merged = len(j) - len(j_out)
    j_d = np.array(j_out)
    V_d = np.array(v_out)
    HFR_d = np.array(hfr_out) if HFR is not None else None
    cond_d = {name: np.array(vals) for name, vals in cond_out.items()} if conditions else None

    return j_d, V_d, HFR_d, cond_d, n_merged


def _dedup_voltage_grid(j, V, HFR, conditions, v_step):
    """Group by voltage proximity — no grid enforced."""
    tol_V = v_step * 0.45  # half the smallest expected step

    # Sort by V descending (OCV first)
    order = np.argsort(V)[::-1]
    V_s, j_s = V[order], j[order]
    HFR_s = HFR[order] if HFR is not None else None
    cond_s = {name: vals[order] for name, vals in conditions.items()} if conditions else None

    # Walk through sorted V and group nearby points
    j_out, v_out, hfr_out = [], [], []
    cond_out = {name: [] for name in (conditions or {})}

    grp_idx = [0]
    for i in range(1, len(V_s)):
        if abs(V_s[i] - np.mean(V_s[grp_idx])) <= tol_V:
            grp_idx.append(i)
        else:
            # Flush group
            j_out.append(np.mean(j_s[grp_idx]))
            v_out.append(np.mean(V_s[grp_idx]))
            if HFR_s is not None:
                hfr_out.append(np.mean(HFR_s[grp_idx]))
            for name in cond_out:
                cond_out[name].append(np.mean(cond_s[name][grp_idx]))
            grp_idx = [i]

    # Flush final group
    j_out.append(np.mean(j_s[grp_idx]))
    v_out.append(np.mean(V_s[grp_idx]))
    if HFR_s is not None:
        hfr_out.append(np.mean(HFR_s[grp_idx]))
    for name in cond_out:
        cond_out[name].append(np.mean(cond_s[name][grp_idx]))

    n_merged = len(j) - len(j_out)

    # Re-sort by ascending j
    j_d = np.array(j_out)
    V_d = np.array(v_out)
    order_j = np.argsort(j_d)
    j_d, V_d = j_d[order_j], V_d[order_j]
    HFR_d = np.array(hfr_out)[order_j] if HFR is not None else None
    cond_d = ({name: np.array(vals)[order_j] for name, vals in cond_out.items()}
              if conditions else None)

    return j_d, V_d, HFR_d, cond_d, n_merged


def enforce_v_monotonicity(j, V, HFR=None, conditions=None):
    """
    Remove points that violate V-monotonicity (V must decrease as j increases).

    At high current density in mass-transport-limited cells, cycle boundary
    artifacts can place a stale dwell point in the wrong sweep, creating
    a local V increase.  This function iteratively removes the worst
    violator until the curve is strictly monotonic.

    Parameters
    ----------
    j, V : array (sorted by ascending j)
    HFR : array or None
    conditions : dict of arrays or None

    Returns
    -------
    j, V, HFR, conditions — cleaned arrays with monotonic V(j)
    """
    if len(j) < 3:
        return j, V, HFR, conditions

    order = np.argsort(j)
    j, V = j[order].copy(), V[order].copy()
    if HFR is not None:
        HFR = HFR[order].copy()
    if conditions:
        conditions = {k: v[order].copy() for k, v in conditions.items()}

    while True:
        # Find all violations: V[i] > V[i-1] when sorted by ascending j
        violations = []
        for i in range(1, len(V)):
            if V[i] > V[i - 1]:
                violations.append(i)
        if not violations:
            break

        # Remove the point that causes the largest V jump
        worst = max(violations, key=lambda i: V[i] - V[i - 1])
        # Decide which point to remove: the one that doesn't fit the local trend
        # If V[worst] > V[worst-1], either worst is too high or worst-1 is too low
        # Check which one is more consistent with neighbors
        if worst < len(V) - 1 and V[worst - 1] >= V[worst + 1]:
            # worst-1 and worst+1 are consistent — remove worst
            remove = worst
        elif worst > 1 and V[worst - 2] >= V[worst]:
            # worst-2 and worst are consistent — remove worst-1
            remove = worst - 1
        else:
            remove = worst

        keep = np.ones(len(j), dtype=bool)
        keep[remove] = False
        j, V = j[keep], V[keep]
        if HFR is not None:
            HFR = HFR[keep]
        if conditions:
            conditions = {k: v[keep] for k, v in conditions.items()}

    return j, V, HFR, conditions
    """Fallback: group by current density proximity."""
    j_out, v_out, hfr_out = [], [], []
    cond_out = {name: [] for name in (conditions or {})}

    group_j = [j[0]]
    group_v = [V[0]]
    group_hfr = [HFR[0]] if HFR is not None else []
    group_cond = {name: [conditions[name][0]] for name in (conditions or {})}

    for i in range(1, len(j)):
        if abs(j[i] - np.mean(group_j)) <= tol:
            group_j.append(j[i])
            group_v.append(V[i])
            if HFR is not None:
                group_hfr.append(HFR[i])
            for name in group_cond:
                group_cond[name].append(conditions[name][i])
        else:
            j_out.append(np.mean(group_j))
            v_out.append(np.mean(group_v))
            if HFR is not None:
                hfr_out.append(np.mean(group_hfr))
            for name in cond_out:
                cond_out[name].append(np.mean(group_cond[name]))
            group_j = [j[i]]
            group_v = [V[i]]
            group_hfr = [HFR[i]] if HFR is not None else []
            group_cond = {name: [conditions[name][i]] for name in (conditions or {})}

    j_out.append(np.mean(group_j))
    v_out.append(np.mean(group_v))
    if HFR is not None:
        hfr_out.append(np.mean(group_hfr))
    for name in cond_out:
        cond_out[name].append(np.mean(group_cond[name]))

    n_merged = len(j) - len(j_out)
    j_d = np.array(j_out)
    V_d = np.array(v_out)
    HFR_d = np.array(hfr_out) if HFR is not None else None
    cond_d = {name: np.array(vals) for name, vals in cond_out.items()} if conditions else None

    return j_d, V_d, HFR_d, cond_d, n_merged


# ═══════════════════════════════════════════════════════════════════════
#  HFR Outlier Filtering
# ═══════════════════════════════════════════════════════════════════════

def filter_hfr_outliers(HFR, geo_area=5.0, iqr_factor=2.0, max_asr=0.5):
    """
    Detect and replace outlier HFR values with linear interpolation.

    Uses two criteria (either triggers replacement):
      1. IQR method: value is beyond Q1 - iqr_factor×IQR or Q3 + iqr_factor×IQR
      2. Absolute ceiling: HFR × geo_area > max_asr (Ω·cm²)

    Parameters
    ----------
    HFR : array or None
        HFR values in Ohm (per-cell).
    geo_area : float
        Geometric area in cm² (for ASR ceiling check).
    iqr_factor : float
        IQR multiplier for outlier detection (default 2.0).
    max_asr : float
        Maximum plausible ASR in Ω·cm² (default 0.5). Points above this
        are flagged as outliers regardless of IQR.

    Returns
    -------
    HFR_clean : array or None
        Cleaned HFR with outliers replaced by interpolation.
    n_outliers : int
        Number of outliers detected and replaced.
    outlier_mask : array of bool
        True where outliers were detected.
    """
    if HFR is None or len(HFR) < 3:
        return HFR, 0, np.zeros(len(HFR) if HFR is not None else 0, dtype=bool)

    HFR_asr = HFR * geo_area

    # ── IQR-based detection ──
    Q1 = np.percentile(HFR_asr, 25)
    Q3 = np.percentile(HFR_asr, 75)
    IQR = Q3 - Q1
    iqr_low = Q1 - iqr_factor * IQR
    iqr_high = Q3 + iqr_factor * IQR

    # ── Combined outlier mask ──
    outlier = (HFR_asr < iqr_low) | (HFR_asr > iqr_high) | (HFR_asr > max_asr)

    # Also flag negative HFR (physically impossible for ohmic resistance)
    outlier |= (HFR < 0)

    n_outliers = int(outlier.sum())
    if n_outliers == 0:
        return HFR.copy(), 0, outlier

    # ── Replace outliers with linear interpolation ──
    HFR_clean = HFR.copy()
    good_idx = np.where(~outlier)[0]

    if len(good_idx) < 2:
        # Almost all points are outliers — use median of good points
        if len(good_idx) == 1:
            HFR_clean[outlier] = HFR[good_idx[0]]
        else:
            HFR_clean[outlier] = np.median(HFR)
        return HFR_clean, n_outliers, outlier

    # Interpolate from good points
    bad_idx = np.where(outlier)[0]
    HFR_clean[bad_idx] = np.interp(bad_idx, good_idx, HFR[good_idx])

    return HFR_clean, n_outliers, outlier


# ═══════════════════════════════════════════════════════════════════════
#  Analysis Functions
# ═══════════════════════════════════════════════════════════════════════

def compute_power_density(j, V):
    """Compute power density (W/cm²) = j × V."""
    return j * V


def compute_ir_free_voltage(V, j, HFR):
    """
    Compute iR-free voltage: V_iRfree = V + j × HFR × A.

    For HFR in Ohm (not area-specific): V_iRfree = V + I × R = V + (j × A) × R
    But since we want per-cm² and j is already per-cm²:
    V_iRfree = V + j × (HFR × geo_area)... no.

    HFR is in Ohm (total cell). ohmic drop = I × R = (j × A) × R.
    Voltage drop per cell = I × R. Since V is cell voltage:
    V_iRfree = V + j × R_ASR  where R_ASR = HFR × A  (Ohm·cm²)

    But we don't know A here. We pass HFR as Ohm and j as A/cm².
    So: V_iRfree = V + j × HFR_ASR

    Convention: if HFR is in Ohm, user must convert. We assume HFR
    passed here is already area-specific (Ohm·cm²).
    """
    return V + j * HFR


def extract_tafel(j, V, j_min=0.01, j_max=0.1):
    """
    Extract Tafel slope from the kinetic region of the polarization curve.

    Fits: V = a - b × log10(j)  in the range [j_min, j_max] A/cm².

    Parameters
    ----------
    j, V : array
        Current density (A/cm²) and voltage (V).
    j_min, j_max : float
        Current density range for Tafel fit (A/cm²).

    Returns
    -------
    dict with Tafel slope, intercept, exchange current density, and fit data.
    """
    mask = (j >= j_min) & (j <= j_max) & (j > 0)
    if mask.sum() < 3:
        return None

    log_j = np.log10(j[mask])
    V_fit = V[mask]

    # V = a + b × log10(j)  →  b is negative (Tafel slope, mV/dec)
    coeffs = np.polyfit(log_j, V_fit, 1)
    b_Vdec = coeffs[0]       # V/decade
    a = coeffs[1]             # V (intercept)

    # Tafel slope in mV/dec (magnitude)
    tafel_slope = abs(b_Vdec) * 1000  # mV/dec

    # Exchange current density: at V = E_rev (~1.23 V for H2/O2)
    # log10(j0) = (E_rev - a) / b
    E_rev = 1.23
    if b_Vdec != 0:
        log_j0 = (E_rev - a) / b_Vdec
        j0 = 10 ** log_j0
    else:
        j0 = np.nan

    # Fit line for plotting
    log_j_line = np.linspace(log_j.min() - 0.5, log_j.max() + 0.5, 100)
    V_line = a + b_Vdec * log_j_line

    return {
        'tafel_slope_mVdec': tafel_slope,
        'b_Vdec': b_Vdec,
        'intercept_V': a,
        'j0_A_cm2': j0,
        'j_min': j_min,
        'j_max': j_max,
        'R_squared': 1.0 - np.sum((V_fit - (a + b_Vdec * log_j))**2) / np.sum((V_fit - V_fit.mean())**2),
        '_log_j_fit': log_j,
        '_V_fit': V_fit,
        '_log_j_line': log_j_line,
        '_V_line': V_line,
    }


def analyze_polcurve(j, V, HFR=None, geo_area=5.0,
                     tafel_j_min=0.01, tafel_j_max=0.10):
    """
    Full polarization curve analysis.

    Parameters
    ----------
    j : array
        Current density in A/cm².
    V : array
        Cell voltage in V.
    HFR : array or None
        HFR in Ohm (per-cell). Will be converted to ASR (Ohm·cm²).
    geo_area : float
        Geometric area in cm².
    tafel_j_min, tafel_j_max : float
        Current density range for Tafel fit (A/cm²).

    Returns
    -------
    dict with all analysis results and plotting data.
    """
    P = compute_power_density(j, V)

    results = {
        'j': j,
        'V': V,
        'P': P,
        'geo_area': geo_area,
        'V_at_1Acm2': np.interp(1.0, j, V) if j.max() >= 1.0 else None,
        'peak_power_W_cm2': P.max(),
        'j_at_peak_power': j[np.argmax(P)],
        'OCV': V[0] if j[0] < 0.01 else np.interp(0, j, V),
    }

    # HFR and iR-free voltage
    if HFR is not None:
        HFR_ASR = HFR * geo_area   # Ohm·cm²
        results['HFR'] = HFR
        results['HFR_ASR'] = HFR_ASR
        results['HFR_mean'] = np.mean(HFR_ASR)

        V_irfree = compute_ir_free_voltage(V, j, HFR_ASR)
        results['V_irfree'] = V_irfree
        results['P_irfree'] = j * V_irfree

        # Tafel on iR-free curve
        tafel = extract_tafel(j, V_irfree, j_min=tafel_j_min, j_max=tafel_j_max)
    else:
        results['HFR'] = None
        results['HFR_ASR'] = None
        results['V_irfree'] = None

        # Tafel on raw curve
        tafel = extract_tafel(j, V, j_min=tafel_j_min, j_max=tafel_j_max)

    results['tafel'] = tafel

    # Loss breakdown at a reference current density
    j_ref_values = [0.5, 1.0, 1.5]
    for j_ref in j_ref_values:
        if j.max() >= j_ref:
            V_at_j = np.interp(j_ref, j, V)
            E_rev = 1.23

            # Ohmic loss
            if HFR is not None:
                hfr_at_j = np.interp(j_ref, j, HFR_ASR)
                eta_ohmic = j_ref * hfr_at_j
            else:
                eta_ohmic = None

            # Kinetic loss (from Tafel extrapolation)
            if tafel is not None:
                V_tafel_at_j = tafel['intercept_V'] + tafel['b_Vdec'] * np.log10(j_ref)
                eta_kinetic = E_rev - V_tafel_at_j
            else:
                eta_kinetic = None

            # Mass transport loss (remainder)
            if eta_ohmic is not None and eta_kinetic is not None:
                eta_mt = E_rev - V_at_j - eta_ohmic - eta_kinetic
                eta_mt = max(eta_mt, 0)  # prevent negative
            else:
                eta_mt = None

            results[f'V_at_{j_ref}'] = V_at_j
            results[f'eta_ohmic_{j_ref}'] = eta_ohmic
            results[f'eta_kinetic_{j_ref}'] = eta_kinetic
            results[f'eta_mt_{j_ref}'] = eta_mt

    return results


# ═══════════════════════════════════════════════════════════════════════
#  Plotting
# ═══════════════════════════════════════════════════════════════════════

def plot_polcurve(results, save_path=None):
    """
    Multi-panel polarization curve analysis plot.
    """
    j = results['j']
    V = results['V']
    P = results['P']
    has_hfr = results['HFR'] is not None
    has_tafel = results['tafel'] is not None

    n_panels = 2 + int(has_hfr) + int(has_tafel)
    fig, axes = plt.subplots(1, n_panels, figsize=(5.5 * n_panels, 5))
    if n_panels == 1:
        axes = [axes]

    panel = 0

    # ── Panel: Polarization curve + power density ──
    ax = axes[panel]
    ax.plot(j, V, 'o-', color='steelblue', ms=4, lw=1.5, label='Cell voltage')
    if results['V_irfree'] is not None:
        ax.plot(j, results['V_irfree'], 's--', color='green', ms=3, lw=1.2,
                alpha=0.7, label='iR-free voltage')
    ax.set_xlabel('Current density (A/cm²)')
    ax.set_ylabel('Voltage (V)')
    ax.set_ylim(bottom=0)
    ax.legend(loc='upper right', fontsize=8)
    ax.grid(True, alpha=0.3)

    ax2 = ax.twinx()
    ax2.plot(j, P * 1000, '^-', color='firebrick', ms=3, lw=1.2, alpha=0.7,
             label='Power density')
    ax2.set_ylabel('Power density (mW/cm²)', color='firebrick')
    ax2.tick_params(axis='y', labelcolor='firebrick')
    ax.set_title('Polarization Curve')

    # Annotation
    txt_lines = [f'OCV = {results["OCV"]:.3f} V',
                 f'Peak P = {results["peak_power_W_cm2"]*1000:.0f} mW/cm²']
    if results['V_at_1Acm2'] is not None:
        txt_lines.append(f'V @ 1 A/cm² = {results["V_at_1Acm2"]:.3f} V')
    ax.text(0.03, 0.03, '\n'.join(txt_lines), transform=ax.transAxes,
            fontsize=8, va='bottom',
            bbox=dict(boxstyle='round,pad=0.4', fc='lightyellow', alpha=0.9))
    panel += 1

    # ── Panel: HFR ──
    if has_hfr:
        ax = axes[panel]
        HFR_ASR = results['HFR_ASR']
        ax.plot(j, HFR_ASR * 1000, 'o-', color='darkorange', ms=4, lw=1.5)
        ax.set_xlabel('Current density (A/cm²)')
        ax.set_ylabel('HFR (mΩ·cm²)')
        ax.set_title(f'HFR — mean = {results["HFR_mean"]*1000:.1f} mΩ·cm²')
        ax.grid(True, alpha=0.3)
        panel += 1

    # ── Panel: Tafel ──
    if has_tafel:
        ax = axes[panel]
        t = results['tafel']
        ax.plot(10**t['_log_j_fit'], t['_V_fit'], 'o', color='steelblue', ms=5,
                label='Data')
        ax.plot(10**t['_log_j_line'], t['_V_line'], '-', color='firebrick', lw=1.5,
                label=f'Fit: {t["tafel_slope_mVdec"]:.0f} mV/dec')
        ax.set_xscale('log')
        ax.set_xlabel('Current density (A/cm²)')
        ylabel = 'iR-free voltage (V)' if results['V_irfree'] is not None else 'Voltage (V)'
        ax.set_ylabel(ylabel)
        ax.set_title(f'Tafel Analysis (R² = {t["R_squared"]:.4f})')
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3, which='both')

        txt = f'b = {t["tafel_slope_mVdec"]:.0f} mV/dec\nj₀ = {t["j0_A_cm2"]:.2e} A/cm²'
        ax.text(0.97, 0.03, txt, transform=ax.transAxes,
                ha='right', va='bottom', fontsize=9,
                bbox=dict(boxstyle='round,pad=0.4', fc='lightyellow', alpha=0.9))
        panel += 1

    # ── Panel: All cycles overlay ──
    ax = axes[panel]
    cycles_raw = results.get('_cycles_raw', [])
    if cycles_raw:
        n_cyc = max(c['cycle_number'] for c in cycles_raw)
        cmap = plt.cm.viridis
        colors = [cmap(i / max(n_cyc - 1, 1)) for i in range(n_cyc)]

        for c in cycles_raw:
            cidx = c['cycle_number'] - 1
            ls = '-' if c['direction'] == 'up' else '--'
            marker = 'o' if c['direction'] == 'up' else 's'
            lbl = f'Cycle {c["cycle_number"]} ({c["direction"]})'
            ax.plot(c['j'], c['V'], marker + ls, color=colors[cidx], ms=3, lw=1.2,
                    alpha=0.8, label=lbl)

        # Highlight representative on top
        ax.plot(j, V, 'o-', color='firebrick', ms=5, lw=2.0, alpha=0.9,
                label='Representative', zorder=10)

        ax.set_xlabel('Current density (A/cm²)')
        ax.set_ylabel('Voltage (V)')
        ax.set_title('All Cycles (dwell-extracted)')
        ax.set_ylim(bottom=0)
        ax.legend(fontsize=7, loc='best')
        ax.grid(True, alpha=0.3)
    else:
        # No cycle data — just plot the representative
        ax.plot(j, V, 'o-', color='steelblue', ms=4, lw=1.5)
        ax.set_xlabel('Current density (A/cm²)')
        ax.set_ylabel('Voltage (V)')
        ax.set_title('Polarization Curve')
        ax.set_ylim(bottom=0)
        ax.grid(True, alpha=0.3)

    from scripts.helpers.conditions import get_condition_label
    cond_label = get_condition_label(
        label=results.get('label', ''),
        conditions=results.get('conditions'))
    title = 'Polarization Curve Analysis (Downswing)'
    if cond_label:
        title += f'\n{cond_label}'
    fig.suptitle(title, fontsize=13, fontweight='bold')
    fig.tight_layout()

    if save_path:
        save_with_sidecar(fig, save_path, plot_type='polcurve_down', dpi=200, bbox_inches='tight')
        print(f'  Saved: {save_path}')
    return fig


def plot_polcurve_overlay(all_results, save_path=None):
    """Overlay multiple polarization curves."""
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))
    n = len(all_results)
    cmap = plt.cm.viridis
    colors = [cmap(i / max(n - 1, 1)) for i in range(n)]

    # ── Left: V vs j ──
    ax = axes[0]
    for i, r in enumerate(all_results):
        lbl = r.get('label', f'File {i+1}')
        ax.plot(r['j'], r['V'], 'o-', color=colors[i], ms=3, lw=1.2, label=lbl)
    ax.set_xlabel('Current density (A/cm²)')
    ax.set_ylabel('Voltage (V)')
    ax.set_title('Polarization Curves')
    ax.set_ylim(bottom=0)
    ax.legend(fontsize=7, loc='best')
    ax.grid(True, alpha=0.3)

    # ── Right: Power density ──
    ax = axes[1]
    for i, r in enumerate(all_results):
        lbl = r.get('label', f'File {i+1}')
        ax.plot(r['j'], r['P'] * 1000, '^-', color=colors[i], ms=3, lw=1.2, label=lbl)
    ax.set_xlabel('Current density (A/cm²)')
    ax.set_ylabel('Power density (mW/cm²)')
    ax.set_title('Power Density')
    ax.legend(fontsize=7, loc='best')
    ax.grid(True, alpha=0.3)

    fig.suptitle('Polarization Curve Overlay (Downswing)', fontsize=13, fontweight='bold')
    fig.tight_layout()

    if save_path:
        save_with_sidecar(fig, save_path, plot_type='polcurve_down_overlay', dpi=200, bbox_inches='tight')
        print(f'  Saved: {save_path}')
    return fig


# ═══════════════════════════════════════════════════════════════════════
#  Print Results
# ═══════════════════════════════════════════════════════════════════════

def print_results(r):
    """Pretty-print polarization curve results."""
    print(f'\n{"═" * 55}')
    print(f'  Geo. area:        {r["geo_area"]:.2f} cm²')

    # Operating conditions
    cond = r.get('conditions')
    if cond:
        print(f'  {"─" * 51}')
        for name, vals in cond.items():
            print(f'  {name + ":":20s} {np.mean(vals):.1f}')
        print(f'  {"─" * 51}')

    print(f'  OCV:              {r["OCV"]:.4f} V')
    if r['V_at_1Acm2'] is not None:
        print(f'  V @ 1 A/cm²:      {r["V_at_1Acm2"]:.4f} V')
    print(f'  Peak power:       {r["peak_power_W_cm2"]*1000:.1f} mW/cm²'
          f' @ {r["j_at_peak_power"]:.2f} A/cm²')

    if r.get('HFR_mean') is not None:
        print(f'  Mean HFR (ASR):   {r["HFR_mean"]*1000:.2f} mΩ·cm²')

    if r['tafel'] is not None:
        t = r['tafel']
        print(f'  {"─" * 51}')
        print(f'  Tafel slope:      {t["tafel_slope_mVdec"]:.1f} mV/dec '
              f'(R² = {t["R_squared"]:.4f})')
        print(f'  j₀:               {t["j0_A_cm2"]:.2e} A/cm²')
        print(f'  Fit range:        {t["j_min"]:.3f} – {t["j_max"]:.3f} A/cm²')

    # Loss breakdown
    has_losses = False
    for j_ref in [0.5, 1.0, 1.5]:
        if f'eta_kinetic_{j_ref}' in r and r[f'eta_kinetic_{j_ref}'] is not None:
            if not has_losses:
                print(f'  {"─" * 51}')
                print(f'  {"Loss @ j (A/cm²)":20s} {"Kinetic":>9s} {"Ohmic":>9s} {"M.T.":>9s} {"Total":>9s}')
                print(f'  {"─" * 51}')
                has_losses = True
            ek = r[f'eta_kinetic_{j_ref}']
            eo = r[f'eta_ohmic_{j_ref}'] or 0
            em = r[f'eta_mt_{j_ref}'] or 0
            et = ek + eo + em
            print(f'  {"j = " + str(j_ref):20s} {ek*1000:>8.0f}  {eo*1000:>8.0f}  {em*1000:>8.0f}  {et*1000:>8.0f} mV')

    print(f'{"═" * 55}\n')


# ═══════════════════════════════════════════════════════════════════════
#  Synthetic Data
# ═══════════════════════════════════════════════════════════════════════

def generate_synthetic_polcurve(geo_area=5.0, noise=0.002):
    """
    Generate a synthetic PEM fuel cell polarization curve with HFR.

    Returns j (A/cm²), V (V), HFR (Ohm).
    """
    j = np.concatenate([
        np.array([0.0]),
        np.linspace(0.01, 0.05, 5),
        np.linspace(0.1, 2.0, 40),
    ])

    E_rev = 1.23
    # Kinetic loss (Tafel)
    alpha_c = 1.0
    j0 = 1e-3
    b = 0.065  # V/dec (Tafel slope)

    with np.errstate(divide='ignore', invalid='ignore'):
        eta_act = np.where(j > 0, b * np.log10(j / j0), 0)

    # Ohmic loss
    R_ohm = 0.050  # Ohm·cm²
    eta_ohm = j * R_ohm

    # Mass transport loss
    j_lim = 2.5
    with np.errstate(divide='ignore', invalid='ignore'):
        eta_mt = np.where(j > 0,
                          -0.03 * np.log(1 - j / j_lim),
                          0)

    V = E_rev - eta_act - eta_ohm - eta_mt

    # HFR with slight current-dependence
    HFR = (R_ohm / geo_area) * (1 + 0.05 * j)  # Ohm

    # Add noise
    V += np.random.normal(0, noise, len(V))
    HFR += np.random.normal(0, noise * 0.01, len(HFR))

    V = np.clip(V, 0, None)

    return j, V, HFR


# ═══════════════════════════════════════════════════════════════════════
#  Excel Output
# ═══════════════════════════════════════════════════════════════════════

def save_single_excel(results, filepath):
    """
    Save single-file analysis to Excel with summary + curve data sheets.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill('solid', fgColor='D9E1F2')

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = 'Summary'
    summary = [
        ('OCV (V)', results['OCV']),
        ('Peak Power (mW/cm²)', results['peak_power_W_cm2'] * 1000),
        ('j @ Peak Power (A/cm²)', results['j_at_peak_power']),
    ]
    if results['V_at_1Acm2'] is not None:
        summary.append(('V @ 1 A/cm² (V)', results['V_at_1Acm2']))
    if results.get('HFR_mean') is not None:
        summary.append(('Mean HFR (mΩ·cm²)', results['HFR_mean'] * 1000))
    if results['tafel'] is not None:
        t = results['tafel']
        summary.append(('Tafel Slope (mV/dec)', t['tafel_slope_mVdec']))
        summary.append(('j₀ (A/cm²)', t['j0_A_cm2']))
        summary.append(('Tafel R²', t['R_squared']))
    cond = results.get('conditions')
    if cond:
        for name, vals in cond.items():
            summary.append((name, np.mean(vals)))

    for i, (label, val) in enumerate(summary, 1):
        ws.cell(row=i, column=1, value=label).font = hdr_font
        ws.cell(row=i, column=2, value=val)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    # ── Sheet 2: Pol Curve Data ──
    ws2 = wb.create_sheet('Pol Curve Data')
    headers = ['j (A/cm²)', 'V (V)']
    has_hfr = results['HFR'] is not None
    if has_hfr:
        headers.append('HFR (mΩ·cm²)')
    if results['V_irfree'] is not None:
        headers.append('V_iR-free (V)')
    headers.append('Power (mW/cm²)')

    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[cell.column_letter].width = 16

    j, V, P = results['j'], results['V'], results['P']
    # Output in V-ascending order (min V first)
    order = np.argsort(V)
    for row_idx, i in enumerate(order):
        col = 1
        ws2.cell(row=row_idx + 2, column=col, value=round(float(j[i]), 6)); col += 1
        ws2.cell(row=row_idx + 2, column=col, value=round(float(V[i]), 6)); col += 1
        if has_hfr:
            ws2.cell(row=row_idx + 2, column=col,
                     value=round(float(results['HFR_ASR'][i]) * 1000, 2)); col += 1
        if results['V_irfree'] is not None:
            ws2.cell(row=row_idx + 2, column=col,
                     value=round(float(results['V_irfree'][i]), 6)); col += 1
        ws2.cell(row=row_idx + 2, column=col, value=round(float(P[i]) * 1000, 2))

    wb.save(filepath)
    print(f'  Excel: {filepath}')


def save_batch_excel(all_results, summary_rows, filepath):
    """
    Save batch analysis to Excel: summary sheet + one data sheet per file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill('solid', fgColor='D9E1F2')

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = 'Summary'
    if summary_rows:
        all_cols = list(summary_rows[0].keys())
        for row in summary_rows[1:]:
            for k in row:
                if k not in all_cols:
                    all_cols.append(k)

        for c, h in enumerate(all_cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
        for r_idx, row in enumerate(summary_rows, 2):
            for c, col_name in enumerate(all_cols, 1):
                val = row.get(col_name)
                if val is not None:
                    ws.cell(row=r_idx, column=c,
                            value=round(val, 6) if isinstance(val, float) else val)
        for c in range(1, len(all_cols) + 1):
            ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 18

    # ── One data sheet per file ──
    for r in all_results:
        label = r.get('label', 'data')
        # Sheet name max 31 chars, no special chars
        sheet_name = label[:31].replace('/', '-').replace('\\', '-').replace('*', '')
        ws2 = wb.create_sheet(sheet_name)

        headers = ['j (A/cm²)', 'V (V)']
        has_hfr = r['HFR'] is not None
        if has_hfr:
            headers.append('HFR (mΩ·cm²)')
        if r['V_irfree'] is not None:
            headers.append('V_iR-free (V)')
        headers.append('Power (mW/cm²)')

        for c, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            ws2.column_dimensions[cell.column_letter].width = 16

        j, V, P = r['j'], r['V'], r['P']
        order = np.argsort(V)
        for row_idx, i in enumerate(order):
            col = 1
            ws2.cell(row=row_idx + 2, column=col, value=round(float(j[i]), 6)); col += 1
            ws2.cell(row=row_idx + 2, column=col, value=round(float(V[i]), 6)); col += 1
            if has_hfr:
                ws2.cell(row=row_idx + 2, column=col,
                         value=round(float(r['HFR_ASR'][i]) * 1000, 2)); col += 1
            if r['V_irfree'] is not None:
                ws2.cell(row=row_idx + 2, column=col,
                         value=round(float(r['V_irfree'][i]), 6)); col += 1
            ws2.cell(row=row_idx + 2, column=col, value=round(float(P[i]) * 1000, 2))

    wb.save(filepath)
    print(f'\n  Excel: {filepath}')


# ═══════════════════════════════════════════════════════════════════════
#  Demo
# ═══════════════════════════════════════════════════════════════════════

def run_demo(save_dir=None):
    """Run full demo with synthetic data."""
    print('\n' + '▓' * 60)
    print('  FUEL CELL POLCURVE ANALYSIS — DEMO')
    print('▓' * 60)

    if save_dir:
        os.makedirs(save_dir, exist_ok=True)

    np.random.seed(42)
    geo_area = 5.0

    # ── Demo 1: Single pol curve with HFR ──
    print('\n[1] Single Polarization Curve + HFR')
    j, V, HFR = generate_synthetic_polcurve(geo_area=geo_area)
    print(f'    {len(j)} points, j: {j.min():.3f} – {j.max():.3f} A/cm²')

    r = analyze_polcurve(j, V, HFR=HFR, geo_area=geo_area)
    print_results(r)

    p1 = os.path.join(save_dir, 'polcurve_demo.png') if save_dir else None
    plot_polcurve(r, save_path=p1)

    # ── Demo 2: Degradation overlay ──
    print('\n[2] Degradation Overlay')
    all_results = []
    for i, (r_scale, j_lim_scale) in enumerate([(1.0, 1.0), (1.1, 0.95),
                                                  (1.25, 0.88), (1.4, 0.80)]):
        j_d, V_d, HFR_d = generate_synthetic_polcurve(geo_area=geo_area)
        # Degrade: increase HFR, reduce limiting current
        V_d -= j_d * 0.050 * (r_scale - 1.0) * geo_area / geo_area
        V_d *= j_lim_scale / 1.0  # crude scaling
        V_d = np.clip(V_d, 0, None)

        rd = analyze_polcurve(j_d, V_d, HFR=HFR_d * r_scale, geo_area=geo_area)
        rd['label'] = f'{i * 250}h'
        all_results.append(rd)
        print(f'    {rd["label"]}: Peak P = {rd["peak_power_W_cm2"]*1000:.0f} mW/cm²')

    p2 = os.path.join(save_dir, 'polcurve_overlay.png') if save_dir else None
    plot_polcurve_overlay(all_results, save_path=p2)

    if save_dir is None:
        plt.show()

    return r


# ═══════════════════════════════════════════════════════════════════════
#  Batch Processing
# ═══════════════════════════════════════════════════════════════════════

def run_batch(filepaths, labels, geo_area,
              delimiter, skip, j_col, v_col, hfr_col,
              current_is_total, tafel_j_min, tafel_j_max,
              condition_cols=None, hfr_scale=1.0, save_dir=None,
              mode_col=None, mode_exclude=None,
              j_scale=1.0, v_scale=1.0, image_ext='png'):
    """
    Batch-process multiple polarization curve files.
    """
    if save_dir:
        os.makedirs(save_dir, exist_ok=True)

    all_results = []
    summary_rows = []

    print(f'\n  Processing {len(filepaths)} files...\n')

    for i, (fp, lbl) in enumerate(zip(filepaths, labels)):
        print(f'  [{i+1}/{len(filepaths)}] {lbl}')
        try:
            fcd = parse_fcd_header(fp)
            file_skip = fcd['skip'] if fcd else skip
            file_j_col = fcd.get('j_col', j_col) if fcd else j_col
            file_v_col = fcd.get('v_col', v_col) if fcd else v_col
            file_hfr_col = fcd.get('hfr_col', hfr_col) if fcd else hfr_col
            file_mode_col = fcd.get('mode_col', mode_col) if fcd else mode_col
            file_cond_cols = fcd.get('condition_cols', condition_cols) if fcd else condition_cols
            j_raw, V_raw, HFR_raw, cond_raw = load_polcurve_data(
                fp, j_col=file_j_col, v_col=file_v_col, hfr_col=file_hfr_col,
                delimiter=delimiter, skip_header=file_skip,
                current_is_total=current_is_total, geo_area=geo_area,
                condition_cols=file_cond_cols, hfr_scale=hfr_scale,
                mode_col=file_mode_col, mode_exclude=mode_exclude,
                j_scale=j_scale, v_scale=v_scale)

            # Extract dwell endpoints
            v_step = detect_voltage_step(V_raw)
            j_dwell, V_dwell, HFR_dwell, cond_dwell = extract_dwell_endpoints(
                j_raw, V_raw, HFR_raw, cond_raw, v_step=v_step)

            # Filter HFR outliers
            if HFR_dwell is not None:
                HFR_dwell, n_out, _ = filter_hfr_outliers(HFR_dwell, geo_area=geo_area)
                if n_out > 0:
                    print(f'         HFR: {n_out} outlier(s) replaced')

            # Detect cycles and use last down-sweep
            cycles = extract_polcurve_cycles(j_dwell, V_dwell, HFR_dwell, cond_dwell, v_step=v_step)
            cycles = enforce_consensus_setpoints(cycles, v_step)
            rep = select_representative_cycle(cycles, choice='last_down')
            j_rep, V_rep, HFR_rep, cond_rep, n_merged = deduplicate_polcurve(
                rep['j'], rep['V'], rep.get('HFR'), rep.get('conditions'),
                v_step=v_step)
            j_rep, V_rep, HFR_rep, cond_rep = enforce_v_monotonicity(
                j_rep, V_rep, HFR_rep, cond_rep)
            n_up = sum(1 for c in cycles if c['direction'] == 'up')
            n_dn = sum(1 for c in cycles if c['direction'] == 'down')

            print(f'         {len(j_raw)} raw → {len(j_dwell)} dwells → '
                  f'{n_up} up + {n_dn} down → {len(j_rep)} pts (last down)'
                  f'{f", {n_merged} deduped" if n_merged else ""}'
                  f'{", HFR included" if HFR_dwell is not None else ""}')

            r = analyze_polcurve(j_rep, V_rep, HFR=HFR_rep, geo_area=geo_area,
                                 tafel_j_min=tafel_j_min, tafel_j_max=tafel_j_max)
            r['label'] = lbl
            r['filepath'] = fp
            r['conditions'] = cond_rep
            r['_cycles_raw'] = cycles
            all_results.append(r)

            peak_p = r['peak_power_W_cm2'] * 1000
            v1 = f'{r["V_at_1Acm2"]:.3f}' if r['V_at_1Acm2'] is not None else '—'
            hfr_str = f'  HFR={r["HFR_mean"]*1000:.1f} mΩ·cm²' if r.get('HFR_mean') else ''
            print(f'         OCV={r["OCV"]:.3f} V  Peak P={peak_p:.0f} mW/cm²  V@1={v1} V{hfr_str}')

            # Print conditions if available
            cond = rep.get('conditions')
            if cond:
                cond_str = '  |  '.join(f'{k}={v[0]:.1f}' for k, v in cond.items())
                print(f'         Conditions: {cond_str}')

            # Individual plot
            if save_dir:
                safe_name = lbl.replace(' ', '_').replace('/', '-').replace('\\', '-')
                plot_polcurve(r, save_path=os.path.join(save_dir, f'polcurve_down_{safe_name}.{image_ext}'))
                plt.close()

            # Summary row
            row = {
                'Label': lbl,
                'File': os.path.basename(fp),
                'OCV (V)': r['OCV'],
                'Peak Power (mW/cm2)': peak_p,
                'j @ Peak P (A/cm2)': r['j_at_peak_power'],
            }
            if r['V_at_1Acm2'] is not None:
                row['V @ 1 A/cm2 (V)'] = r['V_at_1Acm2']
            if r.get('HFR_mean') is not None:
                row['Mean HFR (mOhm·cm2)'] = r['HFR_mean'] * 1000
            if r['tafel'] is not None:
                row['Tafel (mV/dec)'] = r['tafel']['tafel_slope_mVdec']
                row['j0 (A/cm2)'] = r['tafel']['j0_A_cm2']
            # Add conditions to summary
            if cond:
                for k, v in cond.items():
                    row[k] = np.mean(v)  # mean across the sweep
            summary_rows.append(row)

        except Exception as e:
            print(f'         ERROR: {e}')
            continue

    if not all_results:
        print('\n  No files processed successfully.')
        return []

    # ── Summary CSV ──
    if save_dir and summary_rows:
        csv_path = os.path.join(save_dir, 'polcurve_down_batch_summary.csv')
        all_cols = list(summary_rows[0].keys())
        for row in summary_rows[1:]:
            for k in row:
                if k not in all_cols:
                    all_cols.append(k)

        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=all_cols)
            writer.writeheader()
            writer.writerows(summary_rows)
        print(f'\n  Summary CSV: {csv_path}')

    # ── Batch Excel ──
    if save_dir and all_results:
        xlsx_path = os.path.join(save_dir, 'polcurve_down_batch_data.xlsx')
        save_batch_excel(all_results, summary_rows, xlsx_path)

    # ── Combined overlay ──
    if save_dir:
        overlay_path = os.path.join(save_dir, f'polcurve_down_batch_overlay.{image_ext}')
        plot_polcurve_overlay(all_results, save_path=overlay_path)
        plt.close()
    else:
        plot_polcurve_overlay(all_results)
        plt.show()

    # ── Console summary ──
    has_hfr = any(r.get('HFR_mean') is not None for r in all_results)
    has_tafel = any(r['tafel'] is not None for r in all_results)

    print(f'\n  {"═" * 85}')
    hdr = f'  {"Label":25s} {"OCV":>6s} {"V@1A":>7s} {"Peak P":>8s} {"j@Pk":>6s}'
    if has_hfr:
        hdr += f' {"HFR":>8s}'
    if has_tafel:
        hdr += f' {"Tafel":>7s}'
    print(hdr)
    print(f'  {"":25s} {"(V)":>6s} {"(V)":>7s} {"(mW)":>8s} {"(A)":>6s}', end='')
    if has_hfr:
        print(f' {"(mΩ·cm²)":>8s}', end='')
    if has_tafel:
        print(f' {"(mV/d)":>7s}', end='')
    print()
    print(f'  {"─" * 85}')

    for r in all_results:
        v1 = f'{r["V_at_1Acm2"]:.3f}' if r['V_at_1Acm2'] is not None else '—'
        line = (f'  {r["label"]:25s} {r["OCV"]:>6.3f} {v1:>7s}'
                f' {r["peak_power_W_cm2"]*1000:>8.0f} {r["j_at_peak_power"]:>6.2f}')
        if has_hfr:
            if r.get('HFR_mean') is not None:
                line += f' {r["HFR_mean"]*1000:>8.1f}'
            else:
                line += f' {"—":>8s}'
        if has_tafel:
            if r['tafel'] is not None:
                line += f' {r["tafel"]["tafel_slope_mVdec"]:>7.0f}'
            else:
                line += f' {"—":>7s}'
        print(line)

    print(f'  {"═" * 85}\n')

    return all_results


# ═══════════════════════════════════════════════════════════════════════
#  Portal Entry Point
# ═══════════════════════════════════════════════════════════════════════

def run(input_dir: str, output_dir: str, params: dict = None) -> dict:
    """Batch polarization curve DOWNSWING analysis from fuel cell test data.

    Identical pipeline to the standard polcurve analysis except the
    representative cycle is the last DOWN-sweep (OCV → high current) rather
    than the last up-sweep.
    """
    from pathlib import Path
    p = params or {}

    inp = Path(input_dir)
    all_files = sorted([f for f in inp.rglob("*")
                    if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv', '.fcd')])

    if not all_files:
        return {"status": "error", "message": "No data files found"}

    # Filter by polcurve keywords; fall back to all files if none match
    KEYWORDS = ['POLCURVE', 'POL_', 'POL-', 'IV_', 'IV-',
                'POLARIZATION', 'POLDATA']
    filtered = [f for f in all_files
                if any(kw in f.name.upper() for kw in KEYWORDS)]
    files = filtered if filtered else all_files

    has_fcd_files = any(f.suffix.lower() == '.fcd' for f in files)
    has_tab_files = any(f.suffix.lower() in ('.fcd', '.tsv') for f in files)

    filepaths = [str(f) for f in files]
    labels = [f.stem for f in files]

    from scripts.helpers.conditions import img_ext_from_params
    image_ext = img_ext_from_params(p)

    # ── Format detection (matches the standard polcurve analysis) ──
    if has_fcd_files:
        delimiter = '\t'
        skip = 51
        j_col = 1
        v_col = 5
        hfr_col = 20
        hfr_scale = 0.001  # mΩ → Ω
        current_is_total = True
        mode_col = 28
        mode_exclude = {5}
        condition_cols = {
            'T_cell (C)': 13, 'T_anode_dp (C)': 14, 'H2_flow (slpm)': 15,
            'T_cathode_dp (C)': 17, 'Air_flow (slpm)': 18,
        }
    else:
        delimiter = '\t' if has_tab_files else ','
        skip = 1
        j_col = 0
        v_col = 1
        hfr_col = None
        hfr_scale = 1.0
        current_is_total = True
        mode_col = None
        mode_exclude = None
        condition_cols = None

    results = run_batch(
        filepaths, labels,
        geo_area=float(p.get('geo_area', 5.0)),
        delimiter=delimiter,
        skip=skip,
        j_col=j_col,
        v_col=v_col,
        hfr_col=hfr_col,
        hfr_scale=hfr_scale,
        current_is_total=current_is_total,
        mode_col=mode_col,
        mode_exclude=mode_exclude,
        condition_cols=condition_cols,
        tafel_j_min=float(p.get('tafel_j_min', 0.01)),
        tafel_j_max=float(p.get('tafel_j_max', 0.10)),
        save_dir=str(output_dir),
        image_ext=image_ext,
    )

    output_files = [str(f.relative_to(Path(output_dir)))
                    for f in Path(output_dir).rglob('*') if f.is_file()]
    if not output_files:
        raise RuntimeError(
            f'Analysis produced no output. {len(files)} file(s) were found '
            f'but none could be processed. Check file format and parameters.'
        )
    return {"status": "success", "files_processed": len(files), "files_produced": output_files}
