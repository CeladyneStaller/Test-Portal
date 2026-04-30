#!/usr/bin/env python3
"""
Fuel Cell Polarization Curve & HFR Analysis
============================================
Analyzes polarization curve data for PEM fuel cells.

Features:
  - Polarization curve and power density plots
  - HFR overlay and iR-free voltage correction
  - Tafel slope extraction (kinetic region)
  - Loss breakdown: kinetic, ohmic, mass transport
  - Batch processing (folder or file list)
  - Multi-file overlay for degradation tracking

Usage:
  python polcurve_analysis.py                    # interactive mode
  python polcurve_analysis.py --file data.csv    # analyze pol curve
  python polcurve_analysis.py --demo             # run built-in demo
"""

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scripts.helpers.plot_compare import save_with_sidecar
from scipy.optimize import curve_fit
import argparse
import csv
import os
import glob


def run(input_dir: str, output_dir: str, params: dict = None) -> dict:
    """Batch polarization curve analysis from fuel cell test data."""
    from pathlib import Path
    p = params or {}

    inp = Path(input_dir)
    all_files = sorted([f for f in inp.rglob("*")
                    if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv', '.fcd')])

    if not all_files:
        return {"status": "error", "message": "No data files found"}

    # Filter by polcurve keywords; fall back to all files if none match
    KEYWORDS = ['POLCURVE', 'POL_', 'POL-', 'IV_', 'IV-',
                'POLARIZATION', 'POLDATA']
    filtered = [f for f in all_files
                if any(kw in f.name.upper() for kw in KEYWORDS)]
    files = filtered if filtered else all_files

    has_fcd_files = any(f.suffix.lower() == '.fcd' for f in files)
    has_tab_files = any(f.suffix.lower() in ('.fcd', '.tsv') for f in files)

    filepaths = [str(f) for f in files]
    labels = [f.stem for f in files]

    from scripts.helpers.conditions import img_ext_from_params
    image_ext = img_ext_from_params(p)

    # ── Format detection ──
    # For .fcd / Scribner files, use Scribner preset (tab-delimited,
    # multi-line header parsed via parse_fcd_header, HFR in mΩ).
    # For CSV files, use simple defaults that the user can override via params.
    if has_fcd_files:
        # Scribner preset — actual columns/skip will be re-detected per file
        # by parse_fcd_header inside run_batch
        delimiter = '\t'
        skip = 51
        j_col = 1
        v_col = 5
        hfr_col = 20
        hfr_scale = 0.001  # mΩ → Ω
        current_is_total = True
        mode_col = 28
        mode_exclude = {5}
        condition_cols = {
            'T_cell (C)': 13, 'T_anode_dp (C)': 14, 'H2_flow (slpm)': 15,
            'T_cathode_dp (C)': 17, 'Air_flow (slpm)': 18,
        }
    else:
        delimiter = '\t' if has_tab_files else ','
        skip = 1
        j_col = 0
        v_col = 1
        hfr_col = None
        hfr_scale = 1.0
        current_is_total = True
        mode_col = None
        mode_exclude = None
        condition_cols = None

    results = run_batch(
        filepaths, labels,
        geo_area=float(p.get('geo_area', 5.0)),
        delimiter=delimiter,
        skip=skip,
        j_col=j_col,
        v_col=v_col,
        hfr_col=hfr_col,
        hfr_scale=hfr_scale,
        current_is_total=current_is_total,
        mode_col=mode_col,
        mode_exclude=mode_exclude,
        condition_cols=condition_cols,
        tafel_j_min=float(p.get('tafel_j_min', 0.01)),
        tafel_j_max=float(p.get('tafel_j_max', 0.10)),
        save_dir=str(output_dir),
        image_ext=image_ext,
    )

    output_files = [str(f.relative_to(Path(output_dir))) for f in Path(output_dir).rglob('*') if f.is_file()]
    if not output_files:
        raise RuntimeError(
            f'Analysis produced no output. {len(files)} file(s) were found '
            f'but none could be processed. Check file format and parameters.'
        )
    return {"status": "success", "files_processed": len(files), "files_produced": output_files}


# ═══════════════════════════════════════════════════════════════════════
#  Data I/O
# ═══════════════════════════════════════════════════════════════════════

def parse_fcd_header(filepath):
    """
    Parse Scribner .fcd file header to extract skip count and column indices.

    Returns dict with 'skip' and column indices, or None for non-.fcd files.
    Column keys: j_col, v_col, hfr_col, freq_col, zreal_col, zimag_col,
                 mode_col, and condition_cols dict.
    """
    if not filepath.lower().endswith('.fcd'):
        return None
    try:
        with open(filepath, 'r', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                lines.append(line)
                if 'End Comments' in line:
                    skip = i + 1
                    break
            else:
                return None
    except Exception:
        return None

    # Column header line is immediately before "End Comments"
    if len(lines) < 2:
        return None
    header_line = lines[-2].strip()
    cols = header_line.split('\t')

    result = {'skip': skip}
    cond_cols = {}

    for ci, name in enumerate(cols):
        n = name.strip()
        # Current (total, in A) — match "I (A)" but not "I (mA/cm²)"
        if n == 'I (A)':
            result['j_col'] = ci
        # Voltage — stack voltage
        elif n == 'E_Stack (V)':
            result['v_col'] = ci
        # HFR — non-area-normalized (mOhm)
        elif n == 'HFR (mOhm)':
            result['hfr_col'] = ci
        # EIS columns
        elif n.startswith('Z_Freq'):
            result['freq_col'] = ci
        elif n.startswith('Z_Real'):
            result['zreal_col'] = ci
        elif n.startswith('Z_Imag'):
            result['zimag_col'] = ci
        # Control mode
        elif n == 'Ctrl_Mode':
            result['mode_col'] = ci
        # Condition columns
        elif n == 'Cell (C)':
            cond_cols['T_cell (C)'] = ci
        elif n.startswith('Temp_Anode'):
            cond_cols['T_anode_dp (C)'] = ci
        elif n.startswith('Flow_Anode'):
            cond_cols['H2_flow (slpm)'] = ci
        elif n.startswith('Temp_Cathode'):
            cond_cols['T_cathode_dp (C)'] = ci
        elif n.startswith('Flow_Cathode'):
            cond_cols['Air_flow (slpm)'] = ci

    if cond_cols:
        result['condition_cols'] = cond_cols

    return result



def _clean_path(p):
    """Clean a file path from drag-drop artifacts and smart quotes."""
    p = p.strip()
    if p.startswith('& '):
        p = p[2:]
    p = p.strip().strip('"').strip("'")
    p = p.strip('\u2018\u2019\u201c\u201d')
    p = p.strip('\u202a\u200b')
    return p


def load_polcurve_data(filepath, j_col=0, v_col=1, hfr_col=None,
                       delimiter=',', skip_header=1,
                       current_is_total=False, geo_area=5.0,
                       condition_cols=None, hfr_scale=1.0,
                       mode_col=None, mode_exclude=None,
                       j_scale=1.0, v_scale=1.0):
    """
    Load polarization curve data from a CSV/TSV file.

    Parameters
    ----------
    filepath : str
        Path to data file.
    j_col : int
        Column index for current density (A/cm²) or total current (A).
    v_col : int
        Column index for cell voltage (V).
    hfr_col : int or None
        Column index for HFR (Ohm). None if no HFR data.
    delimiter : str
        Column delimiter.
    skip_header : int
        Number of header rows to skip.
    current_is_total : bool
        If True, column is total current (A) and will be divided by geo_area.
    geo_area : float
        Geometric area in cm² (used only if current_is_total=True).
    condition_cols : dict or None
        Mapping of condition name to column index.
    hfr_scale : float
        Multiplier for HFR values (e.g. 0.001 to convert mΩ → Ω).
    mode_col : int or None
        Column index for control mode (e.g. Scribner Ctrl_Mode).
        Used with mode_exclude to filter out non-polcurve data.
    mode_exclude : set/list or None
        Mode values to exclude (e.g. {5} removes HFR measurement points).

    Returns
    -------
    j_raw, V_raw, HFR_raw, conditions_raw
    """
    filepath = filepath.strip()
    if filepath.startswith('& '):
        filepath = filepath[2:]
    filepath = filepath.strip().strip('"').strip("'")
    filepath = filepath.strip('\u2018\u2019\u201c\u201d')
    filepath = filepath.strip('\u202a\u200b')

    if not os.path.isfile(filepath):
        raise FileNotFoundError(
            f'File not found: "{filepath}"\n'
            f'  Check the path and try again. On Windows, right-click the file\n'
            f'  in Explorer → Copy as path, then paste into the terminal.')

    j_data, v_data, hfr_data = [], [], []
    cond_data = {name: [] for name in (condition_cols or {})}
    mode_excl = set(mode_exclude) if mode_exclude else set()

    with open(filepath, 'r', errors='replace') as f:
        reader = csv.reader(f, delimiter=delimiter)
        for _ in range(skip_header):
            next(reader)
        for row in reader:
            try:
                # Filter by control mode if specified
                if mode_col is not None and mode_excl:
                    mode_val = int(float(row[mode_col]))
                    if mode_val in mode_excl:
                        continue
                j_val = float(row[j_col])
                v_val = float(row[v_col])
                j_data.append(j_val)
                v_data.append(v_val)
                if hfr_col is not None:
                    hfr_data.append(float(row[hfr_col]))
                for name, col in (condition_cols or {}).items():
                    cond_data[name].append(float(row[col]))
            except (ValueError, IndexError):
                continue

    j = np.array(j_data) * j_scale
    V = np.array(v_data) * v_scale

    if current_is_total:
        j = j / geo_area

    HFR = np.array(hfr_data) * hfr_scale if hfr_col is not None and len(hfr_data) > 0 else None
    conditions = {name: np.array(vals) for name, vals in cond_data.items()} if condition_cols else None

    return j, V, HFR, conditions


def extract_dwell_endpoints(j_raw, V_raw, HFR_raw=None, conditions_raw=None,
                            min_dwell_pts=5, v_step=None):
    """
    Extract the last data point from each voltage or current dwell.

    When *v_step* is provided (voltage-step protocol detected), groups
    consecutive points by voltage stability — the controlled variable.
    Otherwise, groups by current density with a relative tolerance.

    Parameters
    ----------
    j_raw, V_raw : array
        Time-ordered raw data.
    HFR_raw : array or None
    conditions_raw : dict of arrays or None
    min_dwell_pts : int
        Minimum points for a group to be considered a valid dwell.
    v_step : float or None
        Voltage step size.  When provided, uses V-based grouping (tighter
        tolerance, better separation of adjacent voltage setpoints).

    Returns
    -------
    j, V, HFR, conditions : arrays (one point per dwell)
    """
    if len(j_raw) < 2:
        return j_raw, V_raw, HFR_raw, conditions_raw

    # ── Choose grouping signal: V (controlled) or j (response) ──
    if v_step is not None:
        # Voltage-step protocol: group by voltage stability
        signal = V_raw
        abs_floor = v_step * 0.25
        rel_frac = 0.01

        # Pre-filter: detect HFR impedance measurement zones by high local
        # V variance.  These zones have rapidly alternating V (±50 mV between
        # consecutive points) and contaminate the high-V setpoint region.
        hw = 5  # half-window for rolling std
        n = len(V_raw)
        stable = np.ones(n, dtype=bool)
        var_threshold = v_step * 0.4  # ~10 mV for 25 mV step
        for i in range(n):
            lo, hi = max(0, i - hw), min(n, i + hw + 1)
            if np.std(V_raw[lo:hi]) > var_threshold:
                stable[i] = False
    else:
        # Current-step or unknown: group by current stability
        signal = j_raw
        dj_consecutive = np.abs(np.diff(j_raw))
        noise_est = np.percentile(dj_consecutive, 50)
        abs_floor = max(noise_est * 10, 0.003)
        rel_frac = 0.03
        stable = np.ones(len(j_raw), dtype=bool)

    # ── Group consecutive STABLE points by running-mean comparison ──
    groups = []
    group_start = 0
    group_sum = signal[0]
    group_count = 1

    for i in range(1, len(signal)):
        if not stable[i]:
            # Unstable point (HFR zone): end current group, skip this point
            if group_count > 0:
                groups.append((group_start, i))
            group_start = i + 1
            group_sum = 0
            group_count = 0
            continue
        if group_count == 0:
            # Starting new group after unstable region
            group_start = i
            group_sum = signal[i]
            group_count = 1
            continue
        group_mean = group_sum / group_count
        tol = max(abs_floor, rel_frac * abs(group_mean))
        if abs(signal[i] - group_mean) <= tol:
            group_sum += signal[i]
            group_count += 1
        else:
            groups.append((group_start, i))
            group_start = i
            group_sum = signal[i]
            group_count = 1

    if group_count > 0:
        groups.append((group_start, len(signal)))

    # ── Extract stable current window from each voltage dwell ──
    # V is the controlled variable → use dwell mean (constant by design)
    # j is the response variable → from last 20 stable pts, average those
    #   where V is within 2% of the setpoint voltage
    j_out, v_out, hfr_out = [], [], []
    cond_out = {name: [] for name in (conditions_raw or {})}

    for start, end in groups:
        n_pts = end - start
        if n_pts < min_dwell_pts:
            continue

        # V: mean of entire dwell (controlled variable, should be constant)
        V_setpoint = np.mean(V_raw[start:end])
        v_out.append(V_setpoint)

        # Find stable current window by walking backwards
        cv_threshold = 0.03   # 3% coefficient of variation
        abs_floor = 0.002     # A/cm² absolute std floor for near-zero currents

        stable_start = end - 1
        for k in range(end - 2, start - 1, -1):
            tail = j_raw[k:end]
            tail_mean = np.mean(tail)
            tail_std = np.std(tail)
            threshold = max(cv_threshold * abs(tail_mean), abs_floor)
            if tail_std > threshold:
                break
            stable_start = k

        stable_n = end - stable_start
        if stable_n < 1:
            stable_start = end - 1
            stable_n = 1

        # Take last 20 points of stable window
        n_tail = min(20, stable_n)
        tail_start = end - n_tail
        tail_sl = slice(tail_start, end)

        # Filter: keep only points where V is within 2% of the setpoint
        V_tail = V_raw[tail_sl]
        v_tol = max(abs(V_setpoint) * 0.02, 0.005)  # 2% or 5mV floor
        v_mask = np.abs(V_tail - V_setpoint) <= v_tol

        if v_mask.sum() > 0:
            # Average the V-filtered points
            idx = np.where(v_mask)[0] + tail_start
            j_out.append(np.mean(j_raw[idx]))
            if HFR_raw is not None:
                hfr_out.append(np.mean(HFR_raw[idx]))
            for name in cond_out:
                cond_out[name].append(np.mean(conditions_raw[name][idx]))
        else:
            # Fallback: average last 5 of stable window
            n_fb = min(5, stable_n)
            fb_sl = slice(end - n_fb, end)
            j_out.append(np.mean(j_raw[fb_sl]))
            if HFR_raw is not None:
                hfr_out.append(np.mean(HFR_raw[fb_sl]))
            for name in cond_out:
                cond_out[name].append(np.mean(conditions_raw[name][fb_sl]))

    j = np.array(j_out)
    V = np.array(v_out)
    HFR = np.array(hfr_out) if HFR_raw is not None else None
    conditions = {name: np.array(vals) for name, vals in cond_out.items()} if conditions_raw else None

    return j, V, HFR, conditions


def extract_polcurve_cycles(j, V, HFR=None, conditions=None,
                           min_sweep_pts=4, v_step=None):
    """
    Detect up-sweep and down-sweep cycles from dwell-extracted pol curve data.

    Uses prominence-based local maxima/minima detection to find cycle
    turnarounds.  A peak (local V maximum) or valley (local V minimum)
    is only confirmed when V moves away from it by at least
    min_prominence.  This filters HFR artifact oscillations (±5 mV)
    while catching real turnarounds (±100s mV), and is robust to
    OCV degradation or protocols with partial voltage returns.

    Parameters
    ----------
    j, V : array
        Dwell-extracted data (time-ordered, one point per setpoint).
    HFR : array or None
    conditions : dict of arrays or None
    min_sweep_pts : int
        Minimum number of points for a valid sweep.
    v_step : float or None
        Voltage step size.  Prominence threshold = 5 × v_step.

    Returns
    -------
    cycles : list of dicts
    """
    def _slice_conditions(cond, idx):
        if cond is None:
            return None
        return {name: arr[idx] for name, arr in cond.items()}

    if len(j) < 3:
        direction = 'down' if len(j) < 2 or j[-1] >= j[0] else 'up'
        return [{'j': j, 'V': V, 'HFR': HFR, 'conditions': conditions,
                 'direction': direction, 'cycle_number': 1}]

    # ── Step 1: Find significant peaks and valleys ──
    # Prominence threshold: 5 voltage steps (e.g., 125 mV for 25 mV step)
    # This easily separates real turnarounds (~600 mV swing) from
    # HFR artifacts (~5 mV oscillation)
    if v_step is not None and v_step > 0:
        min_prominence = 5.0 * v_step
    else:
        min_prominence = 0.10  # 100 mV fallback

    # Zigzag algorithm: track running high/low, confirm when V moves
    # away by min_prominence
    peaks = []    # indices of confirmed V maxima
    valleys = []  # indices of confirmed V minima

    running_high = V[0]
    running_high_idx = 0
    running_low = V[0]
    running_low_idx = 0
    state = None  # None = undetermined, 'seeking_valley', 'seeking_peak'

    for i in range(1, len(V)):
        if V[i] > running_high:
            running_high = V[i]
            running_high_idx = i
        if V[i] < running_low:
            running_low = V[i]
            running_low_idx = i

        if state is None:
            # Determine initial direction from first significant move
            if running_high - V[i] >= min_prominence:
                # V dropped significantly from initial high → confirm peak
                peaks.append(running_high_idx)
                state = 'seeking_valley'
                running_low = V[i]
                running_low_idx = i
            elif V[i] - running_low >= min_prominence:
                # V rose significantly from initial low → confirm valley
                valleys.append(running_low_idx)
                state = 'seeking_peak'
                running_high = V[i]
                running_high_idx = i

        elif state == 'seeking_valley':
            if V[i] - running_low >= min_prominence:
                # V rose from the low → confirm the low as a valley
                valleys.append(running_low_idx)
                state = 'seeking_peak'
                running_high = V[i]
                running_high_idx = i

        elif state == 'seeking_peak':
            if running_high - V[i] >= min_prominence:
                # V dropped from the high → confirm the high as a peak
                peaks.append(running_high_idx)
                state = 'seeking_valley'
                running_low = V[i]
                running_low_idx = i

    # Add trailing unconfirmed extremum at end of data
    # If we're seeking a valley but data ends rising, the running high is
    # the final peak; if seeking a peak but data ends falling, the running
    # low is the final valley.
    if state == 'seeking_valley' and running_low_idx > (valleys[-1] if valleys else -1):
        valleys.append(running_low_idx)
    elif state == 'seeking_peak' and running_high_idx > (peaks[-1] if peaks else -1):
        peaks.append(running_high_idx)

    # ── Step 2: Build cycles from alternating peaks and valleys ──
    # Merge peaks and valleys into a sorted sequence of turnarounds
    turnarounds = []
    for idx in peaks:
        turnarounds.append((idx, 'peak'))
    for idx in valleys:
        turnarounds.append((idx, 'valley'))
    turnarounds.sort(key=lambda x: x[0])

    if len(turnarounds) < 2:
        # No significant turnarounds — single sweep
        v_min_idx = np.argmin(V)
        idx_dn = np.arange(0, v_min_idx + 1)
        idx_up = np.arange(v_min_idx, len(V))
        cycles = []
        for idx_arr, direction in [(idx_dn, 'down'), (idx_up, 'up')]:
            if len(idx_arr) < min_sweep_pts:
                continue
            order = np.argsort(j[idx_arr])
            cycles.append({
                'j': j[idx_arr][order], 'V': V[idx_arr][order],
                'HFR': HFR[idx_arr][order] if HFR is not None else None,
                'conditions': _slice_conditions(conditions, idx_arr) if conditions else None,
                'direction': direction, 'cycle_number': 1,
            })
        return cycles if cycles else [{'j': j, 'V': V, 'HFR': HFR,
                'conditions': conditions, 'direction': 'down', 'cycle_number': 1}]

    # ── Step 3: Build sweeps between consecutive turnarounds ──
    # peak → valley = down-sweep, valley → peak = up-sweep
    all_boundaries = turnarounds.copy()

    cycles = []
    down_count = 0

    for seg in range(len(all_boundaries) - 1):
        seg_start_idx, seg_start_type = all_boundaries[seg]
        seg_end_idx, seg_end_type = all_boundaries[seg + 1]

        # Determine direction from turnaround types
        if seg_start_type == 'peak' and seg_end_type == 'valley':
            direction = 'down'
        elif seg_start_type == 'valley' and seg_end_type == 'peak':
            direction = 'up'
        else:
            direction = 'down' if V[seg_start_idx] > V[min(seg_end_idx, len(V) - 1)] else 'up'

        # First sweep extends back to include any initial ramp
        if seg == 0:
            seg_start_idx = 0

        # Last sweep extends to include trailing data
        if seg == len(all_boundaries) - 2:
            seg_end_idx = len(V) - 1

        # Slice inclusive of both endpoints
        sl_start = seg_start_idx
        sl_end = seg_end_idx + 1
        if sl_end - sl_start < min_sweep_pts:
            continue

        idx_arr = np.arange(sl_start, sl_end)
        order = np.argsort(j[idx_arr])

        if direction == 'down':
            down_count += 1
            cycle_num = down_count
        else:
            cycle_num = down_count

        cycles.append({
            'j': j[idx_arr][order], 'V': V[idx_arr][order],
            'HFR': HFR[idx_arr][order] if HFR is not None else None,
            'conditions': _slice_conditions(conditions, idx_arr) if conditions else None,
            'direction': direction,
            'cycle_number': max(cycle_num, 1),
        })

    if not cycles:
        idx = np.argsort(j)
        return [{'j': j[idx], 'V': V[idx],
                 'HFR': HFR[idx] if HFR is not None else None,
                 'conditions': _slice_conditions(conditions, idx),
                 'direction': 'up', 'cycle_number': 1}]

    return cycles


def enforce_consensus_setpoints(cycles, v_step):
    """
    Ensure all cycles share the same voltage setpoints.

    1.  Dedup each cycle independently to find its setpoints.
    2.  Pool and cluster setpoints across all cycles.
    3.  Keep only setpoints present in a majority of cycles (consensus).
    4.  For each cycle, snap raw points to consensus setpoints (one point
        per setpoint, nearest V wins) and discard non-matching points.

    This guarantees every cycle has the same number of points at the
    same voltage levels, eliminating HFR artifacts and boundary noise.

    Parameters
    ----------
    cycles : list of dicts
        Output of extract_polcurve_cycles().
    v_step : float
        Voltage step size for clustering tolerance.

    Returns
    -------
    list of dicts — cleaned cycles with uniform setpoint count.
    """
    if v_step is None or len(cycles) < 2:
        return cycles

    tol_V = v_step * 0.45

    # Step 1: dedup each cycle, collect V setpoints
    deduped_V = []
    for c in cycles:
        _, V_d, _, _, _ = deduplicate_polcurve(c['j'], c['V'], v_step=v_step)
        deduped_V.append(V_d)

    # Step 2: pool all V values and cluster
    v_to_cycles = {}
    for ci, vs in enumerate(deduped_V):
        for v in vs:
            v_to_cycles.setdefault(round(v, 6), set()).add(ci)

    all_V = np.sort(np.concatenate(deduped_V))[::-1]
    clusters = []
    grp = [all_V[0]]
    for i in range(1, len(all_V)):
        if abs(all_V[i] - np.mean(grp)) <= tol_V:
            grp.append(all_V[i])
        else:
            clusters.append(grp)
            grp = [all_V[i]]
    clusters.append(grp)

    # Count unique cycles per cluster
    n_cycles = len(cycles)
    min_cycles = max(2, n_cycles // 2)
    consensus_V = []
    for grp in clusters:
        cyc_set = set()
        for v in grp:
            cyc_set |= v_to_cycles.get(round(v, 6), set())
        if len(cyc_set) >= min_cycles:
            consensus_V.append(np.mean(grp))
    consensus_V = np.array(sorted(consensus_V, reverse=True))

    if len(consensus_V) == 0:
        return cycles

    # Step 3: for each cycle, snap raw points to consensus setpoints
    def _slice(arr, idx):
        return arr[idx] if arr is not None else None

    cleaned = []
    for c in cycles:
        j_c, V_c = c['j'], c['V']
        HFR_c = c['HFR']
        cond_c = c.get('conditions')

        keep_idx = []
        used = set()
        for cv in consensus_V:
            dists = np.abs(V_c - cv)
            candidates = np.argsort(dists)
            for best in candidates:
                if dists[best] > tol_V:
                    break
                if best not in used:
                    keep_idx.append(best)
                    used.add(best)
                    break

        if not keep_idx:
            cleaned.append(c)
            continue

        idx = np.array(sorted(keep_idx))
        order = np.argsort(j_c[idx])
        idx = idx[order]

        cleaned.append({
            'j': j_c[idx], 'V': V_c[idx],
            'HFR': _slice(HFR_c, idx),
            'conditions': ({name: vals[idx] for name, vals in cond_c.items()}
                           if cond_c else None),
            'direction': c['direction'],
            'cycle_number': c['cycle_number'],
        })

    return cleaned


def select_representative_cycle(cycles, choice='last_up'):
    """
    Select a representative cycle from extracted pol curve cycles.

    Parameters
    ----------
    cycles : list of dicts
        Output of extract_polcurve_cycles().
    choice : str
        'last_up'  — last up-sweep (default, most representative)
        'first_up' — first up-sweep
        'last_down' — last down-sweep
        'all'       — return list of all cycles

    Returns
    -------
    dict (single cycle) or list of dicts (if choice='all').
    """
    if choice == 'all':
        return cycles

    up_sweeps = [c for c in cycles if c['direction'] == 'up']
    down_sweeps = [c for c in cycles if c['direction'] == 'down']

    if choice == 'last_up':
        return up_sweeps[-1] if up_sweeps else cycles[-1]
    elif choice == 'first_up':
        return up_sweeps[0] if up_sweeps else cycles[0]
    elif choice == 'last_down':
        return down_sweeps[-1] if down_sweeps else cycles[-1]
    else:
        raise ValueError(f'Unknown choice: {choice}')


def detect_voltage_step(V_raw):
    """
    Auto-detect the voltage setpoint step size from raw time-series voltage.

    Uses histogram peak analysis: each voltage setpoint produces a cluster
    of raw V values. The most common spacing between adjacent peaks gives
    the step size.

    Returns step_size (V) or None if detection fails.
    """
    V_min, V_max = np.min(V_raw), np.max(V_raw)
    bins = np.arange(V_min - 0.005, V_max + 0.01, 0.005)
    if len(bins) < 4:
        return None
    hist, edges = np.histogram(V_raw, bins=bins)
    centers = (edges[:-1] + edges[1:]) / 2

    # Find peaks: bins with more counts than both neighbours
    threshold = np.percentile(hist[hist > 0], 50)
    peaks = []
    for i in range(1, len(hist) - 1):
        if hist[i] > threshold and hist[i] >= hist[i - 1] and hist[i] >= hist[i + 1]:
            peaks.append(centers[i])

    if len(peaks) < 4:
        return None

    peaks = np.sort(peaks)
    gaps = np.diff(peaks)
    # Mode of gaps rounded to nearest 0.005 V
    gaps_rounded = np.round(gaps / 0.005) * 0.005
    unique_gaps, counts = np.unique(gaps_rounded, return_counts=True)
    return float(unique_gaps[np.argmax(counts)])


def deduplicate_polcurve(j, V, HFR=None, conditions=None, v_step=None):
    """
    Average duplicate points that correspond to the same voltage setpoint.

    When *v_step* is provided (auto-detected or manual), deduplication works
    on voltage — the controlled variable — by snapping each point to a
    regular voltage grid and averaging points that map to the same grid
    value.  This is robust regardless of current noise.

    When *v_step* is None, falls back to current-based grouping with a
    fixed tolerance of 0.005 A/cm².

    Parameters
    ----------
    j, V : array
        Sorted by ascending j.
    HFR : array or None
    conditions : dict of arrays or None
    v_step : float or None
        Voltage setpoint spacing (V).  When provided, enables V-based
        snap-to-grid deduplication.  Use detect_voltage_step() on the
        raw voltage time-series to obtain this automatically.

    Returns
    -------
    j, V, HFR, conditions : deduplicated arrays (sorted by ascending j)
    n_merged : int
        Number of points removed by merging.
    """
    if len(j) < 2:
        return j, V, HFR, conditions, 0

    if v_step is not None and v_step > 0:
        return _dedup_voltage_grid(j, V, HFR, conditions, v_step)
    else:
        return _dedup_current(j, V, HFR, conditions, tol=0.005)


def _dedup_current(j, V, HFR, conditions, tol=0.005):
    """Fallback: group by current density proximity."""
    j_out, v_out, hfr_out = [], [], []
    cond_out = {name: [] for name in (conditions or {})}

    group_j = [j[0]]
    group_v = [V[0]]
    group_hfr = [HFR[0]] if HFR is not None else []
    group_cond = {name: [conditions[name][0]] for name in (conditions or {})}

    for i in range(1, len(j)):
        if abs(j[i] - np.mean(group_j)) <= tol:
            group_j.append(j[i])
            group_v.append(V[i])
            if HFR is not None:
                group_hfr.append(HFR[i])
            for name in group_cond:
                group_cond[name].append(conditions[name][i])
        else:
            j_out.append(np.mean(group_j))
            v_out.append(np.mean(group_v))
            if HFR is not None:
                hfr_out.append(np.mean(group_hfr))
            for name in cond_out:
                cond_out[name].append(np.mean(group_cond[name]))
            group_j = [j[i]]
            group_v = [V[i]]
            group_hfr = [HFR[i]] if HFR is not None else []
            group_cond = {name: [conditions[name][i]] for name in (conditions or {})}

    j_out.append(np.mean(group_j))
    v_out.append(np.mean(group_v))
    if HFR is not None:
        hfr_out.append(np.mean(group_hfr))
    for name in cond_out:
        cond_out[name].append(np.mean(group_cond[name]))

    n_merged = len(j) - len(j_out)
    j_d = np.array(j_out)
    V_d = np.array(v_out)
    HFR_d = np.array(hfr_out) if HFR is not None else None
    cond_d = {name: np.array(vals) for name, vals in cond_out.items()} if conditions else None

    return j_d, V_d, HFR_d, cond_d, n_merged


def _dedup_voltage_grid(j, V, HFR, conditions, v_step):
    """Group by voltage proximity — no grid enforced."""
    tol_V = v_step * 0.45  # half the smallest expected step

    # Sort by V descending (OCV first)
    order = np.argsort(V)[::-1]
    V_s, j_s = V[order], j[order]
    HFR_s = HFR[order] if HFR is not None else None
    cond_s = {name: vals[order] for name, vals in conditions.items()} if conditions else None

    # Walk through sorted V and group nearby points
    j_out, v_out, hfr_out = [], [], []
    cond_out = {name: [] for name in (conditions or {})}

    grp_idx = [0]
    for i in range(1, len(V_s)):
        if abs(V_s[i] - np.mean(V_s[grp_idx])) <= tol_V:
            grp_idx.append(i)
        else:
            # Flush group
            j_out.append(np.mean(j_s[grp_idx]))
            v_out.append(np.mean(V_s[grp_idx]))
            if HFR_s is not None:
                hfr_out.append(np.mean(HFR_s[grp_idx]))
            for name in cond_out:
                cond_out[name].append(np.mean(cond_s[name][grp_idx]))
            grp_idx = [i]

    # Flush final group
    j_out.append(np.mean(j_s[grp_idx]))
    v_out.append(np.mean(V_s[grp_idx]))
    if HFR_s is not None:
        hfr_out.append(np.mean(HFR_s[grp_idx]))
    for name in cond_out:
        cond_out[name].append(np.mean(cond_s[name][grp_idx]))

    n_merged = len(j) - len(j_out)

    # Re-sort by ascending j
    j_d = np.array(j_out)
    V_d = np.array(v_out)
    order_j = np.argsort(j_d)
    j_d, V_d = j_d[order_j], V_d[order_j]
    HFR_d = np.array(hfr_out)[order_j] if HFR is not None else None
    cond_d = ({name: np.array(vals)[order_j] for name, vals in cond_out.items()}
              if conditions else None)

    return j_d, V_d, HFR_d, cond_d, n_merged


def enforce_v_monotonicity(j, V, HFR=None, conditions=None):
    """
    Remove points that violate V-monotonicity (V must decrease as j increases).
    """
    if len(j) < 3:
        return j, V, HFR, conditions

    order = np.argsort(j)
    j, V = j[order].copy(), V[order].copy()
    if HFR is not None:
        HFR = HFR[order].copy()
    if conditions:
        conditions = {k: v[order].copy() for k, v in conditions.items()}

    while True:
        violations = []
        for i in range(1, len(V)):
            if V[i] > V[i - 1]:
                violations.append(i)
        if not violations:
            break

        worst = max(violations, key=lambda i: V[i] - V[i - 1])
        if worst < len(V) - 1 and V[worst - 1] >= V[worst + 1]:
            remove = worst
        elif worst > 1 and V[worst - 2] >= V[worst]:
            remove = worst - 1
        else:
            remove = worst

        keep = np.ones(len(j), dtype=bool)
        keep[remove] = False
        j, V = j[keep], V[keep]
        if HFR is not None:
            HFR = HFR[keep]
        if conditions:
            conditions = {k: v[keep] for k, v in conditions.items()}

    return j, V, HFR, conditions


def _curvature_std(V):
    """Compute std of second finite difference (curvature variability)."""
    if len(V) < 3:
        return 0.0
    return float(np.std(np.diff(V, 2)))


def smooth_curve_selection(j, V, HFR=None, conditions=None,
                           max_remove_frac=0.15, improvement_threshold=0.05):
    """
    Iteratively remove points that degrade curve smoothness.

    After monotonicity enforcement, the curve may still have local kinks
    from HFR artifacts or boundary noise.  This function identifies points
    whose removal most reduces curvature variability (std of d²V) and
    removes them if the improvement exceeds a threshold.

    The algorithm:
      1. Sort by j ascending.
      2. Compute curvature std (baseline).
      3. For each interior point, compute the curvature std that would
         result from removing it.
      4. Remove the point whose removal yields the largest fractional
         improvement, if that improvement exceeds improvement_threshold.
      5. Repeat until no point's removal helps enough, or max_remove_frac
         of the original points have been removed.

    Parameters
    ----------
    j, V : array
        Sorted by ascending j, monotonically decreasing V.
    HFR : array or None
    conditions : dict of arrays or None
    max_remove_frac : float
        Maximum fraction of original points to remove (default 15%).
    improvement_threshold : float
        Minimum fractional reduction in curvature std to justify removal
        (default 5%).

    Returns
    -------
    j, V, HFR, conditions — smoothed arrays.
    """
    if len(j) < 5:
        return j, V, HFR, conditions

    order = np.argsort(j)
    j, V = j[order].copy(), V[order].copy()
    if HFR is not None:
        HFR = HFR[order].copy()
    if conditions:
        conditions = {k: v[order].copy() for k, v in conditions.items()}

    n_original = len(j)
    max_removals = max(1, int(n_original * max_remove_frac))
    n_removed = 0

    while len(j) > 5 and n_removed < max_removals:
        baseline = _curvature_std(V)
        if baseline < 1e-10:
            break

        # Compute curvature at each interior point
        d2V = np.diff(V, 2)
        n_d2 = len(d2V)

        # Score each interior point by local anomaly:
        # Compare its curvature to the median of its ±2 neighbors
        best_idx = -1
        best_new_std = baseline

        for i in range(1, len(V) - 1):
            k = i - 1  # index into d2V
            if k < 0 or k >= n_d2:
                continue

            # Local neighborhood: ±2 points in d2V
            lo = max(0, k - 2)
            hi = min(n_d2, k + 3)
            neighbors = np.concatenate([d2V[lo:k], d2V[k+1:hi]])
            if len(neighbors) == 0:
                continue

            local_median = np.median(np.abs(neighbors))
            local_anomaly = abs(d2V[k])

            # Only consider if curvature is > 3× local neighborhood
            if local_median > 1e-10 and local_anomaly < 3.0 * local_median:
                continue

            V_test = np.delete(V, i)
            new_std = _curvature_std(V_test)
            if new_std < best_new_std:
                best_new_std = new_std
                best_idx = i

        # Check if the improvement is significant
        if best_idx < 0:
            break
        improvement = (baseline - best_new_std) / baseline
        if improvement < improvement_threshold:
            break

        # Remove the point
        keep = np.ones(len(j), dtype=bool)
        keep[best_idx] = False
        j, V = j[keep], V[keep]
        if HFR is not None:
            HFR = HFR[keep]
        if conditions:
            conditions = {k: v[keep] for k, v in conditions.items()}
        n_removed += 1

    # Verify monotonicity still holds after smoothing
    j, V, HFR, conditions = enforce_v_monotonicity(j, V, HFR, conditions)

    return j, V, HFR, conditions
    """Fallback: group by current density proximity."""
    j_out, v_out, hfr_out = [], [], []
    cond_out = {name: [] for name in (conditions or {})}

    group_j = [j[0]]
    group_v = [V[0]]
    group_hfr = [HFR[0]] if HFR is not None else []
    group_cond = {name: [conditions[name][0]] for name in (conditions or {})}

    for i in range(1, len(j)):
        if abs(j[i] - np.mean(group_j)) <= tol:
            group_j.append(j[i])
            group_v.append(V[i])
            if HFR is not None:
                group_hfr.append(HFR[i])
            for name in group_cond:
                group_cond[name].append(conditions[name][i])
        else:
            j_out.append(np.mean(group_j))
            v_out.append(np.mean(group_v))
            if HFR is not None:
                hfr_out.append(np.mean(group_hfr))
            for name in cond_out:
                cond_out[name].append(np.mean(group_cond[name]))
            group_j = [j[i]]
            group_v = [V[i]]
            group_hfr = [HFR[i]] if HFR is not None else []
            group_cond = {name: [conditions[name][i]] for name in (conditions or {})}

    j_out.append(np.mean(group_j))
    v_out.append(np.mean(group_v))
    if HFR is not None:
        hfr_out.append(np.mean(group_hfr))
    for name in cond_out:
        cond_out[name].append(np.mean(group_cond[name]))

    n_merged = len(j) - len(j_out)
    j_d = np.array(j_out)
    V_d = np.array(v_out)
    HFR_d = np.array(hfr_out) if HFR is not None else None
    cond_d = {name: np.array(vals) for name, vals in cond_out.items()} if conditions else None

    return j_d, V_d, HFR_d, cond_d, n_merged


# ═══════════════════════════════════════════════════════════════════════
#  HFR Outlier Filtering
# ═══════════════════════════════════════════════════════════════════════

def filter_hfr_outliers(HFR, geo_area=5.0, iqr_factor=2.0, max_asr=0.5):
    """
    Detect and replace outlier HFR values with linear interpolation.

    Uses two criteria (either triggers replacement):
      1. IQR method: value is beyond Q1 - iqr_factor×IQR or Q3 + iqr_factor×IQR
      2. Absolute ceiling: HFR × geo_area > max_asr (Ω·cm²)

    Parameters
    ----------
    HFR : array or None
        HFR values in Ohm (per-cell).
    geo_area : float
        Geometric area in cm² (for ASR ceiling check).
    iqr_factor : float
        IQR multiplier for outlier detection (default 2.0).
    max_asr : float
        Maximum plausible ASR in Ω·cm² (default 0.5). Points above this
        are flagged as outliers regardless of IQR.

    Returns
    -------
    HFR_clean : array or None
        Cleaned HFR with outliers replaced by interpolation.
    n_outliers : int
        Number of outliers detected and replaced.
    outlier_mask : array of bool
        True where outliers were detected.
    """
    if HFR is None or len(HFR) < 3:
        return HFR, 0, np.zeros(len(HFR) if HFR is not None else 0, dtype=bool)

    HFR_asr = HFR * geo_area

    # ── IQR-based detection ──
    Q1 = np.percentile(HFR_asr, 25)
    Q3 = np.percentile(HFR_asr, 75)
    IQR = Q3 - Q1
    iqr_low = Q1 - iqr_factor * IQR
    iqr_high = Q3 + iqr_factor * IQR

    # ── Combined outlier mask ──
    outlier = (HFR_asr < iqr_low) | (HFR_asr > iqr_high) | (HFR_asr > max_asr)

    # Also flag negative HFR (physically impossible for ohmic resistance)
    outlier |= (HFR < 0)

    n_outliers = int(outlier.sum())
    if n_outliers == 0:
        return HFR.copy(), 0, outlier

    # ── Replace outliers with linear interpolation ──
    HFR_clean = HFR.copy()
    good_idx = np.where(~outlier)[0]

    if len(good_idx) < 2:
        # Almost all points are outliers — use median of good points
        if len(good_idx) == 1:
            HFR_clean[outlier] = HFR[good_idx[0]]
        else:
            HFR_clean[outlier] = np.median(HFR)
        return HFR_clean, n_outliers, outlier

    # Interpolate from good points
    bad_idx = np.where(outlier)[0]
    HFR_clean[bad_idx] = np.interp(bad_idx, good_idx, HFR[good_idx])

    return HFR_clean, n_outliers, outlier


# ═══════════════════════════════════════════════════════════════════════
#  Analysis Functions
# ═══════════════════════════════════════════════════════════════════════

def compute_power_density(j, V):
    """Compute power density (W/cm²) = j × V."""
    return j * V


def compute_ir_free_voltage(V, j, HFR):
    """
    Compute iR-free voltage: V_iRfree = V + j × HFR × A.

    For HFR in Ohm (not area-specific): V_iRfree = V + I × R = V + (j × A) × R
    But since we want per-cm² and j is already per-cm²:
    V_iRfree = V + j × (HFR × geo_area)... no.

    HFR is in Ohm (total cell). ohmic drop = I × R = (j × A) × R.
    Voltage drop per cell = I × R. Since V is cell voltage:
    V_iRfree = V + j × R_ASR  where R_ASR = HFR × A  (Ohm·cm²)

    But we don't know A here. We pass HFR as Ohm and j as A/cm².
    So: V_iRfree = V + j × HFR_ASR

    Convention: if HFR is in Ohm, user must convert. We assume HFR
    passed here is already area-specific (Ohm·cm²).
    """
    return V + j * HFR


def extract_tafel(j, V, j_min=0.01, j_max=0.1):
    """
    Extract Tafel slope from the kinetic region of the polarization curve.

    Fits: V = a - b × log10(j)  in the range [j_min, j_max] A/cm².

    Parameters
    ----------
    j, V : array
        Current density (A/cm²) and voltage (V).
    j_min, j_max : float
        Current density range for Tafel fit (A/cm²).

    Returns
    -------
    dict with Tafel slope, intercept, exchange current density, and fit data.
    """
    mask = (j >= j_min) & (j <= j_max) & (j > 0)
    if mask.sum() < 3:
        return None

    log_j = np.log10(j[mask])
    V_fit = V[mask]

    # V = a + b × log10(j)  →  b is negative (Tafel slope, mV/dec)
    coeffs = np.polyfit(log_j, V_fit, 1)
    b_Vdec = coeffs[0]       # V/decade
    a = coeffs[1]             # V (intercept)

    # Tafel slope in mV/dec (magnitude)
    tafel_slope = abs(b_Vdec) * 1000  # mV/dec

    # Exchange current density: at V = E_rev (~1.23 V for H2/O2)
    # log10(j0) = (E_rev - a) / b
    E_rev = 1.23
    if b_Vdec != 0:
        log_j0 = (E_rev - a) / b_Vdec
        j0 = 10 ** log_j0
    else:
        j0 = np.nan

    # Fit line for plotting
    log_j_line = np.linspace(log_j.min() - 0.5, log_j.max() + 0.5, 100)
    V_line = a + b_Vdec * log_j_line

    return {
        'tafel_slope_mVdec': tafel_slope,
        'b_Vdec': b_Vdec,
        'intercept_V': a,
        'j0_A_cm2': j0,
        'j_min': j_min,
        'j_max': j_max,
        'R_squared': 1.0 - np.sum((V_fit - (a + b_Vdec * log_j))**2) / np.sum((V_fit - V_fit.mean())**2),
        '_log_j_fit': log_j,
        '_V_fit': V_fit,
        '_log_j_line': log_j_line,
        '_V_line': V_line,
    }


def analyze_polcurve(j, V, HFR=None, geo_area=5.0,
                     tafel_j_min=0.01, tafel_j_max=0.10):
    """
    Full polarization curve analysis.

    Parameters
    ----------
    j : array
        Current density in A/cm².
    V : array
        Cell voltage in V.
    HFR : array or None
        HFR in Ohm (per-cell). Will be converted to ASR (Ohm·cm²).
    geo_area : float
        Geometric area in cm².
    tafel_j_min, tafel_j_max : float
        Current density range for Tafel fit (A/cm²).

    Returns
    -------
    dict with all analysis results and plotting data.
    """
    P = compute_power_density(j, V)

    results = {
        'j': j,
        'V': V,
        'P': P,
        'geo_area': geo_area,
        'V_at_1Acm2': np.interp(1.0, j, V) if j.max() >= 1.0 else None,
        'peak_power_W_cm2': P.max(),
        'j_at_peak_power': j[np.argmax(P)],
        'OCV': V[0] if j[0] < 0.01 else np.interp(0, j, V),
    }

    # HFR and iR-free voltage
    if HFR is not None:
        HFR_ASR = HFR * geo_area   # Ohm·cm²
        results['HFR'] = HFR
        results['HFR_ASR'] = HFR_ASR
        results['HFR_mean'] = np.mean(HFR_ASR)

        V_irfree = compute_ir_free_voltage(V, j, HFR_ASR)
        results['V_irfree'] = V_irfree
        results['P_irfree'] = j * V_irfree

        # Tafel on iR-free curve
        tafel = extract_tafel(j, V_irfree, j_min=tafel_j_min, j_max=tafel_j_max)
    else:
        results['HFR'] = None
        results['HFR_ASR'] = None
        results['V_irfree'] = None

        # Tafel on raw curve
        tafel = extract_tafel(j, V, j_min=tafel_j_min, j_max=tafel_j_max)

    results['tafel'] = tafel

    # Loss breakdown at a reference current density
    j_ref_values = [0.5, 1.0, 1.5]
    for j_ref in j_ref_values:
        if j.max() >= j_ref:
            V_at_j = np.interp(j_ref, j, V)
            E_rev = 1.23

            # Ohmic loss
            if HFR is not None:
                hfr_at_j = np.interp(j_ref, j, HFR_ASR)
                eta_ohmic = j_ref * hfr_at_j
            else:
                eta_ohmic = None

            # Kinetic loss (from Tafel extrapolation)
            if tafel is not None:
                V_tafel_at_j = tafel['intercept_V'] + tafel['b_Vdec'] * np.log10(j_ref)
                eta_kinetic = E_rev - V_tafel_at_j
            else:
                eta_kinetic = None

            # Mass transport loss (remainder)
            if eta_ohmic is not None and eta_kinetic is not None:
                eta_mt = E_rev - V_at_j - eta_ohmic - eta_kinetic
                eta_mt = max(eta_mt, 0)  # prevent negative
            else:
                eta_mt = None

            results[f'V_at_{j_ref}'] = V_at_j
            results[f'eta_ohmic_{j_ref}'] = eta_ohmic
            results[f'eta_kinetic_{j_ref}'] = eta_kinetic
            results[f'eta_mt_{j_ref}'] = eta_mt

    return results


# ═══════════════════════════════════════════════════════════════════════
#  Plotting
# ═══════════════════════════════════════════════════════════════════════

def plot_polcurve(results, save_path=None):
    """
    Multi-panel polarization curve analysis plot.
    """
    j = results['j']
    V = results['V']
    P = results['P']
    has_hfr = results['HFR'] is not None
    has_tafel = results['tafel'] is not None

    n_panels = 2 + int(has_hfr) + int(has_tafel)
    fig, axes = plt.subplots(1, n_panels, figsize=(5.5 * n_panels, 5))
    if n_panels == 1:
        axes = [axes]

    panel = 0

    # ── Panel: Polarization curve + power density ──
    ax = axes[panel]
    ax.plot(j, V, 'o-', color='steelblue', ms=4, lw=1.5, label='Cell voltage')
    if results['V_irfree'] is not None:
        ax.plot(j, results['V_irfree'], 's--', color='green', ms=3, lw=1.2,
                alpha=0.7, label='iR-free voltage')
    ax.set_xlabel('Current density (A/cm²)')
    ax.set_ylabel('Voltage (V)')
    ax.set_ylim(bottom=0)
    ax.legend(loc='upper right', fontsize=8)
    ax.grid(True, alpha=0.3)

    ax2 = ax.twinx()
    ax2.plot(j, P * 1000, '^-', color='firebrick', ms=3, lw=1.2, alpha=0.7,
             label='Power density')
    ax2.set_ylabel('Power density (mW/cm²)', color='firebrick')
    ax2.tick_params(axis='y', labelcolor='firebrick')
    ax.set_title('Polarization Curve')

    # Annotation
    txt_lines = [f'OCV = {results["OCV"]:.3f} V',
                 f'Peak P = {results["peak_power_W_cm2"]*1000:.0f} mW/cm²']
    if results['V_at_1Acm2'] is not None:
        txt_lines.append(f'V @ 1 A/cm² = {results["V_at_1Acm2"]:.3f} V')
    ax.text(0.03, 0.03, '\n'.join(txt_lines), transform=ax.transAxes,
            fontsize=8, va='bottom',
            bbox=dict(boxstyle='round,pad=0.4', fc='lightyellow', alpha=0.9))
    panel += 1

    # ── Panel: HFR ──
    if has_hfr:
        ax = axes[panel]
        HFR_ASR = results['HFR_ASR']
        ax.plot(j, HFR_ASR * 1000, 'o-', color='darkorange', ms=4, lw=1.5)
        ax.set_xlabel('Current density (A/cm²)')
        ax.set_ylabel('HFR (mΩ·cm²)')
        ax.set_title(f'HFR — mean = {results["HFR_mean"]*1000:.1f} mΩ·cm²')
        ax.grid(True, alpha=0.3)
        panel += 1

    # ── Panel: Tafel ──
    if has_tafel:
        ax = axes[panel]
        t = results['tafel']
        ax.plot(10**t['_log_j_fit'], t['_V_fit'], 'o', color='steelblue', ms=5,
                label='Data')
        ax.plot(10**t['_log_j_line'], t['_V_line'], '-', color='firebrick', lw=1.5,
                label=f'Fit: {t["tafel_slope_mVdec"]:.0f} mV/dec')
        ax.set_xscale('log')
        ax.set_xlabel('Current density (A/cm²)')
        ylabel = 'iR-free voltage (V)' if results['V_irfree'] is not None else 'Voltage (V)'
        ax.set_ylabel(ylabel)
        ax.set_title(f'Tafel Analysis (R² = {t["R_squared"]:.4f})')
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3, which='both')

        txt = f'b = {t["tafel_slope_mVdec"]:.0f} mV/dec\nj₀ = {t["j0_A_cm2"]:.2e} A/cm²'
        ax.text(0.97, 0.03, txt, transform=ax.transAxes,
                ha='right', va='bottom', fontsize=9,
                bbox=dict(boxstyle='round,pad=0.4', fc='lightyellow', alpha=0.9))
        panel += 1

    # ── Panel: All cycles overlay ──
    ax = axes[panel]
    cycles_raw = results.get('_cycles_raw', [])
    if cycles_raw:
        n_cyc = max(c['cycle_number'] for c in cycles_raw)
        cmap = plt.cm.viridis
        colors = [cmap(i / max(n_cyc - 1, 1)) for i in range(n_cyc)]

        for c in cycles_raw:
            cidx = c['cycle_number'] - 1
            ls = '-' if c['direction'] == 'up' else '--'
            marker = 'o' if c['direction'] == 'up' else 's'
            lbl = f'Cycle {c["cycle_number"]} ({c["direction"]})'
            ax.plot(c['j'], c['V'], marker + ls, color=colors[cidx], ms=3, lw=1.2,
                    alpha=0.8, label=lbl)

        # Highlight representative on top
        ax.plot(j, V, 'o-', color='firebrick', ms=5, lw=2.0, alpha=0.9,
                label='Representative', zorder=10)

        ax.set_xlabel('Current density (A/cm²)')
        ax.set_ylabel('Voltage (V)')
        ax.set_title('All Cycles (dwell-extracted)')
        ax.set_ylim(bottom=0)
        ax.legend(fontsize=7, loc='best')
        ax.grid(True, alpha=0.3)
    else:
        # No cycle data — just plot the representative
        ax.plot(j, V, 'o-', color='steelblue', ms=4, lw=1.5)
        ax.set_xlabel('Current density (A/cm²)')
        ax.set_ylabel('Voltage (V)')
        ax.set_title('Polarization Curve')
        ax.set_ylim(bottom=0)
        ax.grid(True, alpha=0.3)

    from scripts.helpers.conditions import get_condition_label
    cond_label = get_condition_label(
        label=results.get('label', ''),
        conditions=results.get('conditions'))
    title = 'Polarization Curve Analysis'
    if cond_label:
        title += f'\n{cond_label}'
    fig.suptitle(title, fontsize=13, fontweight='bold')

    fig.tight_layout()

    if save_path:
        save_with_sidecar(fig, save_path, dpi=200, bbox_inches='tight')
        print(f'  Saved: {save_path}')

        # Write sidecar JSON for comparison feature
        cycles_data = []
        cycles_raw = results.get('_cycles_raw', [])
        if cycles_raw:
            for c in cycles_raw:
                cycles_data.append({
                    'cycle_num': int(c.get('cycle_number', 0)),
                    'mode': c.get('direction', ''),
                    'j': list(c.get('j', [])),
                    'V': list(c.get('V', [])),
                    'HFR_ASR': list(c.get('HFR', [])) if c.get('HFR') is not None else None,
                })
        rep_cycle = {
            'cycle_num': len(cycles_data) + 1 if cycles_data else 1,
            'mode': 'representative',
            'j': list(results['j']),
            'V': list(results['V']),
            'V_irfree': (list(results['V_irfree'])
                         if results.get('V_irfree') is not None else None),
            'HFR_ASR': (list(results['HFR_ASR'])
                        if results.get('HFR_ASR') is not None else None),
        }
        cycles_data.append(rep_cycle)

    return fig


def plot_polcurve_overlay(all_results, save_path=None):
    """Overlay multiple polarization curves."""
    from scripts.helpers.conditions import get_condition_label

    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))
    n = len(all_results)
    cmap = plt.cm.viridis
    colors = [cmap(i / max(n - 1, 1)) for i in range(n)]

    # ── Left: V vs j ──
    ax = axes[0]
    for i, r in enumerate(all_results):
        lbl = r.get('label', f'File {i+1}')
        cond = get_condition_label(label=lbl, conditions=r.get('conditions'), compact=True)
        legend_lbl = f'{lbl}\n  {cond}' if cond else lbl
        ax.plot(r['j'], r['V'], 'o-', color=colors[i], ms=3, lw=1.2, label=legend_lbl)
    ax.set_xlabel('Current density (A/cm²)')
    ax.set_ylabel('Voltage (V)')
    ax.set_title('Polarization Curves')
    ax.set_ylim(bottom=0)
    ax.legend(fontsize=7, loc='best')
    ax.grid(True, alpha=0.3)

    # ── Right: Power density ──
    ax = axes[1]
    for i, r in enumerate(all_results):
        lbl = r.get('label', f'File {i+1}')
        cond = get_condition_label(label=lbl, conditions=r.get('conditions'), compact=True)
        legend_lbl = f'{lbl}\n  {cond}' if cond else lbl
        ax.plot(r['j'], r['P'] * 1000, '^-', color=colors[i], ms=3, lw=1.2, label=legend_lbl)
    ax.set_xlabel('Current density (A/cm²)')
    ax.set_ylabel('Power density (mW/cm²)')
    ax.set_title('Power Density')
    ax.legend(fontsize=7, loc='best')
    ax.grid(True, alpha=0.3)

    fig.suptitle('Polarization Curve Overlay', fontsize=13, fontweight='bold')
    fig.tight_layout()

    if save_path:
        save_with_sidecar(fig, save_path, dpi=200, bbox_inches='tight')
        print(f'  Saved: {save_path}')
    return fig


# ═══════════════════════════════════════════════════════════════════════
#  Print Results
# ═══════════════════════════════════════════════════════════════════════

def print_results(r):
    """Pretty-print polarization curve results."""
    print(f'\n{"═" * 55}')
    print(f'  Geo. area:        {r["geo_area"]:.2f} cm²')

    # Operating conditions
    cond = r.get('conditions')
    if cond:
        print(f'  {"─" * 51}')
        for name, vals in cond.items():
            print(f'  {name + ":":20s} {np.mean(vals):.1f}')
        print(f'  {"─" * 51}')

    print(f'  OCV:              {r["OCV"]:.4f} V')
    if r['V_at_1Acm2'] is not None:
        print(f'  V @ 1 A/cm²:      {r["V_at_1Acm2"]:.4f} V')
    print(f'  Peak power:       {r["peak_power_W_cm2"]*1000:.1f} mW/cm²'
          f' @ {r["j_at_peak_power"]:.2f} A/cm²')

    if r.get('HFR_mean') is not None:
        print(f'  Mean HFR (ASR):   {r["HFR_mean"]*1000:.2f} mΩ·cm²')

    if r['tafel'] is not None:
        t = r['tafel']
        print(f'  {"─" * 51}')
        print(f'  Tafel slope:      {t["tafel_slope_mVdec"]:.1f} mV/dec '
              f'(R² = {t["R_squared"]:.4f})')
        print(f'  j₀:               {t["j0_A_cm2"]:.2e} A/cm²')
        print(f'  Fit range:        {t["j_min"]:.3f} – {t["j_max"]:.3f} A/cm²')

    # Loss breakdown
    has_losses = False
    for j_ref in [0.5, 1.0, 1.5]:
        if f'eta_kinetic_{j_ref}' in r and r[f'eta_kinetic_{j_ref}'] is not None:
            if not has_losses:
                print(f'  {"─" * 51}')
                print(f'  {"Loss @ j (A/cm²)":20s} {"Kinetic":>9s} {"Ohmic":>9s} {"M.T.":>9s} {"Total":>9s}')
                print(f'  {"─" * 51}')
                has_losses = True
            ek = r[f'eta_kinetic_{j_ref}']
            eo = r[f'eta_ohmic_{j_ref}'] or 0
            em = r[f'eta_mt_{j_ref}'] or 0
            et = ek + eo + em
            print(f'  {"j = " + str(j_ref):20s} {ek*1000:>8.0f}  {eo*1000:>8.0f}  {em*1000:>8.0f}  {et*1000:>8.0f} mV')

    print(f'{"═" * 55}\n')


# ═══════════════════════════════════════════════════════════════════════
#  Synthetic Data
# ═══════════════════════════════════════════════════════════════════════

def generate_synthetic_polcurve(geo_area=5.0, noise=0.002):
    """
    Generate a synthetic PEM fuel cell polarization curve with HFR.

    Returns j (A/cm²), V (V), HFR (Ohm).
    """
    j = np.concatenate([
        np.array([0.0]),
        np.linspace(0.01, 0.05, 5),
        np.linspace(0.1, 2.0, 40),
    ])

    E_rev = 1.23
    # Kinetic loss (Tafel)
    alpha_c = 1.0
    j0 = 1e-3
    b = 0.065  # V/dec (Tafel slope)

    with np.errstate(divide='ignore', invalid='ignore'):
        eta_act = np.where(j > 0, b * np.log10(j / j0), 0)

    # Ohmic loss
    R_ohm = 0.050  # Ohm·cm²
    eta_ohm = j * R_ohm

    # Mass transport loss
    j_lim = 2.5
    with np.errstate(divide='ignore', invalid='ignore'):
        eta_mt = np.where(j > 0,
                          -0.03 * np.log(1 - j / j_lim),
                          0)

    V = E_rev - eta_act - eta_ohm - eta_mt

    # HFR with slight current-dependence
    HFR = (R_ohm / geo_area) * (1 + 0.05 * j)  # Ohm

    # Add noise
    V += np.random.normal(0, noise, len(V))
    HFR += np.random.normal(0, noise * 0.01, len(HFR))

    V = np.clip(V, 0, None)

    return j, V, HFR


# ═══════════════════════════════════════════════════════════════════════
#  Excel Output
# ═══════════════════════════════════════════════════════════════════════

def save_single_excel(results, filepath):
    """
    Save single-file analysis to Excel with summary + representative data + cycle data.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill('solid', fgColor='D9E1F2')
    label_fill = PatternFill('solid', fgColor='E2EFDA')

    # ── Sheet 1: Summary + Representative Curve Data ──
    ws = wb.active
    ws.title = 'Pol Curve Summary'
    summary = [
        ('OCV (V)', results['OCV']),
        ('Peak Power (mW/cm²)', results['peak_power_W_cm2'] * 1000),
        ('j @ Peak Power (A/cm²)', results['j_at_peak_power']),
    ]
    if results['V_at_1Acm2'] is not None:
        summary.append(('V @ 1 A/cm² (V)', results['V_at_1Acm2']))
    if results.get('HFR_mean') is not None:
        summary.append(('Mean HFR (mΩ·cm²)', results['HFR_mean'] * 1000))
    if results['tafel'] is not None:
        t = results['tafel']
        summary.append(('Tafel Slope (mV/dec)', t['tafel_slope_mVdec']))
        summary.append(('j₀ (A/cm²)', t['j0_A_cm2']))
        summary.append(('Tafel R²', t['R_squared']))
    cond = results.get('conditions')
    if cond:
        for name, vals in cond.items():
            summary.append((name, np.mean(vals)))

    for i, (label, val) in enumerate(summary, 1):
        ws.cell(row=i, column=1, value=label).font = hdr_font
        ws.cell(row=i, column=2, value=val)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    # Representative curve data below summary
    data_start = len(summary) + 3
    ws.cell(row=data_start - 1, column=1, value='Representative Pol Curve').font = hdr_font

    rep_headers = ['j (A/cm²)', 'j (mA/cm²)', 'V (V)']
    has_hfr = results['HFR'] is not None
    if has_hfr:
        rep_headers.append('HFR (mΩ·cm²)')
    if results['V_irfree'] is not None:
        rep_headers.append('V_iR-free (V)')
    rep_headers.append('Power (mW/cm²)')

    for c, h in enumerate(rep_headers, 1):
        cell = ws.cell(row=data_start, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(c)].width = max(
            ws.column_dimensions[get_column_letter(c)].width or 0, 16)

    j, V, P = results['j'], results['V'], results['P']
    order = np.argsort(V)
    for row_idx, i in enumerate(order):
        row = data_start + 1 + row_idx
        col = 1
        ws.cell(row=row, column=col, value=round(float(j[i]), 6)); col += 1
        ws.cell(row=row, column=col, value=round(float(j[i]) * 1000, 3)); col += 1
        ws.cell(row=row, column=col, value=round(float(V[i]), 6)); col += 1
        if has_hfr:
            ws.cell(row=row, column=col,
                    value=round(float(results['HFR_ASR'][i]) * 1000, 2)); col += 1
        if results['V_irfree'] is not None:
            ws.cell(row=row, column=col,
                    value=round(float(results['V_irfree'][i]), 6)); col += 1
        ws.cell(row=row, column=col, value=round(float(P[i]) * 1000, 2))

    # ── Sheet 2: All Cycles Data ──
    cycles = results.get('_cycles_raw', [])
    if cycles:
        ws2 = wb.create_sheet('Pol Curve Data')
        col = 1
        for ci, cyc in enumerate(cycles):
            direction = cyc.get('direction', '?')
            cycle_num = cyc.get('cycle_number', ci + 1)
            cyc_label = f'Cycle {cycle_num} ({direction})'
            cyc_j = cyc['j']
            cyc_V = cyc['V']
            cyc_hfr = cyc.get('HFR')
            n_cols = 3 + (1 if cyc_hfr is not None else 0)  # j_A, j_mA, V, [HFR]
            col_end = col + n_cols - 1

            # Label row
            cell = ws2.cell(row=1, column=col, value=cyc_label)
            cell.font = hdr_font
            cell.fill = label_fill
            if col_end > col:
                ws2.merge_cells(start_row=1, start_column=col,
                                end_row=1, end_column=col_end)

            # Sub-headers
            hdrs = ['j (A/cm²)', 'j (mA/cm²)', 'V (V)']
            if cyc_hfr is not None:
                hdrs.append('HFR (mΩ·cm²)')
            for hi, h in enumerate(hdrs):
                cell = ws2.cell(row=2, column=col + hi, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
                ws2.column_dimensions[get_column_letter(col + hi)].width = 16

            # Data rows
            for ri in range(len(cyc_j)):
                cc = col
                ws2.cell(row=ri + 3, column=cc, value=round(float(cyc_j[ri]), 6)); cc += 1
                ws2.cell(row=ri + 3, column=cc, value=round(float(cyc_j[ri]) * 1000, 3)); cc += 1
                ws2.cell(row=ri + 3, column=cc, value=round(float(cyc_V[ri]), 6)); cc += 1
                if cyc_hfr is not None:
                    hfr_val = cyc_hfr[ri] if ri < len(cyc_hfr) else 0
                    ws2.cell(row=ri + 3, column=cc,
                             value=round(float(hfr_val) * results.get('geo_area', 5.0) * 1000, 2))

            col = col_end + 2
    else:
        # Fallback: single sheet with representative data
        ws2 = wb.create_sheet('Pol Curve Data')
        for c, h in enumerate(rep_headers, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            ws2.column_dimensions[cell.column_letter].width = 16
        for row_idx, i in enumerate(order):
            col = 1
            ws2.cell(row=row_idx + 2, column=col, value=round(float(j[i]), 6)); col += 1
            ws2.cell(row=row_idx + 2, column=col, value=round(float(j[i]) * 1000, 3)); col += 1
            ws2.cell(row=row_idx + 2, column=col, value=round(float(V[i]), 6)); col += 1
            if has_hfr:
                ws2.cell(row=row_idx + 2, column=col,
                         value=round(float(results['HFR_ASR'][i]) * 1000, 2)); col += 1
            if results['V_irfree'] is not None:
                ws2.cell(row=row_idx + 2, column=col,
                         value=round(float(results['V_irfree'][i]), 6)); col += 1
            ws2.cell(row=row_idx + 2, column=col, value=round(float(P[i]) * 1000, 2))

    wb.save(filepath)
    print(f'  Excel: {filepath}')


def save_batch_excel(all_results, summary_rows, filepath):
    """
    Save batch analysis to Excel: summary + representative data, cycle data per file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill('solid', fgColor='D9E1F2')
    label_fill = PatternFill('solid', fgColor='E2EFDA')

    # ── Sheet 1: Pol Curve Summary (metrics + representative data) ──
    ws = wb.active
    ws.title = 'Pol Curve Summary'
    if summary_rows:
        all_cols = list(summary_rows[0].keys())
        for row in summary_rows[1:]:
            for k in row:
                if k not in all_cols:
                    all_cols.append(k)

        for c, h in enumerate(all_cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
        for r_idx, row in enumerate(summary_rows, 2):
            for c, col_name in enumerate(all_cols, 1):
                val = row.get(col_name)
                if val is not None:
                    ws.cell(row=r_idx, column=c,
                            value=round(val, 6) if isinstance(val, float) else val)
        for c in range(1, len(all_cols) + 1):
            ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 18

    # Representative curve data below the summary table
    data_start = len(summary_rows) + 4
    ws.cell(row=data_start - 1, column=1, value='Representative Pol Curve Data').font = hdr_font

    col = 1
    for r in all_results:
        label = r.get('label', 'data')
        j, V, P = r['j'], r['V'], r['P']
        has_h = r.get('HFR') is not None
        has_irfree = r.get('V_irfree') is not None
        n_cols = 4 + int(has_h) + int(has_irfree)  # j_A, j_mA, V, [HFR], [iRfree], Power
        col_end = col + n_cols - 1

        # Label row
        cell = ws.cell(row=data_start, column=col, value=label)
        cell.font = hdr_font
        cell.fill = label_fill
        if col_end > col:
            ws.merge_cells(start_row=data_start, start_column=col,
                           end_row=data_start, end_column=col_end)

        # Sub-headers
        hdrs = ['j (A/cm²)', 'j (mA/cm²)', 'V (V)']
        if has_h: hdrs.append('HFR (mΩ·cm²)')
        if has_irfree: hdrs.append('V_iR-free (V)')
        hdrs.append('Power (mW/cm²)')
        for ci, h in enumerate(hdrs):
            cell = ws.cell(row=data_start + 1, column=col + ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col + ci)].width = 16

        order = np.argsort(V)
        for row_idx, ri in enumerate(order):
            row = data_start + 2 + row_idx
            cc = col
            ws.cell(row=row, column=cc, value=round(float(j[ri]), 6)); cc += 1
            ws.cell(row=row, column=cc, value=round(float(j[ri]) * 1000, 3)); cc += 1
            ws.cell(row=row, column=cc, value=round(float(V[ri]), 6)); cc += 1
            if has_h:
                ws.cell(row=row, column=cc,
                        value=round(float(r['HFR_ASR'][ri]) * 1000, 2)); cc += 1
            if has_irfree:
                ws.cell(row=row, column=cc,
                        value=round(float(r['V_irfree'][ri]), 6)); cc += 1
            ws.cell(row=row, column=cc, value=round(float(P[ri]) * 1000, 2))

        col = col_end + 2

    # ── Sheet 2: Pol Curve Data (all cycles per file) ──
    ws2 = wb.create_sheet('Pol Curve Data')
    col = 1
    for r in all_results:
        label = r.get('label', 'data')
        cycles = r.get('_cycles_raw', [])
        geo = r.get('geo_area', 5.0)

        if not cycles:
            # Fallback: write representative data
            j, V = r['j'], r['V']
            col_end = col + 2  # j_A, j_mA, V
            cell = ws2.cell(row=1, column=col, value=label)
            cell.font = hdr_font; cell.fill = label_fill
            if col_end > col:
                ws2.merge_cells(start_row=1, start_column=col,
                                end_row=1, end_column=col_end)
            ws2.cell(row=2, column=col, value='j (A/cm²)').font = hdr_font
            ws2.cell(row=2, column=col+1, value='j (mA/cm²)').font = hdr_font
            ws2.cell(row=2, column=col+2, value='V (V)').font = hdr_font
            for ri in range(len(j)):
                ws2.cell(row=ri+3, column=col, value=round(float(j[ri]), 6))
                ws2.cell(row=ri+3, column=col+1, value=round(float(j[ri]) * 1000, 3))
                ws2.cell(row=ri+3, column=col+2, value=round(float(V[ri]), 6))
            col = col_end + 2
            continue

        # File-level label spanning all cycle columns
        total_cycle_cols = sum(3 + (1 if cyc.get('HFR') is not None else 0)
                               for cyc in cycles)
        total_cycle_cols += len(cycles) - 1  # gap columns between cycles
        file_col_end = col + total_cycle_cols - 1
        cell = ws2.cell(row=1, column=col, value=label)
        cell.font = hdr_font; cell.fill = label_fill
        if file_col_end > col:
            ws2.merge_cells(start_row=1, start_column=col,
                            end_row=1, end_column=file_col_end)

        for ci, cyc in enumerate(cycles):
            direction = cyc.get('direction', '?')
            cycle_num = cyc.get('cycle_number', ci + 1)
            cyc_label = f'Cycle {cycle_num} ({direction})'
            cyc_j = cyc['j']
            cyc_V = cyc['V']
            cyc_hfr = cyc.get('HFR')
            n_cols = 3 + (1 if cyc_hfr is not None else 0)  # j_A, j_mA, V, [HFR]
            cyc_col_end = col + n_cols - 1

            # Cycle sub-label
            cell = ws2.cell(row=2, column=col, value=cyc_label)
            cell.font = hdr_font
            if cyc_col_end > col:
                ws2.merge_cells(start_row=2, start_column=col,
                                end_row=2, end_column=cyc_col_end)

            # Column headers
            hdrs = ['j (A/cm²)', 'j (mA/cm²)', 'V (V)']
            if cyc_hfr is not None:
                hdrs.append('HFR (mΩ·cm²)')
            for hi, h in enumerate(hdrs):
                cell = ws2.cell(row=3, column=col + hi, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
                ws2.column_dimensions[get_column_letter(col + hi)].width = 16

            # Data
            for ri in range(len(cyc_j)):
                cc = col
                ws2.cell(row=ri + 4, column=cc,
                         value=round(float(cyc_j[ri]), 6)); cc += 1
                ws2.cell(row=ri + 4, column=cc,
                         value=round(float(cyc_j[ri]) * 1000, 3)); cc += 1
                ws2.cell(row=ri + 4, column=cc,
                         value=round(float(cyc_V[ri]), 6)); cc += 1
                if cyc_hfr is not None and ri < len(cyc_hfr):
                    ws2.cell(row=ri + 4, column=cc,
                             value=round(float(cyc_hfr[ri]) * geo * 1000, 2))

            col = cyc_col_end + 2  # gap between cycles

    wb.save(filepath)
    print(f'\n  Excel: {filepath}')


# ═══════════════════════════════════════════════════════════════════════
#  Demo
# ═══════════════════════════════════════════════════════════════════════

def run_demo(save_dir=None):
    """Run full demo with synthetic data."""
    print('\n' + '▓' * 60)
    print('  FUEL CELL POLCURVE ANALYSIS — DEMO')
    print('▓' * 60)

    if save_dir:
        os.makedirs(save_dir, exist_ok=True)

    np.random.seed(42)
    geo_area = 5.0

    # ── Demo 1: Single pol curve with HFR ──
    print('\n[1] Single Polarization Curve + HFR')
    j, V, HFR = generate_synthetic_polcurve(geo_area=geo_area)
    print(f'    {len(j)} points, j: {j.min():.3f} – {j.max():.3f} A/cm²')

    r = analyze_polcurve(j, V, HFR=HFR, geo_area=geo_area)
    print_results(r)

    p1 = os.path.join(save_dir, 'polcurve_demo.png') if save_dir else None
    plot_polcurve(r, save_path=p1)

    # ── Demo 2: Degradation overlay ──
    print('\n[2] Degradation Overlay')
    all_results = []
    for i, (r_scale, j_lim_scale) in enumerate([(1.0, 1.0), (1.1, 0.95),
                                                  (1.25, 0.88), (1.4, 0.80)]):
        j_d, V_d, HFR_d = generate_synthetic_polcurve(geo_area=geo_area)
        # Degrade: increase HFR, reduce limiting current
        V_d -= j_d * 0.050 * (r_scale - 1.0) * geo_area / geo_area
        V_d *= j_lim_scale / 1.0  # crude scaling
        V_d = np.clip(V_d, 0, None)

        rd = analyze_polcurve(j_d, V_d, HFR=HFR_d * r_scale, geo_area=geo_area)
        rd['label'] = f'{i * 250}h'
        all_results.append(rd)
        print(f'    {rd["label"]}: Peak P = {rd["peak_power_W_cm2"]*1000:.0f} mW/cm²')

    p2 = os.path.join(save_dir, 'polcurve_overlay.png') if save_dir else None
    plot_polcurve_overlay(all_results, save_path=p2)

    if save_dir is None:
        plt.show()

    return r


# ═══════════════════════════════════════════════════════════════════════
#  Batch Processing
# ═══════════════════════════════════════════════════════════════════════

def run_batch(filepaths, labels, geo_area,
              delimiter, skip, j_col, v_col, hfr_col,
              current_is_total, tafel_j_min, tafel_j_max,
              condition_cols=None, hfr_scale=1.0, save_dir=None,
              mode_col=None, mode_exclude=None,
              j_scale=1.0, v_scale=1.0, image_ext='png'):
    """
    Batch-process multiple polarization curve files.
    """
    if save_dir:
        os.makedirs(save_dir, exist_ok=True)

    all_results = []
    summary_rows = []

    print(f'\n  Processing {len(filepaths)} files...\n')

    for i, (fp, lbl) in enumerate(zip(filepaths, labels)):
        print(f'  [{i+1}/{len(filepaths)}] {lbl}')
        try:
            fcd = parse_fcd_header(fp)
            file_skip = fcd['skip'] if fcd else skip
            file_j_col = fcd.get('j_col', j_col) if fcd else j_col
            file_v_col = fcd.get('v_col', v_col) if fcd else v_col
            file_hfr_col = fcd.get('hfr_col', hfr_col) if fcd else hfr_col
            file_mode_col = fcd.get('mode_col', mode_col) if fcd else mode_col
            file_cond_cols = fcd.get('condition_cols', condition_cols) if fcd else condition_cols
            j_raw, V_raw, HFR_raw, cond_raw = load_polcurve_data(
                fp, j_col=file_j_col, v_col=file_v_col, hfr_col=file_hfr_col,
                delimiter=delimiter, skip_header=file_skip,
                current_is_total=current_is_total, geo_area=geo_area,
                condition_cols=file_cond_cols, hfr_scale=hfr_scale,
                mode_col=file_mode_col, mode_exclude=mode_exclude,
                j_scale=j_scale, v_scale=v_scale)

            # Extract dwell endpoints
            v_step = detect_voltage_step(V_raw)
            j_dwell, V_dwell, HFR_dwell, cond_dwell = extract_dwell_endpoints(
                j_raw, V_raw, HFR_raw, cond_raw, v_step=v_step)

            # Filter HFR outliers
            if HFR_dwell is not None:
                HFR_dwell, n_out, _ = filter_hfr_outliers(HFR_dwell, geo_area=geo_area)
                if n_out > 0:
                    print(f'         HFR: {n_out} outlier(s) replaced')

            # Detect cycles and use last up-sweep
            cycles = extract_polcurve_cycles(j_dwell, V_dwell, HFR_dwell, cond_dwell, v_step=v_step)
            cycles = enforce_consensus_setpoints(cycles, v_step)
            rep = select_representative_cycle(cycles, choice='last_up')
            j_rep, V_rep, HFR_rep, cond_rep, n_merged = deduplicate_polcurve(
                rep['j'], rep['V'], rep.get('HFR'), rep.get('conditions'),
                v_step=v_step)
            j_rep, V_rep, HFR_rep, cond_rep = enforce_v_monotonicity(
                j_rep, V_rep, HFR_rep, cond_rep)
            j_rep, V_rep, HFR_rep, cond_rep = smooth_curve_selection(
                j_rep, V_rep, HFR_rep, cond_rep)
            n_up = sum(1 for c in cycles if c['direction'] == 'up')
            n_dn = sum(1 for c in cycles if c['direction'] == 'down')

            print(f'         {len(j_raw)} raw → {len(j_dwell)} dwells → '
                  f'{n_up} up + {n_dn} down → {len(j_rep)} pts (last up)'
                  f'{f", {n_merged} deduped" if n_merged else ""}'
                  f'{", HFR included" if HFR_dwell is not None else ""}')

            r = analyze_polcurve(j_rep, V_rep, HFR=HFR_rep, geo_area=geo_area,
                                 tafel_j_min=tafel_j_min, tafel_j_max=tafel_j_max)
            r['label'] = lbl
            r['filepath'] = fp
            r['conditions'] = cond_rep
            r['_cycles_raw'] = cycles
            all_results.append(r)

            peak_p = r['peak_power_W_cm2'] * 1000
            v1 = f'{r["V_at_1Acm2"]:.3f}' if r['V_at_1Acm2'] is not None else '—'
            hfr_str = f'  HFR={r["HFR_mean"]*1000:.1f} mΩ·cm²' if r.get('HFR_mean') else ''
            print(f'         OCV={r["OCV"]:.3f} V  Peak P={peak_p:.0f} mW/cm²  V@1={v1} V{hfr_str}')

            # Print conditions if available
            cond = rep.get('conditions')
            if cond:
                cond_str = '  |  '.join(f'{k}={v[0]:.1f}' for k, v in cond.items())
                print(f'         Conditions: {cond_str}')

            # Individual plot
            if save_dir and image_ext:
                safe_name = lbl.replace(' ', '_').replace('/', '-').replace('\\', '-')
                plot_polcurve(r, save_path=os.path.join(save_dir, f'polcurve_{safe_name}.{image_ext}'))
                plt.close()

            # Summary row
            row = {
                'Label': lbl,
                'File': os.path.basename(fp),
                'OCV (V)': r['OCV'],
                'Peak Power (mW/cm2)': peak_p,
                'j @ Peak P (A/cm2)': r['j_at_peak_power'],
            }
            if r['V_at_1Acm2'] is not None:
                row['V @ 1 A/cm2 (V)'] = r['V_at_1Acm2']
            if r.get('HFR_mean') is not None:
                row['Mean HFR (mOhm·cm2)'] = r['HFR_mean'] * 1000
            if r['tafel'] is not None:
                row['Tafel (mV/dec)'] = r['tafel']['tafel_slope_mVdec']
                row['j0 (A/cm2)'] = r['tafel']['j0_A_cm2']
            # Add conditions to summary
            if cond:
                for k, v in cond.items():
                    row[k] = np.mean(v)  # mean across the sweep
            summary_rows.append(row)

        except Exception as e:
            print(f'         ERROR: {e}')
            continue

    if not all_results:
        print('\n  No files processed successfully.')
        return []

    # ── Summary CSV ──
    if save_dir and summary_rows:
        csv_path = os.path.join(save_dir, 'polcurve_batch_summary.csv')
        all_cols = list(summary_rows[0].keys())
        for row in summary_rows[1:]:
            for k in row:
                if k not in all_cols:
                    all_cols.append(k)

        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=all_cols)
            writer.writeheader()
            writer.writerows(summary_rows)
        print(f'\n  Summary CSV: {csv_path}')

    # ── Batch Excel ──
    if save_dir and all_results:
        xlsx_path = os.path.join(save_dir, 'polcurve_batch_data.xlsx')
        save_batch_excel(all_results, summary_rows, xlsx_path)

    # ── Combined overlay ──
    if save_dir and image_ext:
        overlay_path = os.path.join(save_dir, f'polcurve_batch_overlay.{image_ext}')
        plot_polcurve_overlay(all_results, save_path=overlay_path)
        plt.close()
    elif not save_dir:
        plot_polcurve_overlay(all_results)
        plt.show()

    # ── Console summary ──
    has_hfr = any(r.get('HFR_mean') is not None for r in all_results)
    has_tafel = any(r['tafel'] is not None for r in all_results)

    print(f'\n  {"═" * 85}')
    hdr = f'  {"Label":25s} {"OCV":>6s} {"V@1A":>7s} {"Peak P":>8s} {"j@Pk":>6s}'
    if has_hfr:
        hdr += f' {"HFR":>8s}'
    if has_tafel:
        hdr += f' {"Tafel":>7s}'
    print(hdr)
    print(f'  {"":25s} {"(V)":>6s} {"(V)":>7s} {"(mW)":>8s} {"(A)":>6s}', end='')
    if has_hfr:
        print(f' {"(mΩ·cm²)":>8s}', end='')
    if has_tafel:
        print(f' {"(mV/d)":>7s}', end='')
    print()
    print(f'  {"─" * 85}')

    for r in all_results:
        v1 = f'{r["V_at_1Acm2"]:.3f}' if r['V_at_1Acm2'] is not None else '—'
        line = (f'  {r["label"]:25s} {r["OCV"]:>6.3f} {v1:>7s}'
                f' {r["peak_power_W_cm2"]*1000:>8.0f} {r["j_at_peak_power"]:>6.2f}')
        if has_hfr:
            if r.get('HFR_mean') is not None:
                line += f' {r["HFR_mean"]*1000:>8.1f}'
            else:
                line += f' {"—":>8s}'
        if has_tafel:
            if r['tafel'] is not None:
                line += f' {r["tafel"]["tafel_slope_mVdec"]:>7.0f}'
            else:
                line += f' {"—":>7s}'
        print(line)

    print(f'  {"═" * 85}\n')

    return all_results


# ═══════════════════════════════════════════════════════════════════════
#  Interactive Mode
# ═══════════════════════════════════════════════════════════════════════

def _prompt(label, default=None, cast=float):
    """Prompt for input with a default value."""
    suffix = f' [{default}]' if default is not None else ''
    raw = input(f'  {label}{suffix}: ').strip()
    if raw == '':
        return default
    if cast is not None:
        return cast(raw)
    return raw


def run_interactive():
    """Walk the user through pol curve analysis step by step."""
    print('\n' + '▓' * 60)
    print('  FUEL CELL POLCURVE ANALYSIS — INTERACTIVE MODE')
    print('▓' * 60)

    # ── Mode ──
    print('\n  Analysis modes:')
    print('    1 = Single pol curve file')
    print('    2 = Batch — folder (analyze all pol curve files)')
    print('    3 = Batch — file list (paste paths one at a time)')
    print('    4 = Run built-in demo')
    mode = int(_prompt('Select mode', default=1, cast=int))

    if mode == 4:
        save = _prompt('Save plots to directory? (path or Enter to show)', default=None, cast=None)
        if save:
            save = _clean_path(save)
        run_demo(save_dir=save if save else None)
        return

    # ── Electrode parameters ──
    print('\n  ── Electrode Parameters ──')
    geo_area = _prompt('Geometric area (cm²)', default=5.0)

    # ── Test stand presets ──
    print('\n  ── Measurement Test Stand ──')
    print('    0 = Scribner')
    print('    1 = FCTS')
    stand = int(_prompt('Test stand', default=0, cast=int))

    STAND_PRESETS = {
        0: {'name': 'Scribner', 'delimiter': '\t', 'skip': 51,
            'j_col': 1, 'v_col': 5, 'hfr_col': 20,
            'current_is_total': True, 'hfr_scale': 0.001,
            'mode_col': 28, 'mode_exclude': {5},
            'condition_cols': {
                'T_cell (C)': 13, 'T_anode_dp (C)': 14, 'H2_flow (slpm)': 15,
                'T_cathode_dp (C)': 17, 'Air_flow (slpm)': 18,
            }},
        1: {'name': 'FCTS',     'delimiter': ',',  'skip': 1,
            'j_col': 0, 'v_col': 1, 'hfr_col': 2,
            'current_is_total': False, 'hfr_scale': 1.0,
            'mode_col': None, 'mode_exclude': None,
            'condition_cols': None},
    }

    if stand in STAND_PRESETS:
        p = STAND_PRESETS[stand]
        delimiter        = p['delimiter']
        skip             = p['skip']
        j_col            = p['j_col']
        v_col            = p['v_col']
        hfr_col          = p['hfr_col']
        current_is_total = p['current_is_total']
        hfr_scale        = p['hfr_scale']
        mode_col         = p.get('mode_col')
        mode_exclude     = p.get('mode_exclude')
        condition_cols   = p['condition_cols']
        delim_name = 'tab' if delimiter == '\t' else 'comma'
        curr_type = 'total current (A)' if current_is_total else 'current density (A/cm²)'
        hfr_unit = 'mΩ' if hfr_scale == 0.001 else 'Ω'
        print(f'  → {p["name"]}: {delim_name}-delimited, {skip} header rows,')
        print(f'    I/j=col {j_col} ({curr_type}), V=col {v_col}, HFR=col {hfr_col} ({hfr_unit})')
        if condition_cols:
            cond_str = ', '.join(f'{k}=col {v}' for k, v in condition_cols.items())
            print(f'    Conditions: {cond_str}')
    else:
        print(f'  Unknown test stand {stand}, using manual entry')
        print('\n  ── File Format ──')
        delim_choice = _prompt('Delimiter: 1=comma  2=tab  3=semicolon', default=1, cast=int)
        delimiter = {1: ',', 2: '\t', 3: ';'}.get(delim_choice, ',')
        skip      = int(_prompt('Header rows to skip', default=1, cast=int))
        j_col     = int(_prompt('Current/j column index (0-based)', default=0, cast=int))
        v_col     = int(_prompt('Voltage column index (0-based)', default=1, cast=int))
        hfr_raw   = _prompt('HFR column index (0-based, Enter for none)', default=None, cast=None)
        hfr_col   = int(hfr_raw) if hfr_raw is not None else None
        print('\n  ── Current type ──')
        print('    1 = Current density (A/cm²)')
        print('    2 = Total current (A)')
        curr_type = int(_prompt('Current type', default=1, cast=int))
        current_is_total = (curr_type == 2)
        print('\n  ── HFR Units ──')
        print('    1 = Ω      2 = mΩ')
        hfr_unit_choice = int(_prompt('HFR unit', default=1, cast=int))
        hfr_scale = 0.001 if hfr_unit_choice == 2 else 1.0
        condition_cols = None
        mode_col = None
        mode_exclude = None

    # ── Tafel range ──
    print('\n  ── Tafel Fit Range ──')
    tafel_j_min = _prompt('Tafel j_min (A/cm²)', default=0.01)
    tafel_j_max = _prompt('Tafel j_max (A/cm²)', default=0.10)

    if mode == 1:
        # ── Single file ──
        print('\n  ── Data File ──')
        print('  Tip: drag-and-drop a file into the terminal to paste its path')
        filepath = _clean_path(_prompt('Pol curve data file path', cast=None))

        print(f'\n  Loading: {filepath}')
        fcd = parse_fcd_header(filepath)
        if fcd:
            skip = fcd['skip']
            j_col = fcd.get('j_col', j_col)
            v_col = fcd.get('v_col', v_col)
            hfr_col = fcd.get('hfr_col', hfr_col)
            mode_col = fcd.get('mode_col', mode_col)
            condition_cols = fcd.get('condition_cols', condition_cols)
        j_raw, V_raw, HFR_raw, cond_raw = load_polcurve_data(
            filepath, j_col=j_col, v_col=v_col, hfr_col=hfr_col,
            delimiter=delimiter, skip_header=skip,
            current_is_total=current_is_total, geo_area=geo_area,
            condition_cols=condition_cols, hfr_scale=hfr_scale,
            mode_col=mode_col, mode_exclude=mode_exclude)
        print(f'  Raw: {len(j_raw)} points')

        # Auto-detect voltage step size for deduplication
        v_step = detect_voltage_step(V_raw)
        if v_step:
            print(f'  Voltage step: {v_step*1000:.1f} mV (auto-detected)')

        # Extract dwell endpoints
        j_dwell, V_dwell, HFR_dwell, cond_dwell = extract_dwell_endpoints(
            j_raw, V_raw, HFR_raw, cond_raw, v_step=v_step)
        print(f'  Dwells: {len(j_dwell)} setpoints extracted'
              f'  (j: {j_dwell.min():.3f} – {j_dwell.max():.3f} A/cm²)')

        # Filter HFR outliers
        if HFR_dwell is not None:
            HFR_dwell, n_out, _ = filter_hfr_outliers(HFR_dwell, geo_area=geo_area)
            if n_out > 0:
                print(f'  HFR: {n_out} outlier(s) replaced by interpolation')

        # Print conditions from first dwell point
        if cond_dwell:
            cond_str = '  |  '.join(f'{k}={v[0]:.1f}' for k, v in cond_dwell.items())
            print(f'  Conditions: {cond_str}')

        # Detect cycles
        cycles = extract_polcurve_cycles(j_dwell, V_dwell, HFR_dwell, cond_dwell, v_step=v_step)
        cycles = enforce_consensus_setpoints(cycles, v_step)
        up_sweeps = [c for c in cycles if c['direction'] == 'up']
        down_sweeps = [c for c in cycles if c['direction'] == 'down']
        print(f'  Cycles: {len(up_sweeps)} up-sweep(s), {len(down_sweeps)} down-sweep(s)')

        # Analyze all cycles
        all_cycle_results = []
        for c in cycles:
            cr = analyze_polcurve(c['j'], c['V'], HFR=c['HFR'], geo_area=geo_area,
                                   tafel_j_min=tafel_j_min, tafel_j_max=tafel_j_max)
            cr['direction'] = c['direction']
            cr['cycle_number'] = c['cycle_number']
            cr['conditions'] = c.get('conditions')
            all_cycle_results.append(cr)
            pk = cr['peak_power_W_cm2'] * 1000
            v1 = f'{cr["V_at_1Acm2"]:.3f}' if cr['V_at_1Acm2'] is not None else '—'
            print(f'    Cycle {c["cycle_number"]} ({c["direction"]:>4s}): '
                  f'{len(c["j"])} pts, OCV = {cr["OCV"]:.3f} V  '
                  f'Peak P = {pk:.0f} mW/cm²  V@1A = {v1} V')

        # Use last up-sweep as representative
        rep = select_representative_cycle(cycles, choice='last_up')
        j_rep, V_rep, HFR_rep, cond_rep, n_merged = deduplicate_polcurve(
            rep['j'], rep['V'], rep.get('HFR'), rep.get('conditions'),
            v_step=v_step)
        j_rep, V_rep, HFR_rep, cond_rep = enforce_v_monotonicity(
            j_rep, V_rep, HFR_rep, cond_rep)
        j_rep, V_rep, HFR_rep, cond_rep = smooth_curve_selection(
            j_rep, V_rep, HFR_rep, cond_rep)
        if n_merged > 0:
            print(f'  Dedup: {n_merged} duplicate point(s) averaged')
        r = analyze_polcurve(j_rep, V_rep, HFR=HFR_rep, geo_area=geo_area,
                             tafel_j_min=tafel_j_min, tafel_j_max=tafel_j_max)
        r['_cycles_raw'] = cycles
        r['conditions'] = cond_rep
        print(f'\n  Representative: last up-sweep (cycle {rep["cycle_number"]}), {len(j_rep)} pts')
        print_results(r)

        save = _prompt('\n  Save plot to directory? (path or Enter to show)', default=None, cast=None)
        if save:
            save = _clean_path(save)
            os.makedirs(save, exist_ok=True)
            plot_polcurve(r, save_path=os.path.join(save, 'polcurve.png'))
            save_single_excel(r, os.path.join(save, 'polcurve_data.xlsx'))
        else:
            plot_polcurve(r)
            plt.show()

    elif mode in (2, 3):
        # ── Batch ──
        save_dir = _clean_path(_prompt('Output directory for results',
                                        default='polcurve_batch_output', cast=None))
        filepaths = []
        labels = []

        if mode == 2:
            folder = _clean_path(_prompt('Folder path', cast=None))
            extensions = ['*.csv', '*.txt', '*.tsv', '*.fcd',
                          '*.CSV', '*.TXT', '*.TSV', '*.FCD']
            for ext in extensions:
                filepaths.extend(glob.glob(os.path.join(folder, ext)))
            filepaths = sorted(set(filepaths))

            # Filter: include files with "polcurve", "pol", "IV", or "polarization"
            keywords = ['POLCURVE', 'POL_', 'POL-', 'IV_', 'IV-',
                        'POLARIZATION', 'POLDATA']
            filepaths = [fp for fp in filepaths
                         if any(kw in os.path.basename(fp).upper() for kw in keywords)]

            if not filepaths:
                print(f'  No polarization curve files found in: {folder}')
                print(f'  (looking for filenames containing: {", ".join(keywords)})')
                return

            print(f'  Found {len(filepaths)} files:')
            for fp in filepaths:
                name = os.path.splitext(os.path.basename(fp))[0]
                labels.append(name)
                print(f'    {name}')

        elif mode == 3:
            print('  Enter file paths one per line. Type "done" when finished.')
            while True:
                raw = input('  File path (or "done"): ').strip()
                if raw.lower() == 'done':
                    break
                fp = _clean_path(raw)
                if not fp:
                    continue
                if not os.path.isfile(fp):
                    print(f'    File not found: {fp}')
                    continue
                default_label = os.path.splitext(os.path.basename(fp))[0]
                lbl = _prompt('  Label', default=default_label, cast=None)
                filepaths.append(fp)
                labels.append(lbl)

            if not filepaths:
                print('  No files entered.')
                return

        run_batch(filepaths, labels, geo_area,
                  delimiter, skip, j_col, v_col, hfr_col,
                  current_is_total, tafel_j_min, tafel_j_max,
                  condition_cols=condition_cols, hfr_scale=hfr_scale,
                  save_dir=save_dir, mode_col=mode_col,
                  mode_exclude=mode_exclude)


# ═══════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='Fuel Cell Polarization Curve Analysis')
    parser.add_argument('--file', type=str, help='Pol curve data file')
    parser.add_argument('--area', type=float, default=5.0,
                        help='Geometric electrode area in cm² (default: 5.0)')
    parser.add_argument('--save-dir', type=str, default=None,
                        help='Directory to save plots')
    parser.add_argument('--demo', action='store_true',
                        help='Run built-in demo')
    parser.add_argument('--interactive', '-i', action='store_true',
                        help='Run interactive guided mode')
    args = parser.parse_args()

    if args.interactive:
        run_interactive()
    elif args.demo:
        run_demo(save_dir=args.save_dir)
    elif args.file:
        j_raw, V_raw, HFR_raw, cond_raw = load_polcurve_data(args.file)
        v_step = detect_voltage_step(V_raw)
        j_dw, V_dw, HFR_dw, cond_dw = extract_dwell_endpoints(
            j_raw, V_raw, HFR_raw, cond_raw, v_step=v_step)
        if HFR_dw is not None:
            HFR_dw, n_out, _ = filter_hfr_outliers(HFR_dw, geo_area=args.area)
            if n_out > 0:
                print(f'  HFR: {n_out} outlier(s) replaced')
        cycles = extract_polcurve_cycles(j_dw, V_dw, HFR_dw, cond_dw, v_step=v_step)
        cycles = enforce_consensus_setpoints(cycles, v_step)
        rep = select_representative_cycle(cycles, choice='last_up')
        j_rep, V_rep, HFR_rep, cond_rep, n_merged = deduplicate_polcurve(
            rep['j'], rep['V'], rep.get('HFR'), rep.get('conditions'),
            v_step=v_step)
        j_rep, V_rep, HFR_rep, cond_rep = enforce_v_monotonicity(
            j_rep, V_rep, HFR_rep, cond_rep)
        j_rep, V_rep, HFR_rep, cond_rep = smooth_curve_selection(
            j_rep, V_rep, HFR_rep, cond_rep)
        print(f'  {len(j_raw)} raw → {len(j_dw)} dwells → '
              f'{sum(1 for c in cycles if c["direction"]=="up")} up cycles → '
              f'{len(j_rep)} pts (last up)'
              f'{f", {n_merged} deduped" if n_merged else ""}')
        r = analyze_polcurve(j_rep, V_rep, HFR=HFR_rep, geo_area=args.area)
        r['_cycles_raw'] = cycles
        print_results(r)
        if args.save_dir:
            os.makedirs(args.save_dir, exist_ok=True)
            plot_polcurve(r, save_path=os.path.join(args.save_dir, 'polcurve.png'))
            save_single_excel(r, os.path.join(args.save_dir, 'polcurve_data.xlsx'))
        else:
            plot_polcurve(r)
            plt.show()
    else:
        run_interactive()


if __name__ == '__main__':
    main()
