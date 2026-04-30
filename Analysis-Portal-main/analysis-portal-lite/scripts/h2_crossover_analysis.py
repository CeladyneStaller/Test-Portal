#!/usr/bin/env python3
"""
H2 Crossover Analysis (CV Method) for PEM Fuel Cells
=====================================================
Analyzes H2/N2 cyclic voltammetry data at 500 mV/s.

The hydrogen crossover current is the average of the anodic and cathodic
sweep currents. Capacitive charging current cancels in the average,
leaving only the faradaic H2 oxidation current + electronic short:

    j_avg(V) = j_xover + V / R_short

Usage:
  python h2_crossover_analysis.py              # interactive mode
  python h2_crossover_analysis.py --demo       # built-in demo
"""

import os, csv, glob, argparse
import numpy as np
from scipy import interpolate, stats
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scripts.helpers.plot_compare import save_with_sidecar


def run(input_dir: str, output_dir: str, params: dict = None) -> dict:
    """Batch H2 crossover analysis from slow-scan CV data."""
    from pathlib import Path
    p = params or {}

    inp = Path(input_dir)
    all_files = sorted([f for f in inp.rglob("*")
                    if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv', '.fcd')])

    if not all_files:
        return {"status": "error", "message": "No data files found"}

    # Filter by crossover keywords; fall back to all files if none match
    KEYWORDS = ['CV-2MVS', 'CROSSOVER', 'XOVER', 'H2XOVER', 'H2X',
                'H2_CROSS', 'PERMEATION']
    filtered = [f for f in all_files
                if any(kw in f.name.upper() for kw in KEYWORDS)]
    files = filtered if filtered else all_files

    has_tab_files = any(f.suffix.lower() in ('.fcd', '.tsv') for f in files)
    delimiter = '\t' if has_tab_files else ','

    membrane = float(p.get('membrane_thickness', 0))

    filepaths = [str(f) for f in files]
    labels = [f.stem for f in files]

    from scripts.helpers.conditions import img_ext_from_params
    image_ext = img_ext_from_params(p)

    results = run_batch(
        filepaths, labels,
        geo_area=float(p.get('geo_area', 5.0)),
        delimiter=delimiter,
        skip=1,
        v_col=0,
        j_col=1,
        current_is_total=True,
        avg_V_min=float(p.get('avg_V_min', 0.35)),
        avg_V_max=float(p.get('avg_V_max', 0.50)),
        membrane_thickness_um=membrane if membrane > 0 else None,
        save_dir=str(output_dir),
        image_ext=image_ext,
    )

    output_files = [str(f.relative_to(Path(output_dir))) for f in Path(output_dir).rglob('*') if f.is_file()]
    if not output_files:
        raise RuntimeError(
            f'Analysis produced no output. {len(files)} file(s) were found '
            f'but none could be processed. Check file format and parameters.'
        )
    return {"status": "success", "files_processed": len(files), "files_produced": output_files}

FARADAY = 96485.3329
N_ELECTRONS = 2
MOLAR_VOL_STP = 22414.0

def parse_fcd_header(filepath):
    """Parse Scribner .fcd header for skip count and column indices."""
    if not filepath.lower().endswith('.fcd'):
        return None
    try:
        with open(filepath, 'r', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                lines.append(line)
                if 'End Comments' in line:
                    skip = i + 1; break
            else:
                return None
    except Exception:
        return None
    if len(lines) < 2:
        return None
    cols = lines[-2].strip().split('\t')
    result = {'skip': skip}
    for ci, name in enumerate(cols):
        n = name.strip()
        if n == 'I (A)': result['j_col'] = ci
        elif n == 'E_Stack (V)': result['v_col'] = ci
        elif n.startswith('Z_Freq'): result['freq_col'] = ci
        elif n.startswith('Z_Real'): result['zreal_col'] = ci
        elif n.startswith('Z_Imag'): result['zimag_col'] = ci
        elif n == 'Ctrl_Mode': result['mode_col'] = ci
    return result


def _clean_path(p):
    if p is None: return None
    p = p.strip()
    if p.startswith('& '): p = p[2:]
    return p.strip().strip('"').strip("'").strip('\u2018\u2019\u201c\u201d\u202a\u200b')

def _prompt(msg, default=None, cast=float):
    tail = f' [{default}]' if default is not None else ''
    raw = input(f'  {msg}{tail}: ').strip()
    if not raw: return default
    return raw if cast is None else cast(raw)


def load_cv_data(filepath, v_col=0, j_col=1, delimiter=',', skip_header=1,
                 current_is_total=False, geo_area=5.0,
                 mode_col=None, mode_exclude=None):
    filepath = _clean_path(filepath)
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f'File not found: "{filepath}"')
    v_data, j_data = [], []
    mode_excl = set(mode_exclude) if mode_exclude else set()
    with open(filepath, 'r', errors='replace') as f:
        reader = csv.reader(f, delimiter=delimiter)
        for _ in range(skip_header): next(reader)
        for row in reader:
            try:
                if mode_col is not None and mode_excl:
                    if int(float(row[mode_col])) in mode_excl: continue
                v_data.append(float(row[v_col]))
                j_data.append(float(row[j_col]))
            except (ValueError, IndexError): continue
    V, j = np.array(v_data), np.array(j_data)
    if current_is_total: j = j / geo_area
    return V, j


def detect_sweeps(V, min_sweep_pts=10):
    dV = np.diff(V)
    signs = np.sign(dV)
    signs[signs == 0] = 1
    sweeps = []
    seg_start, cur_dir = 0, signs[0]
    for i in range(1, len(signs)):
        if signs[i] != cur_dir:
            if i - seg_start >= min_sweep_pts:
                d = 'anodic' if cur_dir > 0 else 'cathodic'
                sweeps.append({'start': seg_start, 'end': i + 1, 'direction': d})
            seg_start, cur_dir = i, signs[i]
    if len(signs) - seg_start >= min_sweep_pts:
        d = 'anodic' if cur_dir > 0 else 'cathodic'
        sweeps.append({'start': seg_start, 'end': len(V), 'direction': d})
    return sweeps


def extract_last_cycle(V, j, sweeps):
    """Extract last complete anodic + cathodic pair (both cover full V range)."""
    anodic = [s for s in sweeps if s['direction'] == 'anodic']
    cathodic = [s for s in sweeps if s['direction'] == 'cathodic']
    if not anodic or not cathodic:
        raise ValueError('Could not detect both anodic and cathodic sweeps')
    V_range = V.max() - V.min()
    # Find last complete cathodic (>80% V range)
    c = cathodic[-1]
    for ci in range(len(cathodic) - 1, -1, -1):
        seg = cathodic[ci]
        V_seg = V[seg['start']:seg['end']]
        if (V_seg.max() - V_seg.min()) > 0.8 * V_range:
            c = seg; break
    # Paired anodic: the one ending just before this cathodic
    a = anodic[-1]
    for ai in range(len(anodic) - 1, -1, -1):
        if anodic[ai]['end'] <= c['start'] + 2:
            a = anodic[ai]; break
    return (V[a['start']:a['end']], j[a['start']:a['end']],
            V[c['start']:c['end']], j[c['start']:c['end']])


def analyze_crossover(V, j, avg_V_min=0.35, avg_V_max=0.50,
                      geo_area=5.0, membrane_thickness_um=None,
                      p_h2_anode_kPa=101.325):
    sweeps = detect_sweeps(V)
    V_a, j_a, V_c, j_c = extract_last_cycle(V, j, sweeps)

    V_lo = max(V_a.min(), V_c.min())
    V_hi = min(V_a.max(), V_c.max())
    V_grid = np.linspace(V_lo, V_hi, 500)

    f_a = interpolate.interp1d(V_a[np.argsort(V_a)], j_a[np.argsort(V_a)],
                                kind='linear', fill_value='extrapolate')
    f_c = interpolate.interp1d(V_c[np.argsort(V_c)], j_c[np.argsort(V_c)],
                                kind='linear', fill_value='extrapolate')
    j_avg = (f_a(V_grid) + f_c(V_grid)) / 2.0

    mask = (V_grid >= avg_V_min) & (V_grid <= avg_V_max)
    if mask.sum() < 5:
        raise ValueError(f'Too few points in [{avg_V_min}, {avg_V_max}] V')
    V_fit, j_fit = V_grid[mask], j_avg[mask]

    slope, intercept, r_value, _, _ = stats.linregress(V_fit, j_fit)
    j_avg_mean = np.mean(j_fit)
    R_short = 1.0 / slope if abs(slope) > 1e-12 else np.inf

    # Primary crossover metric: mean of averaged anodic+cathodic current
    j_xover = abs(j_avg_mean)  # magnitude — sign is convention-dependent
    J_H2 = j_xover / (N_ELECTRONS * FARADAY)
    K_H2 = None
    if membrane_thickness_um and membrane_thickness_um > 0:
        K_H2 = J_H2 * (membrane_thickness_um * 1e-4) / (p_h2_anode_kPa * 1000)

    V_line = np.linspace(0, V_hi, 100)
    j_line = slope * V_line + intercept

    return {
        'V_anodic': V_a, 'j_anodic': j_a,
        'V_cathodic': V_c, 'j_cathodic': j_c,
        'V_grid': V_grid, 'j_avg': j_avg,
        'j_xover_mA_cm2': j_xover * 1000,
        'j_avg_mean_mA_cm2': j_avg_mean * 1000,
        'j_intercept_mA_cm2': intercept * 1000,
        'R_short_ohm_cm2': R_short, 'slope': slope, 'R_squared': r_value**2,
        'J_H2_mol_cm2_s': J_H2,
        'J_H2_nmol_cm2_s': J_H2 * 1e9,
        'J_H2_mL_min_cm2': J_H2 * MOLAR_VOL_STP * 60,
        'K_H2_mol_cm_s_Pa': K_H2,
        'membrane_thickness_um': membrane_thickness_um,
        'p_h2_kPa': p_h2_anode_kPa,
        'avg_V_min': avg_V_min, 'avg_V_max': avg_V_max,
        'geo_area': geo_area,
        '_V_fit': V_fit, '_j_fit': j_fit, '_V_line': V_line, '_j_line': j_line,
    }


def print_results(r):
    w = 55
    print(f'\n  {"=" * w}')
    print(f'  Geo. area:           {r["geo_area"]:.2f} cm2')
    print(f'  Averaging window:    {r["avg_V_min"]:.2f} - {r["avg_V_max"]:.2f} V')
    print(f'  {"-" * w}')
    print(f'  H2 crossover j:     {r["j_xover_mA_cm2"]:.3f} mA/cm2')
    print(f'  {"-" * w}')
    print(f'  H2 flux:             {r["J_H2_nmol_cm2_s"]:.2f} nmol/cm2/s')
    print(f'                       {r["J_H2_mL_min_cm2"]:.4f} mL_STP/min/cm2')
    if r['K_H2_mol_cm_s_Pa'] is not None:
        print(f'  {"-" * w}')
        print(f'  Membrane:            {r["membrane_thickness_um"]:.1f} um')
        print(f'  H2 permeability:     {r["K_H2_mol_cm_s_Pa"]:.2e} mol/cm/s/Pa')
    print(f'  {"=" * w}')


def plot_crossover(r, save_path=None):
    fig, axes = plt.subplots(1, 3, figsize=(16, 5))

    ax = axes[0]
    ax.plot(r['V_anodic'], r['j_anodic'] * 1000, '-', color='firebrick',
            lw=1.2, alpha=0.8, label='Anodic')
    ax.plot(r['V_cathodic'], r['j_cathodic'] * 1000, '-', color='steelblue',
            lw=1.2, alpha=0.8, label='Cathodic')
    ax.set_xlabel('Voltage (V)'); ax.set_ylabel('Current density (mA/cm2)')
    ax.set_title('H2/N2 CV (last cycle)'); ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    ax = axes[1]
    ax.plot(r['V_grid'], r['j_avg'] * 1000, '-', color='purple', lw=1.5,
            label='(anodic + cathodic) / 2')
    ax.axvspan(r['avg_V_min'], r['avg_V_max'], alpha=0.1, color='orange',
               label=f'Window [{r["avg_V_min"]:.2f}, {r["avg_V_max"]:.2f}] V')
    ax.set_xlabel('Voltage (V)'); ax.set_ylabel('Current density (mA/cm2)')
    ax.set_title('Averaged Current'); ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    # Panel 3: DL window zoom with mean
    ax = axes[2]
    mask = (r['V_grid'] >= r['avg_V_min']) & (r['V_grid'] <= r['avg_V_max'])
    ax.plot(r['V_grid'][mask], r['j_avg'][mask] * 1000, '-', color='purple',
            lw=2, label='Averaged j')
    ax.axhline(r['j_avg_mean_mA_cm2'], color='green', ls='-', lw=2,
               label=f'Mean = {r["j_avg_mean_mA_cm2"]:.3f} mA/cm2')
    ax.axhline(0, color='gray', ls=':', lw=0.5)

    txt = (f'|j_xover| = {r["j_xover_mA_cm2"]:.3f} mA/cm2\n'
           f'H2 flux = {r["J_H2_nmol_cm2_s"]:.1f} nmol/cm2/s')
    if r['K_H2_mol_cm_s_Pa'] is not None:
        txt += f'\nK_H2 = {r["K_H2_mol_cm_s_Pa"]:.2e} mol/cm/s/Pa'
    ax.text(0.03, 0.97, txt, transform=ax.transAxes, fontsize=9,
            va='top', bbox=dict(boxstyle='round,pad=0.4', fc='lightyellow', alpha=0.9))
    ax.set_xlabel('Voltage (V)'); ax.set_ylabel('Current density (mA/cm2)')
    ax.set_title(f'DL Window [{r["avg_V_min"]:.2f}, {r["avg_V_max"]:.2f}] V')
    ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    from scripts.helpers.conditions import get_condition_label
    cond_label = get_condition_label(label=r.get('label', ''))
    title = 'H2 Crossover Analysis (CV Method)'
    if cond_label:
        title += f'\n{cond_label}'
    fig.suptitle(title, fontsize=13, fontweight='bold')

    fig.tight_layout()
    if save_path:
        save_with_sidecar(fig, save_path, plot_type='crossover', dpi=200, bbox_inches='tight')
        print(f'  Plot: {save_path}')
    return fig


def save_crossover_excel(r, filepath, label=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    hf, hfill = Font(bold=True), PatternFill('solid', fgColor='D9E1F2')

    ws = wb.active; ws.title = 'Summary'
    rows = [
        ('Label', label or ''),
        ('Geo. Area (cm2)', r['geo_area']),
        ('Averaging Window (V)', f'{r["avg_V_min"]:.2f} - {r["avg_V_max"]:.2f}'),
        ('', ''),
        ('H2 Crossover j (mA/cm2)', r['j_xover_mA_cm2']),
        ('', ''),
        ('H2 Flux (nmol/cm2/s)', r['J_H2_nmol_cm2_s']),
        ('H2 Flux (mL_STP/min/cm2)', r['J_H2_mL_min_cm2']),
    ]
    if r['K_H2_mol_cm_s_Pa'] is not None:
        rows += [('', ''), ('Membrane Thickness (um)', r['membrane_thickness_um']),
                 ('H2 Permeability (mol/cm/s/Pa)', r['K_H2_mol_cm_s_Pa'])]
    for i, (lab, val) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=lab).font = hf
        ws.cell(row=i, column=2, value=val)
    ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 20

    ws2 = wb.create_sheet('Averaged Data')
    for c, h in enumerate(['V (V)', 'j_avg (mA/cm2)'], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[cell.column_letter].width = 16
    for i in range(len(r['V_grid'])):
        ws2.cell(row=i+2, column=1, value=round(float(r['V_grid'][i]), 6))
        ws2.cell(row=i+2, column=2, value=round(float(r['j_avg'][i]) * 1000, 6))

    ws3 = wb.create_sheet('Raw Sweeps')
    for c, h in enumerate(['V_anodic (V)', 'j_anodic (mA/cm2)', '', 'V_cathodic (V)', 'j_cathodic (mA/cm2)'], 1):
        cell = ws3.cell(row=1, column=c, value=h); cell.font = hf; cell.fill = hfill
        ws3.column_dimensions[cell.column_letter].width = 18
    for i in range(len(r['V_anodic'])):
        ws3.cell(row=i+2, column=1, value=round(float(r['V_anodic'][i]), 6))
        ws3.cell(row=i+2, column=2, value=round(float(r['j_anodic'][i]) * 1000, 6))
    for i in range(len(r['V_cathodic'])):
        ws3.cell(row=i+2, column=4, value=round(float(r['V_cathodic'][i]), 6))
        ws3.cell(row=i+2, column=5, value=round(float(r['j_cathodic'][i]) * 1000, 6))
    wb.save(filepath); print(f'  Excel: {filepath}')


def save_batch_crossover_excel(all_results, filepath):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    hf, hfill = Font(bold=True), PatternFill('solid', fgColor='D9E1F2')
    ws = wb.active; ws.title = 'Summary'
    cols = ['Label', 'j_xover (mA/cm2)',
            'H2 Flux (nmol/cm2/s)', 'H2 Flux (mL/min/cm2)']
    has_perm = any(r.get('K_H2_mol_cm_s_Pa') is not None for r in all_results)
    if has_perm: cols += ['Thickness (um)', 'Permeability (mol/cm/s/Pa)']
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 22
    for ri, r in enumerate(all_results, 2):
        vals = [r.get('label',''), r['j_xover_mA_cm2'],
                r['J_H2_nmol_cm2_s'], r['J_H2_mL_min_cm2']]
        if has_perm:
            vals += [r.get('membrane_thickness_um'), r.get('K_H2_mol_cm_s_Pa')]
        for c, v in enumerate(vals, 1):
            if v is not None: ws.cell(row=ri, column=c, value=v)
    for r in all_results:
        name = r.get('label','data')[:31].replace('/','_').replace('\\','_').replace('*','')
        ws2 = wb.create_sheet(name)
        for c, h in enumerate(['V (V)', 'j_avg (mA/cm2)'], 1):
            cell = ws2.cell(row=1, column=c, value=h); cell.font = hf; cell.fill = hfill
            ws2.column_dimensions[cell.column_letter].width = 16
        for i in range(len(r['V_grid'])):
            ws2.cell(row=i+2, column=1, value=round(float(r['V_grid'][i]), 6))
            ws2.cell(row=i+2, column=2, value=round(float(r['j_avg'][i]) * 1000, 6))
    wb.save(filepath); print(f'  Excel: {filepath}')


def run_demo(save_dir=None):
    print('\n' + chr(9619) * 60)
    print('  H2 CROSSOVER ANALYSIS (CV METHOD) - DEMO')
    print(chr(9619) * 60)
    np.random.seed(42)
    # Realistic H2/N2 CV: 0.05-1.2V at 500 mV/s
    n = 120  # ~115 pts per sweep at 10 pts/V
    V_up = np.linspace(0.05, 1.20, n)
    V_dn = np.linspace(1.20, 0.05, n)
    V = np.concatenate([V_up, V_dn])

    j_xover_true = 0.0015    # 1.5 mA/cm2
    C_dl = 0.015              # F/cm2
    sr = 0.500                # V/s
    R_short = 500             # Ohm cm2

    # Capacitive: +C*sr anodic, -C*sr cathodic
    j_cap = np.concatenate([np.full(n, +C_dl*sr), np.full(n, -C_dl*sr)])
    # H_UPD peaks (Gaussian around 0.1V and 0.25V)
    j_hupd_a = 0.080 * np.exp(-(V - 0.12)**2 / (2*0.03**2)) + 0.040 * np.exp(-(V - 0.25)**2 / (2*0.02**2))
    j_hupd_c = -0.070 * np.exp(-(V - 0.10)**2 / (2*0.03**2)) - 0.035 * np.exp(-(V - 0.22)**2 / (2*0.02**2))
    j_hupd = np.concatenate([j_hupd_a[:n], j_hupd_c[n:]])
    # Pt oxide (above 0.8V)
    j_oxide_a = 0.020 * np.where(V > 0.8, (V - 0.8)**1.5, 0)
    j_oxide_c = -0.100 * np.exp(-(V - 0.75)**2 / (2*0.04**2))
    j_oxide = np.concatenate([j_oxide_a[:n], j_oxide_c[n:]])

    j = j_xover_true + V / R_short + j_cap + j_hupd + j_oxide + np.random.normal(0, 2e-4, len(V))

    r = analyze_crossover(V, j, geo_area=5.0, membrane_thickness_um=25.0)
    print_results(r)
    p = os.path.join(save_dir, 'crossover_demo.png') if save_dir else None
    plot_crossover(r, save_path=p)
    if save_dir:
        save_crossover_excel(r, os.path.join(save_dir, 'crossover_demo.xlsx'), label='Demo')
    if not save_dir: plt.show()
    plt.close()


def run_batch(filepaths, labels, geo_area, delimiter, skip, v_col, j_col,
              current_is_total, avg_V_min, avg_V_max,
              membrane_thickness_um=None, p_h2_kPa=101.325,
              mode_col=None, mode_exclude=None, save_dir=None, image_ext='png'):
    if save_dir: os.makedirs(save_dir, exist_ok=True)
    all_results = []
    print(f'\n  Processing {len(filepaths)} files...\n')
    for i, (fp, lbl) in enumerate(zip(filepaths, labels)):
        print(f'  [{i+1}/{len(filepaths)}] {lbl}')
        try:
            fcd = parse_fcd_header(fp)
            f_skip = fcd['skip'] if fcd else skip
            f_v = fcd.get('v_col', v_col) if fcd else v_col
            f_j = fcd.get('j_col', j_col) if fcd else j_col
            f_mode = fcd.get('mode_col', mode_col) if fcd else mode_col
            V, j = load_cv_data(fp, v_col=f_v, j_col=f_j, delimiter=delimiter,
                                skip_header=f_skip, current_is_total=current_is_total,
                                geo_area=geo_area, mode_col=f_mode, mode_exclude=mode_exclude)
            r = analyze_crossover(V, j, avg_V_min=avg_V_min, avg_V_max=avg_V_max,
                                  geo_area=geo_area, membrane_thickness_um=membrane_thickness_um,
                                  p_h2_anode_kPa=p_h2_kPa)
            r['label'] = lbl; all_results.append(r)
            print(f'         j_xover={r["j_xover_mA_cm2"]:.3f} mA/cm2  '
                  f'H2 flux={r["J_H2_nmol_cm2_s"]:.1f} nmol/cm2/s')
            if save_dir and image_ext:
                safe = lbl.replace(' ','_').replace('/','_').replace('\\','_')
                plot_crossover(r, save_path=os.path.join(save_dir, f'crossover_{safe}.{image_ext}'))
                plt.close()
        except Exception as e:
            print(f'         ERROR: {e}')
    if not all_results: print('\n  No files processed.'); return []
    if save_dir:
        save_batch_crossover_excel(all_results, os.path.join(save_dir, 'crossover_batch_data.xlsx'))
    print(f'\n  {"=" * 65}')
    print(f'  {"Label":30s} {"j_xover":>10s} {"H2 flux":>15s}')
    print(f'  {"":30s} {"(mA/cm2)":>10s} {"(nmol/cm2/s)":>15s}')
    print(f'  {"-" * 65}')
    for r in all_results:
        print(f'  {r["label"]:30s} {r["j_xover_mA_cm2"]:10.3f} '
              f'{r["J_H2_nmol_cm2_s"]:15.2f}')
    print(f'  {"=" * 65}')
    return all_results


def run_interactive():
    print('\n' + chr(9619) * 60)
    print('  H2 CROSSOVER ANALYSIS (CV METHOD)')
    print(chr(9619) * 60)
    print('\n  -- Analysis Mode --')
    print('    1 = Single file')
    print('    2 = Batch (folder)')
    print('    3 = Batch (file list)')
    print('    4 = Demo')
    mode = int(_prompt('Mode', default=1, cast=int))
    if mode == 4:
        sd = _clean_path(_prompt('Save directory (Enter to show)', default=None, cast=None))
        if sd: os.makedirs(sd, exist_ok=True)
        run_demo(save_dir=sd); return

    print('\n  -- Test Stand --')
    print('    0 = Scribner'); print('    1 = FCTS')
    stand = int(_prompt('Test stand', default=0, cast=int))
    PRESETS = {
        0: {'name':'Scribner','delimiter':'\t','skip':51,'v_col':5,'j_col':1,
            'current_is_total':True,'mode_col':28,'mode_exclude':{5}},
        1: {'name':'FCTS','delimiter':',','skip':1,'v_col':0,'j_col':1,
            'current_is_total':False,'mode_col':None,'mode_exclude':None},
    }
    if stand in PRESETS:
        p = PRESETS[stand]
        delimiter,skip,v_col,j_col = p['delimiter'],p['skip'],p['v_col'],p['j_col']
        current_is_total = p['current_is_total']
        mode_col,mode_exclude = p.get('mode_col'),p.get('mode_exclude')
        print(f'  -> {p["name"]} preset loaded')
    else:
        delimiter = {1:',',2:'\t',3:';'}[int(_prompt('Delimiter: 1=comma 2=tab 3=semicolon',default=1,cast=int))]
        skip = int(_prompt('Header rows to skip',default=1,cast=int))
        v_col = int(_prompt('Voltage column (0-based)',default=0,cast=int))
        j_col = int(_prompt('Current column (0-based)',default=1,cast=int))
        current_is_total = int(_prompt('1=j (A/cm2) 2=I (A)',default=1,cast=int)) == 2
        mode_col,mode_exclude = None,None

    geo_area = _prompt('Geometric area (cm2)', default=5.0)
    avg_V_min = _prompt('Averaging V_min (V)', default=0.35)
    avg_V_max = _prompt('Averaging V_max (V)', default=0.50)
    print('\n  -- Membrane (optional) --')
    t_raw = _prompt('Membrane thickness (um, Enter to skip)', default=None, cast=None)
    membrane_thickness = float(t_raw) if t_raw else None
    p_h2 = _prompt('H2 partial pressure (kPa)', default=101.325)

    if mode == 1:
        filepath = _clean_path(_prompt('\n  File path', cast=None))
        fcd = parse_fcd_header(filepath)
        if fcd:
            skip = fcd['skip']
            v_col = fcd.get('v_col', v_col)
            j_col = fcd.get('j_col', j_col)
            mode_col = fcd.get('mode_col', mode_col)
        V, j = load_cv_data(filepath, v_col=v_col, j_col=j_col, delimiter=delimiter,
                            skip_header=skip, current_is_total=current_is_total,
                            geo_area=geo_area, mode_col=mode_col, mode_exclude=mode_exclude)
        print(f'  Loaded: {len(V)} points')
        sweeps = detect_sweeps(V)
        na = sum(1 for s in sweeps if s['direction']=='anodic')
        nc = sum(1 for s in sweeps if s['direction']=='cathodic')
        print(f'  Sweeps: {na} anodic, {nc} cathodic (using last cycle)')
        r = analyze_crossover(V, j, avg_V_min=avg_V_min, avg_V_max=avg_V_max,
                              geo_area=geo_area, membrane_thickness_um=membrane_thickness,
                              p_h2_anode_kPa=p_h2)
        print_results(r)
        save = _clean_path(_prompt('\n  Save to directory? (path or Enter to show)',default=None,cast=None))
        if save:
            os.makedirs(save, exist_ok=True)
            plot_crossover(r, save_path=os.path.join(save, 'crossover.png'))
            save_crossover_excel(r, os.path.join(save, 'crossover_data.xlsx'),
                                label=os.path.splitext(os.path.basename(filepath))[0])
        else:
            plot_crossover(r); plt.show()

    elif mode in (2, 3):
        save_dir = _clean_path(_prompt('Output directory',default='crossover_batch_output',cast=None))
        filepaths, labels = [], []
        if mode == 2:
            folder = _clean_path(_prompt('Folder path', cast=None))
            for ext in ['*.csv','*.txt','*.tsv','*.fcd','*.CSV','*.TXT','*.TSV','*.FCD']:
                filepaths.extend(glob.glob(os.path.join(folder, ext)))
            filepaths = sorted(set(filepaths))
            keywords = ['CROSSOVER','XOVER','LSV','H2XOVER','H2X','H2_CROSS','PERMEATION','H2N2','CV-2MVS']
            filepaths = [fp for fp in filepaths
                         if any(kw in os.path.basename(fp).upper() for kw in keywords)]
            if not filepaths:
                print(f'  No crossover files found. (keywords: {", ".join(keywords)})'); return
            print(f'  Found {len(filepaths)} files:')
            for fp in filepaths:
                name = os.path.splitext(os.path.basename(fp))[0]; labels.append(name)
                print(f'    {name}')
        elif mode == 3:
            print('  Paste file paths (one per line, empty to finish):')
            while True:
                line = input('    ').strip()
                if not line: break
                fp = _clean_path(line)
                if os.path.isfile(fp):
                    filepaths.append(fp); labels.append(os.path.splitext(os.path.basename(fp))[0])
        if not filepaths: print('  No files.'); return
        run_batch(filepaths, labels, geo_area, delimiter, skip, v_col, j_col,
                  current_is_total, avg_V_min, avg_V_max,
                  membrane_thickness_um=membrane_thickness, p_h2_kPa=p_h2,
                  mode_col=mode_col, mode_exclude=mode_exclude, save_dir=save_dir)


def main():
    parser = argparse.ArgumentParser(description='H2 Crossover Analysis (CV Method)')
    parser.add_argument('--file', type=str)
    parser.add_argument('--demo', action='store_true')
    parser.add_argument('--save-dir', type=str, default=None)
    parser.add_argument('--area', type=float, default=5.0)
    parser.add_argument('--v-min', type=float, default=0.35)
    parser.add_argument('--v-max', type=float, default=0.50)
    parser.add_argument('--thickness', type=float, default=None)
    args = parser.parse_args()
    if args.demo:
        if args.save_dir: os.makedirs(args.save_dir, exist_ok=True)
        run_demo(save_dir=args.save_dir)
    elif args.file:
        V, j = load_cv_data(args.file)
        r = analyze_crossover(V, j, avg_V_min=args.v_min, avg_V_max=args.v_max,
                              geo_area=args.area, membrane_thickness_um=args.thickness)
        print_results(r)
        if args.save_dir:
            os.makedirs(args.save_dir, exist_ok=True)
            plot_crossover(r, save_path=os.path.join(args.save_dir, 'crossover.png'))
            save_crossover_excel(r, os.path.join(args.save_dir, 'crossover_data.xlsx'))
        else:
            plot_crossover(r); plt.show()
    else:
        run_interactive()


if __name__ == '__main__':
    main()
