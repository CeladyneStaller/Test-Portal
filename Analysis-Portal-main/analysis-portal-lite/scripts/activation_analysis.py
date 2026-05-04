#!/usr/bin/env python3
"""
Activation/Break-in Analysis
=============================
Plots current density vs time and voltage vs time for fuel cell activation
and break-in protocols (which include multiple voltage holds and cycling).

File classification:
  - Scribner (.fcd): files containing 'activation' (case-insensitive)
  - FCTS (.csv):     files containing 'breakin', 'break-in' or 'break_in'

Generates:
  - activation_{label}.png — j vs time with V on twin axis (per file)
  - activation_sequential.png — multi-file stitched timeline (if >1 file)
  - activation_data.xlsx — Summary + Data + Sequential sheets
"""

import os
import glob
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from scripts.helpers.plot_compare import save_with_sidecar


# ═══════════════════════════════════════════════════════════════════════
#  Header Parsing
# ═══════════════════════════════════════════════════════════════════════

def parse_fcd_header(filepath):
    """Parse Scribner .fcd header to extract skip count and column indices."""
    if not filepath.lower().endswith('.fcd'):
        return None
    try:
        with open(filepath, 'r', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                lines.append(line)
                if 'End Comments' in line:
                    skip = i + 1
                    break
            else:
                return None
    except Exception:
        return None
    if len(lines) < 2:
        return None
    cols = lines[-2].strip().split('\t')
    result = {'skip': skip, 'delimiter': '\t'}
    for ci, name in enumerate(cols):
        n = name.strip()
        if n == 'Time (Sec)':
            result['time_col'] = ci
        elif n == 'I (A)':
            result['j_col'] = ci
        elif n == 'E_Stack (V)':
            result['v_col'] = ci
        elif n == 'Ctrl_Mode':
            result['mode_col'] = ci
    return result


def parse_csv_header(filepath):
    """Parse FCTS / generic CSV header to find time, current, voltage columns."""
    try:
        with open(filepath, 'r', errors='replace') as f:
            first_line = f.readline().strip()
    except Exception:
        return None

    if '\t' in first_line:
        delimiter = '\t'
    elif ',' in first_line:
        delimiter = ','
    else:
        return None

    cols = [c.strip() for c in first_line.split(delimiter)]
    if not cols:
        return None

    result = {'skip': 1, 'delimiter': delimiter, 'cols': cols}

    cols_lower = [c.lower() for c in cols]

    # Time column
    for ci, c in enumerate(cols_lower):
        if c == 's' or 'time' in c or c == 'sec':
            result['time_col'] = ci
            break

    # Current column — prefer 'mA', then 'a', then 'current'
    for ci, c in enumerate(cols_lower):
        if c == 'ma':
            result['j_col'] = ci
            result['current_unit'] = 'mA'
            break
    if 'j_col' not in result:
        for ci, c in enumerate(cols_lower):
            if c == 'a' or 'current' in c:
                result['j_col'] = ci
                result['current_unit'] = 'mA' if 'ma' in c or '(ma)' in c else 'A'
                break

    # Voltage column
    for ci, c in enumerate(cols_lower):
        if c == 'mv':
            result['v_col'] = ci
            result['voltage_unit'] = 'mV'
            break
    if 'v_col' not in result:
        for ci, c in enumerate(cols_lower):
            if c == 'v' or 'voltage' in c or c.startswith('e_') or c.startswith('working'):
                result['v_col'] = ci
                result['voltage_unit'] = 'mV' if 'mv' in c or '(mv)' in c else 'V'
                break

    if 'time_col' not in result or 'j_col' not in result or 'v_col' not in result:
        return None

    return result


# ═══════════════════════════════════════════════════════════════════════
#  Data Loading
# ═══════════════════════════════════════════════════════════════════════

def load_activation_data(filepath, geo_area=5.0):
    """Load time, current density, voltage from one activation/break-in file."""
    label = os.path.splitext(os.path.basename(filepath))[0]
    is_fcd = filepath.lower().endswith('.fcd')

    if is_fcd:
        hdr = parse_fcd_header(filepath)
        if hdr is None or 'time_col' not in hdr or 'j_col' not in hdr or 'v_col' not in hdr:
            print(f'    Could not parse FCD header for {os.path.basename(filepath)}')
            return None, None, None, label
        try:
            data = np.genfromtxt(filepath, delimiter=hdr['delimiter'],
                                  skip_header=hdr['skip'],
                                  usecols=(hdr['time_col'], hdr['j_col'], hdr['v_col']),
                                  invalid_raise=False)
        except Exception as e:
            print(f'    Error reading {os.path.basename(filepath)}: {e}')
            return None, None, None, label
        if data is None or data.ndim < 2 or len(data) < 2:
            return None, None, None, label
        # Filter out NaN rows
        mask = ~np.any(np.isnan(data), axis=1)
        data = data[mask]
        if len(data) < 2:
            return None, None, None, label
        time = data[:, 0] - data[0, 0]
        # Scribner I (A) → A/cm²
        j = data[:, 1] / geo_area
        voltage = data[:, 2]
        return time, j, voltage, label

    # CSV / FCTS
    hdr = parse_csv_header(filepath)
    if hdr is None:
        print(f'    Could not parse CSV header for {os.path.basename(filepath)}')
        return None, None, None, label
    try:
        data = np.genfromtxt(filepath, delimiter=hdr['delimiter'],
                              skip_header=hdr['skip'],
                              usecols=(hdr['time_col'], hdr['j_col'], hdr['v_col']),
                              invalid_raise=False)
    except Exception as e:
        print(f'    Error reading {os.path.basename(filepath)}: {e}')
        return None, None, None, label
    if data is None or data.ndim < 2 or len(data) < 2:
        return None, None, None, label
    mask = ~np.any(np.isnan(data), axis=1)
    data = data[mask]
    if len(data) < 2:
        return None, None, None, label

    time = data[:, 0] - data[0, 0]
    if hdr.get('current_unit') == 'mA':
        j = data[:, 1] * 0.001 / geo_area
    else:
        j = data[:, 1] / geo_area
    voltage = data[:, 2]
    if hdr.get('voltage_unit') == 'mV':
        voltage = voltage * 0.001

    return time, j, voltage, label


def resample_activation(time, j, voltage, interval_s=60.0):
    """Box-car average raw data onto a uniform interval_s time grid."""
    if interval_s <= 0 or len(time) < 2:
        return time, j, voltage
    raw_interval = np.median(np.diff(time))
    if raw_interval >= interval_s * 0.9:
        return time, j, voltage

    t_grid = np.arange(time[0], time[-1] + interval_s * 0.01, interval_s)
    half_win = interval_s / 2.0
    j_out = np.empty(len(t_grid))
    v_out = np.empty(len(t_grid))
    for i, t in enumerate(t_grid):
        mask = (time >= t - half_win) & (time < t + half_win)
        if mask.sum() > 0:
            j_out[i] = np.mean(j[mask])
            v_out[i] = np.mean(voltage[mask])
        else:
            j_out[i] = np.interp(t, time, j)
            v_out[i] = np.interp(t, time, voltage)
    return t_grid, j_out, v_out


# ═══════════════════════════════════════════════════════════════════════
#  File Classification
# ═══════════════════════════════════════════════════════════════════════

ACTIVATION_KEYWORDS = ['ACTIVATION', 'BREAKIN', 'BREAK-IN', 'BREAK_IN']
EXCLUDE_KEYWORDS = ['FILTERDATA']
MIN_DURATION_MIN = 50  # minimum duration in minutes


def classify_activation_files(folder):
    """Find activation/break-in files in a folder. Recursive."""
    extensions = ['*.csv', '*.txt', '*.tsv', '*.fcd',
                  '*.CSV', '*.TXT', '*.TSV', '*.FCD']
    all_files = []
    for ext in extensions:
        all_files.extend(glob.glob(os.path.join(folder, '**', ext), recursive=True))
    all_files = sorted(set(all_files))

    all_files = [fp for fp in all_files
                 if not any(ex in os.path.basename(fp).upper() for ex in EXCLUDE_KEYWORDS)]

    matched = []
    for fp in all_files:
        name_upper = os.path.basename(fp).upper()
        if any(kw in name_upper for kw in ACTIVATION_KEYWORDS):
            label = os.path.splitext(os.path.basename(fp))[0]
            matched.append((fp, label))
    return matched


# ═══════════════════════════════════════════════════════════════════════
#  Plotting
# ═══════════════════════════════════════════════════════════════════════

def _select_time_unit(duration_s):
    if duration_s > 7200:
        return 1.0/3600, 'Time (hr)'
    elif duration_s > 120:
        return 1.0/60, 'Time (min)'
    else:
        return 1.0, 'Time (s)'


def plot_activation(time, j, voltage, label, save_path=None):
    """Plot j vs time on left axis, V on right twin axis."""
    fig, ax1 = plt.subplots(figsize=(10, 5.5))
    duration = time[-1] - time[0]
    scale, t_label = _select_time_unit(duration)
    t_plot = time * scale

    color_j = 'tab:blue'
    ax1.plot(t_plot, j, '-', color=color_j, lw=1.4, label='Current density')
    ax1.set_xlabel(t_label)
    ax1.set_ylabel('Current density (A/cm²)', color=color_j)
    ax1.tick_params(axis='y', labelcolor=color_j)
    ax1.grid(True, alpha=0.3)

    ax2 = ax1.twinx()
    color_v = 'tab:red'
    ax2.plot(t_plot, voltage, '-', color=color_v, lw=1.0, alpha=0.7, label='Voltage')
    ax2.set_ylabel('Voltage (V)', color=color_v)
    ax2.tick_params(axis='y', labelcolor=color_v)

    j_max_idx = int(np.argmax(j))
    duration_min = duration / 60.0
    readout = (
        f'Duration = {duration_min:.1f} min\n'
        f'V_start = {voltage[0]:.3f} V\n'
        f'V_end = {voltage[-1]:.3f} V\n'
        f'j_max = {j[j_max_idx]:.4f} A/cm²\n'
        f'V @ j_max = {voltage[j_max_idx]:.3f} V'
    )
    ax1.text(0.02, 0.98, readout, transform=ax1.transAxes,
             fontsize=8, va='top', ha='left',
             bbox=dict(boxstyle='round,pad=0.4', fc='lightyellow', alpha=0.9))

    ax1.set_title(f'Activation — {label}', fontsize=11, fontweight='bold')

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, fontsize=9, loc='lower right')

    fig.tight_layout()
    if save_path:
        save_with_sidecar(fig, save_path, plot_type='activation', dpi=150)
        print(f'    Saved: {save_path}')
    return fig


def plot_activation_sequential(datasets, save_path=None):
    """Stitch multiple activation files end-to-end on one timeline."""
    fig, ax1 = plt.subplots(figsize=(12, 6))

    t_all, j_all, v_all = [], [], []
    boundaries = [0.0]
    t_offset = 0.0
    for time, j, voltage, label in datasets:
        t_all.append(time + t_offset)
        j_all.append(j)
        v_all.append(voltage)
        t_offset += time[-1]
        boundaries.append(t_offset)

    t_stitched = np.concatenate(t_all)
    j_stitched = np.concatenate(j_all)
    v_stitched = np.concatenate(v_all)

    scale, t_label = _select_time_unit(t_offset)
    t_plot = t_stitched * scale

    color_j = 'tab:blue'
    ax1.plot(t_plot, j_stitched, '-', color=color_j, lw=1.2)
    ax1.set_xlabel(t_label)
    ax1.set_ylabel('Current density (A/cm²)', color=color_j)
    ax1.tick_params(axis='y', labelcolor=color_j)
    ax1.grid(True, alpha=0.3)

    ax2 = ax1.twinx()
    color_v = 'tab:red'
    ax2.plot(t_plot, v_stitched, '-', color=color_v, lw=0.9, alpha=0.7)
    ax2.set_ylabel('Voltage (V)', color=color_v)
    ax2.tick_params(axis='y', labelcolor=color_v)

    y_top = ax1.get_ylim()[1]
    for i, (b_start, (_, _, _, lbl)) in enumerate(zip(boundaries[:-1], datasets)):
        ax1.axvline(b_start * scale, color='gray', ls='--', lw=0.8, alpha=0.6)
        b_end = boundaries[i + 1]
        b_mid = (b_start + b_end) / 2.0 * scale
        ax1.text(b_mid, y_top * 0.97, lbl, fontsize=8, ha='center', va='top',
                 color='black',
                 bbox=dict(boxstyle='round,pad=0.2', facecolor='white',
                           edgecolor='gray', alpha=0.8))
    ax1.axvline(boundaries[-1] * scale, color='gray', ls='--', lw=0.8, alpha=0.6)

    total_dur_hr = t_offset / 3600.0
    j_max_idx = int(np.argmax(j_stitched))
    readout = (
        f'Files = {len(datasets)}\n'
        f'Total duration = {total_dur_hr:.2f} hr\n'
        f'V_start = {v_stitched[0]:.3f} V\n'
        f'V_end = {v_stitched[-1]:.3f} V\n'
        f'j_max = {j_stitched[j_max_idx]:.4f} A/cm²'
    )
    ax1.text(0.02, 0.98, readout, transform=ax1.transAxes,
             fontsize=8, va='top', ha='left',
             bbox=dict(boxstyle='round,pad=0.4', fc='lightyellow', alpha=0.9))

    ax1.set_title(f'Activation — {len(datasets)} files stitched sequentially',
                   fontsize=11, fontweight='bold')

    fig.tight_layout()
    if save_path:
        save_with_sidecar(fig, save_path, plot_type='activation_sequential', dpi=150)
        print(f'    Saved: {save_path}')
    return fig


# ═══════════════════════════════════════════════════════════════════════
#  Excel Output
# ═══════════════════════════════════════════════════════════════════════

def _save_activation_excel(datasets, interval_s, geo_area, filepath):
    """Write activation data to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hf = Font(bold=True)
    hfill = PatternFill('solid', fgColor='D9E1F2')
    label_fill = PatternFill('solid', fgColor='E2EFDA')

    ws = wb.active
    ws.title = 'Summary'
    headers = ['Label', 'Duration (min)', 'V_start (V)', 'V_end (V)',
               'j_start (A/cm²)', 'j_end (A/cm²)', 'j_max (A/cm²)',
               'V at j_max (V)', 'Resampling (s)', 'Geo area (cm²)', 'Points']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(c)].width = max(len(h) + 4, 14)

    for i, (time, j, voltage, lbl) in enumerate(datasets, 2):
        j_max_idx = int(np.argmax(j))
        ws.cell(row=i, column=1, value=lbl)
        ws.cell(row=i, column=2, value=round(time[-1] / 60.0, 1))
        ws.cell(row=i, column=3, value=round(float(voltage[0]), 6))
        ws.cell(row=i, column=4, value=round(float(voltage[-1]), 6))
        ws.cell(row=i, column=5, value=round(float(j[0]), 6))
        ws.cell(row=i, column=6, value=round(float(j[-1]), 6))
        ws.cell(row=i, column=7, value=round(float(j[j_max_idx]), 6))
        ws.cell(row=i, column=8, value=round(float(voltage[j_max_idx]), 6))
        ws.cell(row=i, column=9, value=interval_s)
        ws.cell(row=i, column=10, value=geo_area)
        ws.cell(row=i, column=11, value=len(time))
    ws.column_dimensions['A'].width = 40

    ws2 = wb.create_sheet('Data')
    col = 1
    for time, j, voltage, lbl in datasets:
        col_end = col + 2
        for c in range(col, col_end + 1):
            cell = ws2.cell(row=1, column=c, value=lbl if c == col else '')
            cell.font = hf
            cell.fill = label_fill
        ws2.merge_cells(start_row=1, start_column=col,
                        end_row=1, end_column=col_end)

        hdrs = ['Time (min)', 'Current density (A/cm²)', 'Voltage (V)']
        for ci, h in enumerate(hdrs):
            cell = ws2.cell(row=2, column=col + ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            ws2.column_dimensions[get_column_letter(col + ci)].width = 22

        for ri in range(len(time)):
            ws2.cell(row=ri + 3, column=col,
                     value=round(float(time[ri] / 60.0), 4))
            ws2.cell(row=ri + 3, column=col + 1,
                     value=round(float(j[ri]), 6))
            ws2.cell(row=ri + 3, column=col + 2,
                     value=round(float(voltage[ri]), 6))
        col = col_end + 2

    if len(datasets) > 1:
        ws3 = wb.create_sheet('Sequential')
        hdrs = ['Time (min)', 'Current density (A/cm²)', 'Voltage (V)', 'Source file']
        for ci, h in enumerate(hdrs, 1):
            cell = ws3.cell(row=1, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            ws3.column_dimensions[get_column_letter(ci)].width = max(len(h) + 4, 16)
        ws3.column_dimensions['D'].width = 40

        row = 2
        t_offset = 0.0
        for time, j, voltage, lbl in datasets:
            for ri in range(len(time)):
                ws3.cell(row=row, column=1,
                         value=round(float((time[ri] + t_offset) / 60.0), 4))
                ws3.cell(row=row, column=2, value=round(float(j[ri]), 6))
                ws3.cell(row=row, column=3, value=round(float(voltage[ri]), 6))
                ws3.cell(row=row, column=4, value=lbl)
                row += 1
            t_offset += time[-1]

    wb.save(filepath)
    print(f'    Excel: {filepath}')


# ═══════════════════════════════════════════════════════════════════════
#  Batch Processing
# ═══════════════════════════════════════════════════════════════════════

def run_batch(filepaths, labels, save_dir=None, interval_s=60.0, geo_area=5.0,
              image_ext='png'):
    """Load, plot, and summarize multiple activation files."""
    if save_dir:
        os.makedirs(save_dir, exist_ok=True)

    pairs = sorted(zip(filepaths, labels), key=lambda p: os.path.basename(p[0]).lower())
    filepaths = [p[0] for p in pairs]
    labels = [p[1] for p in pairs]

    datasets = []
    print(f'\n  Processing {len(filepaths)} files (sorted by filename)...')
    print(f'  Resampling interval: {interval_s:.0f} s')
    print(f'  Geometric area: {geo_area:.2f} cm²')

    for i, (fp, lbl) in enumerate(zip(filepaths, labels)):
        print(f'\n  [{i+1}/{len(filepaths)}] {lbl}')
        time, j, voltage, label = load_activation_data(fp, geo_area=geo_area)
        if time is None or len(time) < 2:
            print(f'    Skipped: no valid data')
            continue

        duration = time[-1] - time[0]
        duration_min = duration / 60.0
        if duration_min < MIN_DURATION_MIN:
            print(f'    Skipped: {duration_min:.1f} min (minimum {MIN_DURATION_MIN} min)')
            continue

        n_raw = len(time)
        time, j, voltage = resample_activation(time, j, voltage, interval_s=interval_s)
        if duration > 3600:
            dur_str = f'{duration/3600:.1f} hr'
        elif duration > 60:
            dur_str = f'{duration/60:.1f} min'
        else:
            dur_str = f'{duration:.0f} s'
        print(f'    {n_raw} raw -> {len(time)} pts ({interval_s:.0f}s), {dur_str}, '
              f'V: {voltage[0]:.3f}-{voltage[-1]:.3f} V, j_max: {j.max():.3f} A/cm²')
        datasets.append((time, j, voltage, lbl))

        if save_dir:
            safe = lbl.replace(' ', '_').replace('/', '-').replace('\\', '-')
            if len(safe) > 80:
                safe = safe[:80]
            try:
                plot_activation(time, j, voltage, lbl,
                                save_path=os.path.join(save_dir, f'activation_{safe}.{image_ext}'))
            except Exception as e:
                print(f'    Plot failed: {e}')
            plt.close('all')

    if not datasets:
        print('\n  No valid files processed.')
        return []

    if len(datasets) > 1 and save_dir:
        try:
            plot_activation_sequential(
                datasets,
                save_path=os.path.join(save_dir, f'activation_sequential.{image_ext}'))
        except Exception as e:
            print(f'  Sequential plot failed: {e}')
        plt.close('all')

    if save_dir:
        try:
            _save_activation_excel(datasets, interval_s, geo_area,
                                   os.path.join(save_dir, 'activation_data.xlsx'))
        except Exception as e:
            print(f'  Excel export failed: {e}')

    return datasets


# ═══════════════════════════════════════════════════════════════════════
#  Portal Entry Point
# ═══════════════════════════════════════════════════════════════════════

def run(input_dir: str, output_dir: str, params: dict = None) -> dict:
    """Portal entry point: scan input_dir for activation files, run batch."""
    from pathlib import Path
    p = params or {}

    from scripts.helpers.conditions import img_ext_from_params
    image_ext = img_ext_from_params(p)

    matched = classify_activation_files(input_dir)
    if not matched:
        raise RuntimeError(
            f"No activation/break-in files found in {input_dir}. "
            f"Filenames must contain one of: "
            f"{', '.join(ACTIVATION_KEYWORDS)} (case-insensitive)."
        )

    filepaths = [fp for fp, _ in matched]
    labels = [lbl for _, lbl in matched]

    datasets = run_batch(
        filepaths, labels,
        save_dir=str(output_dir),
        interval_s=float(p.get('interval_s', 60.0)),
        geo_area=float(p.get('geo_area', 5.0)),
        image_ext=image_ext,
    )

    output_files = [str(f.relative_to(Path(output_dir)))
                    for f in Path(output_dir).rglob('*') if f.is_file()]

    if not output_files:
        raise RuntimeError(
            f"Activation analysis produced no output. "
            f"Found {len(matched)} candidate file(s) but all were skipped "
            f"(possibly under {MIN_DURATION_MIN} min duration or unparseable)."
        )

    return {
        'status': 'success',
        'files_processed': len(datasets),
        'files_produced': output_files,
    }
