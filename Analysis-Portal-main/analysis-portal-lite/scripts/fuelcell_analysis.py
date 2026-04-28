#!/usr/bin/env python3
"""
Fuel Cell Data Analysis — Batch Orchestrator
=============================================
Scans a folder of fuel cell test data and automatically runs:
  - ECSA analysis (H_UPD / CO stripping)
  - EIS analysis (impedance spectroscopy)
  - H₂ crossover analysis (slow-scan CV)
  - Polarization curve analysis

Each file is classified by filename keywords, then routed to the
appropriate analysis script using Scribner or FCTS presets.

Usage:
  python fuelcell_analysis.py                      # interactive
  python fuelcell_analysis.py --folder /path/to/data --save-dir results
"""

import os, sys, glob, argparse, time
import matplotlib
matplotlib.use("Agg")


def run(input_dir: str, output_dir: str, params: dict = None) -> dict:
    """Batch orchestrator: classify and analyze all fuel cell data in a folder."""
    p = params or {}
    from scripts.helpers.conditions import img_ext_from_params
    image_ext = img_ext_from_params(p)

    info = run_all(input_dir, output_dir,
            geo_area=float(p.get('geo_area', 5.0)),
            loading=float(p.get('loading', 0.2)),
            membrane_thickness=None,
            stand=int(p.get('stand', 0)),
            ocv_interval_s=float(p.get('interval_s', 60.0)),
            image_ext=image_ext)

    from pathlib import Path
    output_files = [str(f.relative_to(Path(output_dir)))
                    for f in Path(output_dir).rglob('*') if f.is_file()]

    if not output_files:
        # Build detailed error message
        info = info or {}
        classified = info.get('classified', {})
        errors = info.get('errors', {})
        total = info.get('total', 0)

        parts = [f"No output produced."]

        if total == 0:
            inp = Path(input_dir)
            data_files = [f for f in inp.rglob('*')
                          if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv', '.fcd')]
            parts.append(f"Found {len(data_files)} data file(s) but none matched "
                         f"analysis keywords (ECSA, EIS, H2X, PolarizationCurve, etc.)")
        else:
            parts.append(f"Classified {total} file(s): {classified}.")
            if errors:
                for label, tb in errors.items():
                    # Get last line of traceback (the actual error)
                    last_line = tb.strip().split('\n')[-1]
                    parts.append(f"{label} FAILED: {last_line}")

        raise RuntimeError('\n'.join(parts))

    return {"status": "success", "files_produced": output_files}

# ═══════════════════════════════════════════════════════════════════════
#  File Classification
# ═══════════════════════════════════════════════════════════════════════

ANALYSIS_TYPES = {
    'ecsa': {
        'keywords': ['ECSA', 'CV-50MVS'],
        'label': 'ECSA',
    },
    'eis': {
        'keywords': ['EIS'],
        'label': 'EIS',
    },
    'crossover': {
        'keywords': ['CV-2MVS', 'CROSSOVER', 'XOVER', 'H2XOVER', 'H2X',
                     'H2_CROSS', 'PERMEATION'],
        'label': 'H₂ Crossover',
    },
    'polcurve': {
        'keywords': ['POLCURVE', 'POL_', 'POL-', 'IV_', 'IV-',
                     'POLARIZATION', 'POLDATA'],
        'label': 'Pol Curve',
    },
    'ocv': {
        'keywords': ['OCV', 'PURGE'],
        'label': 'OCV',
    },
}

# Classification priority (first match wins — OCV last since it's broadest)
CLASSIFICATION_ORDER = ['ecsa', 'crossover', 'eis', 'polcurve', 'ocv']


def parse_fcd_header(filepath):
    """Parse Scribner .fcd header for skip count and column indices."""
    if not filepath.lower().endswith('.fcd'):
        return None
    try:
        with open(filepath, 'r', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                lines.append(line)
                if 'End Comments' in line:
                    skip = i + 1; break
            else:
                return None
    except Exception:
        return None
    if len(lines) < 2:
        return None
    cols = lines[-2].strip().split('\t')
    result = {'skip': skip}
    cond_cols = {}
    for ci, name in enumerate(cols):
        n = name.strip()
        if n == 'I (A)': result['j_col'] = ci; result['i_col'] = ci
        elif n == 'E_Stack (V)': result['v_col'] = ci
        elif n == 'HFR (mOhm)': result['hfr_col'] = ci
        elif n.startswith('Z_Freq'): result['freq_col'] = ci
        elif n.startswith('Z_Real'): result['zreal_col'] = ci
        elif n.startswith('Z_Imag'): result['zimag_col'] = ci
        elif n == 'Ctrl_Mode': result['mode_col'] = ci
        elif n == 'Cell (C)': cond_cols['T_cell (C)'] = ci
        elif n.startswith('Temp_Anode'): cond_cols['T_anode_dp (C)'] = ci
        elif n.startswith('Flow_Anode'): cond_cols['H2_flow (slpm)'] = ci
        elif n.startswith('Temp_Cathode'): cond_cols['T_cathode_dp (C)'] = ci
        elif n.startswith('Flow_Cathode'): cond_cols['Air_flow (slpm)'] = ci
    if cond_cols: result['condition_cols'] = cond_cols
    return result


def classify_files(folder):
    """
    Scan folder (recursively) and classify files by analysis type.

    Returns dict: {type_name: [(filepath, label), ...]}
    """
    from pathlib import Path
    p = Path(folder)
    all_files = sorted(set(
        str(f) for f in p.rglob('*')
        if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv', '.fcd')
    ))

    # Exclude filter/diagnostic files
    all_files = [fp for fp in all_files
                 if 'FILTERDATA' not in os.path.basename(fp).upper()]

    classified = {t: [] for t in ANALYSIS_TYPES}
    unclassified = []
    used = set()

    for atype in CLASSIFICATION_ORDER:
        keywords = ANALYSIS_TYPES[atype]['keywords']
        for fp in all_files:
            if fp in used:
                continue
            name_upper = os.path.basename(fp).upper()
            if any(kw in name_upper for kw in keywords):
                label = os.path.splitext(os.path.basename(fp))[0]
                classified[atype].append((fp, label))
                used.add(fp)

    for fp in all_files:
        if fp not in used:
            unclassified.append(fp)

    return classified, unclassified


def _clean_path(p):
    if p is None:
        return None
    p = p.strip()
    if p.startswith('& '):
        p = p[2:]
    return p.strip().strip('"').strip("'").strip('\u2018\u2019\u201c\u201d\u202a\u200b')


def _prompt(msg, default=None, cast=float):
    tail = f' [{default}]' if default is not None else ''
    raw = input(f'  {msg}{tail}: ').strip()
    if not raw:
        return default
    return raw if cast is None else cast(raw)


# ═══════════════════════════════════════════════════════════════════════
#  Analysis Runners
# ═══════════════════════════════════════════════════════════════════════

def run_ecsa_batch(files, geo_area, loading, save_dir, stand=0, image_ext="png"):
    """Run ECSA analysis on classified files. Returns list of result dicts."""
    try:
        from scripts.ecsa_analysis import run_batch
    except ImportError:
        print('    ERROR: ecsa_analysis.py not found in current directory')
        return []

    filepaths = [f[0] for f in files]
    labels = [f[1] for f in files]

    PRESETS = {
        0: {'delimiter': '\t', 'skip': 76, 'v_col': 5, 'i_col': 1,
            'v_low': 0.08, 'v_high': 0.40, 'i_scale': 1.0},
        1: {'delimiter': ',', 'skip': 1, 'v_col': 2, 'i_col': 3,
            'v_low': 0.08, 'v_high': 0.40, 'i_scale': 1.0},
    }
    p = PRESETS[stand]
    fcd = parse_fcd_header(filepaths[0])
    skip = fcd['skip'] if fcd else p['skip']
    v_col = fcd.get('v_col', p['v_col']) if fcd else p['v_col']
    i_col = fcd.get('i_col', p['i_col']) if fcd else p['i_col']
    scan_rate = 0.050

    return run_batch(filepaths, labels, scan_rate=scan_rate, geo_area=geo_area,
                     loading=loading, delimiter=p['delimiter'], skip=skip,
                     v_col=v_col, i_col=i_col,
                     v_low=p['v_low'], v_high=p['v_high'],
                     i_scale=p['i_scale'], cycle=2, save_dir=save_dir, image_ext=image_ext) or []


def run_eis_batch(files, geo_area, save_dir, stand=0, image_ext="png"):
    """Run EIS analysis on classified files. Returns list of result dicts."""
    try:
        from scripts.eis_analysis import run_batch
    except ImportError:
        print('    ERROR: eis_analysis.py not found in current directory')
        return []

    filepaths = [f[0] for f in files]
    labels = [f[1] for f in files]

    PRESETS = {
        0: {'delimiter': '\t', 'skip': 59, 'freq_col': 22,
            'zreal_col': 23, 'zimag_col': 24, 'zimag_sign': 'negative'},
        1: {'delimiter': ',', 'skip': 1, 'freq_col': 3,
            'zreal_col': 6, 'zimag_col': 7, 'zimag_sign': 'negative'},
    }
    p = PRESETS[stand]
    fcd = parse_fcd_header(filepaths[0])
    skip = fcd['skip'] if fcd else p['skip']
    freq_col = fcd.get('freq_col', p['freq_col']) if fcd else p['freq_col']
    zreal_col = fcd.get('zreal_col', p['zreal_col']) if fcd else p['zreal_col']
    zimag_col = fcd.get('zimag_col', p['zimag_col']) if fcd else p['zimag_col']

    return run_batch(filepaths, labels, model_name='R-RC-RC', geo_area=geo_area,
                     delimiter=p['delimiter'], skip=skip,
                     freq_col=freq_col, zreal_col=zreal_col,
                     zimag_col=zimag_col, zimag_sign=p['zimag_sign'],
                     save_dir=save_dir, image_ext=image_ext) or []


def run_crossover_batch(files, geo_area, membrane_thickness, save_dir, stand=0, image_ext="png"):
    """Run H₂ crossover analysis on classified files. Returns list of result dicts."""
    try:
        from scripts.h2_crossover_analysis import run_batch
    except ImportError:
        print('    ERROR: h2_crossover_analysis.py not found in current directory')
        return []

    filepaths = [f[0] for f in files]
    labels = [f[1] for f in files]

    PRESETS = {
        0: {'delimiter': '\t', 'skip': 51, 'v_col': 5, 'j_col': 1,
            'current_is_total': True, 'mode_col': 28, 'mode_exclude': {5}},
        1: {'delimiter': ',', 'skip': 1, 'v_col': 2, 'j_col': 3,
            'current_is_total': True, 'mode_col': None, 'mode_exclude': None},
    }
    p = PRESETS[stand]
    fcd = parse_fcd_header(filepaths[0])
    skip = fcd['skip'] if fcd else p['skip']
    v_col = fcd.get('v_col', p['v_col']) if fcd else p['v_col']
    j_col = fcd.get('j_col', p['j_col']) if fcd else p['j_col']
    mode_col = fcd.get('mode_col', p.get('mode_col')) if fcd else p.get('mode_col')

    return run_batch(filepaths, labels, geo_area=geo_area,
                     delimiter=p['delimiter'], skip=skip,
                     v_col=v_col, j_col=j_col,
                     current_is_total=p['current_is_total'],
                     avg_V_min=0.35, avg_V_max=0.50,
                     membrane_thickness_um=membrane_thickness,
                     mode_col=p.get('mode_col'), mode_exclude=p.get('mode_exclude'),
                     save_dir=save_dir, image_ext=image_ext) or []


def run_polcurve_batch(files, geo_area, save_dir, stand=0, image_ext="png"):
    """Run polarization curve analysis on classified files. Returns list of result dicts."""
    try:
        from scripts.polcurve_analysis import run_batch
    except ImportError:
        print('    ERROR: polcurve_analysis.py not found in current directory')
        return []

    filepaths = [f[0] for f in files]
    labels = [f[1] for f in files]

    PRESETS = {
        0: {'delimiter': '\t', 'skip': 51, 'j_col': 1, 'v_col': 5,
            'hfr_col': 20, 'current_is_total': True, 'hfr_scale': 0.001,
            'mode_col': 28, 'mode_exclude': {5},
            'j_scale': 1.0, 'v_scale': 1.0,
            'condition_cols': {
                'T_cell (C)': 13, 'T_anode_dp (C)': 14, 'H2_flow (slpm)': 15,
                'T_cathode_dp (C)': 17, 'Air_flow (slpm)': 18,
            }},
        1: {'delimiter': ',', 'skip': 1, 'j_col': 2, 'v_col': 1,
            'hfr_col': None, 'current_is_total': True, 'hfr_scale': 1.0,
            'mode_col': None, 'mode_exclude': None,
            'condition_cols': None, 'j_scale': 0.001, 'v_scale': 0.001},
    }
    p = PRESETS[stand]
    fcd = parse_fcd_header(filepaths[0])
    skip = fcd['skip'] if fcd else p['skip']
    j_col = fcd.get('j_col', p['j_col']) if fcd else p['j_col']
    v_col = fcd.get('v_col', p['v_col']) if fcd else p['v_col']
    hfr_col = fcd.get('hfr_col', p['hfr_col']) if fcd else p['hfr_col']
    mode_col = fcd.get('mode_col', p.get('mode_col')) if fcd else p.get('mode_col')
    cond_cols = fcd.get('condition_cols', p['condition_cols']) if fcd else p['condition_cols']

    return run_batch(filepaths, labels, geo_area=geo_area,
                     delimiter=p['delimiter'], skip=skip,
                     j_col=j_col, v_col=v_col, hfr_col=hfr_col,
                     current_is_total=p['current_is_total'],
                     tafel_j_min=0.01, tafel_j_max=0.10,
                     condition_cols=cond_cols,
                     hfr_scale=p['hfr_scale'],
                     mode_col=mode_col, mode_exclude=p.get('mode_exclude'),
                     j_scale=p.get('j_scale', 1.0), v_scale=p.get('v_scale', 1.0),
                     save_dir=save_dir, image_ext=image_ext) or []


def run_ocv_batch(files, save_dir, interval_s=60.0, image_ext="png"):
    """Run OCV analysis on classified files. Returns list of datasets."""
    try:
        from scripts.ocv_analysis import run_batch
    except ImportError:
        print('    ERROR: ocv_analysis.py not found')
        return []

    filepaths = [f[0] for f in files]
    labels = [f[1] for f in files]

    return run_batch(filepaths, labels, save_dir=save_dir,
                     interval_s=interval_s, image_ext=image_ext) or []


# ═══════════════════════════════════════════════════════════════════════
#  Consolidated Excel
# ═══════════════════════════════════════════════════════════════════════

def save_consolidated_excel(results, filepath, geo_area=5.0):
    """
    Write a single Excel workbook summarizing all analyses.
    Summary sheet + one data sheet per analysis type (all samples on same tab).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    import numpy as np

    wb = Workbook()
    hf = Font(bold=True)
    hfill = PatternFill('solid', fgColor='D9E1F2')
    label_fill = PatternFill('solid', fgColor='E2EFDA')
    wb.remove(wb.active)

    def _write_header(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(c)].width = max(len(str(h)) + 4, 14)

    def _write_val(ws, row, col, val):
        if val is not None:
            ws.cell(row=row, column=col, value=val)

    def _write_label_row(ws, row, col_start, col_end, label):
        """Write a merged label row spanning columns."""
        cell = ws.cell(row=row, column=col_start, value=label)
        cell.font = hf
        cell.fill = label_fill
        if col_end > col_start:
            ws.merge_cells(start_row=row, start_column=col_start,
                           end_row=row, end_column=col_end)

    # ══════════════════════════════════════════════════════════════
    #  ECSA
    # ══════════════════════════════════════════════════════════════
    ecsa = results.get('ecsa', [])
    if ecsa:
        # Summary
        ws = wb.create_sheet('ECSA Summary')
        headers = ['Label', 'Avg ECSA (cm²)', 'Avg ECSA (m²/g)', 'Avg RF',
                   'Anodic ECSA (cm²)', 'Cathodic ECSA (cm²)',
                   'Anodic Q (mC/cm²)', 'Cathodic Q (mC/cm²)']
        _write_header(ws, headers)
        for i, r in enumerate(ecsa, 2):
            _write_val(ws, i, 1, r.get('label', ''))
            _write_val(ws, i, 2, r.get('average_ECSA_cm2'))
            _write_val(ws, i, 3, r.get('average_ECSA_m2_per_g'))
            _write_val(ws, i, 4, r.get('average_RF'))
            _write_val(ws, i, 5, r.get('anodic_ECSA_cm2'))
            _write_val(ws, i, 6, r.get('cathodic_ECSA_cm2'))
            _write_val(ws, i, 7, r.get('anodic_Q_mC_cm2'))
            _write_val(ws, i, 8, r.get('cathodic_Q_mC_cm2'))

        # Data: all CV cycles for all samples, side by side
        ws2 = wb.create_sheet('ECSA Data')
        col = 1
        for r in ecsa:
            label = r.get('label', 'data')
            cycles = r.get('_cycles', [])
            if not cycles:
                continue
            n_cyc = len(cycles)
            # Label row spanning all columns for this sample
            col_end = col + n_cyc * 2 - 1
            _write_label_row(ws2, 1, col, col_end, label)
            # Sub-headers: V/I per cycle
            for ci, (V_c, I_c) in enumerate(cycles):
                c1 = col + ci * 2
                h1 = ws2.cell(row=2, column=c1, value=f'V cyc{ci+1} (V)')
                h2 = ws2.cell(row=2, column=c1 + 1, value=f'I cyc{ci+1} (A)')
                for h in [h1, h2]:
                    h.font = hf; h.fill = hfill; h.alignment = Alignment(horizontal='center')
                ws2.column_dimensions[get_column_letter(c1)].width = 14
                ws2.column_dimensions[get_column_letter(c1 + 1)].width = 14
                for ri in range(len(V_c)):
                    ws2.cell(row=ri + 3, column=c1, value=round(float(V_c[ri]), 6))
                    ws2.cell(row=ri + 3, column=c1 + 1, value=round(float(I_c[ri]), 6))
            # Gap column
            col = col_end + 2

    # ══════════════════════════════════════════════════════════════
    #  EIS
    # ══════════════════════════════════════════════════════════════
    eis = results.get('eis', [])
    if eis:
        # Summary
        ws = wb.create_sheet('EIS Summary')
        headers = ['Label', 'HFR (mΩ·cm²)', 'R_ct (mΩ·cm²)',
                   'R_total (mΩ·cm²)', 'R²']
        _write_header(ws, headers)
        for i, r in enumerate(eis, 2):
            _write_val(ws, i, 1, r.get('label', ''))
            _write_val(ws, i, 2, r.get('HFR', 0) * 1000 if r.get('HFR') else None)
            _write_val(ws, i, 3, r.get('R_ct', 0) * 1000 if r.get('R_ct') else None)
            _write_val(ws, i, 4, r.get('R_total', 0) * 1000 if r.get('R_total') else None)
            _write_val(ws, i, 5, r.get('R_squared'))

        # Data: freq, Z', -Z'' for all samples side by side (Ω·cm²)
        ws2 = wb.create_sheet('EIS Data')
        col = 1
        for r in eis:
            label = r.get('label', 'data')
            freq = r.get('_freq')
            Zr = r.get('_Z_real')
            Zi = r.get('_Z_imag')
            if freq is None or Zr is None:
                continue
            ga = r.get('geo_area', geo_area)
            col_end = col + 2
            _write_label_row(ws2, 1, col, col_end, label)
            hdrs = ['Freq (Hz)', "Z' (Ω·cm²)", "-Z'' (Ω·cm²)"]
            for ci, h in enumerate(hdrs):
                cell = ws2.cell(row=2, column=col + ci, value=h)
                cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center')
                ws2.column_dimensions[get_column_letter(col + ci)].width = 16
            for ri in range(len(freq)):
                ws2.cell(row=ri + 3, column=col, value=float(freq[ri]))
                ws2.cell(row=ri + 3, column=col + 1, value=round(float(Zr[ri]) * ga, 6))
                ws2.cell(row=ri + 3, column=col + 2, value=round(float(-Zi[ri]) * ga, 6))
            col = col_end + 2

    # ══════════════════════════════════════════════════════════════
    #  H₂ Crossover
    # ══════════════════════════════════════════════════════════════
    xover = results.get('crossover', [])
    if xover:
        # Summary
        ws = wb.create_sheet('Crossover Summary')
        headers = ['Label', 'j_xover (mA/cm²)', 'H₂ Flux (nmol/cm²/s)',
                   'H₂ Flux (mL/min/cm²)']
        has_perm = any(r.get('K_H2_mol_cm_s_Pa') is not None for r in xover)
        if has_perm:
            headers += ['Thickness (µm)', 'Permeability (mol/cm/s/Pa)']
        _write_header(ws, headers)
        for i, r in enumerate(xover, 2):
            c = 1
            _write_val(ws, i, c, r.get('label', '')); c += 1
            _write_val(ws, i, c, r.get('j_xover_mA_cm2')); c += 1
            _write_val(ws, i, c, r.get('J_H2_nmol_cm2_s')); c += 1
            _write_val(ws, i, c, r.get('J_H2_mL_min_cm2')); c += 1
            if has_perm:
                _write_val(ws, i, c, r.get('membrane_thickness_um')); c += 1
                _write_val(ws, i, c, r.get('K_H2_mol_cm_s_Pa'))

        # Data: raw last CV cycle (anodic + cathodic) for all samples
        ws2 = wb.create_sheet('Crossover Data')
        col = 1
        for r in xover:
            label = r.get('label', 'data')
            V_a = r.get('V_anodic')
            j_a = r.get('j_anodic')
            V_c = r.get('V_cathodic')
            j_c = r.get('j_cathodic')
            if V_a is None:
                continue
            col_end = col + 3
            _write_label_row(ws2, 1, col, col_end, label)
            hdrs = ['V_anodic (V)', 'j_anodic (mA/cm²)',
                    'V_cathodic (V)', 'j_cathodic (mA/cm²)']
            for ci, h in enumerate(hdrs):
                cell = ws2.cell(row=2, column=col + ci, value=h)
                cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center')
                ws2.column_dimensions[get_column_letter(col + ci)].width = 18
            for ri in range(len(V_a)):
                ws2.cell(row=ri + 3, column=col, value=round(float(V_a[ri]), 6))
                ws2.cell(row=ri + 3, column=col + 1, value=round(float(j_a[ri]) * 1000, 4))
            for ri in range(len(V_c)):
                ws2.cell(row=ri + 3, column=col + 2, value=round(float(V_c[ri]), 6))
                ws2.cell(row=ri + 3, column=col + 3, value=round(float(j_c[ri]) * 1000, 4))
            col = col_end + 2

    # ══════════════════════════════════════════════════════════════
    #  Pol Curve
    # ══════════════════════════════════════════════════════════════
    polcurve = results.get('polcurve', [])
    if polcurve:
        # Summary + representative data
        ws = wb.create_sheet('Pol Curve Summary')
        headers = ['Label', 'OCV (V)', 'V @ 1 A/cm² (V)',
                   'Peak Power (mW/cm²)', 'j @ Peak (A/cm²)']
        has_hfr = any(r.get('HFR_mean') is not None for r in polcurve)
        has_tafel = any(r.get('tafel') is not None for r in polcurve)
        if has_hfr:
            headers.append('HFR (mΩ·cm²)')
        if has_tafel:
            headers += ['Tafel (mV/dec)', 'j₀ (A/cm²)']
        _write_header(ws, headers)
        for i, r in enumerate(polcurve, 2):
            c = 1
            _write_val(ws, i, c, r.get('label', '')); c += 1
            _write_val(ws, i, c, r.get('OCV')); c += 1
            _write_val(ws, i, c, r.get('V_at_1Acm2')); c += 1
            _write_val(ws, i, c, r.get('peak_power_W_cm2', 0) * 1000); c += 1
            _write_val(ws, i, c, r.get('j_at_peak_power')); c += 1
            if has_hfr:
                hfr = r.get('HFR_mean')
                _write_val(ws, i, c, hfr * 1000 if hfr else None); c += 1
            if has_tafel:
                t = r.get('tafel')
                if t:
                    _write_val(ws, i, c, t.get('tafel_slope_mVdec')); c += 1
                    _write_val(ws, i, c, t.get('j0_A_cm2'))
                else:
                    c += 2

        # Representative curve data below the summary metrics
        data_start = len(polcurve) + 4
        ws.cell(row=data_start - 1, column=1,
                value='Representative Pol Curve Data').font = hf

        col = 1
        for r in polcurve:
            label = r.get('label', 'data')
            j, V, P = r['j'], r['V'], r['P']
            has_h = r.get('HFR') is not None
            has_irfree = r.get('V_irfree') is not None
            n_cols = 4 + int(has_h) + int(has_irfree)  # j_A, j_mA, V, [HFR], [iRfree], Power
            col_end = col + n_cols - 1
            _write_label_row(ws, data_start, col, col_end, label)

            hdrs = ['j (A/cm²)', 'j (mA/cm²)', 'V (V)']
            if has_h: hdrs.append('HFR (mΩ·cm²)')
            if has_irfree: hdrs.append('V_iR-free (V)')
            hdrs.append('Power (mW/cm²)')
            for ci, h in enumerate(hdrs):
                cell = ws.cell(row=data_start + 1, column=col + ci, value=h)
                cell.font = hf; cell.fill = hfill
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(col + ci)].width = 16

            order = np.argsort(V)
            for row_idx, ri in enumerate(order):
                row = data_start + 2 + row_idx
                cc = col
                ws.cell(row=row, column=cc, value=round(float(j[ri]), 6)); cc += 1
                ws.cell(row=row, column=cc, value=round(float(j[ri]) * 1000, 3)); cc += 1
                ws.cell(row=row, column=cc, value=round(float(V[ri]), 6)); cc += 1
                if has_h:
                    ws.cell(row=row, column=cc,
                            value=round(float(r['HFR_ASR'][ri]) * 1000, 2)); cc += 1
                if has_irfree:
                    ws.cell(row=row, column=cc,
                            value=round(float(r['V_irfree'][ri]), 6)); cc += 1
                ws.cell(row=row, column=cc, value=round(float(P[ri]) * 1000, 2))

            col = col_end + 2

        # Cycle data: all up/down cycles per file
        ws2 = wb.create_sheet('Pol Curve Data')
        col = 1
        for r in polcurve:
            label = r.get('label', 'data')
            cycles = r.get('_cycles_raw', [])
            ga = r.get('geo_area', geo_area)

            if not cycles:
                # Fallback: write representative data
                j, V = r['j'], r['V']
                col_end = col + 2  # j_A, j_mA, V
                _write_label_row(ws2, 1, col, col_end, label)
                ws2.cell(row=2, column=col, value='j (A/cm²)').font = hf
                ws2.cell(row=2, column=col+1, value='j (mA/cm²)').font = hf
                ws2.cell(row=2, column=col+2, value='V (V)').font = hf
                for ri in range(len(j)):
                    ws2.cell(row=ri+3, column=col, value=round(float(j[ri]), 6))
                    ws2.cell(row=ri+3, column=col+1, value=round(float(j[ri]) * 1000, 3))
                    ws2.cell(row=ri+3, column=col+2, value=round(float(V[ri]), 6))
                col = col_end + 2
                continue

            # File label spanning all cycle columns
            total_cols = sum(3 + (1 if cyc.get('HFR') is not None else 0)
                             for cyc in cycles)
            total_cols += len(cycles) - 1  # gap columns
            file_col_end = col + total_cols - 1
            _write_label_row(ws2, 1, col, file_col_end, label)

            for ci, cyc in enumerate(cycles):
                direction = cyc.get('direction', '?')
                cycle_num = cyc.get('cycle_number', ci + 1)
                cyc_label = f'Cycle {cycle_num} ({direction})'
                cyc_j = cyc['j']
                cyc_V = cyc['V']
                cyc_hfr = cyc.get('HFR')
                n_c = 3 + (1 if cyc_hfr is not None else 0)  # j_A, j_mA, V, [HFR]
                cyc_col_end = col + n_c - 1

                # Cycle sub-label
                cell = ws2.cell(row=2, column=col, value=cyc_label)
                cell.font = hf
                if cyc_col_end > col:
                    ws2.merge_cells(start_row=2, start_column=col,
                                    end_row=2, end_column=cyc_col_end)

                hdrs = ['j (A/cm²)', 'j (mA/cm²)', 'V (V)']
                if cyc_hfr is not None:
                    hdrs.append('HFR (mΩ·cm²)')
                for hi, h in enumerate(hdrs):
                    cell = ws2.cell(row=3, column=col + hi, value=h)
                    cell.font = hf; cell.fill = hfill
                    cell.alignment = Alignment(horizontal='center')
                    ws2.column_dimensions[get_column_letter(col + hi)].width = 16

                for ri in range(len(cyc_j)):
                    cc = col
                    ws2.cell(row=ri+4, column=cc,
                             value=round(float(cyc_j[ri]), 6)); cc += 1
                    ws2.cell(row=ri+4, column=cc,
                             value=round(float(cyc_j[ri]) * 1000, 3)); cc += 1
                    ws2.cell(row=ri+4, column=cc,
                             value=round(float(cyc_V[ri]), 6)); cc += 1
                    if cyc_hfr is not None and ri < len(cyc_hfr):
                        ws2.cell(row=ri+4, column=cc,
                                 value=round(float(cyc_hfr[ri]) * ga * 1000, 2))

                col = cyc_col_end + 2

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet('No Results')
        ws['A1'] = 'No analyses produced results.'

    wb.save(filepath)
    print(f'\n  Consolidated Excel: {filepath}')


# ═══════════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════════

def run_all(folder, save_dir, geo_area=5.0, loading=0.2,
            membrane_thickness=None, stand=0, ocv_interval_s=60.0,
            image_ext='png'):
    """
    Scan folder, classify files, and run all applicable analyses.
    """
    banner = chr(9619) * 65
    print(f'\n{banner}')
    print('  FUEL CELL DATA ANALYSIS — BATCH ORCHESTRATOR')
    print(f'{banner}')

    # Auto-detect test stand from file extensions if stand=0 (default)
    # FCD files → Scribner, CSV files → FCTS
    from pathlib import Path
    data_files = [f for f in Path(folder).rglob('*')
                  if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv', '.fcd')]
    has_fcd = any(f.suffix.lower() == '.fcd' for f in data_files)
    has_csv_only = all(f.suffix.lower() in ('.csv', '.txt', '.tsv') for f in data_files)

    if has_csv_only and not has_fcd and stand == 0:
        stand = 1
        print(f'  Auto-detected FCTS (CSV files, no FCD files)')

    print(f'\n  Folder:    {folder}')
    print(f'  Output:    {save_dir}')
    print(f'  Geo area:  {geo_area} cm²')
    print(f'  Stand:     {"Scribner" if stand == 0 else "FCTS"}')

    # Classify
    classified, unclassified = classify_files(folder)

    total = sum(len(v) for v in classified.values())
    print(f'\n  Found {total} classifiable files:')
    for atype in CLASSIFICATION_ORDER:
        files = classified[atype]
        label = ANALYSIS_TYPES[atype]['label']
        if files:
            print(f'    {label:15s}  {len(files)} file(s)')
            for fp, name in files:
                print(f'      - {name}')
    if unclassified:
        print(f'    {"Unclassified":15s}  {len(unclassified)} file(s)')
        for fp in unclassified[:5]:
            print(f'      - {os.path.basename(fp)}')
        if len(unclassified) > 5:
            print(f'      ... and {len(unclassified) - 5} more')

    if total == 0:
        print('\n  No files matched any analysis keywords.')
        return {'classified': {}, 'errors': {}, 'total': 0}

    os.makedirs(save_dir, exist_ok=True)
    t0 = time.time()
    all_results = {}
    errors = {}

    # Run each analysis type
    for atype in CLASSIFICATION_ORDER:
        files = classified[atype]
        if not files:
            continue

        label = ANALYSIS_TYPES[atype]['label']
        sub_dir = os.path.join(save_dir, atype)
        os.makedirs(sub_dir, exist_ok=True)

        print(f'\n{"=" * 65}')
        print(f'  {label} — {len(files)} file(s)')
        print(f'{"=" * 65}')

        try:
            if atype == 'ecsa':
                all_results['ecsa'] = run_ecsa_batch(
                    files, geo_area, loading, sub_dir, stand, image_ext=image_ext)
            elif atype == 'eis':
                all_results['eis'] = run_eis_batch(
                    files, geo_area, sub_dir, stand, image_ext=image_ext)
            elif atype == 'crossover':
                all_results['crossover'] = run_crossover_batch(
                    files, geo_area, membrane_thickness, sub_dir, stand, image_ext=image_ext)
            elif atype == 'polcurve':
                all_results['polcurve'] = run_polcurve_batch(
                    files, geo_area, sub_dir, stand, image_ext=image_ext)
            elif atype == 'ocv':
                all_results['ocv'] = run_ocv_batch(
                    files, sub_dir, interval_s=ocv_interval_s, image_ext=image_ext)

            # Check if the analysis actually produced output files
            from pathlib import Path
            sub_output = list(Path(sub_dir).rglob('*'))
            sub_files = [f for f in sub_output if f.is_file()]
            if not sub_files:
                stand_name = "Scribner" if stand == 0 else "FCTS"
                errors[label] = (
                    f"{label}: processed {len(files)} file(s) but produced no output. "
                    f"Stand={stand_name}. Check that the test stand matches "
                    f"your data format (Scribner=tab-delimited FCD, FCTS=comma-delimited CSV)."
                )
        except Exception as e:
            import traceback
            errors[label] = traceback.format_exc()
            print(f'\n  ERROR in {label}: {e}')
            traceback.print_exc()

    # Consolidated Excel
    if any(all_results.values()):
        xlsx_path = os.path.join(save_dir, 'fuelcell_summary.xlsx')
        save_consolidated_excel(all_results, xlsx_path, geo_area=geo_area)

    elapsed = time.time() - t0
    print(f'\n{banner}')
    print(f'  COMPLETE — {elapsed:.1f}s')
    print(f'  Results in: {os.path.abspath(save_dir)}')
    print(f'{banner}\n')

    # Return info for the portal
    classified_summary = {
        ANALYSIS_TYPES[atype]['label']: len(classified[atype])
        for atype in CLASSIFICATION_ORDER if classified[atype]
    }
    return {
        'classified': classified_summary,
        'errors': errors,
        'total': total,
        'unclassified': len(unclassified),
    }


def run_interactive():
    banner = chr(9619) * 65
    print(f'\n{banner}')
    print('  FUEL CELL DATA ANALYSIS — BATCH ORCHESTRATOR')
    print(f'{banner}')

    folder = _clean_path(_prompt('\n  Data folder path', cast=None))
    if not folder or not os.path.isdir(folder):
        print(f'  Folder not found: {folder}')
        return

    save_dir = _clean_path(_prompt('  Output directory',
                                    default='fuelcell_results', cast=None))

    print('\n  — Test Stand —')
    print('    0 = Scribner')
    print('    1 = FCTS')
    stand = int(_prompt('  Test stand', default=0, cast=int))

    geo_area = _prompt('  Geometric area (cm²)', default=5.0)

    # Show classification before running
    classified, unclassified = classify_files(folder)
    total = sum(len(v) for v in classified.values())

    print(f'\n  File classification:')
    for atype in CLASSIFICATION_ORDER:
        files = classified[atype]
        label = ANALYSIS_TYPES[atype]['label']
        if files:
            print(f'    {label:15s}  {len(files)} file(s)')

    if total == 0:
        print('    No files matched. Check folder path and filenames.')
        return

    # ECSA-specific
    loading = 0.2
    if classified['ecsa']:
        loading = _prompt('  Cathode Pt loading (mg/cm², for ECSA)', default=0.2)

    # Crossover-specific
    membrane_thickness = None
    if classified['crossover']:
        t_raw = _prompt('  Membrane thickness (µm, Enter to skip)', default=None, cast=None)
        membrane_thickness = float(t_raw) if t_raw else None

    proceed = _prompt('\n  Run all analyses? (y/n)', default='y', cast=None)
    if proceed.lower() != 'y':
        print('  Cancelled.')
        return

    run_all(folder, save_dir, geo_area=geo_area, loading=loading,
            membrane_thickness=membrane_thickness, stand=stand)


def main():
    parser = argparse.ArgumentParser(
        description='Fuel Cell Data Analysis — Batch Orchestrator')
    parser.add_argument('--folder', type=str, help='Data folder path')
    parser.add_argument('--save-dir', type=str, default='fuelcell_results')
    parser.add_argument('--area', type=float, default=5.0)
    parser.add_argument('--loading', type=float, default=0.2,
                        help='Cathode Pt loading (mg/cm²)')
    parser.add_argument('--thickness', type=float, default=None,
                        help='Membrane thickness (µm)')
    parser.add_argument('--stand', type=int, default=0,
                        choices=[0, 1], help='0=Scribner, 1=FCTS')
    args = parser.parse_args()

    if args.folder:
        run_all(args.folder, args.save_dir, geo_area=args.area,
                loading=args.loading, membrane_thickness=args.thickness,
                stand=args.stand)
    else:
        run_interactive()


if __name__ == '__main__':
    main()
