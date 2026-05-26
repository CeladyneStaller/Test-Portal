#!/usr/bin/env python3
"""
Electrode Cleaning Cycle Analysis (500 mV/s)
=============================================

Analyzes H2/N2 CV cleaning cycles run at high scan rate (default 500 mV/s).
Tracks cycle-by-cycle convergence to determine when cleaning is complete
and reports final ECSA from the converged region.

File classification:
  - Scribner (.fcd): files containing 'CV-500mVs' (case-insensitive)
  - FCTS (.csv):     files containing 'EClean'/'E-Clean'/'E_Clean' (case-insensitive)

Generates per file:
  - cleaning_cycles_<label>.png — 3 panels (all cycles, final, evolution)
  - cleaning_diagnostics_<label>.png — 3 panels (Q vs cycle, DL log-y, normalized overlay)
  - cleaning_data.xlsx — Summary + per-cycle metrics + CV data
"""

import os
import glob
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy.integrate import trapezoid

from scripts.helpers.plot_compare import save_with_sidecar


# ═══════════════════════════════════════════════════════════════════════
#  Header Parsing
# ═══════════════════════════════════════════════════════════════════════

def parse_fcd_header(filepath):
    """Parse Scribner .fcd header for skip count and column indices."""
    if not filepath.lower().endswith('.fcd'):
        return None
    try:
        with open(filepath, 'r', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                lines.append(line)
                if 'End Comments' in line:
                    skip = i + 1
                    break
            else:
                return None
    except Exception:
        return None
    if len(lines) < 2:
        return None
    cols = lines[-2].strip().split('\t')
    result = {'skip': skip, 'delimiter': '\t',
              'current_unit': 'A', 'voltage_unit': 'V'}
    for ci, name in enumerate(cols):
        n = name.strip()
        if n == 'Time (Sec)':
            result['time_col'] = ci
        elif n == 'I (A)':
            result['j_col'] = ci
        elif n == 'E_Stack (V)':
            result['v_col'] = ci
    return result


def parse_csv_header(filepath):
    """Parse FCTS / generic CSV header for column indices."""
    try:
        with open(filepath, 'r', errors='replace') as f:
            first_line = f.readline().strip()
    except Exception:
        return None
    if '\t' in first_line:
        delimiter = '\t'
    elif ',' in first_line:
        delimiter = ','
    else:
        return None
    cols = [c.strip() for c in first_line.split(delimiter)]
    if not cols:
        return None
    result = {'skip': 1, 'delimiter': delimiter}
    cols_lower = [c.lower() for c in cols]
    # Voltage
    for ci, c in enumerate(cols_lower):
        if 'working electrode voltage' in c or c == 'v' or c == 'voltage':
            result['v_col'] = ci
            result['voltage_unit'] = 'mV' if 'mv' in c else 'V'
            break
    if 'v_col' not in result:
        for ci, c in enumerate(cols_lower):
            if c.startswith('e_') or c.startswith('e ') or 'cell voltage' in c:
                result['v_col'] = ci
                result['voltage_unit'] = 'V'
                break
    # Current — prefer mA/cm² for FCTS
    for ci, c in enumerate(cols_lower):
        if 'ma/cm' in c or 'ma·cm' in c:
            result['j_col'] = ci
            result['current_unit'] = 'mA/cm2'
            break
    if 'j_col' not in result:
        for ci, c in enumerate(cols_lower):
            if c == 'current' or c == 'ma':
                result['j_col'] = ci
                result['current_unit'] = 'mA'
                break
    if 'j_col' not in result:
        for ci, c in enumerate(cols_lower):
            if c == 'a' or c == 'current (a)':
                result['j_col'] = ci
                result['current_unit'] = 'A'
                break
    if 'v_col' not in result or 'j_col' not in result:
        return None
    return result


# ═══════════════════════════════════════════════════════════════════════
#  Data Loading
# ═══════════════════════════════════════════════════════════════════════

def load_cleaning_data(filepath, geo_area=5.0):
    """Load V (V) and j (A/cm²) from a cleaning CV file."""
    label = os.path.splitext(os.path.basename(filepath))[0]

    fcd = parse_fcd_header(filepath)
    if fcd is not None and 'v_col' in fcd and 'j_col' in fcd:
        try:
            data = np.genfromtxt(filepath, delimiter=fcd['delimiter'],
                                  skip_header=fcd['skip'],
                                  usecols=[fcd['v_col'], fcd['j_col']],
                                  invalid_raise=False)
        except Exception as e:
            print(f'    Error reading {os.path.basename(filepath)}: {e}')
            return None, None, label
        if data is None or data.ndim < 2:
            return None, None, label
        valid = ~np.isnan(data).any(axis=1)
        data = data[valid]
        if len(data) == 0:
            return None, None, label
        V = data[:, 0]
        j = data[:, 1] / geo_area  # Scribner total current → A/cm²
        return V, j, label

    hdr = parse_csv_header(filepath)
    if hdr is None:
        print(f'    Could not parse header for {os.path.basename(filepath)}')
        return None, None, label
    try:
        data = np.genfromtxt(filepath, delimiter=hdr['delimiter'],
                              skip_header=hdr['skip'],
                              usecols=[hdr['v_col'], hdr['j_col']],
                              invalid_raise=False)
    except Exception as e:
        print(f'    Error reading {os.path.basename(filepath)}: {e}')
        return None, None, label
    if data is None or data.ndim < 2:
        return None, None, label
    valid = ~np.isnan(data).any(axis=1)
    data = data[valid]
    if len(data) == 0:
        return None, None, label
    V = data[:, 0]
    if hdr.get('voltage_unit') == 'mV':
        V = V * 0.001
    unit = hdr.get('current_unit', 'A')
    if unit == 'mA/cm2':
        j = data[:, 1] * 0.001
    elif unit == 'mA':
        j = data[:, 1] * 0.001 / geo_area
    else:
        j = data[:, 1] / geo_area
    return V, j, label


# ═══════════════════════════════════════════════════════════════════════
#  Cycle Extraction
# ═══════════════════════════════════════════════════════════════════════

def extract_cv_cycles(V, j, min_pts_per_cycle=20):
    """Split continuous CV data into individual cycles, starting at V minima."""
    V = np.asarray(V)
    j = np.asarray(j)
    if len(V) < min_pts_per_cycle * 2:
        return []

    # Find V-minima by looking for local minima in a smoothed window.
    # A minimum is where V[i] <= V[i±k] for k=1..window.
    starts = []
    window = max(3, min_pts_per_cycle // 4)
    for i in range(window, len(V) - window):
        left_max = V[i - window:i].max()
        right_max = V[i + 1:i + window + 1].max()
        # Check this is a strict local minimum (lower than both surrounding maxes)
        if V[i] < left_max - 1e-6 and V[i] < right_max - 1e-6:
            # Avoid double-counting: skip if too close to previous start
            if not starts or i - starts[-1] >= min_pts_per_cycle:
                starts.append(i)

    if not starts:
        # No minima found — treat the whole trace as one cycle
        return [(V, j)]

    cycles = []
    # Optional pre-segment (data before first cycle start)
    if starts[0] > min_pts_per_cycle:
        cycles.append((V[:starts[0]], j[:starts[0]]))
    for i in range(len(starts) - 1):
        s, e = starts[i], starts[i + 1]
        if e - s >= min_pts_per_cycle:
            cycles.append((V[s:e], j[s:e]))
    # Final segment after last start
    if len(V) - starts[-1] >= min_pts_per_cycle:
        cycles.append((V[starts[-1]:], j[starts[-1]:]))
    return cycles


def split_anodic_cathodic(V_cycle, j_cycle):
    """Split one cycle into anodic (V increasing) and cathodic legs."""
    if len(V_cycle) < 3:
        return None, None, None, None
    # Anodic = V increasing leg starting from V_min
    i_max = int(np.argmax(V_cycle))
    V_a = V_cycle[:i_max + 1]
    j_a = j_cycle[:i_max + 1]
    V_c = V_cycle[i_max:][::-1]
    j_c = j_cycle[i_max:][::-1]
    return V_a, j_a, V_c, j_c


def _integrate_hupd(V_sweep, j_sweep, scan_rate,
                    v_low=0.05, v_high=0.40, baseline_window=(0.40, 0.50)):
    """Integrate H_UPD charge with double-layer baseline subtraction.

    Returns charge in mC/cm².
    """
    if len(V_sweep) < 5:
        return 0.0
    # DL baseline from V in baseline_window
    bl_mask = (V_sweep >= baseline_window[0]) & (V_sweep <= baseline_window[1])
    if bl_mask.sum() < 3:
        baseline = 0.0
    else:
        baseline = float(np.mean(j_sweep[bl_mask]))
    # Integrate in H_UPD window
    hupd_mask = (V_sweep >= v_low) & (V_sweep <= v_high)
    if hupd_mask.sum() < 3:
        return 0.0
    V_int = V_sweep[hupd_mask]
    j_int = j_sweep[hupd_mask] - baseline
    # Sort by V (anodic = ascending; cathodic = ascending after we reversed)
    order = np.argsort(V_int)
    V_int = V_int[order]
    j_int = j_int[order]
    # Q (C/cm²) = integral j dV / scan_rate
    q = trapezoid(j_int, V_int) / scan_rate
    return float(abs(q) * 1000.0)  # mC/cm²


def compute_cycle_metrics(V_cycle, j_cycle, scan_rate,
                          v_hupd_low=0.05, v_hupd_high=0.40,
                          v_dl_low=0.40, v_dl_high=0.50):
    """Compute Q_HUPD anodic & cathodic + DL currents for one cycle."""
    V_a, j_a, V_c, j_c = split_anodic_cathodic(V_cycle, j_cycle)
    if V_a is None or V_c is None or len(V_a) < 5 or len(V_c) < 5:
        return None

    Q_a = _integrate_hupd(V_a, j_a, scan_rate,
                          v_low=v_hupd_low, v_high=v_hupd_high,
                          baseline_window=(v_dl_low, v_dl_high))
    Q_c = _integrate_hupd(V_c, j_c, scan_rate,
                          v_low=v_hupd_low, v_high=v_hupd_high,
                          baseline_window=(v_dl_low, v_dl_high))
    ratio = (Q_a / Q_c) if Q_c > 0 else 0.0

    # DL currents in the DL window
    dl_a_mask = (V_a >= v_dl_low) & (V_a <= v_dl_high)
    dl_c_mask = (V_c >= v_dl_low) & (V_c <= v_dl_high)
    dl_a = float(np.mean(j_a[dl_a_mask]) * 1000) if dl_a_mask.sum() >= 3 else 0.0
    dl_c = float(np.mean(j_c[dl_c_mask]) * 1000) if dl_c_mask.sum() >= 3 else 0.0
    dl_mean = (abs(dl_a) + abs(dl_c)) / 2.0
    dl_asym = abs(abs(dl_a) - abs(dl_c))

    return {
        'Q_anodic_mC_cm2': Q_a,
        'Q_cathodic_mC_cm2': Q_c,
        'Q_ratio': ratio,
        'DL_anodic_mA_cm2': dl_a,
        'DL_cathodic_mA_cm2': dl_c,
        'DL_mean_mA_cm2': dl_mean,
        'DL_asymmetry_mA_cm2': dl_asym,
    }


def find_convergence_cycle(metrics, q_tol=0.01, dl_tol=0.05, window=3):
    """Find the first cycle where Q_cathodic and DL are stable over `window` cycles."""
    n = len(metrics)
    if n < window + 1:
        return None
    Q_c = np.array([m['Q_cathodic_mC_cm2'] for m in metrics])
    DL = np.abs(np.array([m['DL_mean_mA_cm2'] for m in metrics]))
    for i in range(window, n - window + 1):
        ref_Qc = Q_c[i]
        ref_DL = DL[i]
        if ref_Qc <= 0 or ref_DL <= 0:
            continue
        Q_stable = np.all(np.abs(Q_c[i:i + window] - ref_Qc) / max(ref_Qc, 1e-6) < q_tol)
        DL_stable = np.all(np.abs(DL[i:i + window] - ref_DL) / max(ref_DL, 1e-6) < dl_tol)
        if Q_stable and DL_stable:
            return i
    return None


# ═══════════════════════════════════════════════════════════════════════
#  Plotting
# ═══════════════════════════════════════════════════════════════════════

def plot_cv_overlays(cycles, conv_idx, metrics, label, scan_rate, save_path=None):
    """3-panel: all cycles, final cycle, evolution (cycle 1 / convergence / final)."""
    fig, axes = plt.subplots(1, 3, figsize=(16, 5))
    n = len(cycles)

    # Panel (a): all cycles
    ax = axes[0]
    cmap = plt.cm.viridis
    for i, (V_c, j_c) in enumerate(cycles):
        color = cmap(i / max(n - 1, 1))
        ax.plot(V_c, j_c * 1000, '-', color=color, lw=0.6, alpha=0.7)
    ax.plot([], [], '-', color=cmap(0.0), lw=2, label='Cycle 1')
    ax.plot([], [], '-', color=cmap(1.0), lw=2, label=f'Cycle {n}')
    ax.set_xlabel('Voltage (V)')
    ax.set_ylabel('Current density (mA/cm²)')
    ax.set_title(f'All cycles overlaid ({n} cycles)')
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=9, loc='best')

    # Panel (b): final cycle
    ax = axes[1]
    V_f, j_f = cycles[-1]
    ax.plot(V_f, j_f * 1000, '-', color='firebrick', lw=1.5)
    ax.set_xlabel('Voltage (V)')
    ax.set_ylabel('Current density (mA/cm²)')
    ax.set_title(f'Final cycle (cycle {n})')
    ax.grid(True, alpha=0.3)

    # Add summary readout on the final-cycle panel for the comparison feature
    n_avg = min(5, n)
    final_Qa = float(np.mean([m['Q_anodic_mC_cm2'] for m in metrics[-n_avg:]]))
    final_Qc = float(np.mean([m['Q_cathodic_mC_cm2'] for m in metrics[-n_avg:]]))
    final_dl = float(np.mean([m['DL_mean_mA_cm2'] for m in metrics[-n_avg:]]))
    init_dl = metrics[0]['DL_mean_mA_cm2']
    dl_red = (1.0 - final_dl / init_dl) * 100 if init_dl > 0 else 0.0
    conv_str = f'cycle {conv_idx + 1}' if conv_idx is not None else 'not reached'
    readout = (f'Cycles = {n}\n'
               f'Convergence = {conv_str}\n'
               f'Q_anodic_final = {final_Qa:.3f} mC/cm²\n'
               f'Q_cathodic_final = {final_Qc:.3f} mC/cm²\n'
               f'DL_initial = {init_dl:.3f} mA/cm²\n'
               f'DL_final = {final_dl:.3f} mA/cm²\n'
               f'DL_reduction = {dl_red:.1f}%')
    ax.text(0.03, 0.97, readout, transform=ax.transAxes,
             fontsize=8, va='top', ha='left', family='monospace',
             bbox=dict(boxstyle='round,pad=0.4', fc='lightyellow', alpha=0.92))

    # Panel (c): cycle 1, convergence, final
    ax = axes[2]
    ax.plot(cycles[0][0], cycles[0][1] * 1000, '-',
            color='tab:blue', lw=1.4, label='Cycle 1', alpha=0.85)
    if conv_idx is not None and 0 < conv_idx < n - 1:
        ax.plot(cycles[conv_idx][0], cycles[conv_idx][1] * 1000, '-',
                color='tab:orange', lw=1.4,
                label=f'Convergence (cycle {conv_idx + 1})', alpha=0.85)
    ax.plot(cycles[-1][0], cycles[-1][1] * 1000, '-',
            color='firebrick', lw=1.4, label=f'Final (cycle {n})', alpha=0.85)
    ax.set_xlabel('Voltage (V)')
    ax.set_ylabel('Current density (mA/cm²)')
    ax.set_title('Cycle evolution')
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=9, loc='best')

    fig.suptitle(f'Electrode Cleaning — {label} ({scan_rate*1000:.0f} mV/s)',
                 fontweight='bold', y=1.02)
    fig.tight_layout()

    if save_path:
        save_with_sidecar(fig, save_path, plot_type='cleaning_cycles', dpi=150)
        print(f'    Saved: {save_path}')
    return fig


def plot_diagnostics(cycles, metrics, conv_idx, label, scan_rate, save_path=None):
    """3-panel: Q vs cycle, DL log-y vs cycle, normalized overlay of last 5."""
    fig, axes = plt.subplots(1, 3, figsize=(16, 5))
    n = len(metrics)
    cycle_nums = np.arange(1, n + 1)
    Q_a = np.array([m['Q_anodic_mC_cm2'] for m in metrics])
    Q_c = np.array([m['Q_cathodic_mC_cm2'] for m in metrics])
    DL_mean = np.array([m['DL_mean_mA_cm2'] for m in metrics])

    # Panel (a): Q_HUPD vs cycle
    ax = axes[0]
    ax.plot(cycle_nums, Q_a, 'o-', color='tab:blue', lw=1.2, ms=5, label='Q anodic')
    ax.plot(cycle_nums, Q_c, 's-', color='tab:red', lw=1.2, ms=5, label='Q cathodic')
    if conv_idx is not None:
        ax.axvline(conv_idx + 1, color='gray', ls='--', lw=1,
                   label=f'Convergence = cycle {conv_idx + 1}')
    ax.set_xlabel('Cycle number')
    ax.set_ylabel('Q_H_UPD (mC/cm²)')
    ax.set_title('H_UPD charge vs cycle')
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=9, loc='best')

    # Add summary readout (same fields, so comparisons of either plot work)
    n_avg = min(5, n)
    final_Qa = float(np.mean(Q_a[-n_avg:]))
    final_Qc = float(np.mean(Q_c[-n_avg:]))
    final_dl = float(np.mean(DL_mean[-n_avg:]))
    init_dl = float(DL_mean[0])
    dl_red = (1.0 - final_dl / init_dl) * 100 if init_dl > 0 else 0.0
    conv_str = f'cycle {conv_idx + 1}' if conv_idx is not None else 'not reached'
    readout = (f'Cycles = {n}\n'
               f'Convergence = {conv_str}\n'
               f'Q_anodic_final = {final_Qa:.3f} mC/cm²\n'
               f'Q_cathodic_final = {final_Qc:.3f} mC/cm²\n'
               f'DL_reduction = {dl_red:.1f}%')
    ax.text(0.97, 0.97, readout, transform=ax.transAxes,
            fontsize=8, va='top', ha='right', family='monospace',
            bbox=dict(boxstyle='round,pad=0.4', fc='lightyellow', alpha=0.92))

    # Panel (b): DL current vs cycle, log-y
    ax = axes[1]
    ax.semilogy(cycle_nums, np.abs(DL_mean), 'o-', color='tab:green', lw=1.2, ms=5)
    if conv_idx is not None:
        ax.axvline(conv_idx + 1, color='gray', ls='--', lw=1,
                   label=f'Convergence = cycle {conv_idx + 1}')
        ax.legend(fontsize=9, loc='best')
    ax.set_xlabel('Cycle number')
    ax.set_ylabel('Mean |DL current| (mA/cm²)')
    ax.set_title('Double-layer current vs cycle (log-y)')
    ax.grid(True, alpha=0.3, which='both')

    # Panel (c): j/ν normalized last 5 cycles
    ax = axes[2]
    n_overlay = min(5, len(cycles))
    last_cycles = cycles[-n_overlay:]
    cmap = plt.cm.Reds
    for i, (V_c, j_c) in enumerate(last_cycles):
        cap = j_c / scan_rate * 1e6  # j/ν in µF/cm²
        cycle_num = n - n_overlay + i + 1
        color = cmap(0.4 + 0.6 * i / max(n_overlay - 1, 1))
        ax.plot(V_c, cap, '-', color=color, lw=1.0, alpha=0.85,
                label=f'Cycle {cycle_num}')
    ax.set_xlabel('Voltage (V)')
    ax.set_ylabel('j/ν (µF/cm²)')
    ax.set_title(f'Scan-rate normalized — last {n_overlay} cycles')
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=8, loc='best')

    fig.suptitle(f'Cleaning Diagnostics — {label}', fontweight='bold', y=1.02)
    fig.tight_layout()

    if save_path:
        save_with_sidecar(fig, save_path, plot_type='cleaning_diagnostics', dpi=150)
        print(f'    Saved: {save_path}')
    return fig


# ═══════════════════════════════════════════════════════════════════════
#  Excel Output
# ═══════════════════════════════════════════════════════════════════════

def _save_cleaning_excel(results, scan_rate, geo_area, filepath):
    """Write Summary + per-cycle Metrics + CV Data sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hf = Font(bold=True)
    hfill = PatternFill('solid', fgColor='D9E1F2')
    label_fill = PatternFill('solid', fgColor='E2EFDA')

    # Summary
    ws = wb.active
    ws.title = 'Summary'
    headers = ['Label', 'Total cycles', 'Convergence cycle',
               'Final Q anodic (mC/cm²)', 'Final Q cathodic (mC/cm²)',
               'Final Q ratio', 'Final DL (mA/cm²)',
               'Initial DL (mA/cm²)', 'DL reduction (%)',
               'Scan rate (mV/s)', 'Geo area (cm²)', 'Converged']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(c)].width = max(len(h) + 2, 14)

    for i, r in enumerate(results, 2):
        metrics = r['metrics']
        n = len(metrics)
        conv = r['conv_idx']
        n_avg = min(5, n)
        final_Qa = float(np.mean([m['Q_anodic_mC_cm2'] for m in metrics[-n_avg:]]))
        final_Qc = float(np.mean([m['Q_cathodic_mC_cm2'] for m in metrics[-n_avg:]]))
        final_ratio = float(np.mean([m['Q_ratio'] for m in metrics[-n_avg:]]))
        final_dl = float(np.mean([m['DL_mean_mA_cm2'] for m in metrics[-n_avg:]]))
        init_dl = metrics[0]['DL_mean_mA_cm2']
        dl_red = (1.0 - final_dl / init_dl) * 100 if init_dl > 0 else 0.0
        ws.cell(row=i, column=1, value=r['label'])
        ws.cell(row=i, column=2, value=n)
        ws.cell(row=i, column=3, value=(conv + 1) if conv is not None else 'Not converged')
        ws.cell(row=i, column=4, value=round(final_Qa, 4))
        ws.cell(row=i, column=5, value=round(final_Qc, 4))
        ws.cell(row=i, column=6, value=round(final_ratio, 4))
        ws.cell(row=i, column=7, value=round(final_dl, 4))
        ws.cell(row=i, column=8, value=round(init_dl, 4))
        ws.cell(row=i, column=9, value=round(dl_red, 2))
        ws.cell(row=i, column=10, value=round(scan_rate * 1000, 1))
        ws.cell(row=i, column=11, value=geo_area)
        ws.cell(row=i, column=12, value='Yes' if conv is not None else 'No')
    ws.column_dimensions['A'].width = 40

    # Metrics (per-cycle for all samples)
    ws2 = wb.create_sheet('Metrics')
    col = 1
    for r in results:
        lbl = r['label']
        metrics = r['metrics']
        n = len(metrics)
        n_cols = 8
        col_end = col + n_cols - 1
        cell = ws2.cell(row=1, column=col, value=lbl)
        cell.font = hf
        cell.fill = label_fill
        cell.alignment = Alignment(horizontal='center')
        ws2.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col_end)
        hdrs = ['Cycle', 'Q_anodic (mC/cm²)', 'Q_cathodic (mC/cm²)',
                'Q_ratio', 'DL_anodic (mA/cm²)', 'DL_cathodic (mA/cm²)',
                'DL_mean (mA/cm²)', 'DL_asymmetry (mA/cm²)']
        for ci, h in enumerate(hdrs):
            cell = ws2.cell(row=2, column=col + ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            ws2.column_dimensions[get_column_letter(col + ci)].width = max(len(h) + 2, 14)
        for i, m in enumerate(metrics):
            ws2.cell(row=i + 3, column=col, value=i + 1)
            ws2.cell(row=i + 3, column=col + 1, value=round(m['Q_anodic_mC_cm2'], 4))
            ws2.cell(row=i + 3, column=col + 2, value=round(m['Q_cathodic_mC_cm2'], 4))
            ws2.cell(row=i + 3, column=col + 3, value=round(m['Q_ratio'], 4))
            ws2.cell(row=i + 3, column=col + 4, value=round(m['DL_anodic_mA_cm2'], 4))
            ws2.cell(row=i + 3, column=col + 5, value=round(m['DL_cathodic_mA_cm2'], 4))
            ws2.cell(row=i + 3, column=col + 6, value=round(m['DL_mean_mA_cm2'], 4))
            ws2.cell(row=i + 3, column=col + 7, value=round(m['DL_asymmetry_mA_cm2'], 4))
        col = col_end + 2

    # CV Data sheet: cycle 1, convergence, last 5 — for each sample
    ws3 = wb.create_sheet('CV Data')
    col = 1
    for r in results:
        lbl = r['label']
        cycles = r['cycles']
        n = len(cycles)
        conv = r['conv_idx']
        select_indices = [0]
        if conv is not None and 0 < conv < n - 1:
            select_indices.append(conv)
        for idx in range(max(0, n - 5), n):
            if idx not in select_indices:
                select_indices.append(idx)
        for sel in select_indices:
            V_c, j_c = cycles[sel]
            if sel == 0:
                tag = 'Cycle 1'
            elif conv is not None and sel == conv:
                tag = f'Convergence (cycle {conv + 1})'
            else:
                tag = f'Cycle {sel + 1}'
            col_end = col + 1
            cell = ws3.cell(row=1, column=col, value=f'{lbl} — {tag}')
            cell.font = hf
            cell.fill = label_fill
            cell.alignment = Alignment(horizontal='center')
            ws3.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col_end)
            for ci, h in enumerate(['V (V)', 'j (mA/cm²)']):
                c2 = ws3.cell(row=2, column=col + ci, value=h)
                c2.font = hf
                c2.fill = hfill
                c2.alignment = Alignment(horizontal='center')
                ws3.column_dimensions[get_column_letter(col + ci)].width = 14
            for k in range(len(V_c)):
                ws3.cell(row=k + 3, column=col, value=round(float(V_c[k]), 5))
                ws3.cell(row=k + 3, column=col + 1, value=round(float(j_c[k]) * 1000, 5))
            col = col_end + 2

    wb.save(filepath)
    print(f'    Excel: {filepath}')


# ═══════════════════════════════════════════════════════════════════════
#  File Classification
# ═══════════════════════════════════════════════════════════════════════

CLEANING_KEYWORDS = ['CV-500MVS', 'ECLEAN', 'E-CLEAN', 'E_CLEAN']
EXCLUDE_KEYWORDS = ['FILTERDATA']


def classify_cleaning_files(folder):
    """Find electrode-cleaning CV files in a folder (recursive)."""
    extensions = ['*.csv', '*.txt', '*.tsv', '*.fcd',
                  '*.CSV', '*.TXT', '*.TSV', '*.FCD']
    all_files = []
    for ext in extensions:
        all_files.extend(glob.glob(os.path.join(folder, '**', ext), recursive=True))
    all_files = sorted(set(all_files))
    all_files = [fp for fp in all_files
                 if not any(ex in os.path.basename(fp).upper() for ex in EXCLUDE_KEYWORDS)]
    matched = []
    for fp in all_files:
        name_upper = os.path.basename(fp).upper()
        if any(kw in name_upper for kw in CLEANING_KEYWORDS):
            label = os.path.splitext(os.path.basename(fp))[0]
            matched.append((fp, label))
    return matched


# ═══════════════════════════════════════════════════════════════════════
#  Batch Processing
# ═══════════════════════════════════════════════════════════════════════

def run_batch(filepaths, labels, save_dir=None, scan_rate=0.5, geo_area=5.0,
              v_hupd_low=0.05, v_hupd_high=0.40,
              v_dl_low=0.40, v_dl_high=0.50,
              image_ext='png'):
    """Process electrode-cleaning CV files: extract cycles, compute metrics, plot."""
    if save_dir:
        os.makedirs(save_dir, exist_ok=True)

    print(f'\n  Processing {len(filepaths)} files...')
    print(f'  Scan rate: {scan_rate*1000:.0f} mV/s')
    print(f'  Geometric area: {geo_area:.2f} cm²')
    print(f'  H_UPD window: {v_hupd_low:.3f}–{v_hupd_high:.3f} V')
    print(f'  DL window:    {v_dl_low:.3f}–{v_dl_high:.3f} V\n')

    results = []
    for i, (fp, lbl) in enumerate(zip(filepaths, labels)):
        print(f'  [{i+1}/{len(filepaths)}] {lbl}')
        V, j, label = load_cleaning_data(fp, geo_area=geo_area)
        if V is None or len(V) < 100:
            print(f'    Skipped: no valid data')
            continue

        cycles = extract_cv_cycles(V, j)
        if len(cycles) < 2:
            print(f'    Skipped: only {len(cycles)} cycle(s)')
            continue

        metrics = []
        for V_c, j_c in cycles:
            m = compute_cycle_metrics(V_c, j_c, scan_rate,
                                       v_hupd_low, v_hupd_high,
                                       v_dl_low, v_dl_high)
            metrics.append(m if m is not None else
                          {'Q_anodic_mC_cm2': 0.0, 'Q_cathodic_mC_cm2': 0.0,
                           'Q_ratio': 0.0, 'DL_anodic_mA_cm2': 0.0,
                           'DL_cathodic_mA_cm2': 0.0, 'DL_mean_mA_cm2': 0.0,
                           'DL_asymmetry_mA_cm2': 0.0})

        conv_idx = find_convergence_cycle(metrics)
        conv_str = f'cycle {conv_idx + 1}' if conv_idx is not None else 'NOT REACHED'
        n_avg = min(5, len(metrics))
        final_Qc = float(np.mean([m['Q_cathodic_mC_cm2'] for m in metrics[-n_avg:]]))
        init_dl = metrics[0]['DL_mean_mA_cm2']
        final_dl = float(np.mean([m['DL_mean_mA_cm2'] for m in metrics[-n_avg:]]))
        print(f'    {len(cycles)} cycles, convergence: {conv_str}')
        print(f'    Final Q_cathodic: {final_Qc:.3f} mC/cm²')
        print(f'    DL: {init_dl:.3f} → {final_dl:.3f} mA/cm²')

        results.append({
            'label': lbl, 'cycles': cycles, 'metrics': metrics,
            'conv_idx': conv_idx,
        })

        if save_dir:
            safe = lbl.replace(' ', '_').replace('/', '-').replace('\\', '-')
            if len(safe) > 60:
                safe = safe[:60]
            try:
                plot_cv_overlays(cycles, conv_idx, metrics, lbl, scan_rate,
                                  save_path=os.path.join(save_dir,
                                                          f'cleaning_cycles_{safe}.{image_ext}'))
            except Exception as e:
                print(f'    Plot failed: {e}')
            plt.close('all')
            try:
                plot_diagnostics(cycles, metrics, conv_idx, lbl, scan_rate,
                                  save_path=os.path.join(save_dir,
                                                          f'cleaning_diagnostics_{safe}.{image_ext}'))
            except Exception as e:
                print(f'    Diagnostics failed: {e}')
            plt.close('all')

    if not results:
        print('\n  No valid files processed.')
        return []

    if save_dir:
        try:
            _save_cleaning_excel(results, scan_rate, geo_area,
                                  os.path.join(save_dir, 'cleaning_data.xlsx'))
        except Exception as e:
            print(f'  Excel export failed: {e}')

    return results


# ═══════════════════════════════════════════════════════════════════════
#  Portal Entry Point
# ═══════════════════════════════════════════════════════════════════════

def run(input_dir: str, output_dir: str, params: dict = None) -> dict:
    """Portal entry point: scan input_dir for electrode-cleaning files."""
    from pathlib import Path
    p = params or {}

    from scripts.helpers.conditions import img_ext_from_params
    image_ext = img_ext_from_params(p)

    matched = classify_cleaning_files(input_dir)
    if not matched:
        raise RuntimeError(
            f"No electrode-cleaning files found in {input_dir}. "
            f"Filenames must contain one of: "
            f"{', '.join(CLEANING_KEYWORDS)} (case-insensitive)."
        )

    filepaths = [fp for fp, _ in matched]
    labels = [lbl for _, lbl in matched]

    results = run_batch(
        filepaths, labels,
        save_dir=str(output_dir),
        scan_rate=float(p.get('scan_rate', 0.5)),
        geo_area=float(p.get('geo_area', 5.0)),
        v_hupd_low=float(p.get('v_hupd_low', 0.05)),
        v_hupd_high=float(p.get('v_hupd_high', 0.40)),
        v_dl_low=float(p.get('v_dl_low', 0.40)),
        v_dl_high=float(p.get('v_dl_high', 0.50)),
        image_ext=image_ext,
    )

    output_files = [str(f.relative_to(Path(output_dir)))
                    for f in Path(output_dir).rglob('*') if f.is_file()]

    if not output_files:
        raise RuntimeError(
            f"Electrode cleaning analysis produced no output. "
            f"Found {len(matched)} candidate file(s) but all were skipped."
        )

    return {
        'status': 'success',
        'files_processed': len(results),
        'files_produced': output_files,
    }
