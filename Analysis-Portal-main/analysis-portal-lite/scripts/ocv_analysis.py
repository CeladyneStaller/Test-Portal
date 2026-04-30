#!/usr/bin/env python3
"""
OCV vs Time Analysis
====================

Plots open-circuit voltage vs time for fuel cell data.
Supports Scribner (.fcd) and FCTS (.csv) file formats.

Auto-classifies files by keyword (OCV, Purge) and excludes
files with fewer than 10 data points.

Usage:
  python ocv_analysis.py                          # interactive
  python ocv_analysis.py --folder /path/to/data   # batch from CLI
"""

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scripts.helpers.plot_compare import save_with_sidecar
import csv
import os
import glob
import argparse


def run(input_dir: str, output_dir: str, params: dict = None) -> dict:
    """Batch OCV vs time analysis from fuel cell test data."""
    from pathlib import Path
    p = params or {}

    inp = Path(input_dir)
    all_files = sorted([f for f in inp.rglob('*')
                        if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv', '.fcd')])

    if not all_files:
        return {"status": "error", "message": "No data files found"}

    # Filter by OCV keywords; fall back to all files if none match
    filtered = [f for f in all_files
                if any(kw in f.name.upper() for kw in OCV_KEYWORDS)
                and not any(ex in f.name.upper() for ex in EXCLUDE_KEYWORDS)
                and not any(pk in f.name.upper() for pk in POLCURVE_KEYWORDS)]
    files_to_process = filtered if filtered else all_files

    filepaths = [str(f) for f in files_to_process]
    labels = [f.stem for f in files_to_process]
    interval_s = float(p.get('interval_s', 60.0))

    from scripts.helpers.conditions import img_ext_from_params
    image_ext = img_ext_from_params(p)

    datasets = run_batch(filepaths, labels, save_dir=str(output_dir),
                         interval_s=interval_s, image_ext=image_ext)

    output_files = [str(f.relative_to(Path(output_dir)))
                    for f in Path(output_dir).rglob('*') if f.is_file()]
    if not output_files:
        raise RuntimeError(
            f"Analysis produced no output. {len(files_to_process)} file(s) were found "
            f"but none could be processed. Check file format and parameters."
        )
    return {"status": "success", "files_processed": len(files_to_process),
            "files_produced": output_files}


# ═══════════════════════════════════════════════════════════════════════
#  Header Parsing
# ═══════════════════════════════════════════════════════════════════════

def parse_fcd_header(filepath):
    """
    Parse Scribner .fcd file header to extract skip count and column indices.
    Returns dict with 'skip' and column indices, or None for non-.fcd files.
    """
    if not filepath.lower().endswith('.fcd'):
        return None
    try:
        with open(filepath, 'r', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                lines.append(line)
                if 'End Comments' in line:
                    skip = i + 1
                    break
            else:
                return None
    except Exception:
        return None
    if len(lines) < 2:
        return None
    cols = lines[-2].strip().split('\t')
    result = {'skip': skip, 'delimiter': '\t'}
    for ci, name in enumerate(cols):
        n = name.strip()
        if n == 'Time (Sec)':
            result['time_col'] = ci
        elif n == 'I (A)':
            result['j_col'] = ci
        elif n == 'E_Stack (V)':
            result['v_col'] = ci
        elif n == 'HFR (mOhm)':
            result['hfr_col'] = ci
        elif n.startswith('Z_Freq'):
            result['freq_col'] = ci
        elif n.startswith('Z_Real'):
            result['zreal_col'] = ci
        elif n.startswith('Z_Imag'):
            result['zimag_col'] = ci
        elif n == 'Ctrl_Mode':
            result['mode_col'] = ci
        elif n == 'Cell (C)':
            result.setdefault('condition_cols', {})['T_cell (C)'] = ci
        elif n.startswith('Temp_Anode'):
            result.setdefault('condition_cols', {})['T_anode_dp (C)'] = ci
        elif n.startswith('Flow_Anode'):
            result.setdefault('condition_cols', {})['H2_flow (slpm)'] = ci
        elif n.startswith('Temp_Cathode'):
            result.setdefault('condition_cols', {})['T_cathode_dp (C)'] = ci
        elif n.startswith('Flow_Cathode'):
            result.setdefault('condition_cols', {})['Air_flow (slpm)'] = ci
    return result


def parse_csv_header(filepath):
    """
    Parse CSV header to find time and voltage columns by name.
    Returns dict with column indices, or None if parsing fails.
    """
    try:
        with open(filepath, 'r', errors='replace') as f:
            first_line = f.readline().strip()
    except Exception:
        return None

    # Detect delimiter
    if '\t' in first_line:
        delimiter = '\t'
    elif ',' in first_line:
        delimiter = ','
    else:
        return None

    cols = [c.strip() for c in first_line.split(delimiter)]
    result = {'skip': 1, 'delimiter': delimiter}

    for ci, name in enumerate(cols):
        n = name.lower()
        # Time column
        if n in ('time stamp', 'time (sec)', 'time (s)', 'time', 'elapsed time',
                 'elapsed_time', 'timestamp', 'time_s', 'time_sec'):
            result['time_col'] = ci
        # Voltage column — prefer stack/working electrode voltage
        if n in ('e_stack (v)', 'voltage (v)', 'working electrode voltage',
                 'voltage', 'cell voltage', 'cell_voltage', 'v',
                 'e_stack', 'vcell', 'v_cell'):
            result['v_col'] = ci
        # Counter electrode (for FCTS — not used for OCV but detected)
        if n in ('counter electrode voltage',):
            result['counter_v_col'] = ci
        # Current
        if n in ('i (a)', 'current', 'current (a)', 'i'):
            result['j_col'] = ci

    return result if 'time_col' in result and 'v_col' in result else None


# ═══════════════════════════════════════════════════════════════════════
#  Data Loading
# ═══════════════════════════════════════════════════════════════════════

def load_ocv_data(filepath):
    """
    Load time and voltage from a data file.
    Auto-detects format (Scribner .fcd or CSV/FCTS).

    Returns
    -------
    time : ndarray — time in seconds (zeroed to start)
    voltage : ndarray — voltage in V
    label : str — filename stem for plot labels
    """
    label = os.path.splitext(os.path.basename(filepath))[0]

    # Try Scribner .fcd first
    fcd = parse_fcd_header(filepath)
    if fcd is not None:
        time_col = fcd.get('time_col', 0)
        v_col = fcd.get('v_col', 5)
        data = np.genfromtxt(filepath, delimiter=fcd['delimiter'],
                             skip_header=fcd['skip'],
                             usecols=[time_col, v_col])
        if data.ndim == 1:
            data = data.reshape(1, -1)
        # Remove NaN rows
        valid = ~np.isnan(data).any(axis=1)
        data = data[valid]
        if len(data) == 0:
            return None, None, label
        time = data[:, 0]
        voltage = data[:, 1]
        # Zero-reference time
        time = time - time[0]
        return time, voltage, label

    # Try CSV/FCTS
    hdr = parse_csv_header(filepath)
    if hdr is not None:
        time_col = hdr['time_col']
        v_col = hdr['v_col']
        data = np.genfromtxt(filepath, delimiter=hdr['delimiter'],
                             skip_header=hdr['skip'],
                             usecols=[time_col, v_col])
        if data.ndim == 1:
            data = data.reshape(1, -1)
        valid = ~np.isnan(data).any(axis=1)
        data = data[valid]
        if len(data) == 0:
            return None, None, label
        time = data[:, 0]
        voltage = data[:, 1]
        time = time - time[0]
        return time, voltage, label

    # Fallback: assume first two numeric columns are time, voltage
    try:
        for delim in ['\t', ',']:
            try:
                data = np.genfromtxt(filepath, delimiter=delim, skip_header=1,
                                     usecols=[0, 1])
                if data.ndim == 2 and data.shape[1] >= 2:
                    valid = ~np.isnan(data).any(axis=1)
                    data = data[valid]
                    if len(data) > 0:
                        time = data[:, 0] - data[0, 0]
                        voltage = data[:, 1]
                        return time, voltage, label
            except Exception:
                continue
    except Exception:
        pass

    return None, None, label


def resample_ocv(time, voltage, interval_s=60.0):
    """
    Resample OCV data to a uniform time grid using windowed averaging.

    For each target time point on the output grid, all raw data within
    ±interval/2 are averaged.  This acts as a box-car low-pass filter:
    it uses every raw measurement (no information discarded) and
    suppresses sub-interval noise without introducing phase lag.

    Parameters
    ----------
    time : ndarray — raw time in seconds (zeroed)
    voltage : ndarray — raw voltage in V
    interval_s : float — desired interval between output points (seconds)

    Returns
    -------
    time_out, voltage_out : ndarrays — uniformly spaced, smoothed data
    """
    if interval_s <= 0 or len(time) < 2:
        return time, voltage

    raw_interval = np.median(np.diff(time))
    # If raw data is already sparser than requested, return as-is
    if raw_interval >= interval_s * 0.9:
        return time, voltage

    # Build uniform output grid
    t_start = time[0]
    t_end = time[-1]
    t_grid = np.arange(t_start, t_end + interval_s * 0.01, interval_s)

    half_win = interval_s / 2.0
    v_out = np.empty(len(t_grid))

    for i, t in enumerate(t_grid):
        mask = (time >= t - half_win) & (time < t + half_win)
        if mask.sum() > 0:
            v_out[i] = np.mean(voltage[mask])
        else:
            # No points in window — interpolate from nearest neighbors
            v_out[i] = np.interp(t, time, voltage)

    return t_grid, v_out


# ═══════════════════════════════════════════════════════════════════════
#  File Classification
# ═══════════════════════════════════════════════════════════════════════

OCV_KEYWORDS = ['OCV', 'PURGE']
EXCLUDE_KEYWORDS = ['FILTERDATA']
POLCURVE_KEYWORDS = ['IV_', 'IV-', 'POLCURVE', 'POL_', 'POL-', 'POLARIZATION', 'POLDATA']
MIN_DURATION_MIN = 50  # minimum duration in minutes to include a file


def classify_ocv_files(folder):
    """
    Find files containing OCV/Purge keywords in a folder (recursively).
    Excludes files with FILTERDATA or pol curve keywords in the name.

    Returns list of (filepath, label) tuples.
    """
    from pathlib import Path
    p = Path(folder)
    all_files = sorted(set(
        str(f) for f in p.rglob('*')
        if f.is_file() and f.suffix.lower() in ('.csv', '.txt', '.tsv', '.fcd')
    ))

    # Exclude FILTERDATA
    all_files = [fp for fp in all_files
                 if not any(ex in os.path.basename(fp).upper() for ex in EXCLUDE_KEYWORDS)]

    matched = []
    for fp in all_files:
        name_upper = os.path.basename(fp).upper()
        # Skip pol curve files (contain OCV in voltage range like "OCV-0o3V")
        if any(kw in name_upper for kw in POLCURVE_KEYWORDS):
            continue
        if any(kw in name_upper for kw in OCV_KEYWORDS):
            label = os.path.splitext(os.path.basename(fp))[0]
            matched.append((fp, label))

    return matched


# ═══════════════════════════════════════════════════════════════════════
#  Plotting
# ═══════════════════════════════════════════════════════════════════════

def plot_ocv(time, voltage, label, save_path=None):
    """Plot a single OCV vs time curve."""
    fig, ax = plt.subplots(figsize=(8, 5))

    # Choose time unit
    duration = time[-1] - time[0]
    if duration > 7200:
        t_plot = time / 3600
        t_label = 'Time (hr)'
    elif duration > 120:
        t_plot = time / 60
        t_label = 'Time (min)'
    else:
        t_plot = time
        t_label = 'Time (s)'

    ax.plot(t_plot, voltage, '-', color='steelblue', lw=1.5)
    ax.set_xlabel(t_label)
    ax.set_ylabel('Voltage (V)')
    ax.set_title(f'OCV vs Time — {label}')

    from scripts.helpers.conditions import get_condition_label
    cond_label = get_condition_label(label=label)
    if cond_label:
        ax.set_title(f'OCV vs Time — {label}\n{cond_label}', fontsize=11)

    ax.grid(True, alpha=0.3)

    # Annotate start and end voltage
    ax.annotate(f'{voltage[0]:.4f} V', xy=(t_plot[0], voltage[0]),
                fontsize=9, color='green', ha='left', va='bottom')
    ax.annotate(f'{voltage[-1]:.4f} V', xy=(t_plot[-1], voltage[-1]),
                fontsize=9, color='firebrick', ha='right', va='top')

    fig.tight_layout()
    if save_path:
        save_with_sidecar(fig, save_path, dpi=150, bbox_inches='tight')
        print(f'  Saved: {save_path}')
    return fig


def plot_ocv_overlay(datasets, save_path=None):
    """Plot multiple OCV curves overlaid on one figure."""
    from scripts.helpers.conditions import get_condition_label

    fig, ax = plt.subplots(figsize=(10, 6))
    colors = plt.cm.tab10(np.linspace(0, 1, max(len(datasets), 1)))

    for i, (time, voltage, label) in enumerate(datasets):
        duration = time[-1] - time[0]
        if duration > 7200:
            t_plot = time / 3600
        elif duration > 120:
            t_plot = time / 60
        else:
            t_plot = time
        cond = get_condition_label(label=label, compact=True)
        legend_lbl = f'{label} ({voltage[-1]:.4f} V)\n  {cond}' if cond else f'{label} ({voltage[-1]:.4f} V)'
        ax.plot(t_plot, voltage, '-', color=colors[i], lw=1.5, alpha=0.8,
                label=legend_lbl)

    # Use time unit from longest dataset
    max_dur = max(d[0][-1] for d in datasets)
    if max_dur > 7200:
        ax.set_xlabel('Time (hr)')
    elif max_dur > 120:
        ax.set_xlabel('Time (min)')
    else:
        ax.set_xlabel('Time (s)')

    ax.set_ylabel('Voltage (V)')
    ax.set_title('OCV vs Time — Overlay')
    ax.legend(fontsize=8, loc='best')
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    if save_path:
        save_with_sidecar(fig, save_path, dpi=150, bbox_inches='tight')
        print(f'  Saved: {save_path}')
    return fig


# ═══════════════════════════════════════════════════════════════════════
#  Batch Processing
# ═══════════════════════════════════════════════════════════════════════

def run_batch(filepaths, labels, save_dir=None, interval_s=60.0, image_ext='png'):
    """
    Load, plot, and summarize multiple OCV files.
    Excludes files shorter than MIN_DURATION_MIN minutes.
    Resamples data to uniform interval_s spacing with windowed averaging.
    """
    if save_dir:
        os.makedirs(save_dir, exist_ok=True)

    datasets = []
    print(f'\n  Processing {len(filepaths)} files...')
    print(f'  Resampling interval: {interval_s:.0f} s\n')

    for i, (fp, lbl) in enumerate(zip(filepaths, labels)):
        print(f'  [{i+1}/{len(filepaths)}] {lbl}')
        time, voltage, label = load_ocv_data(fp)

        if time is None or len(time) < 2:
            print(f'         Skipped: no valid data')
            continue

        duration = time[-1] - time[0]
        duration_min = duration / 60.0
        if duration_min < MIN_DURATION_MIN:
            print(f'         Skipped: {duration_min:.1f} min (minimum {MIN_DURATION_MIN} min)')
            continue

        # Resample to uniform grid
        n_raw = len(time)
        time, voltage = resample_ocv(time, voltage, interval_s=interval_s)

        if duration > 3600:
            dur_str = f'{duration/3600:.1f} hr'
        elif duration > 60:
            dur_str = f'{duration/60:.1f} min'
        else:
            dur_str = f'{duration:.0f} s'

        print(f'         {n_raw} raw → {len(time)} pts ({interval_s:.0f}s), {dur_str}, '
              f'V: {voltage[0]:.4f} → {voltage[-1]:.4f} V '
              f'(ΔV={voltage[-1]-voltage[0]:+.4f})')

        datasets.append((time, voltage, lbl))

        # Individual plot
        if save_dir and image_ext:
            safe = lbl.replace(' ', '_').replace('/', '-').replace('\\', '-')
            plot_ocv(time, voltage, lbl,
                     save_path=os.path.join(save_dir, f'ocv_{safe}.{image_ext}'))
            plt.close()

    if not datasets:
        print('\n  No valid files processed.')
        return []

    # Overlay plot
    if len(datasets) > 1:
        if save_dir and image_ext:
            plot_ocv_overlay(datasets,
                             save_path=os.path.join(save_dir, f'ocv_overlay.{image_ext}'))
            plt.close()
        elif not save_dir:
            plot_ocv_overlay(datasets)
            plt.show()

    # Summary table
    print(f'\n  {"═" * 80}')
    print(f'  {"Label":40s} {"Duration":>10s} {"V_start":>10s} {"V_end":>10s} {"ΔV":>10s}')
    print(f'  {"─" * 80}')
    for time, voltage, lbl in datasets:
        duration = time[-1]
        if duration > 3600:
            dur_str = f'{duration/3600:.1f} hr'
        elif duration > 60:
            dur_str = f'{duration/60:.1f} min'
        else:
            dur_str = f'{duration:.0f} s'
        dv = voltage[-1] - voltage[0]
        print(f'  {lbl:40s} {dur_str:>10s} {voltage[0]:>10.4f} {voltage[-1]:>10.4f} {dv:>+10.4f}')
    print(f'  {"═" * 80}\n')

    # Excel output
    if save_dir:
        xlsx_path = os.path.join(save_dir, 'ocv_data.xlsx')
        _save_ocv_excel(datasets, interval_s, xlsx_path)

    return datasets


def _save_ocv_excel(datasets, interval_s, filepath):
    """Write OCV data to Excel with Summary and Data sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hf = Font(bold=True)
    hfill = PatternFill('solid', fgColor='D9E1F2')
    label_fill = PatternFill('solid', fgColor='E2EFDA')

    # ── Summary sheet ──
    ws = wb.active
    ws.title = 'Summary'
    headers = ['Label', 'Duration (min)', 'V_start (V)', 'V_end (V)', 'ΔV (V)',
               'Resampling (s)', 'Points']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(c)].width = max(len(h) + 4, 14)

    for i, (time, voltage, lbl) in enumerate(datasets, 2):
        ws.cell(row=i, column=1, value=lbl)
        ws.cell(row=i, column=2, value=round(time[-1] / 60.0, 1))
        ws.cell(row=i, column=3, value=round(float(voltage[0]), 6))
        ws.cell(row=i, column=4, value=round(float(voltage[-1]), 6))
        ws.cell(row=i, column=5, value=round(float(voltage[-1] - voltage[0]), 6))
        ws.cell(row=i, column=6, value=interval_s)
        ws.cell(row=i, column=7, value=len(time))
    ws.column_dimensions['A'].width = 40

    # ── Data sheet — all samples side by side ──
    ws2 = wb.create_sheet('Data')
    col = 1
    for time, voltage, lbl in datasets:
        # Label row (merged)
        col_end = col + 1
        for c in range(col, col_end + 1):
            cell = ws2.cell(row=1, column=c, value=lbl if c == col else '')
            cell.font = hf
            cell.fill = label_fill
        ws2.merge_cells(start_row=1, start_column=col,
                        end_row=1, end_column=col_end)

        # Sub-headers
        h1 = ws2.cell(row=2, column=col, value='Time (min)')
        h2 = ws2.cell(row=2, column=col + 1, value='Voltage (V)')
        for h in [h1, h2]:
            h.font = hf
            h.fill = hfill
            h.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[get_column_letter(col)].width = 14
        ws2.column_dimensions[get_column_letter(col + 1)].width = 14

        # Data rows
        for ri in range(len(time)):
            ws2.cell(row=ri + 3, column=col,
                     value=round(float(time[ri] / 60.0), 4))
            ws2.cell(row=ri + 3, column=col + 1,
                     value=round(float(voltage[ri]), 6))

        # Gap column
        col = col_end + 2

    wb.save(filepath)
    print(f'  Excel: {filepath}')


# ═══════════════════════════════════════════════════════════════════════
#  Interactive Mode
# ═══════════════════════════════════════════════════════════════════════

def _prompt(label, default=None, cast=float):
    suffix = f' [{default}]' if default is not None else ''
    raw = input(f'  {label}{suffix}: ').strip()
    if raw == '':
        return default
    if cast is None:
        return raw
    return cast(raw)


def _clean_path(p):
    p = p.strip()
    # Strip PowerShell call operator prefix: & 'path' or & "path"
    if p.startswith('& '):
        p = p[2:].strip()
    return p.strip('"').strip("'").strip()


def run_interactive():
    print('\n' + '▓' * 60)
    print('  OCV vs TIME ANALYSIS')
    print('▓' * 60)

    print('\n  Analysis modes:')
    print('    1 = Single file')
    print('    2 = Batch — folder (all OCV/Purge files)')
    print('    3 = Batch — file list')
    mode = int(_prompt('Select mode', default=1, cast=int))

    # Data resampling interval
    interval_s = _prompt('\n  Resampling interval (seconds)', default=60.0)

    if mode == 1:
        filepath = _clean_path(_prompt('\n  File path', cast=None))
        time, voltage, label = load_ocv_data(filepath)
        if time is None or len(time) < 2:
            print(f'  Error: no valid data')
            return

        duration = time[-1]
        duration_min = duration / 60.0
        if duration_min < MIN_DURATION_MIN:
            print(f'  Skipped: {duration_min:.1f} min (minimum {MIN_DURATION_MIN} min)')
            return

        # Resample
        n_raw = len(time)
        time, voltage = resample_ocv(time, voltage, interval_s=interval_s)

        if duration > 3600:
            dur_str = f'{duration/3600:.1f} hr'
        elif duration > 60:
            dur_str = f'{duration/60:.1f} min'
        else:
            dur_str = f'{duration:.0f} s'

        print(f'\n  Loaded: {n_raw} raw → {len(time)} pts ({interval_s:.0f}s), {dur_str}')
        print(f'  V: {voltage[0]:.4f} → {voltage[-1]:.4f} V (ΔV={voltage[-1]-voltage[0]:+.4f})')

        save = _prompt('\n  Save plot to directory? (path or Enter to show)',
                       default=None, cast=None)
        if save:
            save = _clean_path(save)
            os.makedirs(save, exist_ok=True)
            safe = label.replace(' ', '_').replace('/', '-')
            plot_ocv(time, voltage, label,
                     save_path=os.path.join(save, f'ocv_{safe}.png'))
            _save_ocv_excel([(time, voltage, label)], interval_s,
                            os.path.join(save, 'ocv_data.xlsx'))
        else:
            plot_ocv(time, voltage, label)
            plt.show()

    elif mode == 2:
        folder = _clean_path(_prompt('\n  Folder path', cast=None))
        save = _prompt('  Save results to directory? (path or Enter to show)',
                       default=None, cast=None)
        if save:
            save = _clean_path(save)

        files = classify_ocv_files(folder)
        if not files:
            print(f'  No OCV/Purge files found in {folder}')
            return

        print(f'\n  Found {len(files)} OCV/Purge file(s):')
        for fp, lbl in files:
            print(f'    {lbl}')

        filepaths = [f[0] for f in files]
        labels = [f[1] for f in files]
        run_batch(filepaths, labels, save_dir=save if save else None,
                  interval_s=interval_s)

    elif mode == 3:
        n = int(_prompt('\n  Number of files', default=1, cast=int))
        filepaths = []
        labels = []
        for i in range(n):
            fp = _clean_path(_prompt(f'  File {i+1} path', cast=None))
            lbl = _prompt(f'  File {i+1} label',
                          default=os.path.splitext(os.path.basename(fp))[0], cast=None)
            filepaths.append(fp)
            labels.append(lbl)

        save = _prompt('\n  Save results to directory? (path or Enter to show)',
                       default=None, cast=None)
        if save:
            save = _clean_path(save)
        run_batch(filepaths, labels, save_dir=save if save else None,
                  interval_s=interval_s)


# ═══════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='OCV vs Time Analysis')
    parser.add_argument('--file', type=str, help='Single file to analyze')
    parser.add_argument('--folder', type=str, help='Folder to batch-process')
    parser.add_argument('--save-dir', type=str, default=None,
                        help='Directory to save plots')
    parser.add_argument('--interval', type=float, default=60.0,
                        help='Resampling interval in seconds (default: 60)')
    args = parser.parse_args()

    if args.file:
        time, voltage, label = load_ocv_data(args.file)
        if time is not None and len(time) >= 2:
            duration_min = (time[-1] - time[0]) / 60.0
            if duration_min < MIN_DURATION_MIN:
                print(f'Skipped: {duration_min:.1f} min (minimum {MIN_DURATION_MIN} min)')
            else:
                time, voltage = resample_ocv(time, voltage, interval_s=args.interval)
                if args.save_dir:
                    os.makedirs(args.save_dir, exist_ok=True)
                    safe = label.replace(' ', '_').replace('/', '-')
                    plot_ocv(time, voltage, label,
                             save_path=os.path.join(args.save_dir, f'ocv_{safe}.png'))
                    _save_ocv_excel([(time, voltage, label)], args.interval,
                                    os.path.join(args.save_dir, 'ocv_data.xlsx'))
                else:
                    plot_ocv(time, voltage, label)
                    plt.show()
        else:
            print(f'Error: no valid data')

    elif args.folder:
        files = classify_ocv_files(args.folder)
        if files:
            fps = [f[0] for f in files]
            lbls = [f[1] for f in files]
            run_batch(fps, lbls, save_dir=args.save_dir, interval_s=args.interval)
        else:
            print(f'No OCV/Purge files found in {args.folder}')

    else:
        run_interactive()
